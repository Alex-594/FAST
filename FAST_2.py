
import os
import sys
import json
import time
import re
import math
import openpyxl
import shutil
import traceback
from PyQt6.QtCore import (Qt, QDate, QDateTime, QTime, QTimer, QEvent, 
                          QSettings, pyqtSignal, QObject, QUrl, QThread)
from PyQt6.QtGui import (QIntValidator, QColor)
from PyQt6.QtWidgets import (QApplication, QMainWindow, QTabWidget, 
                            QWidget, QVBoxLayout, QLabel, QPushButton,
                            QTableWidget, QTableWidgetItem, QHeaderView, 
                            QMessageBox, QLineEdit, QScrollArea, QHBoxLayout, 
                            QCheckBox, QDialog, QDialogButtonBox, QRadioButton, QButtonGroup,
                            QComboBox, QDateEdit, QListWidget, QListWidgetItem, QFileDialog,
                            QGroupBox, QSplitter, QTimeEdit, QFormLayout, QSizePolicy, QProgressBar,
                            QInputDialog, QTextEdit, QFrame, QAbstractItemView, QTextBrowser)
from PyQt6.QtMultimedia import QSoundEffect
from PyQt6.QtGui import QFont

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import COLOR_INDEX

VERSION = "v2.0.3.1"
CHANGELOG_TEXT = f"""
## История изменений FAST {VERSION}
**v2.0.3.1 (2025-04-24)**
* Исправлено несоотвествие заголовков содержимому вкладка проверка, полный вид
* Таймер в 2 строки при общем старте

**v2.0.3 (2025-04-21)**
* Добавлена прокрутка этапов в настройках логики
* Выключена бесполезная возможность редактрования в таблицах
* В окно удаления экипажа добавлен штурман

**v2.0.2.1 (2025-04-11)**
* Добавлен звук на таймере после сборки

**v2.0.2 (2025-04-11)**
* Добавлено диалоговое окно "История изменений" (эта функция!).
* Разрешено редактирование времени СКП для финишировавших экипажей (с предупреждением в диалоге).
* Исправлена сортировка строк КП во вкладке "КП" -> "Этапы" (теперь 1, 2, ..., 10, 11).
* Скорректирована ширина столбцов этапов/КП во вкладке "КП" -> "Этапы".
* Добавлена детализация по времени СКП во вкладку "Проверка КП" (Полный вид).
* Реализован выбор только одного этапа для КП с помощью RadioButton во вкладке "КП" -> "Этапы".
* Изменен шрифт в экспорте Excel на Times New Roman.
* Добавлена возможность сортировки таблиц Финиш/Не финиш по клику на заголовки.
* Изменена логика раскрашивания красным при DNF во вкладке финиш

**v2.0.1 (2025-04-02)**
* Запуск и тестирвоание
"""

#pyinstaller --onefile --windowed --icon=i.ico FAST_2.py
# сборка в один файл без консоли
#pyinstaller --onefile --icon=i.ico FAST_2.py


# По мере внесения осмысленных изменений, повторяйте цикл:
# Подготовьте изменения: git add . (или укажите конкретные файлы)
# Сделайте коммит: git commit -m "Описание ваших изменений"
# Отправьте новые коммиты на GitHub: git push origin main
############## ТАЙМЕР ОТСЧЕТА ##############

from PyQt6.QtWidgets import (QMainWindow, QLabel, QPushButton, QVBoxLayout, 
                            QWidget, QMessageBox, QHBoxLayout)
from PyQt6.QtCore import QTimer, QTime, QSettings, pyqtSignal, QObject, QUrl
from PyQt6.QtMultimedia import QSoundEffect
import os

# Класс SoundPlayer:

# Класс SoundPlayer (обновленный):
class SoundPlayer(QObject):
    sound_finished = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.effect = QSoundEffect()
        self.effect.setVolume(1.0)
        self.muted = False
        # --- Используем ОБА сигнала для лучшей диагностики ---
        self.effect.statusChanged.connect(self._handle_status_changed)
        self.effect.playingChanged.connect(self._handle_playing_changed) # <-- Добавили этот
        # --------------------------------------------------------
        self._last_played_valid = False # Флаг, что звук был успешно загружен и готов к игре/играл

    def _handle_status_changed(self):
         """Обрабатывает изменение статуса QSoundEffect."""
         current_status = self.effect.status()
         source_name = os.path.basename(self.effect.source().path()) if self.effect.source().isValid() else "N/A"
         # Логгируем ВСЕ изменения статуса
         print(f"[SoundPlayer Status] '{source_name}' Status: {current_status}")

         if current_status == QSoundEffect.Status.Ready:
             # Запоминаем, что звук был успешно загружен (или уже был готов)
             self._last_played_valid = True
             # Не отправляем finished здесь, т.к. Ready срабатывает и до начала игры
         elif current_status == QSoundEffect.Status.Error:
             print(f"!!! [SoundPlayer] Ошибка QSoundEffect: {source_name}")
             self._last_played_valid = False # Сбрасываем флаг при ошибке
             self.sound_finished.emit() # Сигналим о завершении при ошибке

    def _handle_playing_changed(self):
        """Обрабатывает начало и окончание воспроизведения."""
        source_name = os.path.basename(self.effect.source().path()) if self.effect.source().isValid() else "N/A"
        is_playing = self.effect.isPlaying()
        # Логгируем изменение состояния воспроизведения
        print(f"[SoundPlayer Playing] '{source_name}' isPlaying: {is_playing}")

        if not is_playing:
            # Звук перестал играть. Отправляем сигнал finished, ТОЛЬКО если
            # мы знаем, что он до этого был успешно загружен/воспроизводился.
            if self._last_played_valid:
                print(f"[SoundPlayer] Звук '{source_name}' ЗАВЕРШЕН (playingChanged -> False)")
                self.sound_finished.emit()
                self._last_played_valid = False # Сбрасываем флаг после завершения
        else:
            # Звук начал играть, значит он точно был 'Ready'
            self._last_played_valid = True

    def play(self, file_path):
        """Воспроизводит звуковой файл по указанному пути."""
        print(f"[SoundPlayer.play] Получен путь: {file_path}")
        try:
            if self.muted:
                print("[SoundPlayer] Звук отключен (muted).")
                return

            if not os.path.exists(file_path):
                print(f"!!! [SoundPlayer] Файл не найден (в методе play): {file_path}")
                return
            if not os.access(file_path, os.R_OK):
                print(f"!!! [SoundPlayer] Нет прав на чтение файла (в методе play): {file_path}")
                return

            url = QUrl.fromLocalFile(file_path)
            if not url.isValid():
                print(f"!!! [SoundPlayer] Неверный URL для файла: {file_path}")
                return

            # Сбрасываем флаг валидности перед загрузкой нового звука
            self._last_played_valid = False

            # Останавливаем предыдущий, если он играет
            if self.effect.isPlaying():
                print(f"[SoundPlayer] Остановка предыдущего звука для: {os.path.basename(file_path)}")
                self.effect.stop()

            print(f"[SoundPlayer] Установка источника: {url.fileName()}")
            self.effect.setSource(url)

            # Небольшая задержка перед вызовом play(), чтобы дать время на обработку setSource()
            # Используем новый вспомогательный метод для вызова play()
            QTimer.singleShot(20, self._initiate_play) # Задержка 20 мс

        except Exception as e:
            print(f"!!! [SoundPlayer] Исключение в методе play: {str(e)}")
            traceback.print_exc()

    def _initiate_play(self):
        """Вызывает self.effect.play() после небольшой задержки."""
        source_name = os.path.basename(self.effect.source().path()) if self.effect.source().isValid() else "N/A"
        current_status = self.effect.status()
        print(f"[SoundPlayer._initiate_play] Вызов play() для '{source_name}' (статус сейчас: {current_status})")

        if current_status == QSoundEffect.Status.Error:
            print(f"!!! [SoundPlayer] Ошибка ПЕРЕД вызовом play() для {source_name}")
            self._last_played_valid = False
            self.sound_finished.emit() # Сигналим, если ошибка произошла до play
        else:
            # В любом другом состоянии (Ready, Loading, Buffering) - вызываем play
            # QSoundEffect сам начнет играть, когда будет готов
            self.effect.play()

class CountdownWindow(QMainWindow):

    def __init__(self, parent=None, start_time="", common_start=True):
        super().__init__(parent)
        self.start_time = QTime.fromString(start_time, "HH:mm")
        if not self.start_time.isValid():
            self.start_time = QTime.currentTime().addSecs(60)
        self.common_start = common_start
        self.sound_player = SoundPlayer()
        self.played_sounds = set()
        self.current_cycle = 0  # Счетчик циклов
        self.setup_ui()
        self.setup_timers()
        self.play_test_sound()

    def play_test_sound(self):
        print("Тестирование звука...")  # Для отладки
        self.play_sound("проверка.wav")

    def play_start_sound(self):
        self.play_sound("старт.wav")

    def toggle_sound(self):
        self.sound_player.muted = self.mute_btn.isChecked()
        self.save_settings()

    def load_settings(self):
        settings = QSettings("RaceTimer", "SoundSettings")
        self.sound_player.muted = settings.value("muted", False, type=bool)
        self.mute_btn.setChecked(self.sound_player.muted)

    def save_settings(self):
        settings = QSettings("RaceTimer", "SoundSettings")
        settings.setValue("muted", self.sound_player.muted)

    def setup_ui(self):
        self.setWindowTitle("Таймер старта")
        self.setMinimumSize(800, 400)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        self.timer_label = QLabel("00:00:00")
        self.set_style("white")  # Инициализация стиля
        
        layout.addWidget(self.timer_label)
        
        btn_layout = QHBoxLayout()
        self.test_btn = QPushButton("Проверка звука")
        self.test_btn.clicked.connect(self.play_test_sound)
        btn_layout.addWidget(self.test_btn)
        
        self.mute_btn = QPushButton("Отключить звук")
        self.mute_btn.setCheckable(True)
        self.mute_btn.clicked.connect(self.toggle_sound)
        btn_layout.addWidget(self.mute_btn)
        
        layout.addLayout(btn_layout)

    def set_style(self, color):
        """Устанавливает стиль для таймера с указанным цветом текста"""
        self.timer_label.setStyleSheet(f"""
            font-size: 72pt;
            font-weight: bold;
            color: {color};
            background-color: #000000;
            padding: 50px;
            qproperty-alignment: AlignCenter;
        """)

    def clear_last_played(self, point):
        """Очищаем отметку о воспроизведении"""
        if point in self.last_played:
            del self.last_played[point]

    def update_display(self, seconds):
        """Обновляет отображение таймера"""
        abs_seconds = abs(seconds)
        sign = "-" if seconds < 0 else ""
        time_str = f"{sign}{abs_seconds//3600:02d}:{(abs_seconds%3600)//60:02d}:{abs_seconds%60:02d}"
        
        if seconds > 0:
            self.timer_label.setText(f"До старта: {time_str}")
            self.set_style("white")
        elif seconds == 0:
            self.timer_label.setText("СТАРТ!")
            self.set_style("red")
        else:
            self.timer_label.setText(f"После старта: {time_str}")
            self.set_style("blue")

    def reset_sound_flags(self):
        """Сброс всех флагов воспроизведения"""
        self.played_sounds.clear()

    def setup_timers(self):
        self.main_timer = QTimer()
        self.main_timer.timeout.connect(self.update_countdown)
        self.main_timer.start(100)  # 10 обновлений в секунду

        # Таймер для циклического режима
        self.cycle_timer = QTimer()
        self.cycle_timer.timeout.connect(self.start_new_cycle)
        
        if not self.common_start:
            # Запускаем первый цикл сразу
            QTimer.singleShot(0, self.start_new_cycle)

    def start_new_cycle(self):
        """Начинает новый минутный цикл"""
        if not self.common_start:
            self.current_cycle += 1
            print(f"Начало цикла #{self.current_cycle}")
            self.start_time = QTime.currentTime().addSecs(60)  # Следующая минута
            self.played_sounds.clear()
            
            # Перезапускаем таймер цикла
            self.cycle_timer.start(60000)  # Каждые 60 секунд

    def check_sounds(self, seconds):
        sound_triggers = {
            45: "45 сек.wav",
            30: "30 сек.wav", 
            15: "15 сек.wav",
            10: "10 сек.wav",
            5: "секунды.wav",
            4: "секунды.wav",
            3: "секунды.wav",
            2: "секунды.wav",
            1: "секунды.wav",
            0: "старт.wav"
        }
        
        if seconds in sound_triggers:
            sound_key = f"{self.current_cycle}_{seconds}_{sound_triggers[seconds]}"
            if sound_key not in self.played_sounds:
                self.played_sounds.add(sound_key)
                self.play_sound(sound_triggers[seconds])
                print(f"Цикл {self.current_cycle}: {seconds} сек - {sound_triggers[seconds]}")

    def closeEvent(self, event):
        self.main_timer.stop()
        self.cycle_timer.stop()
        super().closeEvent(event)

    def play_sound(self, filename):
        """Воспроизводит звук, ища папку 'sounds' рядом с .py или .exe"""
        try:
            # --- НАЧАЛО ИСПРАВЛЕНИЯ: Путь относительно .exe или .py ---
            if getattr(sys, 'frozen', False):
                # Запущено из собранного .exe
                # Папка 'sounds' должна лежать РЯДОМ с .exe файлом
                base_dir = os.path.dirname(sys.executable)
                runtime_env = f"PyInstaller (.exe): {base_dir}"
            elif __file__:
                # Обычный запуск .py скрипта
                # Папка 'sounds' должна лежать РЯДОМ со скриптом .py
                base_dir = os.path.dirname(os.path.abspath(__file__))
                runtime_env = f".py script: {base_dir}"
            else:
                # Резервный вариант (маловероятен для основного приложения)
                base_dir = os.getcwd()
                runtime_env = f"Fallback (CWD): {base_dir}"

            print(f"[CountdownWindow.play_sound] Runtime env: {runtime_env}")

            # Собираем путь к файлу звука (папка 'sounds' ожидается рядом)
            sound_path = os.path.join(base_dir, "sounds", filename)
            # --- КОНЕЦ ИСПРАВЛЕНИЯ ---

            print(f"[CountdownWindow] Попытка воспроизвести: {sound_path}")

            # --- Проверки файла ---
            if not os.path.exists(sound_path):
                print(f"!!! [CountdownWindow] Файл не найден: {sound_path}")
                try:
                    print(f"    Содержимое папки приложения ({base_dir}):")
                    # Выводим только первые N элементов, если папка большая
                    dir_list = os.listdir(base_dir)
                    for i, item in enumerate(dir_list):
                        print(f"      - {item}")
                        if i > 20: # Ограничим вывод
                             print("      ... (и другие)")
                             break
                except Exception as list_err:
                    print(f"    Не удалось прочитать содержимое {base_dir}: {list_err}")
                return

            if not os.access(sound_path, os.R_OK):
                print(f"!!! [CountdownWindow] Нет прав на чтение файла: {sound_path}")
                return

            # --- Воспроизведение ---
            if hasattr(self, 'sound_player') and isinstance(self.sound_player, SoundPlayer):
                 self.sound_player.play(sound_path)
            else:
                 print("!!! [CountdownWindow] Ошибка: Экземпляр SoundPlayer (self.sound_player) не найден.")

        except Exception as e:
            print(f"!!! [CountdownWindow] Исключение в play_sound для {filename}: {str(e)}")
            traceback.print_exc()

    def update_countdown(self):
        current_time = QTime.currentTime()
        # Время до старта в секундах (может быть отрицательным)
        seconds = current_time.secsTo(self.start_time)

        # Обновление отображения
        abs_seconds = abs(seconds)
        # Форматируем абсолютное время в ЧЧ:ММ:СС
        time_str = f"{abs_seconds//3600:02d}:{(abs_seconds%3600)//60:02d}:{abs_seconds%60:02d}"

        # --- НАЧАЛО ИЗМЕНЕНИЯ ---
        if self.common_start:
            # Режим ОБЩЕГО старта
            if seconds > 0:
                # Время до старта
                # Убираем "Цикл...", добавляем перенос строки перед временем
                self.timer_label.setText(f"До старта:\n{time_str}")
                self.set_style("white")
            elif seconds == 0:
                # Момент старта
                self.timer_label.setText("СТАРТ!")
                self.set_style("red") # Используем красный цвет для "СТАРТ!"
            else: # seconds < 0
                # Время после старта
                # Убираем "Цикл...", добавляем перенос строки и знак "-"
                self.timer_label.setText(f"Старт прошел:\n-{time_str}")
                self.set_style("blue")
        else:
            # Режим ЦИКЛИЧЕСКОГО старта (оставляем как было)
            if seconds >= 0:
                self.timer_label.setText(f"Цикл {self.current_cycle} | До старта: {time_str}")
                self.set_style("white")
            else: # seconds < 0
                # Добавляем знак "-" вручную, т.к. time_str использует abs_seconds
                self.timer_label.setText(f"Цикл {self.current_cycle} | Старт прошел: -{time_str}")
                self.set_style("blue")
        # --- КОНЕЦ ИЗМЕНЕНИЯ ---

        # Проверка и воспроизведение звуков (эта часть остается без изменений)
        self.check_sounds(seconds)



class NumericTableWidgetItem(QTableWidgetItem):
    """Класс элемента таблицы для корректной сортировки чисел."""
    def __lt__(self, other):
        # Попытка преобразовать текст в число (float для универсальности)
        try:
            self_num = float(self.text().replace(',', '.')) # Учитываем возможные запятые
        except ValueError:
            return True # Нечисловые значения считаем "меньше" числовых

        try:
            other_num = float(other.text().replace(',', '.'))
        except ValueError:
            return False # Числовые значения считаем "больше" нечисловых

        return self_num < other_num

class TimeTableWidgetItem(QTableWidgetItem):
    """Класс элемента таблицы для корректной сортировки времени ЧЧ:ММ:СС."""
    def __lt__(self, other):
        # Попытка преобразовать текст во время QTime
        self_time = QTime.fromString(self.text(), "HH:mm:ss")
        other_time = QTime.fromString(other.text(), "HH:mm:ss")

        # Обработка невалидных времен
        if not self_time.isValid():
            return True # Невалидное время "меньше" валидного
        if not other_time.isValid():
            return False # Валидное время "больше" невалидного

        return self_time < other_time


















class RaceApp(QMainWindow):



    def __init__(self):
        super().__init__()
        self._load_call_count = 0 # Счетчик вызовов load_latest_data
        print("--- Запуск __init__ ---")

        # --- Определение путей ---
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
            print(f"Определен режим .exe. Путь к exe: {sys.executable}")
        elif __file__:
            application_path = os.path.dirname(os.path.abspath(__file__))
            print(f"Определен режим .py. Путь к скрипту: {__file__}")
        else:
            application_path = os.getcwd()
            print(f"Не удалось определить путь. Используется CWD: {application_path}")

        self.app_dir = application_path
        self.data_dir = os.path.join(application_path, "data")
        self.sounds_dir = os.path.join(application_path, "sounds") # Пример

        print(f"Итоговый путь к папке данных (self.data_dir): {self.data_dir}")
        print(f"Существует ли папка data? {'Да' if os.path.isdir(self.data_dir) else 'НЕТ!'}")

        # --- Загрузка данных (ТОЛЬКО ОДИН РАЗ) ---
        self.current_file = None
        print("--> Вызов load_latest_data...")
        self.load_latest_data()
        print("<-- ВЕРНУЛИСЬ из load_latest_data.")

        # --- Проверка данных после загрузки ---
        if hasattr(self, 'data') and isinstance(self.data, dict):
            print(f"    Ключи верхнего уровня в self.data: {list(self.data.keys())}")
            print(f"    self.data['meta'] = {self.data.get('meta', 'ОТСУТСТВУЕТ')}")
            print(f"    self.data['params'] = {self.data.get('params', 'ОТСУТСТВУЕТ')}")
        else:
            print("!!! ОШИБКА: self.data не существует или не является словарем после load_latest_data !!!")

        # --- Настройка UI (ТОЛЬКО ОДИН РАЗ) ---
        print("--> Вызов setup_ui...")
        try:
            self.setup_ui() # Этот метод создаст self.tabs и т.д.
            print("<-- setup_ui УСПЕШНО завершен.")
        except Exception as ui_error:
            print(f"!!! КРИТИЧЕСКАЯ ОШИБКА ВО ВРЕМЯ setup_ui: {ui_error} !!!")
            traceback.print_exc()
            # ... (обработка ошибки UI) ...

        # --- Настройка ОКНА и СОБЫТИЙ (ПОСЛЕ setup_ui) ---
        self.resize(1000, 700)
        self.update_window_title()
        # Подключаем обработчик изменения вкладки (self.tabs уже создан в setup_ui)
        if hasattr(self, 'tabs'):
             self.tabs.currentChanged.connect(self.on_tab_changed)
        else:
             print("!!! ОШИБКА: self.tabs не был создан в setup_ui !!!")

        # --- Запуск ВСЕХ ТАЙМЕРОВ (с уникальными именами!) ---
        print("Запуск таймеров...")

        # Таймер для обновления вкладки Старт (пример)
        self.start_tab_update_timer = QTimer(self) # Уникальное имя
        self.start_tab_update_timer.timeout.connect(self.update_start_tab) # Подключаем к методу обновления UI Старта
        self.start_tab_update_timer.start(1000)

        # Таймер для обновления вкладки Финиш (пример)
        self.finish_tab_update_timer = QTimer(self) # Уникальное имя
        self.finish_tab_update_timer.timeout.connect(self.update_finish_tab) # Подключаем к методу обновления UI Финиша
        self.finish_tab_update_timer.start(1000)

        # Таймер для авто-финиша по закрытию трассы
        self.auto_finish_done = False # Этот флаг, возможно, больше не нужен, если логика в check_track_closing
        self.track_closing_timer = QTimer(self) # Уникальное имя
        self.track_closing_timer.timeout.connect(self.check_track_closing) # Проверка закрытия трассы
        self.track_closing_timer.start(1000)

        # Таймер для проверки DNF по превышению времени трассы (если используется)
        self.track_timeout_timer = QTimer(self) # Уникальное имя
        self.track_timeout_timer.timeout.connect(self.check_track_timeout) # Проверка превышения лимита
        self.track_timeout_timer.start(1000)

        # Таймер для обратного отсчета старта (если используется)
        self.start_countdown_display_timer = QTimer(self) # Уникальное имя
        self.start_countdown_display_timer.timeout.connect(self._tick_start_tab_updates) # Обновляет только UI таймера
        self.start_countdown_display_timer.start(1000)

        # Таймер текущего времени (если нужен глобально)
        self.current_time_display_timer = QTimer(self) # Уникальное имя
        self.current_time_display_timer.timeout.connect(self.update_current_time)
        self.current_time_display_timer.start(1000)

        # Таймер для проверки времени общего старта
        self.general_start_check_timer = QTimer(self) # Уникальное имя
        self.general_start_check_timer.timeout.connect(self.check_start_time) # Проверяет, не пора ли общему старту
        self.general_start_check_timer.start(1000)

        # Таймеры для вкладки Этапы (v2)
        self.start_stages_timers_v2() # Запускает свои внутренние таймеры

        # Инициализация других переменных
        self.cp_input = None
        print("--- Завершение __init__ ---")

    def _extract_cp_number(self, cp_name_obj):
        """Извлекает число из имени КП для сортировки."""
        cp_name = ""
        if isinstance(cp_name_obj, dict): # Если передали весь словарь КП
            cp_name = cp_name_obj.get('name', '')
        elif isinstance(cp_name_obj, str): # Если передали только имя
            cp_name = cp_name_obj

        if not cp_name:
            return float('inf') # Обработка пустого имени

        # Ищем первую последовательность цифр в строке
        match = re.search(r'\d+', cp_name)
        if match:
            try:
                return int(match.group(0))
            except ValueError:
                return float('inf') # Если не смогли преобразовать (маловероятно)
        return float('inf') # Если цифр нет, ставим в конец

    def update_window_title(self):
        """Обновляем заголовок окна"""
        if hasattr(self, 'data') and 'meta' in self.data:
            name = self.data['meta'].get('name', 'Новое соревнование')
            date = self.data['meta'].get('date', '')
            self.setWindowTitle(f"{name} - {date}")
        else:
            self.setWindowTitle("Автосоревнования")

    def setup_ui(self):
        print("  --- Начало setup_ui ---")
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout(central_widget)

        self.tabs = QTabWidget() # Создаем вкладки здесь
        main_layout.addWidget(self.tabs)

        print("    --> Вызов setup_params_tab...")
        try:
            self.setup_params_tab() # <--- ОСТАВЛЯЕМ ТОЛЬКО ОДИН ВЫЗОВ
            print("    <-- setup_params_tab завершен.")
        except Exception as e:
            print(f"    !!! ОШИБКА внутри setup_params_tab: {e} !!!")
            traceback.print_exc()

        # Добавляем все остальные вкладки ПОСЛЕ создания self.tabs
        print("    Вызов setup_logic_tab...")
        self.setup_logic_tab()
        print("    Вызов setup_checkpoints_tab...")
        self.setup_checkpoints_tab()
        print("    Вызов setup_members_tab...")
        self.setup_members_tab()
        print("    Вызов setup_registration_tab...")
        self.setup_registration_tab()
        print("    Вызов setup_start_tab...")
        self.setup_start_tab()
        print("    Вызов setup_stages_tab_v2...")
        self.setup_stages_tab_v2() # Убедитесь, что используете новую версию
        print("    Вызов setup_finish_tab...")
        self.setup_finish_tab()
        print("    Вызов setup_check_tab...")
        self.setup_check_tab()
        print("    Вызов setup_results_tab...")
        self.setup_results_tab()

        print("  --- Завершение setup_ui ---")

    def on_tab_changed(self, index):
        """Обработчик изменения вкладки"""
        # Скрываем таблицу КП при переключении вкладок
        if self.tabs.tabText(index) != "КП":
            self.checkpoints_table.setVisible(False)
            self.stages_table.setVisible(False)
            self.btn_show.setText("Показать КП")
        
        # Если переключились на вкладку "Параметры" - обновляем ее
        if self.tabs.tabText(index) == "Параметры":
            self.update_params_fields()

        if self.tabs.tabText(index) == "Проверка КП":
            self.update_check_tab()  # Полное обновление данных

        if self.tabs.tabText(index) == "Старт":
            print("Переключение на вкладку Старт, вызываем update_start_tab") # Отладка
            if hasattr(self, 'update_start_tab'): # Проверка
                 self.update_start_tab() # Вызываем ОБЫЧНОЕ обновление UI, БЕЗ перезагрузки данных
            else:
                 print("ПРЕДУПРЕЖДЕНИЕ: Метод update_start_tab не найден!")

    def load_latest_data(self):
        """Загружаем последнюю версию файла данных (с отладкой повторных вызовов)."""
        # --- Начало отладки повторного вызова ---
        self._load_call_count += 1
        print(f"\n\n===== Начало load_latest_data (ВЫЗОВ № {self._load_call_count}) - Время: {time.time()} =====")
        if self._load_call_count > 1:
            print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            print("!!! ПРЕДУПРЕЖДЕНИЕ: load_latest_data вызван ПОВТОРНО !!!")
            print("--- СТЕК ВЫЗОВОВ (кто вызвал): ---")
            traceback.print_stack() # <-- Печатаем стек, чтобы найти виновника
            print("----------------------------------")
            print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        # --- Конец отладки повторного вызова ---

        try:
            print(f"    Проверка/создание папки данных: {self.data_dir}")
            # Убедимся, что data_dir вообще определен
            if not hasattr(self, 'data_dir') or not self.data_dir:
                 print("!!! КРИТИЧЕСКАЯ ОШИБКА: self.data_dir не определен перед использованием!")
                 self._initialize_default_data("Путь к данным не определен")
                 print("<-- Завершен метод load_latest_data (ошибка пути)")
                 return

            os.makedirs(self.data_dir, exist_ok=True)

            print(f"    Поиск json файлов в: {self.data_dir}")
            try:
                all_files = os.listdir(self.data_dir)
                json_files = [f for f in all_files
                              if f.startswith('race_v') and f.endswith('.json')]
                print(f"    Найденные файлы (*.json): {json_files if json_files else 'НЕТ'}")
            except OSError as list_err:
                 print(f"!!! ОШИБКА OSError при чтении содержимого папки {self.data_dir}: {list_err}")
                 self._initialize_default_data("Ошибка чтения папки данных")
                 print("<-- Завершен метод load_latest_data (после ошибки listdir)")
                 return

            if json_files:
                try:
                    print(f"    Сортировка файлов: {json_files}")
                    def get_version_num(filename):
                        try: return int(filename.split('_v')[1].split('.')[0])
                        except: return -1
                    json_files.sort(key=get_version_num)
                    # Убираем файлы с некорректным именем из рассмотрения
                    json_files = [f for f in json_files if get_version_num(f) != -1]
                    print(f"    Отсортированные корректные файлы: {json_files}")
                except Exception as sort_err:
                    print(f"!!! ОШИБКА при сортировке файлов: {sort_err}")
                    self._initialize_default_data("Ошибка сортировки файлов")
                    print("<-- Завершен метод load_latest_data (после ошибки сортировки)")
                    return

                if not json_files:
                     print("!!! Не найдено корректных файлов версии после сортировки.")
                     self._initialize_default_data("Нет корректных файлов версии")
                     print("<-- Завершен метод load_latest_data (нет корректных файлов)")
                     return

                latest_file = json_files[-1]
                self.current_file = os.path.join(self.data_dir, latest_file)
                print(f"    Выбран последний файл: {self.current_file}")

                print(f"    Открываем файл: {self.current_file}")
                try:
                    with open(self.current_file, 'r', encoding='utf-8') as f:
                        print(f"      Файл открыт. Пытаемся загрузить JSON...")
                        self.data = json.load(f) # <--- Загрузка JSON
                        print(f"      JSON успешно загружен из {latest_file}") # <--- Сообщение, которое вы видели
                        # Дополнительная проверка ключей сразу после загрузки
                        if isinstance(self.data, dict):
                             print(f"      Проверка ключей после загрузки: {list(self.data.keys())}")
                        else:
                             print(f"      !!! ПРЕДУПРЕЖДЕНИЕ: Загруженные данные не являются словарем!")
                             raise TypeError("Загруженные данные не являются словарем")

                except FileNotFoundError:
                    print(f"!!! ОШИБКА FileNotFoundError при открытии '{self.current_file}'")
                    self._initialize_default_data(f"Файл {latest_file} не найден")
                except PermissionError:
                    print(f"!!! ОШИБКА PermissionError: Нет прав на чтение файла '{self.current_file}'")
                    self._initialize_default_data(f"Нет прав на чтение {latest_file}")
                except json.JSONDecodeError as jde:
                    print(f"!!! ОШИБКА JSONDecodeError в файле '{self.current_file}': {jde}")
                    print(f"!!! Файл '{latest_file}' поврежден!")
                    self._initialize_default_data(f"Ошибка чтения JSON в {latest_file}")
                except Exception as read_err:
                    print(f"!!! НЕПРЕДВИДЕННАЯ ОШИБКА при чтении/загрузке JSON из '{self.current_file}': {read_err}")
                    traceback.print_exc()
                    self._initialize_default_data(f"Ошибка при чтении {latest_file}")

            else:
                print("JSON файлы версии не найдены, создаем данные по умолчанию.")
                self._initialize_default_data("Файлы не найдены")
                print("  Пытаемся сохранить данные по умолчанию...")
                save_ok = self.save_data() # Используем вашу (упрощенную?) save_data
                print(f"  Результат сохранения данных по умолчанию: {'Успех' if save_ok else 'Неудача'}")

            print("--> Завершение блока try в load_latest_data")

        except OSError as os_err:
             print(f"!!! ОШИБКА OSError в load_latest_data (папка {self.data_dir}?): {os_err}")
             self._initialize_default_data(f"Ошибка доступа к папке {self.data_dir}")
        except Exception as e:
             print(f"!!! НЕПРЕДВИДЕННАЯ ОШИБКА ВНЕ try/except в load_latest_data: {e}")
             traceback.print_exc()
             self._initialize_default_data("Неизвестная ошибка при загрузке")

        print(f"<-- Завершен метод load_latest_data (ВЫЗОВ № {self._load_call_count})")

    def _initialize_default_data(self, reason="Причина не указана"):
        """Вспомогательный метод для установки данных по умолчанию при ошибке загрузки."""
        # ... (код метода без изменений) ...
        print(f"!!! Инициализация данных по умолчанию. Причина: {reason} !!!")
        self.data = {
            "meta": {"version": 0, "created_at": "", "name": "Ошибка Загрузки Данных", "date": ""},
            "params": {},
            "logic_params": {},
            "checkpoints": [],
            "members": [],
            "registration": [],
            "start": [],
            "finish": []
        }
        self.current_file = None
        print(f"    Установлены данные по умолчанию: {list(self.data.keys())}")

    def save_data(self):
        """Сохраняем данные в новый файл (Упрощенная версия с возвратом True/False)"""
        # Проверяем наличие базовой структуры данных
        if not hasattr(self, 'data') or not isinstance(self.data, dict):
            print("[ERROR] save_data: Атрибут self.data отсутствует или не является словарем!")
            return False # <-- Возвращаем False
        if 'meta' not in self.data or not isinstance(self.data.get('meta'), dict) or 'version' not in self.data['meta']:
            print("[ERROR] save_data: Структура self.data['meta'] некорректна или отсутствует ключ 'version'!")
            return False # <-- Возвращаем False

        current_version = 0 # Инициализируем на случай ошибки ниже
        new_filename = "" # Инициализируем на случай ошибки ниже
        try:
            current_version = int(self.data["meta"].get("version", 0))
            new_version = current_version + 1
            self.data["meta"]["version"] = new_version # Обновляем версию в данных
            new_filename = os.path.join(self.data_dir, f"race_v{new_version}.json")
            print(f"[save_data] Попытка сохранения в: {new_filename} (Версия {new_version})")

            # Убедимся, что директория 'data' существует
            try:
                os.makedirs(self.data_dir, exist_ok=True)
            except OSError as oe:
                 print(f"[ERROR] save_data: Не удалось создать директорию {self.data_dir}: {oe}")
                 self.data["meta"]["version"] = current_version # Откат версии
                 return False # <-- Возвращаем False

            # Пытаемся записать файл
            with open(new_filename, 'w', encoding='utf-8') as f:
                try:
                    # Добавляем default=str на случай несериализуемых типов
                    json.dump(self.data, f, indent=4, ensure_ascii=False, default=str)
                    print(f"[SUCCESS] save_data: Данные успешно записаны в {new_filename}")
                except TypeError as te:
                    print(f"[ERROR] save_data: Ошибка сериализации данных в JSON: {te}")
                    self.data["meta"]["version"] = current_version # Откат версии
                    # Попытка удалить некорректный файл
                    try:
                        f.close()
                        os.remove(new_filename)
                        print(f"[INFO] save_data: Поврежденный файл {new_filename} удален.")
                    except Exception: pass
                    return False # <--- Возвращаем False при ошибке сериализации

            # Если запись прошла успешно
            self.current_file = new_filename
            self.update_window_title()
            return True # <--- ВОЗВРАЩАЕМ TRUE ПРИ УСПЕХЕ

        except KeyError as ke:
            print(f"[ERROR] save_data: Отсутствует ключ в self.data['meta']: {ke}")
            if 'meta' in self.data and isinstance(self.data['meta'], dict): self.data["meta"]["version"] = current_version
            return False # <-- Возвращаем False
        except IOError as ioe:
            print(f"[ERROR] save_data: Ошибка ввода-вывода при записи файла {new_filename}: {ioe}")
            if 'meta' in self.data and isinstance(self.data['meta'], dict): self.data["meta"]["version"] = current_version
            return False # <-- Возвращаем False
        except Exception as e:
            print(f"[ERROR] save_data: Непредвиденная ошибка сохранения данных: {e}")
            traceback.print_exc()
            if 'meta' in self.data and isinstance(self.data['meta'], dict): self.data["meta"]["version"] = current_version
            return False # <-- Возвращаем False

    def closeEvent(self, event):
        """Обработчик события закрытия главного окна."""
        print("Получено событие закрытия окна...")

        # Останавливаем НОВЫЕ таймеры вкладки Этапы
        if hasattr(self, 'stop_stages_timers_v2'):
            try:
                print("Останавливаем таймеры вкладки Этапы (v2)...")
                self.stop_stages_timers_v2()
            except Exception as e:
                print(f"Ошибка при остановке таймеров этапов (v2): {e}")

        if hasattr(self, 'start_tab_timer'):
            try:
                self.start_tab_timer.stop()
                print("Таймер 'start_tab_timer' остановлен.")
            except Exception as e:
                print(f"Ошибка при остановке 'start_tab_timer': {e}")

        # !!! ВАЖНО: Останавливаем ВСЕ ОСТАЛЬНЫЕ таймеры, которые вы запускали !!!
        # Пройдитесь по вашему коду (особенно __init__) и найдите все self.имя_таймера.start()
        # Для каждого добавьте здесь self.имя_таймера.stop()

        # Останавливаем ВСЕ ОСТАЛЬНЫЕ таймеры по их УНИКАЛЬНЫМ именам
        print("Останавливаем остальные таймеры приложения...")
        timers_to_stop = [
            'start_tab_update_timer',   # Новое имя
            'finish_tab_update_timer',  # Новое имя
            'track_closing_timer',      # Новое имя
            'track_timeout_timer',      # Оставили старое или дайте новое
            'start_countdown_display_timer', # Новое имя
            'current_time_display_timer', # Новое имя
            'general_start_check_timer', # Новое имя
            # Удалите или переименуйте старые 'timer', 'timeout_timer' и т.д., если они больше не нужны
        ]
        for timer_name in timers_to_stop:
            if hasattr(self, timer_name):
                timer_instance = getattr(self, timer_name)
                if isinstance(timer_instance, QTimer): # Убедимся, что это таймер
                    try:
                        timer_instance.stop()
                        print(f"Таймер '{timer_name}' остановлен.")
                    except Exception as e:
                        print(f"Ошибка при остановке таймера '{timer_name}': {e}")
                # else: # На случай, если имя 'timer' переиспользовалось не для таймера
                #    print(f"Атрибут '{timer_name}' не является QTimer.")

        print("Завершение работы...")
        event.accept()










############## ПАРАМЕТРЫ ##############


    def setup_params_tab(self):
        """Вкладка параметров соревнования"""
        print("      --- Начало setup_params_tab ---")
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Словарь для хранения текстовых полей
        self.params_fields = {}
        
        # Параметры с подсказками и валидацией
        params_config = {
            "Время регистрации": {
                "hint": "Формат: ЧЧ:ММ (например: 09:30)",
                "validator": self.validate_time,
                "required": True
            },
            "Время брифинга": {
                "hint": "Формат: ЧЧ:ММ (например: 10:00)",
                "validator": self.validate_time,
                "required": True
            },
            "Время старта": {
                "hint": "Формат: ЧЧ:ММ (например: 11:00)",
                "validator": self.validate_time,
                "required": True
            },
            "Время закрытия трассы": {
                "hint": "Формат: ЧЧ:ММ (например: 17:00)",
                "validator": self.validate_time,
                "required": True
            },
            "Время награждения": {
                "hint": "Формат: ЧЧ:ММ (например: 18:00)",
                "validator": self.validate_time,
                "required": True
            },
            "Количество КП": {
                "hint": "Целое число (например: 10)",
                "validator": self.validate_int,
                "required": True
            },
            "Бальность КП": {
                "hint": "Число (например: 5)",
                "validator": self.validate_float,
                "required": True
            },
            "Зачет1": {
                "hint": "Название зачета (например: Спорт) СТРОГОЕ СООТВЕТСТВИЕ ЗАЯВКАМ",
                "validator": self.validate_classification,
                "required": True
            },
            "Зачет2": {
                "hint": "Название зачета (например: Дебют)",
                "validator": self.validate_classification,
                "required": True
            },
            "Зачет3": {
                "hint": "Название зачета (необязательное поле)",
                "validator": None,  # Убрана валидация
                "required": False  # Поле необязательное
            },
            "Ссылка на таблицу регистрации": {
                "hint": "URL или путь к файлу",
                "validator": None,
                "required": False
            }
        }
        
        # Создаем поля для каждого параметра
        for param, config in params_config.items():
            layout.addWidget(QLabel(param + ":" + ("" if not config["required"] else " *")))
            field = QLineEdit()
            field.setPlaceholderText(config["hint"])
            
            # Заполняем поле значением из данных, если оно есть
            if hasattr(self, 'data') and 'params' in self.data:
                key = param.lower().replace(' ', '_')
                field.setText(str(self.data['params'].get(key, "")))
            
            self.params_fields[param] = field
            layout.addWidget(field)

        meta_group = QGroupBox("Информация о соревновании и программе")
        meta_layout = QVBoxLayout()
        if hasattr(self, 'data') and 'meta' in self.data:
             meta_layout.addWidget(QLabel(f"Название: {self.data['meta'].get('name', '')}"))
             meta_layout.addWidget(QLabel(f"Дата: {self.data['meta'].get('date', '')}"))
             meta_layout.addWidget(QLabel(f"Версия данных: {self.data['meta'].get('version', 1)}"))
        meta_layout.addWidget(QLabel(f"Версия программы: FAST {VERSION}")) # Укажите актуальную версию

        # ---> НАЧАЛО ИЗМЕНЕНИЯ: Добавляем кнопку Истории <---
        btn_changelog = QPushButton("История изменений программы")
        btn_changelog.setToolTip("Показать список изменений в этой версии программы")
        btn_changelog.clicked.connect(self._show_changelog_dialog)
        meta_layout.addWidget(btn_changelog, alignment=Qt.AlignmentFlag.AlignLeft) # Добавляем кнопку
        # ---> КОНЕЦ ИЗМЕНЕНИЯ <---

        meta_group.setLayout(meta_layout)
        layout.addWidget(meta_group)


        # Кнопка сохранения параметров
        btn_save = QPushButton("Сохранить параметры")
        btn_save.clicked.connect(self.confirm_save_params)
        layout.addWidget(btn_save)

        layout.addStretch() # Добавляем растяжение внизу

        self.tabs.addTab(tab, "Параметры")
        self._params_tab_index = self.tabs.indexOf(tab)
        self.params_tab_dirty = False
        print("      <--- setup_params_tab завершен.")
    
    def validate_time(self, time_str):
        """Проверка формата времени ЧЧ:ММ"""
        try:
            hours, minutes = map(int, time_str.split(':'))
            return 0 <= hours < 24 and 0 <= minutes < 60
        except ValueError:
            return False
    
    def validate_int(self, value):
        """Проверка что значение целое число"""
        try:
            int(value)
            return True
        except ValueError:
            return False
    
    def validate_float(self, value):
        """Проверка что значение число"""
        try:
            float(value)
            return True
        except ValueError:
            return False
    
    def validate_classification(self, value):
        """Проверка что зачет начинается с большой буквы"""
        return value and value[0].isupper()
    
    def confirm_save_params(self):
        """Подтверждение сохранения параметров"""
        # Первое подтверждение
        reply = QMessageBox.question(
            self, 'Подтверждение',
            'Вы уверены, что хотите сохранить параметры?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Второе подтверждение
            reply = QMessageBox.question(
                self, 'Последнее подтверждение',
                'Все данные будут перезаписаны. Продолжить?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.save_params()

    def save_params(self):
        """Сохраняем параметры после проверки"""
        errors = []
        
        # Проверяем все обязательные поля
        for param, field in self.params_fields.items():
            validator = None
            required = False
            
            # Определяем валидатор и обязательность поля
            if param in ["Время регистрации", "Время брифинга", "Время старта", 
                        "Время закрытия трассы", "Время награждения"]:
                validator = self.validate_time
                required = True
            elif param == "Количество КП":
                validator = self.validate_int
                required = True
            elif param == "Бальность КП":
                validator = self.validate_float
                required = True
            elif param in ["Зачет1", "Зачет2"]:
                validator = self.validate_classification
                required = True
            elif param == "Зачет3":
                # Для Зачета3 валидация не требуется
                if field.text() and not self.validate_classification(field.text()):
                    errors.append(f"Некорректное значение для '{param}' (должно начинаться с большой буквы)")
                continue
            
            if required and not field.text():
                errors.append(f"Поле '{param}' обязательно для заполнения")
            elif validator and field.text() and not validator(field.text()):
                errors.append(f"Некорректное значение для '{param}'")
        
        if errors:
            QMessageBox.critical(self, "Ошибки", "\n".join(errors))
            return
        
        # Если ошибок нет, сохраняем
        try:
            params = {
                'время_регистрации': self.params_fields["Время регистрации"].text(),
                'время_брифинга': self.params_fields["Время брифинга"].text(),
                'время_старта': self.params_fields["Время старта"].text(),
                'время_закрытия_трассы': self.params_fields["Время закрытия трассы"].text(),
                'время_награждения': self.params_fields["Время награждения"].text(),
                'количество_кп': self.params_fields["Количество КП"].text(),
                'бальность_кп': self.params_fields["Бальность КП"].text(),
                'зачет1': self.params_fields["Зачет1"].text(),
                'зачет2': self.params_fields["Зачет2"].text()
            }
            
            # Добавляем необязательные поля только если они заполнены
            if self.params_fields["Зачет3"].text():
                params['зачет3'] = self.params_fields["Зачет3"].text()
            if self.params_fields["Ссылка на таблицу регистрации"].text():
                params['ссылка_на_таблицу_регистрации'] = self.params_fields["Ссылка на таблицу регистрации"].text()
            
            self.data['params'] = params
            self.save_data()
            QMessageBox.information(self, "Сохранено", "Параметры успешно сохранены!")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения: {str(e)}")
    
    def update_params_fields(self):
        """Обновляем поля ввода и метаданные на вкладке параметров"""
        if not hasattr(self, 'data'):
            return
        
        # 1. Обновляем текстовые поля параметров
        if 'params' in self.data:
            for param, field in self.params_fields.items():
                key = param.lower().replace(' ', '_')
                if key in self.data['params']:
                    field.setText(str(self.data['params'][key]))
                else:
                    field.clear()
        
        # 2. Обновляем метаданные (название, дата, версия)
        if 'meta' in self.data:
            # Находим все QLabel на вкладке
            for i in range(self.tabs.widget(0).layout().count()):  # 0 - индекс вкладки "Параметры"
                widget = self.tabs.widget(0).layout().itemAt(i).widget()
                if isinstance(widget, QLabel):
                    text = widget.text()
                    if text.startswith("Название:"):
                        widget.setText(f"Название: {self.data['meta'].get('name', '')}")
                    elif text.startswith("Дата:"):
                        widget.setText(f"Дата: {self.data['meta'].get('date', '')}")
                    elif text.startswith("Версия данных:"):
                        widget.setText(f"Версия данных: {self.data['meta'].get('version', 1)}")

    def _show_changelog_dialog(self):
        """Отображает диалоговое окно с историей изменений программы."""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("История изменений программы")
            dialog.setMinimumSize(600, 450) # Задаем размер окна

            layout = QVBoxLayout(dialog)

            text_browser = QTextBrowser() # Используем QTextBrowser для поддержки форматирования
            text_browser.setReadOnly(True)
            text_browser.setOpenExternalLinks(True) # Чтобы ссылки (если будут) открывались
            text_browser.setMarkdown(CHANGELOG_TEXT) # Устанавливаем текст из константы

            layout.addWidget(text_browser)

            # Кнопка Закрыть
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
            button_box.accepted.connect(dialog.accept) # Close привязан к accept по умолчанию
            button_box.rejected.connect(dialog.reject) # На всякий случай
            layout.addWidget(button_box)

            dialog.exec() # Показываем модальное окно

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось показать историю изменений:\n{e}")
            traceback.print_exc()



############## ЛОГИКА ##############

    def setup_logic_tab(self):
        """Вкладка логики проведения соревнования с настройками СКП"""
        self.logic_tab = QWidget()
        layout = QVBoxLayout(self.logic_tab)
        
        # 1. Общий старт
        self.common_start_cb = QCheckBox("Общий старт")
        layout.addWidget(self.common_start_cb)
        
        # 2. Штраф за опоздание на финиш
        penalty_layout = QHBoxLayout()
        penalty_layout.addWidget(QLabel("Штраф за опоздание на финиш:"))
        
        self.penalty_type_combo = QComboBox()
        self.penalty_type_combo.addItems(["Нет", "Баллы за 1 мин", "DNF"])
        self.penalty_type_combo.currentTextChanged.connect(self.update_penalty_ui)
        penalty_layout.addWidget(self.penalty_type_combo)
        
        self.penalty_value_edit = QLineEdit()
        self.penalty_value_edit.setPlaceholderText("Значение штрафа")
        self.penalty_value_edit.setValidator(QIntValidator(0, 1000))
        self.penalty_value_edit.setFixedWidth(80)
        penalty_layout.addWidget(self.penalty_value_edit)
        
        layout.addLayout(penalty_layout)
        
        # 3. Штраф за ложное КП
        false_cp_layout = QHBoxLayout()
        false_cp_layout.addWidget(QLabel("Штраф за ложное КП (Не работает. Ставить во вкладке кп отрицательные баллы)"))
        
        self.false_cp_penalty_edit = QLineEdit()
        self.false_cp_penalty_edit.setValidator(QIntValidator(0, 1000))
        self.false_cp_penalty_edit.setFixedWidth(80)
        false_cp_layout.addWidget(self.false_cp_penalty_edit)
        false_cp_layout.addWidget(QLabel("баллов"))
        
        layout.addLayout(false_cp_layout)
        
        # 4. Контрольное время (рассчитывается от старта до закрытия)
        self.route_time_group = QGroupBox("Контрольное время")
        route_time_layout = QVBoxLayout()
        
        self.route_time_info = QLabel("Рассчитывается автоматически\nот старта до закрытия трассы")
        self.route_time_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        route_time_layout.addWidget(self.route_time_info)
        
        self.route_time_group.setLayout(route_time_layout)
        layout.addWidget(self.route_time_group)
        

        # 5. Этапность - доработанная версия
        self.staged_cb = QCheckBox("Этапность")
        self.staged_cb.stateChanged.connect(self.toggle_stage_settings)
        layout.addWidget(self.staged_cb)

        # Контейнер для настроек этапов (основной GroupBox)
        self.stage_settings_group = QGroupBox("Настройки этапов")
        self.stage_settings_layout = QVBoxLayout() # Layout для GroupBox
        self.stage_settings_group.setLayout(self.stage_settings_layout)
        self.stage_settings_group.setVisible(False) # Изначально скрываем весь GroupBox
        # Добавляем GroupBox В ОСНОВНОЙ layout вкладки
        layout.addWidget(self.stage_settings_group)

        # --- Элементы ВНУТРИ GroupBox (добавляем в self.stage_settings_layout) ---

        # Количество этапов (остается сверху в GroupBox)
        stages_layout = QHBoxLayout()
        stages_layout.addWidget(QLabel("Количество этапов:"))
        self.stages_count_combo = QComboBox()
        self.stages_count_combo.addItems([str(i) for i in range(1, 11)]) # Увеличил макс. до 10 для теста
        self.stages_count_combo.currentIndexChanged.connect(self.update_stages_ui)
        stages_layout.addWidget(self.stages_count_combo)
        # Добавляем в layout Группы Настроек
        self.stage_settings_layout.addLayout(stages_layout)

        # Настройки нейтрализации на СКП (остается в GroupBox)
        neutralization_group = QGroupBox("Нейтрализация на СКП")
        neutralization_layout = QVBoxLayout()
        # ... (весь код для neutralization_layout без изменений) ...
        # Тип нейтрализации
        self.neutralization_type_group = QButtonGroup()
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("Тип нейтрализации:"))
        rb_no_neutral = QRadioButton("Нет")
        rb_total_neutral = QRadioButton("Суммарная")
        rb_per_skp_neutral = QRadioButton("На каждом СКП")
        self.neutralization_type_group.addButton(rb_no_neutral)
        self.neutralization_type_group.addButton(rb_total_neutral)
        self.neutralization_type_group.addButton(rb_per_skp_neutral)
        rb_no_neutral.setChecked(True)
        type_layout.addWidget(rb_no_neutral)
        type_layout.addWidget(rb_total_neutral)
        type_layout.addWidget(rb_per_skp_neutral)
        neutralization_layout.addLayout(type_layout)
        # Максимальное время нейтрализации (только для суммарной)
        self.max_neutral_time_layout = QHBoxLayout()
        self.max_neutral_time_layout.addWidget(QLabel("Макс. время нейтрализации:"))
        self.max_neutral_time_edit = QLineEdit()
        self.max_neutral_time_edit.setValidator(QIntValidator(0, 1000))
        self.max_neutral_time_edit.setPlaceholderText("минут")
        self.max_neutral_time_edit.setFixedWidth(80)
        self.max_neutral_time_edit.setEnabled(False)
        self.max_neutral_time_layout.addWidget(self.max_neutral_time_edit)
        neutralization_layout.addLayout(self.max_neutral_time_layout)
        # ... (конец кода для neutralization_layout) ...
        neutralization_group.setLayout(neutralization_layout)
        # Добавляем в layout Группы Настроек
        self.stage_settings_layout.addWidget(neutralization_group)
        # Обработчики для нейтрализации (остаются как есть)
        rb_no_neutral.toggled.connect(self.update_neutralization_ui)
        rb_total_neutral.toggled.connect(self.update_neutralization_ui)
        rb_per_skp_neutral.toggled.connect(self.update_neutralization_ui)

        # --- НАЧАЛО ИЗМЕНЕНИЯ ---
        # Контейнер для самих этапов и СКП (этот виджет будет расти)
        self.stages_container = QWidget()
        self.stages_container_layout = QVBoxLayout() # Layout для растущего контента
        self.stages_container.setLayout(self.stages_container_layout)

        # Создаем QScrollArea
        self.logic_stages_scroll_area = QScrollArea()
        self.logic_stages_scroll_area.setWidgetResizable(True) # Важно для растягивания контента по ширине
        self.logic_stages_scroll_area.setWidget(self.stages_container) # Устанавливаем наш контейнер внутрь скролла
        self.logic_stages_scroll_area.setMinimumHeight(200) # Задаем минимальную высоту для ScrollArea (по желанию)

        # Добавляем QScrollArea в layout Группы Настроек Этапов
        self.stage_settings_layout.addWidget(self.logic_stages_scroll_area)
        # --- КОНЕЦ ИЗМЕНЕНИЯ ---

        # Кнопка сохранения (остается в основном layout вкладки)
        save_btn = QPushButton("Сохранить параметры логики") # Изменил текст для ясности
        save_btn.clicked.connect(self.save_logic_params)
        layout.addWidget(save_btn)

        layout.addStretch() # Добавляем растяжитель В КОНЕЦ основного layout вкладки

        self.tabs.insertTab(1, self.logic_tab, "Логика") # Эта строка остается как есть
        self.update_logic_tab() # Эта строка остается как есть
   
    def update_logic_tab(self):
        """Обновляет данные на вкладке Логика с учетом настроек СКП"""
        if not hasattr(self, 'data'):
            return

        # Расчет времени трассы
        self.calculate_route_time_from_params()

        params = self.data.setdefault('logic_params', {})
        
        # Основные параметры
        self.common_start_cb.setChecked(params.get('common_start', False))
        
        penalty_type = params.get('penalty_type', 'Нет')
        self.penalty_type_combo.setCurrentText(penalty_type)
        self.penalty_value_edit.setText(str(params.get('penalty_value', 0)))
        self.update_penalty_ui(penalty_type)
        
        self.false_cp_penalty_edit.setText(str(params.get('false_cp_penalty', 0)))
        
        # Параметры этапности
        self.staged_cb.setChecked(params.get('staged', False))
        self.toggle_stage_settings(params.get('staged', False))
        
        if params.get('staged'):
            # Тип нейтрализации
            neutral_type = params.get('neutralization_type', 'Нет')
            for btn in self.neutralization_type_group.buttons():
                if btn.text() == neutral_type:
                    btn.setChecked(True)
                    break
            
            # Устанавливаем время нейтрализации в зависимости от типа
            if neutral_type == 'Суммарная':
                self.max_neutral_time_edit.setText(str(params.get('total_max_neutral_time', 0)))
            else:
                self.max_neutral_time_edit.setText(str(params.get('max_neutral_time', 0)))
            
            self.max_neutral_time_edit.setEnabled(neutral_type == 'Суммарная')
            
            # Количество этапов
            self.stages_count_combo.setCurrentText(str(params.get('stages_count', 1)))
            self.update_stages_ui()
            
            # Восстановление этапов
            for i, stage in enumerate(params.get('stages', [])):
                if i < len(self.stage_widgets):
                    self.stage_widgets[i]['name_edit'].setText(stage.get('name', ''))
                    self.stage_widgets[i]['time_limit_edit'].setText(stage.get('time_limit', ''))
            
            # Восстановление СКП (для всех типов нейтрализации)
            for i, skp in enumerate(params.get('skp_settings', [])):
                if i < len(self.skp_widgets):
                    # Время работы
                    self.skp_widgets[i]['open_time_edit'].setTime(
                        QTime.fromString(skp.get('open_time', '00:00'), "HH:mm"))
                    self.skp_widgets[i]['close_time_edit'].setTime(
                        QTime.fromString(skp.get('close_time', '23:59'), "HH:mm"))
                    
                    # Действие при опоздании
                    late_action = skp.get('late_action', 'Перенос на след. этап')
                    for btn in self.skp_widgets[i]['late_action_group'].buttons():
                        if btn.text() == late_action:
                            btn.setChecked(True)
                            break
                    
                    # Время нейтрализации
                    self.skp_widgets[i]['max_time_edit'].setText(str(skp.get('max_neutral_time', 0)))

    def save_logic_params(self):
        """Сохраняет параметры логики с учетом настроек СКП"""
        try:
            # Основные параметры
            params = {
                'common_start': self.common_start_cb.isChecked(),
                'penalty_type': self.penalty_type_combo.currentText(),
                'penalty_value': int(self.penalty_value_edit.text() or 0) if self.penalty_value_edit.isEnabled() else 0,
                'false_cp_penalty': int(self.false_cp_penalty_edit.text() or 0),
                'staged': self.staged_cb.isChecked(),
                'neutralization_type': self.neutralization_type_group.checkedButton().text()
            }

            # Для суммарной нейтрализации сохраняем общее время
            if params['neutralization_type'] == 'Суммарная':
                params['total_max_neutral_time'] = int(self.max_neutral_time_edit.text() or 0)
            else:
                params['max_neutral_time'] = int(self.max_neutral_time_edit.text() or 0) if self.max_neutral_time_edit.isEnabled() else 0

            if self.staged_cb.isChecked():
                # Сохраняем этапы
                params['stages_count'] = int(self.stages_count_combo.currentText())
                params['stages'] = []
                params['skp_settings'] = []  # Настройки СКП

                # Сохраняем данные этапов
                for i in range(params['stages_count']):
                    if i < len(self.stage_widgets):
                        stage_data = {
                            'name': self.stage_widgets[i]['name_edit'].text(),
                            'time_limit': self.stage_widgets[i]['time_limit_edit'].text()
                        }
                        params['stages'].append(stage_data)

                # Сохраняем настройки СКП (для всех типов нейтрализации)
                for i in range(len(self.skp_widgets)):
                    skp_data = {
                        'number': i+1,  # Добавляем номер СКП (1-based)
                        'open_time': self.skp_widgets[i]['open_time_edit'].time().toString("HH:mm"),
                        'close_time': self.skp_widgets[i]['close_time_edit'].time().toString("HH:mm"),
                        'late_action': self.skp_widgets[i]['late_action_group'].checkedButton().text(),
                        'max_neutral_time': int(self.skp_widgets[i]['max_time_edit'].text() or 0) if self.skp_widgets[i]['max_time_edit'].isEnabled() else 0
                    }
                    params['skp_settings'].append(skp_data)

            # Сохраняем в данные
            self.data['logic_params'] = params
            self.save_data()
            
            QMessageBox.information(self, "Сохранено", "Параметры логики успешно сохранены!")
            self.update_logic_tab()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить параметры: {str(e)}")

    def update_stages_ui(self):
        """Обновляет UI для редактирования этапов и СКП с новыми настройками"""
        # Очищаем предыдущие виджеты
        for i in reversed(range(self.stages_container_layout.count())): 
            self.stages_container_layout.itemAt(i).widget().setParent(None)
        
        self.stage_widgets = []
        self.skp_widgets = []
        stages_count = int(self.stages_count_combo.currentText())
        
        for i in range(stages_count):
            # Добавляем этап
            stage_group = QGroupBox(f"Этап {i+1}")
            stage_layout = QVBoxLayout()
            
            # Название этапа
            name_layout = QHBoxLayout()
            name_layout.addWidget(QLabel("Название:"))
            name_edit = QLineEdit()
            name_edit.setPlaceholderText(f"Этап {i+1}")
            name_layout.addWidget(name_edit)
            stage_layout.addLayout(name_layout)
            
            # Макс. время этапа
            time_limit_layout = QHBoxLayout()
            time_limit_layout.addWidget(QLabel("Макс. время (мин):"))
            time_limit_edit = QLineEdit()
            time_limit_edit.setPlaceholderText("не ограничено")
            time_limit_edit.setValidator(QIntValidator(1, 1000))
            time_limit_layout.addWidget(time_limit_edit)
            stage_layout.addLayout(time_limit_layout)
            
            stage_group.setLayout(stage_layout)
            self.stages_container_layout.addWidget(stage_group)
            
            self.stage_widgets.append({
                'name_edit': name_edit,
                'time_limit_edit': time_limit_edit
            })
            
            # Добавляем СКП после этапа (кроме последнего)
            if i < stages_count - 1:  # Убрана проверка на тип нейтрализации
                skp_group = QGroupBox(f"СКП {i+1} (между этапами {i+1} и {i+2})")
                skp_layout = QVBoxLayout()
                
                # Время работы СКП
                time_work_layout = QHBoxLayout()
                time_work_layout.addWidget(QLabel("Время работы:"))
                
                # Открытие СКП
                skp_open_edit = QTimeEdit()
                skp_open_edit.setDisplayFormat("HH:mm")
                skp_open_edit.setTime(QTime(0, 0))
                time_work_layout.addWidget(QLabel("Открытие:"))
                time_work_layout.addWidget(skp_open_edit)
                
                # Закрытие СКП
                skp_close_edit = QTimeEdit()
                skp_close_edit.setDisplayFormat("HH:mm")
                skp_close_edit.setTime(QTime(23, 59))
                time_work_layout.addWidget(QLabel("Закрытие:"))
                time_work_layout.addWidget(skp_close_edit)
                
                skp_layout.addLayout(time_work_layout)
                
                # Действие при опоздании
                late_action_layout = QHBoxLayout()
                late_action_layout.addWidget(QLabel("Действие при опоздании:"))
                
                skp_late_action_group = QButtonGroup()
                rb_move_next = QRadioButton("Перенос на след. этап")
                rb_dnq = QRadioButton("DNQ (дисквалификация)")
                skp_late_action_group.addButton(rb_move_next)
                skp_late_action_group.addButton(rb_dnq)
                rb_move_next.setChecked(True)
                
                late_action_layout.addWidget(rb_move_next)
                late_action_layout.addWidget(rb_dnq)
                skp_layout.addLayout(late_action_layout)
                
                # Макс. время нейтрализации на этом СКП
                time_layout = QHBoxLayout()
                time_layout.addWidget(QLabel("Макс. время нейтрализации:"))
                
                skp_time_edit = QLineEdit()
                skp_time_edit.setValidator(QIntValidator(0, 1000))
                skp_time_edit.setPlaceholderText("минут")
                skp_time_edit.setFixedWidth(80)
                
                # Блокировка поля в зависимости от типа нейтрализации
                neutral_type = self.neutralization_type_group.checkedButton().text()
                if neutral_type == "Суммарная" or neutral_type == "Нет":
                    skp_time_edit.setEnabled(False)
                else:
                    skp_time_edit.setEnabled(True)
                
                time_layout.addWidget(skp_time_edit)
                skp_layout.addLayout(time_layout)
                skp_group.setLayout(skp_layout)
                self.stages_container_layout.addWidget(skp_group)
                
                self.skp_widgets.append({
                    'open_time_edit': skp_open_edit,
                    'close_time_edit': skp_close_edit,
                    'late_action_group': skp_late_action_group,
                    'max_time_edit': skp_time_edit
                })

    def update_neutralization_ui(self):
        """Обновляет UI в зависимости от выбранного типа нейтрализации"""
        neutral_type = self.neutralization_type_group.checkedButton().text()
        
        # Для суммарной нейтрализации - одно поле времени
        if neutral_type == "Суммарная":
            self.max_neutral_time_edit.setEnabled(True)
            self.max_neutral_time_edit.setToolTip("Общее время нейтрализации для всех СКП")
        else:
            self.max_neutral_time_edit.setEnabled(False)
            self.max_neutral_time_edit.setToolTip("")
        
        # Блокировка полей времени нейтрализации в СКП
        for skp in self.skp_widgets:
            if neutral_type == "На каждом СКП":
                skp['max_time_edit'].setEnabled(True)
                skp['max_time_edit'].setToolTip("Макс. время для этого СКП")
            else:
                skp['max_time_edit'].setEnabled(False)
                skp['max_time_edit'].setToolTip("")
        
        # Обновляем контейнер с этапами/СКП
        self.update_stages_ui()         

    def calculate_remaining_time(self, member, skp):
        """Вычисляет оставшееся время нейтрализации с учетом типа"""
        if not member.get("skp_entries"):
            return "0:00"
        
        neutral_type = self.data.get("logic_params", {}).get("neutralization_type", "Нет")
        
        if neutral_type == "Суммарная":
            # Для суммарной нейтрализации - общее время для всех СКП
            total_time = self.data["logic_params"].get("total_max_neutral_time", 0) * 60
            used_time = sum(
                self.calculate_actual_duration(entry)
                for entry in member.get("skp_entries", [])
            )
        else:
            # Для других типов - время только для текущего СКП
            total_time = skp.get("max_neutral_time", 0) * 60
            used_time = sum(
                self.calculate_actual_duration(entry)
                for entry in member.get("skp_entries", [])
                if entry.get("skp") == skp.get("number", 0)
            )
        
        remaining = max(0, total_time - used_time)
        return f"{remaining//60}:{remaining%60:02d}"            

    def calculate_route_time_from_params(self):
        """Рассчитывает КВ от старта до закрытия из параметров"""
        if not hasattr(self, 'data') or 'params' not in self.data:
            self.route_time_info.setText("Нет данных о параметрах")
            return
        
        try:
            start_time_str = self.data['params'].get('время_старта', '')
            closing_time_str = self.data['params'].get('время_закрытия_трассы', '')
            
            if not start_time_str or not closing_time_str:
                self.route_time_info.setText("Не задано время старта/закрытия")
                return
            
            # Добавляем секунды, если их нет
            if len(start_time_str.split(':')) == 2:
                start_time_str += ":00"
            if len(closing_time_str.split(':')) == 2:
                closing_time_str += ":00"
            
            start_time = QTime.fromString(start_time_str, "HH:mm:ss")
            closing_time = QTime.fromString(closing_time_str, "HH:mm:ss")
            
            if not start_time.isValid() or not closing_time.isValid():
                self.route_time_info.setText("Некорректное время в параметрах")
                return
            
            # Если закрытие на следующий день (например, старт 20:00, закрытие 02:00)
            if closing_time < start_time:
                closing_time = closing_time.addSecs(24 * 3600)  # Добавляем 24 часа
                
            seconds = start_time.secsTo(closing_time)
            
            if seconds <= 0:
                self.route_time_info.setText("Ошибка: время закрытия раньше старта")
                return
                
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            seconds = seconds % 60
            
            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            self.route_time_info.setText(f"Контрольное время: {time_str}\n(от старта до закрытия НЕ вкл. нейтр.)")
            
            # Сохраняем расчетное время в параметры
            self.data['params']['расчетное_время_трассы'] = time_str
            self.save_data()
            
        except Exception as e:
            self.route_time_info.setText(f"Ошибка расчета: {str(e)}")

    def toggle_stage_settings(self, checked):
        """Показывает/скрывает настройки этапов"""
        self.stage_settings_group.setVisible(checked)
        if checked:
            self.update_stages_ui()

    def update_penalty_ui(self, penalty_type):
        """Обновляет UI в зависимости от выбранного типа штрафа"""
        if penalty_type in ["Нет", "DNF"]:
            self.penalty_value_edit.setEnabled(False)
            self.penalty_value_edit.clear()
        else:
            self.penalty_value_edit.setEnabled(True)

    def update_skp_penalty_ui(self, penalty_type):
        """Обновляет UI штрафа за опоздание на СКП"""
        sender = self.sender()
        for widget in self.stage_widgets:
            if widget['skp_penalty_combo'] == sender:
                if penalty_type in ["Нет", "DNQ"]:
                    widget['skp_penalty_value_edit'].setEnabled(False)
                    widget['skp_penalty_value_edit'].clear()
                else:
                    widget['skp_penalty_value_edit'].setEnabled(True)
                break














############## КП ##############


    def setup_checkpoints_tab(self):
        """Вкладка контрольных пунктов с исправленным скрытием таблиц"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Верхняя панель с кнопками
        top_panel = QWidget()
        top_layout = QHBoxLayout(top_panel)
        
        btn_generate = QPushButton("Генерировать КП")
        btn_generate.clicked.connect(self.generate_checkpoints)
        top_layout.addWidget(btn_generate)
        
        self.btn_show = QPushButton("Показать КП")
        self.btn_show.clicked.connect(self.show_checkpoints_table)
        top_layout.addWidget(self.btn_show)
        
        btn_add = QPushButton("Добавить КП")
        btn_add.clicked.connect(self.add_checkpoint_manual)
        top_layout.addWidget(btn_add)
        
        # btn_false_cps = QPushButton("Ложные КП")
        # btn_false_cps.clicked.connect(self.show_false_cps_table)
        # top_layout.addWidget(btn_false_cps)
        
        btn_stages = QPushButton("Этапы")
        btn_stages.clicked.connect(self.show_stages_table1)
        top_layout.addWidget(btn_stages)
        
        layout.addWidget(top_panel)
        
        # Основная таблица КП (изначально скрыта)
        self.checkpoints_table = QTableWidget()
        self.checkpoints_table.setVisible(False)
        self.checkpoints_table.doubleClicked.connect(self.edit_checkpoint_dialog)
        self.checkpoints_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.checkpoints_table)
        
        # Таблица ложных КП (изначально скрыта)
        self.false_cps_table = QTableWidget()
        self.false_cps_table.setVisible(False)
        self.false_cps_table.doubleClicked.connect(self.edit_checkpoint_dialog)
        layout.addWidget(self.false_cps_table)
        
        # Таблица этапов (изначально скрыта)
        self.stages_table = QTableWidget()
        self.stages_table.setVisible(False)
        self.stages_table.doubleClicked.connect(self.edit_checkpoint_dialog)
        layout.addWidget(self.stages_table)
        
        # Кнопка сохранения внизу
        btn_save = QPushButton("Сохранить изменения")
        btn_save.clicked.connect(self.save_checkpoints) # <--- Убедитесь, что эта строка есть
        layout.addWidget(btn_save)
        
        self.tabs.addTab(tab, "КП")

    def update_checkpoints_table(self):
        """Обновляем данные в таблице КП (упрощенная версия)"""
        # Определяем какие зачеты есть
        classifications = []
        if hasattr(self, 'data') and 'params' in self.data:
            if self.data['params'].get('зачет1'):
                classifications.append(self.data['params']['зачет1'])
            if self.data['params'].get('зачет2'):
                classifications.append(self.data['params']['зачет2'])
            if self.data['params'].get('зачет3'):
                classifications.append(self.data['params']['зачет3'])
        
        # Настраиваем таблицу (только название, зачеты и бальность)
        headers = ["Название"] + classifications + ["Баллы"]
        self.checkpoints_table.setColumnCount(len(headers))
        self.checkpoints_table.setHorizontalHeaderLabels(headers)
        
        # Заполняем данными из self.data
        if hasattr(self, 'data') and 'checkpoints' in self.data:
            self.checkpoints_table.setRowCount(len(self.data['checkpoints']))
            
            for row, cp in enumerate(self.data['checkpoints']):
                # Колонка "Название"
                self.checkpoints_table.setItem(row, 0, QTableWidgetItem(cp.get('name', '')))
                
                # Колонки с зачетами (чекбоксы)
                for col, classif in enumerate(classifications, 1):
                    checkbox = QCheckBox()
                    checkbox.setChecked(cp.get('classifications', {}).get(classif, False))
                    checkbox.stateChanged.connect(lambda state, r=row, c=col: self.on_checkbox_changed(r, c, state))
                    self.checkpoints_table.setCellWidget(row, col, checkbox)
                
                # Колонка с бальностью
                offset = 1 + len(classifications)
                self.checkpoints_table.setItem(row, offset, QTableWidgetItem(str(cp.get('score', ''))))
        
        # Настраиваем растягивание столбцов
        self.checkpoints_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.checkpoints_table.resizeColumnsToContents()

    def toggle_all_false_cps(self, state):
        """Массовое выделение/снятие всех ложных КП"""
        if not hasattr(self, 'data') or 'params' not in self.data:
            return
            
        classifications = []
        if self.data['params'].get('зачет1'):
            classifications.append(self.data['params']['зачет1'])
        if self.data['params'].get('зачет2'):
            classifications.append(self.data['params']['зачет2'])
        if self.data['params'].get('зачет3'):
            classifications.append(self.data['params']['зачет3'])
            
        if not classifications:
            return
            
        first_false_col = 1  # После колонки "Название КП"
        
        for row in range(self.false_cps_table.rowCount()):
            for col in range(first_false_col, first_false_col + len(classifications)):
                widget = self.false_cps_table.cellWidget(row, col)
                if widget and isinstance(widget, QCheckBox):
                    widget.setChecked(state == Qt.CheckState.Checked.value)

    def on_false_cp_changed(self, row, col, state):
        """Обработка изменения чекбокса ложного КП"""
        if not hasattr(self, 'data') or 'checkpoints' not in self.data:
            return
            
        # Определяем какой зачет изменился
        classifications = []
        if self.data['params'].get('зачет1'):
            classifications.append(self.data['params']['зачет1'])
        if self.data['params'].get('зачет2'):
            classifications.append(self.data['params']['зачет2'])
        if self.data['params'].get('зачет3'):
            classifications.append(self.data['params']['зачет3'])
            
        if col-1 < len(classifications):
            classif = classifications[col-1]
            cp_name = self.false_cps_table.item(row, 0).text()
            
            # Находим КП в данных
            for cp in self.data['checkpoints']:
                if cp.get('name') == cp_name:
                    if 'false_for' not in cp:
                        cp['false_for'] = []
                    
                    if state == Qt.CheckState.Checked.value:
                        if classif not in cp['false_for']:
                            cp['false_for'].append(classif)
                    else:
                        if classif in cp['false_for']:
                            cp['false_for'].remove(classif)
                    break

    def refresh_checkpoints_from_file(self):
        """Обновление данных из файла"""
        try:
            # Загружаем данные заново
            print(">>> СОБИРАЮСЬ ВЫЗВАТЬ load_latest_data из refresh_checkpoints_from_file")
            self.load_latest_data()
            
            # Если таблица видима - обновляем ее
            if self.checkpoints_table.isVisible():
                self.update_checkpoints_table()
                
            QMessageBox.information(self, "Обновлено", "Данные загружены из файла")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить данные: {str(e)}")

    def toggle_all_classifications(self, state):
        """Массовое выделение/снятие всех чекбоксов"""
        if not hasattr(self, 'data') or 'params' not in self.data:
            return
            
        classifications = []
        if self.data['params'].get('зачет1'):
            classifications.append(self.data['params']['зачет1'])
        if self.data['params'].get('зачет2'):
            classifications.append(self.data['params']['зачет2'])
        if self.data['params'].get('зачет3'):
            classifications.append(self.data['params']['зачет3'])
            
        if not classifications:
            return
            
        first_classif_col = 1  # После колонки "Название"
        
        for row in range(self.checkpoints_table.rowCount()):
            for col in range(first_classif_col, first_classif_col + len(classifications)):
                widget = self.checkpoints_table.cellWidget(row, col)
                if widget and isinstance(widget, QCheckBox):
                    widget.setChecked(state == Qt.CheckState.Checked.value)
    
    def on_checkbox_changed(self, row, col, state):
        """Обработка изменения чекбокса"""
        if not hasattr(self, 'data') or 'checkpoints' not in self.data:
            return
            
        # Определяем какой зачет изменился
        classifications = []
        if self.data['params'].get('зачет1'):
            classifications.append(self.data['params']['зачет1'])
        if self.data['params'].get('зачет2'):
            classifications.append(self.data['params']['зачет2'])
        if self.data['params'].get('зачет3'):
            classifications.append(self.data['params']['зачет3'])
            
        if col-1 < len(classifications):
            classif = classifications[col-1]
            self.data['checkpoints'][row]['classifications'][classif] = (state == Qt.CheckState.Checked.value)

    def save_checkpoints(self):
        """Сохранение изменений в КП"""
        if not hasattr(self, 'data'):
            QMessageBox.critical(self, "Ошибка", "Нет данных для сохранения")
            return

        try:
            # Данные уже обновляются через on_checkbox_changed (для основных КП)
            # и _on_stage_radio_button_toggled (для этапов)
            # Просто сохраняем текущее состояние self.data
            if self.save_data(): # Вызываем save_data и проверяем результат
                 QMessageBox.information(self, "Сохранено", "Изменения в КП успешно сохранены")
                 # Обновляем текущую видимую таблицу (на всякий случай)
                 if self.checkpoints_table.isVisible():
                     self.update_checkpoints_table()
                 elif self.stages_table.isVisible():
                      self.show_stages_table1() # Обновляем вид этапов
                 elif self.false_cps_table.isVisible():
                      self.show_false_cps_table()

            else:
                 QMessageBox.critical(self, "Ошибка Сохранения", "Не удалось сохранить данные КП в файл!")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения КП: {str(e)}")
            traceback.print_exc()

    def generate_checkpoints(self):
        """Генерация КП с подтверждением"""
        if not hasattr(self, 'data') or 'params' not in self.data:
            QMessageBox.critical(self, "Ошибка", "Сначала задайте параметры соревнования")
            return
        
        try:
            # Получаем количество КП из параметров
            kp_count = int(self.data['params'].get('количество_кп', 0))
            if kp_count <= 0:
                QMessageBox.critical(self, "Ошибка", "Неверное количество КП в параметрах")
                return
            
            # Получаем бальность КП из параметров
            score = str(self.data['params'].get('бальность_кп', 10))
            
            # Запрос подтверждения
            reply = QMessageBox.question(
                self, 'Подтверждение',
                f'Вы хотите сгенерировать КП? Будет сгенерировано {kp_count} КП.\n'
                f'Бальность каждого КП: {score}',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # Генерируем новые КП
                existing_count = len(self.data.get('checkpoints', []))
                for i in range(1, kp_count + 1):
                    new_cp = {
                        "name": f"КП {existing_count + i}",  # Изменено на пробел вместо дефиса
                        "classifications": {},
                        "score": score
                    }
                    
                    # Заполняем зачеты если они есть
                    if self.data['params'].get('зачет1'):
                        new_cp['classifications'][self.data['params']['зачет1']] = True
                    if self.data['params'].get('зачет2'):
                        new_cp['classifications'][self.data['params']['зачет2']] = True
                    if self.data['params'].get('зачет3'):
                        new_cp['classifications'][self.data['params']['зачет3']] = False
                    
                    self.data['checkpoints'].append(new_cp)
                
                # Сохраняем изменения
                self.save_data()
                
                # Обновляем таблицы если они видимы
                if hasattr(self, 'checkpoints_table') and self.checkpoints_table.isVisible():
                    self.update_checkpoints_table()
                if hasattr(self, 'false_cps_table') and self.false_cps_table.isVisible():
                    self.show_false_cps_table()
                if hasattr(self, 'stages_table') and self.stages_table.isVisible():
                    self.show_stages_table1()
                
                QMessageBox.information(
                    self, "Готово", 
                    f"Успешно сгенерировано {kp_count} контрольных пунктов\n"
                    f"Бальность: {score}"
                )
        
        except ValueError as ve:
            QMessageBox.critical(self, "Ошибка значения", f"Некорректные числовые параметры: {str(ve)}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Неизвестная ошибка при генерации КП: {str(e)}")

    def confirm_delete_checkpoint(self, row, dialog):
        """Подтверждение удаления КП"""
        cp_name = self.data['checkpoints'][row].get('name', 'этот КП')
        
        reply = QMessageBox.question(
            self, 
            'Подтверждение удаления',
            f'Вы уверены, что хотите удалить {cp_name}?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.delete_checkpoint(row)
            dialog.accept()

    def delete_checkpoint(self, row):
        """Удаление КП по индексу"""
        try:
            if 0 <= row < len(self.data['checkpoints']):
                # Удаляем из данных
                del self.data['checkpoints'][row]
                
                # Сохраняем изменения
                self.save_data()
                
                # Обновляем все таблицы
                self.update_checkpoints_table()
                if self.false_cps_table.isVisible():
                    self.show_false_cps_table()
                if self.stages_table.isVisible():
                    self.show_stages_table1()
                    
                QMessageBox.information(self, "Успех", "КП успешно удален!")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось удалить КП: {str(e)}")

    def show_checkpoints_table(self):
        """Показывает таблицу КП и скрывает остальные"""
        self.checkpoints_table.setVisible(True)
        self.false_cps_table.setVisible(False)
        self.stages_table.setVisible(False)
        
        # Обновляем данные в таблице
        self.update_checkpoints_table()
        self.btn_show.setText("Обновить КП")

    def add_checkpoint_manual(self):
        """Добавление нового КП вручную с сохранением в файл"""
        if not hasattr(self, 'data'):
            QMessageBox.critical(self, "Ошибка", "Данные не загружены")
            return
        
        try:
            # Получаем количество КП из параметров
            kp_count = len(self.data.get('checkpoints', []))
            
            # Создаем новый КП с базовыми значениями
            new_cp = {
                "name": f"КП {kp_count + 1}",
                "classifications": {},
                "score": str(self.data.get('params', {}).get('бальность_кп', 10)),
                "false_for": [],
                "stages": []
            }
            
            # Заполняем зачеты если они есть
            if 'params' in self.data:
                if self.data['params'].get('зачет1'):
                    new_cp['classifications'][self.data['params']['зачет1']] = True
                if self.data['params'].get('зачет2'):
                    new_cp['classifications'][self.data['params']['зачет2']] = False
                if self.data['params'].get('зачет3'):
                    new_cp['classifications'][self.data['params']['зачет3']] = False
            
            # Добавляем в данные
            self.data.setdefault('checkpoints', []).append(new_cp)
            
            # Сохраняем изменения в файл
            if not self.save_data():
                raise Exception("Не удалось сохранить данные")
            
            # Обновляем таблицы если они видимы
            if hasattr(self, 'checkpoints_table') and self.checkpoints_table.isVisible():
                self.update_checkpoints_table()
            if hasattr(self, 'false_cps_table') and self.false_cps_table.isVisible():
                self.show_false_cps_table()
            if hasattr(self, 'stages_table') and self.stages_table.isVisible():
                self.show_stages_table1()
            
            QMessageBox.information(self, "Добавлено", f"Добавлен новый КП: {new_cp['name']}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось добавить КП: {str(e)}")

    def edit_checkpoint_dialog(self, index):
        """Диалог редактирования КП с поддержкой отрицательных значений и выбором ОДНОГО этапа."""
        try:
            row = index.row()
            if row < 0 or row >= len(self.data['checkpoints']):
                return

            cp = self.data['checkpoints'][row]
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Редактирование КП: {cp.get('name', '')}")
            dialog.setMinimumWidth(500)

            layout = QVBoxLayout(dialog)

            # Основная информация (без изменений)
            info_group = QGroupBox("Основная информация")
            info_layout = QVBoxLayout()
            # ... (код для name_edit, score_edit как и был) ...
            # Название КП
            name_layout = QHBoxLayout()
            name_layout.addWidget(QLabel("Название:"))
            name_edit = QLineEdit(cp.get('name', ''))
            name_layout.addWidget(name_edit)
            info_layout.addLayout(name_layout)

            # Баллы (с возможностью отрицательных значений)
            score_layout = QHBoxLayout()
            score_layout.addWidget(QLabel("Баллы:"))
            score_edit = QLineEdit(str(cp.get('score', 10)))
            score_edit.setValidator(QIntValidator(-1000, 1000)) # Валидатор для целых чисел
            score_layout.addWidget(score_edit)
            info_layout.addLayout(score_layout)

            info_group.setLayout(info_layout)
            layout.addWidget(info_group)

            # Доступность для зачетов (без изменений)
            classifications = self.get_available_classifications() # Используем вспомогательную функцию
            if classifications:
                classif_group = QGroupBox("Доступность для зачетов")
                classif_layout = QVBoxLayout()
                self.dialog_classif_checkboxes = {} # Используем атрибут диалога
                for classif in classifications:
                    cb = QCheckBox(classif)
                    # Убедимся, что 'classifications' существует в cp перед доступом
                    cb.setChecked(cp.get('classifications', {}).get(classif, False))
                    self.dialog_classif_checkboxes[classif] = cb
                    classif_layout.addWidget(cb)
                classif_group.setLayout(classif_layout)
                layout.addWidget(classif_group)

            # Ложность для зачетов (без изменений)
            # (код для false_checkboxes как и был)
            if classifications:
                 false_group = QGroupBox("Ложный КП для зачетов") # Убрали (не работает)
                 false_layout = QVBoxLayout()
                 self.dialog_false_checkboxes = {} # Используем атрибут диалога
                 for classif in classifications:
                     cb = QCheckBox(classif)
                     cb.setChecked(classif in cp.get('false_for', []))
                     self.dialog_false_checkboxes[classif] = cb
                     false_layout.addWidget(cb)
                 false_group.setLayout(false_layout)
                 false_group.setEnabled(False) # Отключаем группу, т.к. не работает
                 layout.addWidget(false_group)


            # ---> НАЧАЛО ИЗМЕНЕНИЯ: Принадлежность к этапам (Радиокнопки) <---
            self.dialog_stage_button_group = None # Сбрасываем группу для этого диалога
            if (hasattr(self, 'data') and 'logic_params' in self.data and
                    self.data['logic_params'].get('staged') and
                    'stages' in self.data['logic_params']):

                stages = [stage.get('name', f'Этап {i+1}')
                          for i, stage in enumerate(self.data['logic_params'].get('stages', []))]

                if stages: # Показываем блок, только если есть этапы
                    stages_group = QGroupBox("Принадлежность к этапу (выберите один)")
                    stages_layout = QVBoxLayout()
                    self.dialog_stage_button_group = QButtonGroup(dialog) # Группа радиокнопок

                    # Кнопка "Без этапа"
                    rb_none = QRadioButton("Без этапа")
                    self.dialog_stage_button_group.addButton(rb_none)
                    stages_layout.addWidget(rb_none)

                    # Кнопки для каждого этапа
                    current_stage = None
                    cp_stages = cp.get('stages', [])
                    if cp_stages:
                        current_stage = cp_stages[0] # Берем первый (и единственный) этап

                    for stage_name in stages:
                        rb = QRadioButton(stage_name)
                        self.dialog_stage_button_group.addButton(rb)
                        stages_layout.addWidget(rb)
                        if stage_name == current_stage:
                            rb.setChecked(True) # Отмечаем текущий этап

                    # Если ни один этап не отмечен, выбираем "Без этапа"
                    if current_stage is None:
                        rb_none.setChecked(True)

                    stages_group.setLayout(stages_layout)
                    layout.addWidget(stages_group)
            # ---> КОНЕЦ ИЗМЕНЕНИЯ <---


            # Кнопки (без изменений)
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save |
                                          QDialogButtonBox.StandardButton.Cancel)

            delete_button = QPushButton("Удалить")
            delete_button.setStyleSheet("background-color: #ff4444; color: white;")
            # Убедитесь, что роль правильная или используйте стандартные кнопки, если нужно
            button_box.addButton(delete_button, QDialogButtonBox.ButtonRole.DestructiveRole)

            # Подключаем сохранение к кнопке Save (OK), отмену и удаление
            button_box.button(QDialogButtonBox.StandardButton.Save).setText("Сохранить") # Меняем текст кнопки OK
            button_box.accepted.connect(lambda: self.save_checkpoint_changes(
                row, name_edit.text(), score_edit.text(), dialog))
            button_box.rejected.connect(dialog.reject)
            delete_button.clicked.connect(lambda: self.confirm_delete_checkpoint(row, dialog))

            layout.addWidget(button_box)
            dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при открытии диалога: {str(e)}")
            traceback.print_exc() # Выводим traceback в консоль для отладки

    def save_checkpoint_changes(self, row, name, score, dialog):
        """Сохранение изменений КП с проверкой и обработкой выбора ОДНОГО этапа."""
        try:
            # --- Валидация ---
            if not name:
                QMessageBox.warning(dialog, "Ошибка", "Название КП не может быть пустым!")
                return
            try:
                score_int = int(score) # Преобразуем в число (может быть отрицательным)
            except ValueError:
                QMessageBox.warning(dialog, "Ошибка", "Баллы должны быть целым числом!")
                return

            # --- Сохранение данных ---
            cp = self.data['checkpoints'][row]
            cp['name'] = name
            cp['score'] = str(score_int) # Сохраняем как строку

            # Обновляем доступность для зачетов
            if hasattr(self, 'dialog_classif_checkboxes'):
                cp.setdefault('classifications', {}) # Создаем словарь, если его нет
                for classif, cb in self.dialog_classif_checkboxes.items():
                     cp['classifications'][classif] = cb.isChecked()

            # Обновляем ложные КП
            if hasattr(self, 'dialog_false_checkboxes'):
                cp['false_for'] = [
                    classif for classif, cb in self.dialog_false_checkboxes.items()
                    if cb.isChecked()
                ]

            # ---> НАЧАЛО ИЗМЕНЕНИЯ: Сохранение этапа (одного) <---
            if hasattr(self, 'dialog_stage_button_group') and self.dialog_stage_button_group:
                checked_button = self.dialog_stage_button_group.checkedButton()
                if checked_button:
                    button_text = checked_button.text()
                    if button_text == "Без этапа":
                        cp['stages'] = [] # Пустой список, если не принадлежит этапу
                    else:
                        cp['stages'] = [button_text] # Список с одним именем этапа
                else:
                    # На всякий случай, если ни одна кнопка не выбрана (не должно происходить)
                    cp['stages'] = []
            elif 'stages' in cp:
                 # Если этапность отключена или нет этапов, но поле stages было, очищаем его
                 cp['stages'] = []
            # ---> КОНЕЦ ИЗМЕНЕНИЯ <---

            # Сохраняем основной файл данных
            if self.save_data():
                # Обновляем таблицы, если они видимы
                if self.checkpoints_table.isVisible():
                    self.update_checkpoints_table()
                if self.stages_table.isVisible():
                    self.show_stages_table1() # Обновляем информационную таблицу этапов
                if self.false_cps_table.isVisible():
                     self.show_false_cps_table() # Если вы ее используете

                dialog.accept() # Закрываем диалог только после успешного сохранения
            else:
                 QMessageBox.critical(dialog, "Ошибка сохранения", "Не удалось сохранить данные в файл!")


        except Exception as e:
            QMessageBox.critical(dialog, "Ошибка", f"Не удалось сохранить изменения: {str(e)}")
            traceback.print_exc()

    def _on_stage_radio_button_toggled(self, is_checked, cp_name, stage_name_or_none):
        """
        Обрабатывает изменение состояния радиокнопки выбора этапа в таблице.
        cp_name: Имя контрольного пункта.
        stage_name_or_none: Имя выбранного этапа или None, если выбрано "Без этапа".
        """
        # Реагируем только на ВЫБОР кнопки (checked = True)
        if not is_checked:
            return

        # Находим КП в данных
        if not hasattr(self, 'data') or 'checkpoints' not in self.data:
            return

        found_cp = None
        for cp in self.data['checkpoints']:
            if cp.get('name') == cp_name:
                found_cp = cp
                break

        if not found_cp:
            print(f"Предупреждение: КП '{cp_name}' не найден в данных при обработке радиокнопки.")
            return

        # Обновляем поле 'stages' для этого КП
        if stage_name_or_none is None: # Выбрано "Без этапа"
            if found_cp.get('stages'): # Обновляем только если было что-то другое
                 print(f"Этап для КП '{cp_name}' снят.")
                 found_cp['stages'] = []
                 # Здесь НЕ вызываем save_data(), позволяем сохранить через кнопку "Сохранить изменения"
        else: # Выбран конкретный этап
            if found_cp.get('stages') != [stage_name_or_none]: # Обновляем только если этап изменился
                 print(f"КП '{cp_name}' назначен этапу '{stage_name_or_none}'.")
                 found_cp['stages'] = [stage_name_or_none]
                 # Здесь НЕ вызываем save_data()

        # ВАЖНО: Не сохраняем данные здесь автоматически.
        # Пользователь должен нажать кнопку "Сохранить изменения" на вкладке "КП",
        # которая вызовет self.save_checkpoints() (а тот уже self.save_data()).
        # Это предотвращает лишние операции сохранения при каждом клике.

# В классе RaceApp

# В классе RaceApp

    def show_stages_table1(self):
        """
        Показывает таблицу принадлежности КП к этапам с РАДИОКНОПКАМИ для быстрого редактирования.
        Исправлена сортировка строк КП и ширина колонок этапов/КП.
        """
        # ... (весь код до настройки ширины остается как в предыдущем ответе) ...
        is_staged = self.data.get('logic_params', {}).get('staged', False)
        stages_info = self.data['logic_params'].get('stages', []) if is_staged else []

        if not is_staged or not stages_info:
            self.stages_table.setVisible(False)
            self.checkpoints_table.setVisible(True)
            self.update_checkpoints_table()
            self.btn_show.setText("Обновить КП")
            return

        self.checkpoints_table.setVisible(False)
        self.false_cps_table.setVisible(False)
        self.stages_table.setVisible(True)

        stages_names = [stage.get('name', f'Этап {i+1}') for i, stage in enumerate(stages_info)]
        headers = ["Название КП", "Без этапа"] + stages_names
        self.stages_table.setColumnCount(len(headers))
        self.stages_table.setHorizontalHeaderLabels(headers)
        self.stages_table.setRowCount(0)

        self.stages_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.stages_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.stages_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)

        if hasattr(self, 'data') and 'checkpoints' in self.data:
            all_cps = sorted(self.data['checkpoints'], key=self._extract_cp_number)
            self.stages_table.setRowCount(len(all_cps))
            self._row_button_groups = []

            for row, cp in enumerate(all_cps):
                 # ... (код заполнения строки радиокнопками как в прошлый раз) ...
                 cp_name = cp.get('name', '')
                 name_item = QTableWidgetItem(cp_name)
                 name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                 self.stages_table.setItem(row, 0, name_item)

                 row_button_group = QButtonGroup(self.stages_table)
                 self._row_button_groups.append(row_button_group)
                 current_stage = cp.get('stages', [])[0] if cp.get('stages') else None

                 col_no_stage = 1
                 rb_no_stage = QRadioButton()
                 widget_no_stage = QWidget()
                 layout_no_stage = QHBoxLayout(widget_no_stage)
                 layout_no_stage.addWidget(rb_no_stage)
                 layout_no_stage.setAlignment(Qt.AlignmentFlag.AlignCenter)
                 layout_no_stage.setContentsMargins(0,0,0,0)
                 widget_no_stage.setLayout(layout_no_stage)
                 self.stages_table.setCellWidget(row, col_no_stage, widget_no_stage)
                 row_button_group.addButton(rb_no_stage)
                 rb_no_stage.toggled.connect(lambda checked, cn=cp_name, sn=None:
                                             self._on_stage_radio_button_toggled(checked, cn, sn))
                 if current_stage is None:
                     rb_no_stage.setChecked(True)

                 for stage_idx, stage_name in enumerate(stages_names):
                     stage_col_index = col_no_stage + 1 + stage_idx
                     rb_stage = QRadioButton()
                     widget_stage = QWidget()
                     layout_stage = QHBoxLayout(widget_stage)
                     layout_stage.addWidget(rb_stage)
                     layout_stage.setAlignment(Qt.AlignmentFlag.AlignCenter)
                     layout_stage.setContentsMargins(0,0,0,0)
                     widget_stage.setLayout(layout_stage)
                     self.stages_table.setCellWidget(row, stage_col_index, widget_stage)
                     row_button_group.addButton(rb_stage)
                     rb_stage.toggled.connect(lambda checked, cn=cp_name, sn=stage_name:
                                              self._on_stage_radio_button_toggled(checked, cn, sn))
                     if stage_name == current_stage:
                         rb_stage.setChecked(True)


        # ---> НАЧАЛО ИЗМЕНЕНИЯ: Настройка ширины столбцов <---
        header = self.stages_table.horizontalHeader()

        # Устанавливаем режим изменения размера для всех (Interactive позволяет ручное изменение)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        # ---> ИЗМЕНЕНИЕ: Устанавливаем режим Interactive для первого столбца (Название КП)
        # И затем подгоняем его ширину по содержимому ОДИН РАЗ
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.stages_table.resizeColumnToContents(0)
        # Опционально: можно установить минимальную ширину, если подгонка дает слишком мало
        # current_width = header.sectionSize(0)
        # header.resizeSection(0, max(120, current_width)) # Например, минимум 120px

        # Устанавливаем УЗКУЮ ширину для колонок с радиокнопками
        fixed_radio_width = 70 # Ширина в пикселях (можно подстроить)
        for col in range(1, self.stages_table.columnCount()):
            # Устанавливаем начальную ширину
            self.stages_table.setColumnWidth(col, fixed_radio_width)
            # Оставляем режим Interactive, чтобы пользователь мог изменить, если захочет
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)


        # Отключаем растягивание последнего столбца
        header.setStretchLastSection(False)
        # ---> КОНЕЦ ИЗМЕНЕНИЯ <---

############## УЧАСТНИКИ ##############

    def setup_members_tab(self):
        """Вкладка участников"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Верхняя панель с кнопками
        button_panel = QWidget()
        button_layout = QHBoxLayout(button_panel)
        
        # Создаем кнопки
        self.btn_load_xlsx = QPushButton("Загрузить xlsx")
        self.btn_load_xlsx.clicked.connect(self.import_from_excel)
        button_layout.addWidget(self.btn_load_xlsx)
        
        self.btn_edit_crew = QPushButton("Редактировать экипаж")
        self.btn_edit_crew.clicked.connect(self.show_edit_crew_dialog)  # Изменено подключение
        button_layout.addWidget(self.btn_edit_crew)
        
        self.btn_toggle_view = QPushButton("Полная таблица")
        self.btn_toggle_view.setCheckable(True)
        self.btn_toggle_view.clicked.connect(self.toggle_members_view)
        button_layout.addWidget(self.btn_toggle_view)
        
        self.btn_add_crew = QPushButton("Добавить экипаж")
        self.btn_add_crew.clicked.connect(self.show_add_crew_dialog)
        button_layout.addWidget(self.btn_add_crew)
        
        self.btn_delete_crew = QPushButton("Удалить экипаж")
        self.btn_delete_crew.clicked.connect(self.show_delete_crew_dialog)  # Изменено подключение
        button_layout.addWidget(self.btn_delete_crew)
        
        layout.addWidget(button_panel)

        # Создаем контейнер с прокруткой для таблицы
        table_container = QScrollArea()
        table_container.setWidgetResizable(True)
        
        # Создаем таблицу внутри контейнера (ТОЛЬКО ОДИН РАЗ)
        self.members_table = QTableWidget()
        self.members_table.setSizeAdjustPolicy(QTableWidget.SizeAdjustPolicy.AdjustToContents)
        self.members_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # Настройки прокрутки
        self.members_table.setVerticalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        self.members_table.setHorizontalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        
        # Добавляем таблицу в контейнер
        table_container.setWidget(self.members_table)
        layout.addWidget(table_container)
        
        # По умолчанию показываем сокращенную таблицу
        self.is_full_view = False
        self.setup_short_table_headers()
        
        self.members_table.setSortingEnabled(True)  # Разрешаем сортировку по клику на заголовок

        # Заполняем таблицу данными
        if hasattr(self, 'data') and 'members' in self.data:
            self.update_members_table()
        
        # УДАЛЕНО: layout.addWidget(self.members_table) - таблица уже добавлена через контейнер
        
        self.tabs.addTab(tab, "Участники")

    def setup_short_table_headers(self):
        """Устанавливаем заголовки для сокращенного вида"""
        headers = ["Номер", "Зачет", "Пилот", "Штурман", "Авто", "Гос. номер"]
        self.members_table.setColumnCount(len(headers))
        self.members_table.setHorizontalHeaderLabels(headers)
        self.is_full_view = False
        self.btn_toggle_view.setText("Полная таблица")
    
        # Настраиваем ширину столбцов
        self.members_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.members_table.horizontalHeader().setStretchLastSection(True)

    def setup_full_table_headers(self):
        """Устанавливаем заголовки для полного вида"""
        headers = [
            "Адрес электронной почты",
            "Капитан",
            "Субъект",
            "Номер",
            "Пилот",
            "Пол пилота",
            "Дата рождения пилота",
            "Контактный телефон пилота",
            "Штурман",
            "Пол штурмана",
            "Дата рождения штурмана",
            "Контактный телефон штурмана",
            "Пассажиры",
            "Общее количество",
            "До 18 лет",
            "Авто",
            "Гос. номер",
            "Привод автомобиля",
            "Зачет"
        ]
        self.members_table.setColumnCount(len(headers))
        self.members_table.setHorizontalHeaderLabels(headers)
        self.is_full_view = True
        self.btn_toggle_view.setText("Сокращенная таблица")

        # Настраиваем ширину столбцов
        self.members_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.members_table.horizontalHeader().setStretchLastSection(False)
        
        # Устанавливаем минимальную ширину для некоторых колонок (опционально)
        for col in [3, 4, 8, 15, 16]:  # Номера колонок, которые нужно сделать шире
            self.members_table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            self.members_table.setColumnWidth(col, 120)  # Ширина в пикселях

    def toggle_members_view(self):
        """Переключение между полной и сокращенной таблицей"""
        if self.is_full_view:
            self.setup_short_table_headers()
        else:
            self.setup_full_table_headers()
        self.update_members_table()

    def update_members_table(self):
        """Обновление данных в таблице участников (универсальный для обоих форматов)"""
        if not hasattr(self, 'data') or 'members' not in self.data:
            return
        
        self.members_table.setRowCount(len(self.data['members']))
        
        for row, member in enumerate(self.data['members']):
            if self.is_full_view:
                # Заполняем полную таблицу (универсально для обоих форматов)
                self.members_table.setItem(row, 0, QTableWidgetItem(member.get('адрес_электронной_почты') or member.get('email', '')))
                self.members_table.setItem(row, 1, QTableWidgetItem(member.get('капитан') or member.get('captain', '')))
                self.members_table.setItem(row, 2, QTableWidgetItem(member.get('субъект') or member.get('subject', '')))
                self.members_table.setItem(row, 3, QTableWidgetItem(str(member.get('номер') or member.get('number', ''))))
                self.members_table.setItem(row, 4, QTableWidgetItem(member.get('пилот') or member.get('driver', '')))
                self.members_table.setItem(row, 5, QTableWidgetItem(member.get('пол_пилота') or member.get('driver_gender', '')))
                self.members_table.setItem(row, 6, QTableWidgetItem(member.get('дата_рождения_пилота') or member.get('driver_birth', '')))
                self.members_table.setItem(row, 7, QTableWidgetItem(member.get('контактный_телефон_пилота') or member.get('phone', '')))
                self.members_table.setItem(row, 8, QTableWidgetItem(member.get('штурман') or member.get('navigator', '')))
                self.members_table.setItem(row, 9, QTableWidgetItem(member.get('пол_штурмана') or member.get('navigator_gender', '')))
                self.members_table.setItem(row, 10, QTableWidgetItem(member.get('дата_рождения_штурмана') or member.get('navigator_birth', '')))
                self.members_table.setItem(row, 11, QTableWidgetItem(member.get('контактный_телефон_штурмана') or member.get('navigator_phone', '')))
                self.members_table.setItem(row, 12, QTableWidgetItem(member.get('пассажиры') or member.get('passengers', '')))
                self.members_table.setItem(row, 13, QTableWidgetItem(member.get('общее_количество') or member.get('total_count', '')))
                self.members_table.setItem(row, 14, QTableWidgetItem(member.get('до_18_лет') or member.get('under_18', '')))
                self.members_table.setItem(row, 15, QTableWidgetItem(member.get('авто') or member.get('car', '')))
                self.members_table.setItem(row, 16, QTableWidgetItem(member.get('гос.номер') or member.get('plate', '')))
                self.members_table.setItem(row, 17, QTableWidgetItem(member.get('привод_автомобиля') or member.get('drive_type', '')))
                self.members_table.setItem(row, 18, QTableWidgetItem(member.get('зачет') or member.get('classification', '')))
            else:
                # Сокращенная таблица
                self.members_table.setItem(row, 0, QTableWidgetItem(str(member.get('номер') or member.get('number', ''))))
                self.members_table.setItem(row, 1, QTableWidgetItem(member.get('зачет') or member.get('classification', '')))
                self.members_table.setItem(row, 2, QTableWidgetItem(member.get('пилот') or member.get('driver', '')))
                self.members_table.setItem(row, 3, QTableWidgetItem(member.get('штурман') or member.get('navigator', '')))
                self.members_table.setItem(row, 4, QTableWidgetItem(member.get('авто') or member.get('car', '')))
                self.members_table.setItem(row, 5, QTableWidgetItem(member.get('гос.номер') or member.get('plate', '')))
        
        self.members_table.resizeColumnsToContents()
        
        # Устанавливаем минимальную ширину для важных колонок
        if self.is_full_view:
            self.members_table.setColumnWidth(3, 50)   # Номер
            self.members_table.setColumnWidth(4, 150)  # Пилот
            self.members_table.setColumnWidth(8, 150)  # Штурман
        else:
            self.members_table.setColumnWidth(0, 50)   # Номер
            self.members_table.setColumnWidth(2, 150)  # Пилот
            self.members_table.setColumnWidth(3, 150)  # Штурман

    def show_add_crew_dialog(self):
        """Показывает диалог добавления нового экипажа со всеми полями"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить новый экипаж")
        dialog.setMinimumWidth(600)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QVBoxLayout(content)
        
        # Словарь для хранения всех полей ввода
        self.fields = {}
        
        # Обязательные поля (отмечены *)
        basic_fields = [
            "Номер*",
            "Пилот*",
            "Пол пилота",
            "Дата рождения пилота",
            "Контактный телефон пилота*",
            "Штурман",
            "Пол штурмана",
            "Дата рождения штурмана",
            "Контактный телефон штурмана"
        ]
        
        # Добавляем выбор капитана
        layout.addWidget(QLabel("Капитан экипажа:"))
        self.captain_group = QButtonGroup()
        captain_pilot = QRadioButton("Пилот")
        captain_navigator = QRadioButton("Штурман")
        self.captain_group.addButton(captain_pilot)
        self.captain_group.addButton(captain_navigator)
        
        captain_layout = QHBoxLayout()
        captain_layout.addWidget(captain_pilot)
        captain_layout.addWidget(captain_navigator)
        captain_pilot.setChecked(True)  # По умолчанию выбран пилот
        layout.addLayout(captain_layout)
        
        # Остальные поля
        other_fields = [
            "Адрес электронной почты",
            "Субъект",
            "Пассажиры",
            "Общее количество",
            "До 18 лет",
            "Авто*",
            "Гос. номер*",
            "Привод автомобиля"
        ]
        
        # Создаем поля для ввода
        for field in basic_fields + other_fields:
            row = QHBoxLayout()
            label = QLabel(field.replace("*", "") + (":" if not field.endswith("*") else "*:"))
            
            if "Пол" in field or "Привод" in field:
                # Для полей с ограниченным выбором используем комбобоксы
                combo = QComboBox()
                if "Пол" in field:
                    combo.addItems(["Мужской", "Женский"])
                else:  # Привод автомобиля
                    combo.addItems(["Передний", "Задний", "Полный"])
                self.fields[field.replace("*", "")] = combo
                row.addWidget(label)
                row.addWidget(combo)
            elif "Дата" in field:
                # Для дат используем виджет выбора даты
                date_edit = QDateEdit()
                date_edit.setCalendarPopup(True)
                date_edit.setDate(QDate.currentDate())
                self.fields[field] = date_edit
                row.addWidget(label)
                row.addWidget(date_edit)
            else:
                # Обычные текстовые поля
                edit = QLineEdit()
                self.fields[field.replace("*", "")] = edit
                row.addWidget(label)
                row.addWidget(edit)
            
            layout.addLayout(row)
        
        # Добавляем выбор зачета
        layout.addWidget(QLabel("Зачет:"))
        self.classification_group = QButtonGroup()
        classifications = self.get_available_classifications()
        
        classif_layout = QHBoxLayout()
        for classif in classifications:
            rb = QRadioButton(classif)
            self.classification_group.addButton(rb)
            classif_layout.addWidget(rb)
        layout.addLayout(classif_layout)
        
        # Кнопки сохранения/отмены
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(lambda: self.validate_and_add_crew(dialog))
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        scroll.setWidget(content)
        dialog_layout = QVBoxLayout(dialog)
        dialog_layout.addWidget(scroll)
        
        dialog.exec()

    def get_available_classifications(self):
        """Возвращает список доступных зачетов"""
        classifications = []
        if hasattr(self, 'data') and 'params' in self.data:
            if self.data['params'].get('зачет1'):
                classifications.append(self.data['params']['зачет1'])
            if self.data['params'].get('зачет2'):
                classifications.append(self.data['params']['зачет2'])
            if self.data['params'].get('зачет3'):
                classifications.append(self.data['params']['зачет3'])
        return classifications

    def validate_and_add_crew(self, dialog):
        """Проверяет данные и добавляет новый экипаж (в формате как у импортированных)"""
        # Проверяем обязательные поля
        required_fields = [
            "Номер", "Пилот", "Контактный телефон пилота", 
            "Авто", "Гос. номер"
        ]
        
        for field in required_fields:
            if isinstance(self.fields[field], QLineEdit) and not self.fields[field].text():
                QMessageBox.warning(self, "Ошибка", f"Поле '{field}' обязательно для заполнения!")
                return
        
        # Проверяем уникальность номера
        number = self.fields["Номер"].text()
        if any(str(m.get('номер')) == number for m in self.data.get('members', [])):
            QMessageBox.warning(self, "Ошибка", "Экипаж с таким номером уже существует!")
            return
        
        # Получаем выбранный зачет
        selected_classif = None
        for btn in self.classification_group.buttons():
            if btn.isChecked():
                selected_classif = btn.text()
                break
        
        if not selected_classif:
            QMessageBox.warning(self, "Ошибка", "Выберите зачет!")
            return
        
        # Собираем все данные в формате как у импортированных
        new_member = {
            'адрес_электронной_почты': self.fields["Адрес электронной почты"].text(),
            'капитан': "Пилот" if self.captain_group.buttons()[0].isChecked() else "Штурман",
            'субъект': self.fields["Субъект"].text(),
            'номер': number,
            'пилот': self.fields["Пилот"].text(),
            'пол_пилота': self.fields["Пол пилота"].currentText() if isinstance(self.fields["Пол пилота"], QComboBox) else "",
            'дата_рождения_пилота': self.fields["Дата рождения пилота"].date().toString("yyyy-MM-dd") + " 00:00:00",
            'контактный_телефон_пилота': self.fields["Контактный телефон пилота"].text(),
            'штурман': self.fields["Штурман"].text(),
            'пол_штурмана': self.fields["Пол штурмана"].currentText() if isinstance(self.fields["Пол штурмана"], QComboBox) else "",
            'дата_рождения_штурмана': self.fields["Дата рождения штурмана"].date().toString("yyyy-MM-dd") + " 00:00:00",
            'контактный_телефон_штурмана': self.fields["Контактный телефон штурмана"].text(),
            'пассажиры': self.fields["Пассажиры"].text(),
            'общее_количество': self.fields["Общее количество"].text(),
            'до_18_лет': self.fields["До 18 лет"].text(),
            'авто': self.fields["Авто"].text(),
            'гос.номер': self.fields["Гос. номер"].text(),
            'привод_автомобиля': self.fields["Привод автомобиля"].currentText() if isinstance(self.fields["Привод автомобиля"], QComboBox) else "",
            'зачет': selected_classif
        }
        
        # Добавляем в данные
        if 'members' not in self.data:
            self.data['members'] = []
        self.data['members'].append(new_member)
        
        # Сохраняем и обновляем
        self.save_data()
        self.update_members_table()
        dialog.accept()
        QMessageBox.information(self, "Успех", "Экипаж успешно добавлен!")

    def show_delete_crew_dialog(self):
        """Диалог удаления экипажа - версия со штурманом и без редактирования"""
        if not hasattr(self, 'data') or not self.data.get('members'):
            QMessageBox.warning(self, "Ошибка", "Нет экипажей для удаления!")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Удаление экипажа")
        dialog.setMinimumWidth(600) # Немного увеличим ширину для нового столбца

        layout = QVBoxLayout(dialog)

        # Таблица для выбора
        table = QTableWidget()
        # --- ИЗМЕНЕНИЕ: 4 столбца ---
        table.setColumnCount(4)
        # --- ИЗМЕНЕНИЕ: Добавлен заголовок "Штурман" ---
        table.setHorizontalHeaderLabels(["Номер", "Пилот", "Штурман", "Авто"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        # --- ИЗМЕНЕНИЕ: Запрещаем редактирование ячеек таблицы по умолчанию ---
        # Пользователь все равно сможет выбрать строку, но не сможет изменить текст в ячейках
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Заполняем таблицу
        table.setRowCount(len(self.data['members']))
        for row, member in enumerate(self.data['members']):
            # Универсальное получение данных
            number = str(member.get('номер') or member.get('number', ''))
            driver = member.get('пилот') or member.get('driver', '')
            # --- ИЗМЕНЕНИЕ: Получаем штурмана ---
            navigator = member.get('штурман') or member.get('navigator', '')
            car = member.get('авто') or member.get('car', '')

            # Создаем элементы QTableWidgetItem
            item_num = QTableWidgetItem(number)
            item_driver = QTableWidgetItem(driver)
            item_navigator = QTableWidgetItem(navigator) # --- ИЗМЕНЕНИЕ: Создаем элемент для штурмана ---
            item_car = QTableWidgetItem(car)

            # --- ДОБАВЛЕНО: Блокировка редактирования для КАЖДОЙ ячейки ---
            # Мы используем флаги, чтобы убрать возможность редактирования
            # item_num.setFlags(item_num.flags() & ~Qt.ItemFlag.ItemIsEditable)
            # item_driver.setFlags(item_driver.flags() & ~Qt.ItemFlag.ItemIsEditable)
            # item_navigator.setFlags(item_navigator.flags() & ~Qt.ItemFlag.ItemIsEditable)
            # item_car.setFlags(item_car.flags() & ~Qt.ItemFlag.ItemIsEditable)
            # Примечание: Строки выше закомментированы, так как мы уже установили
            # table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers),
            # что является более общим способом запретить редактирование для ВСЕЙ таблицы.
            # Оставляю их здесь как пример для объяснения ниже.

            # Добавляем элементы в таблицу
            table.setItem(row, 0, item_num)
            table.setItem(row, 1, item_driver)
            # --- ИЗМЕНЕНИЕ: Добавляем штурмана в столбец 2 ---
            table.setItem(row, 2, item_navigator)
            # --- ИЗМЕНЕНИЕ: Сдвигаем авто в столбец 3 ---
            table.setItem(row, 3, item_car)

        table.resizeColumnsToContents() # Подгоняем ширину столбцов
        layout.addWidget(table)

        # Кнопки (без изменений)
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        # Используем lambda, чтобы передать выбранную строку в функцию удаления
        btn_box.accepted.connect(lambda: self.delete_crew_by_index(table.currentRow(), dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)

        dialog.exec()

    def delete_crew_by_index(self, row, dialog):
        """Удаляет экипаж по индексу строки в таблице"""
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж для удаления!")
            return
        
        # Двойное подтверждение
        confirm = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите удалить выбранный экипаж?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if confirm == QMessageBox.StandardButton.Yes:
            try:
                # Удаляем по индексу
                if 0 <= row < len(self.data['members']):
                    deleted_crew = self.data['members'].pop(row)
                    
                    # Сохраняем изменения
                    self.save_data()
                    
                    # Обновляем таблицу
                    self.update_members_table()
                    
                    dialog.accept()
                    
                    # Показываем подтверждение
                    number = str(deleted_crew.get('номер') or deleted_crew.get('number', 'Без номера'))
                    QMessageBox.information(
                        self,
                        "Успех",
                        f"Экипаж {number} успешно удален!"
                    )
                else:
                    QMessageBox.warning(self, "Ошибка", "Неверный индекс экипажа!")
                    
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Ошибка",
                    f"Не удалось удалить экипаж: {str(e)}"
                )

    def confirm_crew_deletion(self, dialog):
        """Подтверждает удаление выбранного экипажа"""
        selected_items = self.crew_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж для удаления!")
            return
        
        # Первое подтверждение
        reply = QMessageBox.question(
            self, 'Подтверждение',
            'Вы уверены, что хотите удалить выбранный экипаж?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Второе подтверждение
            reply = QMessageBox.question(
                self, 'Последнее подтверждение',
                'Экипаж будет полностью удален из системы. Продолжить?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.delete_selected_crew(selected_items[0], dialog)

    def delete_selected_crew(self, item, dialog):
        """Удаляет выбранный экипаж"""
        crew_number = item.data(Qt.ItemDataRole.UserRole)
        
        # Находим и удаляем экипаж
        for i, member in enumerate(self.data['members']):
            if str(member.get('number')) == str(crew_number):
                del self.data['members'][i]
                break
        
        # Сохраняем изменения
        self.save_data()
        
        # Обновляем таблицу
        self.update_members_table()
        
        dialog.accept()
        QMessageBox.information(self, "Успех", "Экипаж успешно удален!")

    def import_from_excel(self):
        """Импорт данных из Excel файла"""
        # Открываем диалог выбора файла
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выбете файл с участниками", "", 
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        
        if not file_path:
            return  # Пользователь отменил выбор
        
        try:
            # Загружаем книгу Excel
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Собираем заголовки из Excel (первая строка)
            excel_headers = []
            for col in range(1, 21):  # Колонки с 1 по 20
                excel_headers.append(str(sheet.cell(row=1, column=col).value or ""))
            
            # Сопоставление столбцов (программа -> excel)
            column_mapping = {
                "Адрес электронной почты": 2,
                "Капитан": 3,
                "Субъект": 4,
                "Номер": 5,
                "Пилот": 6,
                "Пол пилота": 7,
                "Дата рождения пилота": 8,
                "Контактный телефон пилота": 9,
                "Штурман": 10,
                "Пол штурмана": 11,
                "Дата рождения штурмана": 12,
                "Контактный телефон штурмана": 13,
                "Пассажиры": 14,
                "Общее количество": 15,
                "До 18 лет": 16,
                "Авто": 17,
                "Гос.номер": 18,
                "Привод автомобиля": 19,
                "Зачет": 20
            }
            
            # Показываем диалог подтверждения соответствия
            if not self.show_column_mapping_dialog(column_mapping, excel_headers):
                return
            
            # Читаем данные из Excel
            new_members = []
            for row in range(2, sheet.max_row + 1):  # Начинаем со второй строки
                member = {}
                for field, col in column_mapping.items():
                    value = sheet.cell(row=row, column=col).value
                    member[field.lower().replace(" ", "_")] = str(value) if value is not None else ""
                
                # Преобразуем специальные поля
                if member.get("капитан"):
                    member["капитан"] = "Пилот" if member["капитан"].lower() == "пилот" else "Штурман"
                
                new_members.append(member)
            
            # Добавляем новых участников
            if 'members' not in self.data:
                self.data['members'] = []
            
            added_count = 0
            if 'members' not in self.data:
                self.data['members'] = []
            self.data['members'].extend(new_members)  # Добавляем всех
            added_count = len(new_members)  # Считаем всех как добавленных
            
            # Сохраняем и обновляем
            self.save_data()
            self.update_members_table()
            
            QMessageBox.information(
                self, "Импорт завершен", 
                f"Успешно добавлено {added_count} новых экипажей!\n"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка импорта", 
                f"Произошла ошибка при чтении файла:\n{str(e)}"
            )

    def show_column_mapping_dialog(self, column_mapping, excel_headers):
        """Показывает диалог с соответствием столбцов"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Проверка соответствия столбцов")
        dialog.setMinimumWidth(600)
        
        layout = QVBoxLayout(dialog)
        
        # Таблица соответствия
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Поле в программе", "Столбец в Excel"])
        table.setRowCount(len(column_mapping))
        
        # Заполняем таблицу
        for row, (field, col) in enumerate(column_mapping.items()):
            table.setItem(row, 0, QTableWidgetItem(field))
            
            excel_col = excel_headers[col-1] if col-1 < len(excel_headers) else "НЕТ ДАННЫХ"
            table.setItem(row, 1, QTableWidgetItem(f"{col} ({excel_col})"))
        
        table.resizeColumnsToContents()
        table.horizontalHeader().setStretchLastSection(True)
        
        layout.addWidget(QLabel("Пожалуйста, проверьте соответствие столбцов:"))
        layout.addWidget(table)
        layout.addWidget(QLabel("Убедитесь, что столбцы в Excel соответствуют указанным номерам"))
        
        # Кнопки подтверждения
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        return dialog.exec() == QDialog.DialogCode.Accepted

    def show_edit_crew_dialog(self):
        """Диалог редактирования экипажа"""
        if not hasattr(self, 'data') or not self.data.get('members'):
            QMessageBox.warning(self, "Ошибка", "Нет экипажей для редактирования!")
            return
        
        # Диалог выбора экипажа (аналогично удалению)
        dialog = QDialog(self)
        dialog.setWindowTitle("Выбор экипажа для редактирования")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Номер", "Пилот", "Штурман"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # Заполняем таблицу
        table.setRowCount(len(self.data['members']))
        for row, member in enumerate(self.data['members']):
            table.setItem(row, 0, QTableWidgetItem(str(member.get('номер', ''))))
            table.setItem(row, 1, QTableWidgetItem(member.get('пилот', '')))
            table.setItem(row, 2, QTableWidgetItem(member.get('штурман', '')))
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.edit_selected_crew(table.currentRow(), dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.exec()

    def edit_selected_crew(self, row, selection_dialog):
        """Редактирование выбранного экипажа со всеми полями"""
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж для редактирования!")
            return
        
        member = self.data['members'][row]
        
        edit_dialog = QDialog(self)
        edit_dialog.setWindowTitle(f"Редактирование экипажа №{member.get('номер', '')}")
        edit_dialog.setMinimumWidth(700)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QVBoxLayout(content)
        
        # Словарь для хранения полей редактирования
        fields = {}

        # Создаем все поля для редактирования
        fields_layout = [
            ("Адрес электронной почты", QLineEdit(member.get('адрес_электронной_почты', ''))),
            ("Субъект", QLineEdit(member.get('субъект', ''))),
            ("Номер", QLineEdit(str(member.get('номер', '')))),
            ("Пилот", QLineEdit(member.get('пилот', ''))),
            ("Контактный телефон пилота", QLineEdit(member.get('контактный_телефон_пилота', ''))),
            ("Штурман", QLineEdit(member.get('штурман', ''))),
            ("Контактный телефон штурмана", QLineEdit(member.get('контактный_телефон_штурмана', ''))),
            ("Пассажиры", QLineEdit(member.get('пассажиры', ''))),
            ("Общее количество", QLineEdit(str(member.get('общее_количество', '')))),
            ("До 18 лет", QLineEdit(str(member.get('до_18_лет', '')))),
            ("Авто", QLineEdit(member.get('авто', ''))),
            ("Гос.номер", QLineEdit(member.get('гос.номер', ''))),
        ]

        # Добавляем основные текстовые поля
        for label, widget in fields_layout:
            row_layout = QHBoxLayout()
            row_layout.addWidget(QLabel(f"{label}:"))
            row_layout.addWidget(widget)
            layout.addLayout(row_layout)
            fields[label.lower().replace(' ', '_')] = widget

        # Поля с выбором из списка
        gender_layout = QHBoxLayout()
        gender_layout.addWidget(QLabel("Пол пилота:"))
        gender_driver = QComboBox()
        gender_driver.addItems(["Мужской", "Женский"])
        gender_driver.setCurrentText(member.get('пол_пилота', 'Мужской'))
        gender_layout.addWidget(gender_driver)
        layout.addLayout(gender_layout)
        fields['пол_пилота'] = gender_driver

        gender_layout = QHBoxLayout()
        gender_layout.addWidget(QLabel("Пол штурмана:"))
        gender_navigator = QComboBox()
        gender_navigator.addItems(["Мужской", "Женский"])
        gender_navigator.setCurrentText(member.get('пол_штурмана', 'Мужской'))
        gender_layout.addWidget(gender_navigator)
        layout.addLayout(gender_layout)
        fields['пол_штурмана'] = gender_navigator

        # Поля с датами
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Дата рождения пилота:"))
        birth_driver = QDateEdit()
        birth_driver.setCalendarPopup(True)
        if member.get('дата_рождения_пилота'):
            birth_driver.setDate(QDate.fromString(member['дата_рождения_пилота'].split()[0], "yyyy-MM-dd"))
        date_layout.addWidget(birth_driver)
        layout.addLayout(date_layout)
        fields['дата_рождения_пилота'] = birth_driver

        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Дата рождения штурмана:"))
        birth_navigator = QDateEdit()
        birth_navigator.setCalendarPopup(True)
        if member.get('дата_рождения_штурмана'):
            birth_navigator.setDate(QDate.fromString(member['дата_рождения_штурмана'].split()[0], "yyyy-MM-dd"))
        date_layout.addWidget(birth_navigator)
        layout.addLayout(date_layout)
        fields['дата_рождения_штурмана'] = birth_navigator

        # Выбор привода автомобиля
        drive_layout = QHBoxLayout()
        drive_layout.addWidget(QLabel("Привод автомобиля:"))
        drive_type = QComboBox()
        drive_type.addItems(["Передний", "Задний", "Полный"])
        drive_type.setCurrentText(member.get('привод_автомобиля', 'Передний'))
        drive_layout.addWidget(drive_type)
        layout.addLayout(drive_layout)
        fields['привод_автомобиля'] = drive_type

        # Капитан экипажа
        captain_layout = QHBoxLayout()
        captain_layout.addWidget(QLabel("Капитан экипажа:"))
        captain_group = QButtonGroup()
        captain_pilot = QRadioButton("Пилот")
        captain_navigator = QRadioButton("Штурман")
        captain_group.addButton(captain_pilot)
        captain_group.addButton(captain_navigator)
        captain_layout.addWidget(captain_pilot)
        captain_layout.addWidget(captain_navigator)
        if member.get('капитан', 'Пилот') == 'Пилот':
            captain_pilot.setChecked(True)
        else:
            captain_navigator.setChecked(True)
        layout.addLayout(captain_layout)
        fields['капитан'] = captain_group

        # Выбор зачета
        classif_layout = QHBoxLayout()
        classif_layout.addWidget(QLabel("Зачет:"))
        classifications = self.get_available_classifications()
        classif_group = QButtonGroup()
        
        classif_buttons_layout = QHBoxLayout()
        for classif in classifications:
            rb = QRadioButton(classif)
            classif_group.addButton(rb)
            classif_buttons_layout.addWidget(rb)
            if member.get('зачет', '') == classif:
                rb.setChecked(True)
        
        classif_layout.addLayout(classif_buttons_layout)
        layout.addLayout(classif_layout)
        fields['зачет'] = classif_group

        # Кнопки сохранения/отмены
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.save_crew_changes(row, fields, edit_dialog, selection_dialog))
        btn_box.rejected.connect(edit_dialog.reject)
        layout.addWidget(btn_box)
        
        scroll.setWidget(content)
        edit_layout = QVBoxLayout(edit_dialog)
        edit_layout.addWidget(scroll)
        
        edit_dialog.exec()

    def save_crew_changes(self, row, fields, edit_dialog, selection_dialog):
        """Сохранение всех измененных полей экипажа"""
        member = self.data['members'][row]
        
        # Обновляем текстовые поля
        text_fields = [
            'адрес_электронной_почты', 'субъект', 'номер', 'пилот', 
            'контактный_телефон_пилота', 'штурман', 'контактный_телефон_штурмана',
            'пассажиры', 'общее_количество', 'до_18_лет', 'авто', 'гос.номер'
        ]
        
        for field in text_fields:
            member[field] = fields[field].text()
        
        # Обновляем поля с выбором
        member['пол_пилота'] = fields['пол_пилота'].currentText()
        member['пол_штурмана'] = fields['пол_штурмана'].currentText()
        member['привод_автомобиля'] = fields['привод_автомобиля'].currentText()
        
        # Обновляем даты
        member['дата_рождения_пилота'] = fields['дата_рождения_пилота'].date().toString("yyyy-MM-dd") + " 00:00:00"
        member['дата_рождения_штурмана'] = fields['дата_рождения_штурмана'].date().toString("yyyy-MM-dd") + " 00:00:00"
        
        # Обновляем капитана и зачет
        member['капитан'] = "Пилот" if fields['капитан'].checkedButton().text() == "Пилот" else "Штурман"
        member['зачет'] = fields['зачет'].checkedButton().text()
        
        # Сохраняем изменения
        self.save_data()
        self.update_members_table()
        
        edit_dialog.accept()
        selection_dialog.accept()
        QMessageBox.information(self, "Сохранено", "Все изменения успешно сохранены!")













############## РЕГИСТРАЦИЯ ##############


    def setup_registration_tab(self):
        """Настройка вкладки регистрации с добавлением штурмана"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Панель поиска
        search_panel = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Введите номер или фамилию и нажмите Enter")
        self.search_input.returnPressed.connect(self.search_crew)
        search_panel.addWidget(self.search_input)
        layout.addLayout(search_panel)
        
        # Разделение на две таблицы
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Таблица незарегистрированных (теперь 5 столбцов)
        self.unregistered_table = QTableWidget()
        self.unregistered_table.setColumnCount(5)
        self.unregistered_table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Штурман", "Авто"])
        self.unregistered_table.doubleClicked.connect(self.show_crew_details)
        self.unregistered_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.unregistered_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Таблица зарегистрированных (теперь 6 столбцов)
        self.registered_table = QTableWidget()
        self.registered_table.setColumnCount(6)
        self.registered_table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Штурман", "Авто", "Время"])
        self.registered_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.registered_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Настройка прокрутки и размеров столбцов
        for table in [self.unregistered_table, self.registered_table]:
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setStretchLastSection(True)
            table.setColumnWidth(3, 150)  # Фиксированная ширина для столбца "Штурман"
        
        splitter.addWidget(self.create_table_group("Незарегистрированные", self.unregistered_table))
        splitter.addWidget(self.create_table_group("Зарегистрированные", self.registered_table))
        layout.addWidget(splitter)
        
        # Загрузка данных
        self.load_unregistered_crews()
        self.load_registered_crews()
        
        self.tabs.addTab(tab, "Регистрация")

    def search_crew(self):
        """Поиск экипажа с учетом штурмана"""
        search_term = self.search_input.text().strip().lower()
        if not search_term:
            self.load_unregistered_crews()
            return
            
        unregistered = [m for m in self.data.get('members', []) if not m.get('registered', False)]
        results = []
        
        for member in unregistered:
            if (search_term in str(member.get('номер', '')).lower() or
                search_term in member.get('пилот', '').lower() or
                search_term in member.get('штурман', '').lower()):
                results.append(member)
        
        self.unregistered_table.setRowCount(0)
        for row, crew in enumerate(results):
            self.unregistered_table.insertRow(row)
            self.unregistered_table.setItem(row, 0, QTableWidgetItem(str(crew.get('номер', ''))))
            self.unregistered_table.setItem(row, 1, QTableWidgetItem(crew.get('зачет', '')))
            self.unregistered_table.setItem(row, 2, QTableWidgetItem(crew.get('пилот', '')))
            self.unregistered_table.setItem(row, 3, QTableWidgetItem(crew.get('штурман', '')))
            self.unregistered_table.setItem(row, 4, QTableWidgetItem(crew.get('авто', '')))

    def create_table_group(self, title, table):
        """Создает группу с заголовком для таблицы"""
        group = QGroupBox(title)
        layout = QVBoxLayout()
        layout.addWidget(table)
        group.setLayout(layout)
        return group

    def on_crew_selected_from_search(self, row, crews, dialog):
        """Обработчик выбора экипажа из результатов поиска"""
        if row >= 0:
            self.show_crew_details(crews[row])
            dialog.accept()

    def save_crew_edits(self, crew_data, fields, dialog):
        """Сохраняет изменения в данных экипажа"""
        try:
            # Текстовые поля
            for field, widget in fields.items():
                if isinstance(widget, QLineEdit):
                    crew_data[field] = widget.text()
                elif isinstance(widget, QComboBox):
                    crew_data[field] = widget.currentText()
                elif isinstance(widget, QDateEdit):
                    crew_data[field] = widget.date().toString("yyyy-MM-dd") + " 00:00:00"
                elif isinstance(widget, QButtonGroup):
                    if field == 'капитан':
                        crew_data[field] = widget.checkedButton().text()
                    elif field == 'зачет':
                        crew_data[field] = widget.checkedButton().text()
            
            # Сохраняем изменения
            self.save_data()
            QMessageBox.information(self, "Сохранено", "Изменения успешно сохранены!")
            dialog.accept()
            return True
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить изменения: {str(e)}")
            return False

    def show_edit_dialog1(self, crew_data):
        """Диалоговое окно редактирования экипажа"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Редактирование экипажа №{crew_data.get('номер', '')}")
        dialog.setMinimumWidth(700)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QVBoxLayout(content)
        
        fields = {}
        
        # Основные текстовые поля
        text_fields = [
            ("Номер", "номер"),
            ("Пилот", "пилот"),
            ("Штурман", "штурман"),
            ("Авто", "авто"),
            ("Гос.номер", "гос.номер"),
            ("Телефон пилота", "контактный_телефон_пилота"),
            ("Телефон штурмана", "контактный_телефон_штурмана"),
            ("Пассажиры", "пассажиры")
        ]
        
        for label, field in text_fields:
            row = QHBoxLayout()
            row.addWidget(QLabel(f"{label}:"))
            edit = QLineEdit(str(crew_data.get(field, '')))
            fields[field] = edit
            row.addWidget(edit)
            layout.addLayout(row)
        
        # Выпадающие списки
        combo_fields = [
            ("Пол пилота", "пол_пилота", ["Мужской", "Женский"]),
            ("Пол штурмана", "пол_штурмана", ["Мужской", "Женский"]),
            ("Привод", "привод_автомобиля", ["Передний", "Задний", "Полный"])
        ]
        
        for label, field, items in combo_fields:
            row = QHBoxLayout()
            row.addWidget(QLabel(f"{label}:"))
            combo = QComboBox()
            combo.addItems(items)
            combo.setCurrentText(crew_data.get(field, items[0]))
            fields[field] = combo
            row.addWidget(combo)
            layout.addLayout(row)
        
        # Даты
        date_fields = [
            ("Дата рождения пилота", "дата_рождения_пилота"),
            ("Дата рождения штурмана", "дата_рождения_штурмана")
        ]
        
        for label, field in date_fields:
            row = QHBoxLayout()
            row.addWidget(QLabel(f"{label}:"))
            date_edit = QDateEdit()
            date_edit.setCalendarPopup(True)
            if crew_data.get(field):
                date_edit.setDate(QDate.fromString(crew_data[field].split()[0], "yyyy-MM-dd"))
            fields[field] = date_edit
            row.addWidget(date_edit)
            layout.addLayout(row)
        
        # Капитан экипажа
        captain_layout = QHBoxLayout()
        captain_layout.addWidget(QLabel("Капитан:"))
        captain_group = QButtonGroup()
        captain_pilot = QRadioButton("Пилот")
        captain_navigator = QRadioButton("Штурман")
        captain_group.addButton(captain_pilot)
        captain_group.addButton(captain_navigator)
        captain_layout.addWidget(captain_pilot)
        captain_layout.addWidget(captain_navigator)
        
        if crew_data.get('капитан', 'Пилот') == 'Пилот':
            captain_pilot.setChecked(True)
        else:
            captain_navigator.setChecked(True)
        fields['капитан'] = captain_group
        layout.addLayout(captain_layout)
        
        # Зачет
        classif_layout = QHBoxLayout()
        classif_layout.addWidget(QLabel("Зачет:"))
        classifications = self.get_available_classifications()
        classif_group = QButtonGroup()
        
        for classif in classifications:
            rb = QRadioButton(classif)
            classif_group.addButton(rb)
            classif_layout.addWidget(rb)
            if crew_data.get('зачет', '') == classif:
                rb.setChecked(True)
        
        fields['зачет'] = classif_group
        layout.addLayout(classif_layout)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.save_crew_edits(crew_data, fields, dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        scroll.setWidget(content)
        dialog_layout = QVBoxLayout(dialog)
        dialog_layout.addWidget(scroll)
        
        btn_box.button(QDialogButtonBox.StandardButton.Ok).clicked.disconnect()
        btn_box.button(QDialogButtonBox.StandardButton.Ok).clicked.connect(
            lambda: self.validate_and_save_edit(crew_data, fields, dialog)
        )
        
        dialog.exec()
        return dialog.result() == QDialog.DialogCode.Accepted

    def edit_crew(self, crew, parent_dialog=None):
        """Редактирование данных экипажа с последующей регистрацией"""
        # Создаем копию для безопасного редактирования
        crew_copy = crew.copy()
        
        # Открываем диалог редактирования
        if not self.show_edit_dialog1(crew_copy):
            return  # Редактирование отменено
        
        # Если данные изменились
        if crew_copy != crew:
            crew.update(crew_copy)
            self.save_data()
            
            # Обновляем отображение
            self.load_unregistered_crews()
            self.load_registered_crews()
            
            # Предлагаем зарегистрировать
            if not crew.get('registered', False):
                self.offer_registration(crew)
        
        if parent_dialog:
            parent_dialog.accept()  

    def offer_print(self, crew):
        """Предлагает распечатать заявку с проверкой"""
        if not crew.get('registered', False):
            return
            
        reply = QMessageBox.question(
            self, 'Печать заявки',
            'Распечатать заявку для экипажа?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.print_crew(crew)

    def print_crew(self, crew):
        """Заглушка для функции печати"""
        QMessageBox.information(
            self, "Печать", 
            f"Здесь будет печать заявки для экипажа №{crew.get('номер', '')}"
        )

    def offer_registration(self, crew):
        """Предлагает зарегистрировать экипаж после редактирования"""
        reply = QMessageBox.question(
            self, 'Регистрация',
            'Хотите зарегистрировать этот экипаж сейчас?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.register_crew(crew)
            
            # После регистрации предлагаем печать
            # self.offer_print(crew)

    def validate_and_save_edit(self, crew_data, fields, dialog):
        """Проверяет данные перед сохранением"""
        # Проверка выбора зачета
        if not fields['зачет'].checkedButton():
            QMessageBox.warning(dialog, "Ошибка", "Не выбран зачет!")
            return
            
        # Проверка обязательных полей
        required_fields = {
            'номер': "Номер экипажа",
            'пилот': "ФИО пилота", 
            'авто': "Марка автомобиля",
            'гос.номер': "Гос. номер"
        }
        
        errors = []
        for field, name in required_fields.items():
            value = fields[field].text() if isinstance(fields[field], QLineEdit) else None
            if not value:
                errors.append(name)
        
        if errors:
            QMessageBox.warning(
                dialog, 
                "Ошибка", 
                f"Не заполнены обязательные поля:\n{', '.join(errors)}"
            )
            return
            
        # Если все проверки пройдены
        self.save_crew_edits(crew_data, fields, dialog)
    
    def register_crew_dialog(self, crew):
        """Диалог регистрации экипажа"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Регистрация экипажа №{crew.get('номер', '')}")
        
        layout = QVBoxLayout(dialog)
        
        # Информация об экипаже
        info = QLabel(
            f"Пилот: {crew.get('пилот', '')}\n"
            f"Штурман: {crew.get('штурман', '')}\n"
            f"Авто: {crew.get('авто', '')}\n"
            f"Гос.номер: {crew.get('гос.номер', '')}"
        )
        layout.addWidget(info)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.register_crew(crew, dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.exec()

    def show_crew_details(self, index):
        """Показывает детальную информацию об экипаже с кнопками действий"""
        table = self.sender()
        selected_row = table.currentRow()
        
        if selected_row < 0:
            return
            
        # Находим экипаж в данных
        number = table.item(selected_row, 0).text()
        crew = None
        for member in self.data["members"]:
            if str(member.get('номер', '')) == number:
                crew = member
                break
                
        if not crew:
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Экипаж №{crew.get('номер', '')}")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        # Отображение информации об экипаже
        info_group = QGroupBox("Информация об экипаже")
        info_layout = QFormLayout()
        
        info_layout.addRow("Номер:", QLabel(str(crew.get('номер', ''))))
        info_layout.addRow("Зачет:", QLabel(crew.get('зачет', '')))
        info_layout.addRow("Пилот:", QLabel(crew.get('пилот', '')))
        info_layout.addRow("Штурман:", QLabel(crew.get('штурман', '')))
        info_layout.addRow("Авто:", QLabel(crew.get('авто', '')))
        info_layout.addRow("Гос.номер:", QLabel(crew.get('гос.номер', '')))
        
        if crew.get('registered', False):
            info_layout.addRow("Статус:", QLabel("Зарегистрирован"))
            info_layout.addRow("Время регистрации:", QLabel(crew.get('registration_time', '')))
        else:
            info_layout.addRow("Статус:", QLabel("Не зарегистрирован"))
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # Кнопки действий
        button_box = QDialogButtonBox()
        
        if not crew.get('registered', False):
            # Для незарегистрированных - кнопки Редактировать и Зарегистрировать
            edit_button = button_box.addButton("Редактировать", QDialogButtonBox.ButtonRole.ActionRole)
            register_button = button_box.addButton("Зарегистрировать", QDialogButtonBox.ButtonRole.AcceptRole)
            register_button.setStyleSheet("background-color: #4CAF50; color: white;")
            
            edit_button.clicked.connect(lambda: self.edit_and_register(crew, dialog))
            register_button.clicked.connect(lambda: self.register_crew(crew, dialog))
        else:
            # Для зарегистрированных - только кнопка Закрыть
            close_button = button_box.addButton("Закрыть", QDialogButtonBox.ButtonRole.RejectRole)
            close_button.clicked.connect(dialog.reject)
        
        cancel_button = button_box.addButton("Отмена", QDialogButtonBox.ButtonRole.RejectRole)
        cancel_button.clicked.connect(dialog.reject)
        
        layout.addWidget(button_box)
        dialog.exec()

    def edit_and_register(self, crew, parent_dialog):
        """Редактирует экипаж и предлагает зарегистрировать"""
        parent_dialog.accept()  # Закрываем текущий диалог
        
        # Редактируем экипаж
        if self.edit_crew(crew):
            # После успешного редактирования предлагаем зарегистрировать
            self.offer_registration(crew)    

    def load_unregistered_crews(self):
        """Загружает список незарегистрированных экипажей с штурманом"""
        self.unregistered_table.setRowCount(0)
        if not hasattr(self, 'data') or 'members' not in self.data:
            return
            
        row = 0
        for member in self.data['members']:
            if not member.get('registered', False):
                self.unregistered_table.insertRow(row)
                self.unregistered_table.setItem(row, 0, QTableWidgetItem(str(member.get('номер', ''))))
                self.unregistered_table.setItem(row, 1, QTableWidgetItem(member.get('зачет', '')))
                self.unregistered_table.setItem(row, 2, QTableWidgetItem(member.get('пилот', '')))
                self.unregistered_table.setItem(row, 3, QTableWidgetItem(member.get('штурман', '')))
                self.unregistered_table.setItem(row, 4, QTableWidgetItem(member.get('авто', '')))
                row += 1

    def load_registered_crews(self):
        """Загружает список зарегистрированных экипажей с штурманом"""
        self.registered_table.setRowCount(0)
        if not hasattr(self, 'data') or 'members' not in self.data:
            return
            
        row = 0
        for member in self.data['members']:
            if member.get('registered', False):
                self.registered_table.insertRow(row)
                self.registered_table.setItem(row, 0, QTableWidgetItem(str(member.get('номер', ''))))
                self.registered_table.setItem(row, 1, QTableWidgetItem(member.get('зачет', '')))
                self.registered_table.setItem(row, 2, QTableWidgetItem(member.get('пилот', '')))
                self.registered_table.setItem(row, 3, QTableWidgetItem(member.get('штурман', '')))
                self.registered_table.setItem(row, 4, QTableWidgetItem(member.get('авто', '')))
                self.registered_table.setItem(row, 5, QTableWidgetItem(member.get('registration_time', '')))
                row += 1

    def register_crew(self, crew, dialog=None):
        """Регистрация экипажа с дополнительными проверками"""
        try:
            # 1. Проверка обязательных полей
            required_fields = {
                'номер': "Номер экипажа",
                'пилот': "ФИО пилота",
                'зачет': "Категория зачета"
            }
            
            missing_fields = [name for field, name in required_fields.items() if not crew.get(field)]
            if missing_fields:
                QMessageBox.warning(
                    self, "Ошибка регистрации",
                    f"Нельзя зарегистрировать - отсутствуют:\n• " + "\n• ".join(missing_fields)
                )
                return False

            # 2. Проверка уникальности номера
            crew_number = str(crew['номер'])
            existing_crew = next(
                (m for m in self.data['members'] 
                if str(m.get('номер')) == crew_number and m.get('registered')),
                None
            )
            
            if existing_crew:
                QMessageBox.warning(
                    self, "Ошибка регистрации",
                    f"Экипаж с номером {crew_number} уже зарегистрирован!\n"
                    f"Пилот: {existing_crew.get('пилот')}\n"
                    f"Время регистрации: {existing_crew.get('registration_time')}"
                )
                return False

            # 3. Проверка минимальных требований к данным
            if not all(crew.get(field) for field in ['авто', 'гос.номер']):
                QMessageBox.warning(
                    self, "Ошибка регистрации",
                    "Для регистрации необходимо указать:\n"
                    "• Марку автомобиля\n"
                    "• Государственный номер"
                )
                return False

            # Устанавливаем параметры регистрации
            crew.update({
                'registered': True,
                'registration_time': QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm")
            })

            # Обновляем данные в списке members
            for i, m in enumerate(self.data['members']):
                if str(m.get('номер')) == crew_number:
                    self.data['members'][i] = crew
                    break

            # Сохраняем данные
            if not self.save_data():
                QMessageBox.critical(self, "Ошибка", "Не удалось сохранить данные в файл!")
                return False

            # Обновляем интерфейс
            self.load_unregistered_crews()
            self.load_registered_crews()

            # Успешное завершение
            if dialog:
                dialog.accept()
                
            QMessageBox.information(
                self, "Успех", 
                f"Экипаж №{crew_number} успешно зарегистрирован!\n"
                f"Пилот: {crew.get('пилот')}\n"
                f"Время: {crew['registration_time']}"
            )
            return True

        except Exception as e:
            error_msg = f"Ошибка регистрации: {str(e)}"
            print(f"[ERROR] {error_msg}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Ошибка", error_msg)
            return False






































############## СТАРТ ##############

    def setup_start_tab(self):
        """Вкладка старта участников с новыми функциями"""
        self.start_tab = QWidget()
        layout = QVBoxLayout(self.start_tab)

        # Верхняя панель управления
        top_panel = QHBoxLayout()
        
        # Таймер до старта
        self.countdown_label = QLabel("До старта: 00:00:00")
        self.countdown_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.countdown_label.setStyleSheet("font-size: 24pt; font-weight: bold;")
        top_panel.addWidget(self.countdown_label)
        
        # Текущее время
        self.current_time_label = QLabel()
        self.current_time_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.current_time_label.setStyleSheet("font-size: 14pt;")
        top_panel.addWidget(self.current_time_label)
        
        # Короткое поле для ручного старта
        self.manual_start_input = QLineEdit()
        self.manual_start_input.setPlaceholderText("№ экипажа")
        self.manual_start_input.setFixedWidth(100)
        self.manual_start_input.returnPressed.connect(self.handle_manual_start)
        top_panel.addWidget(self.manual_start_input)
        
        # Кнопки управления
        self.edit_start_btn = QPushButton("Редактировать старт")
        self.edit_start_btn.clicked.connect(self.show_edit_start_dialog)
        top_panel.addWidget(self.edit_start_btn)
        
        self.cancel_start_btn = QPushButton("Отменить старт")
        self.cancel_start_btn.clicked.connect(self.show_cancel_start_dialog)
        top_panel.addWidget(self.cancel_start_btn)
        
        # Чекбокс блокировки двойного клика
        self.lock_dblclick_cb = QCheckBox("Блокировать старт двойным кликом")
        top_panel.addWidget(self.lock_dblclick_cb)

        # Добавляем кнопку таймера
        timer_btn = QPushButton("Таймер старта")
        timer_btn.clicked.connect(self.show_countdown_window)
        top_panel.addWidget(timer_btn)
        
        layout.addLayout(top_panel)

        

        # Разделение на две таблицы
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Левая таблица: не стартовавшие (5 столбцов)
        self.not_started_table = QTableWidget()
        self.not_started_table.setColumnCount(5)
        self.not_started_table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Штурман", "Авто"])
        self.not_started_table.doubleClicked.connect(self.start_crew_by_dblclick)
        self.not_started_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.not_started_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Правая таблица: стартовавшие (6 столбцов)
        self.started_table = QTableWidget()
        self.started_table.setColumnCount(7)  # Было 6
        self.started_table.setHorizontalHeaderLabels([
            "Номер", "Зачет", "Пилот", "Штурман", 
            "Авто", "Время старта", "Этап"  # Новая колонка
        ])
        self.started_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.started_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Настройка таблиц
        for table in [self.not_started_table, self.started_table]:
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setStretchLastSection(True)
            table.setColumnWidth(3, 150)  # Фиксированная ширина для штурмана

        splitter.addWidget(self.create_table_group("Не стартовавшие", self.not_started_table))
        splitter.addWidget(self.create_table_group("Стартовавшие", self.started_table))
        layout.addWidget(splitter)

        
        self.started_table.setSortingEnabled(True)  # Разрешаем сортировку по клику на заголовок
        self.started_table.sortByColumn(5, Qt.SortOrder.AscendingOrder)  # Сортировка по времени старта по умолчанию

        # Для не стартовавших:
        self.not_started_table.sortByColumn(0, Qt.SortOrder.AscendingOrder)  # Сортировка по номеру


        # Таймеры
        self.start_check_timer = QTimer()
        self.start_check_timer.timeout.connect(self.check_start_time)
        self.start_check_timer.start(1000)
        
        self.current_time_timer = QTimer()
        self.current_time_timer.timeout.connect(self.update_current_time)
        self.current_time_timer.start(1000)

        self.tabs.addTab(self.start_tab, "Старт")
        self.update_start_tab()

    def show_countdown_window(self):
        """Метод для показа окна таймера"""
        try:
            # Получаем параметры из данных
            start_time = self.data["params"]["время_старта"]
            common_start = self.data["logic_params"].get("common_start", False)
            
            # Показываем уведомление
            if common_start:
                msg = "Включен общий старт. Таймер проиграет звуки 1 раз в указанное время."
            else:
                msg = "Общий старт отключен. Таймер будет циклически проигрывать звуки каждую минуту."
                
            QMessageBox.information(self, "Режим таймера", msg)
            
            # Создаем и показываем окно
            self.countdown_window = CountdownWindow(
                self, 
                start_time=start_time,
                common_start=common_start
            )
            self.countdown_window.show()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть таймер: {str(e)}")

    def update_tables(self):
        """Заполняет таблицы данными"""
        self.not_started_table.setRowCount(0)
        self.started_table.setRowCount(0)
        
        for member in self.data.get("members", []):
            if not member.get("registered", False):
                continue  # Пропускаем незарегистрированных
            
            # Получаем данные экипажа
            number = str(member.get("номер", ""))
            classification = member.get("зачет", "")
            driver = member.get("пилот", "")
            navigator = member.get("штурман", "")  # Добавляем штурмана
            car = member.get("авто", "")
            start_time = member.get("start_time", "")

            if member.get("started", False):
                # Стартовавшие
                row = self.started_table.rowCount()
                self.started_table.insertRow(row)
                self.started_table.setItem(row, 0, QTableWidgetItem(number))
                self.started_table.setItem(row, 1, QTableWidgetItem(classification))
                self.started_table.setItem(row, 2, QTableWidgetItem(driver))
                self.started_table.setItem(row, 3, QTableWidgetItem(navigator))  # Штурман
                self.started_table.setItem(row, 4, QTableWidgetItem(car))
                self.started_table.setItem(row, 5, QTableWidgetItem(start_time))
            else:
                # Не стартовавшие
                row = self.not_started_table.rowCount()
                self.not_started_table.insertRow(row)
                self.not_started_table.setItem(row, 0, QTableWidgetItem(number))
                self.not_started_table.setItem(row, 1, QTableWidgetItem(classification))
                self.not_started_table.setItem(row, 2, QTableWidgetItem(driver))
                self.not_started_table.setItem(row, 3, QTableWidgetItem(navigator))  # Штурман
                self.not_started_table.setItem(row, 4, QTableWidgetItem(car))

    def handle_general_start(self):
        if not self.data.get("logic_params", {}).get("common_start", False):
            return
            
        start_time = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
        if QTime.currentTime() >= start_time:
            for member in self.data["members"]:
                if member.get("registered") and not member.get("started"):
                    member["started"] = True
                    member["start_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
            self.save_data()
            self.update_start_tab()  # Принудительное обновление

    def handle_manual_start(self):
        """Обработка ручного ввода номера экипажа"""
        number = self.manual_start_input.text()
        found = False
        for member in self.data["members"]:
            if str(member.get("номер", "")) == number and not member.get("started", False):
                member["started"] = True
                member["start_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
                self.save_data()
                self.show_start_confirmation(member)
                found = True
                break
        
        if not found:
            QMessageBox.warning(self, "Ошибка", "Экипаж с таким номером не найден или уже стартовал!")
        self.manual_start_input.clear()
        self.update_start_tab()

        if not member.get("registered", False):
            QMessageBox.warning(self, "Ошибка", "Экипаж не зарегистрирован!")
            return

    def show_start_confirmation(self, member):
        """Показывает окно подтверждения старта"""
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText(
            f"Экипаж №{member['номер']} стартовал!\n"
            f"Пилот: {member['пилот']}\n"
            f"Авто: {member['авто']}\n"
            f"Время: {member['start_time']}"
        )
        msg.setWindowTitle("Старт зарегистрирован")
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()
        # Автоматическое закрытие через 3 секунды
        QTimer.singleShot(1000, msg.close)

    def show_general_start_notification(self, count, time):
        """Показывает уведомление об автоматическом старте"""
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText(
            f"Автоматический старт {count} экипажей!\n"
            f"Время: {time}\n\n"
            f"Таблицы будут обновлены..."
        )
        msg.setWindowTitle("Общий старт")
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def start_crew_by_dblclick(self, index):
        """Старт экипажа по двойному клику (если не заблокировано)"""
        if self.lock_dblclick_cb.isChecked():
            return
            
        row = index.row()
        if row >= 0:
            number = self.not_started_table.item(row, 0).text()
            for member in self.data["members"]:
                if str(member.get('номер', '')) == number and not member.get('started', False):
                    self.register_crew_start(member)
                    break

    def register_crew_start(self, member):
        """Регистрирует старт экипажа"""
        try:
            member["started"] = True
            member["start_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
            
            # Если включена этапность - устанавливаем первый этап
            if self.data.get("logic_params", {}).get("staged", False):
                member["current_stage"] = 1
                member["stage_history"] = [{
                    "stage": 1,
                    "start_time": member["start_time"]
                }]
            
            self.save_data()
            
            # Показываем подтверждение
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setText(
                f"Экипаж №{member['номер']} стартовал!\n"
                f"Пилот: {member['пилот']}\n"
                f"Штурман: {member['штурман']}\n"
                f"Авто: {member['авто']}\n"
                f"Время: {member['start_time']}"
            )
            msg.setWindowTitle("Старт зарегистрирован")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()
            
            self.update_start_tab()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось зарегистрировать старт: {str(e)}")

    def apply_start_time_changes(self, row, crews, dialog):
        """Применяет изменения времени старта"""
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж!")
            return
            
        new_time = self.new_start_time_edit.time().toString("HH:mm:ss")
        crew_number = str(crews[row].get('номер', ''))
        
        for member in self.data["members"]:
            if str(member.get('номер', '')) == crew_number:
                old_time = member['start_time']
                member['start_time'] = new_time
                self.save_data()
                self.update_start_tab()
                dialog.accept()
                
                # Показываем подтверждение с информацией об экипаже
                QMessageBox.information(
                    self, 
                    "Сохранено", 
                    f"Время старта обновлено для экипажа №{crew_number}:\n"
                    f"Пилот: {member['пилот']}\n"
                    f"Штурман: {member['штурман']}\n"
                    f"Старое время: {old_time}\n"
                    f"Новое время: {new_time}"
                )
                return

    def show_cancel_start_dialog(self):
        """Диалог отмены старта"""
        if not hasattr(self, 'data') or not any(m.get('started') for m in self.data.get('members', [])):
            QMessageBox.warning(self, "Ошибка", "Нет стартовавших экипажей для отмены!")
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle("Отмена старта экипажа")
        dialog.setMinimumWidth(600)
        
        layout = QVBoxLayout(dialog)
        
        # Таблица стартовавших экипажей
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Штурман", "Авто", "Время старта"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Заполняем таблицу
        started_crews = [m for m in self.data.get('members', []) if m.get('started', False)]
        table.setRowCount(len(started_crews))
        
        for row, crew in enumerate(started_crews):
            table.setItem(row, 0, QTableWidgetItem(str(crew.get('номер', ''))))
            table.setItem(row, 1, QTableWidgetItem(crew.get('зачет', '')))
            table.setItem(row, 2, QTableWidgetItem(crew.get('пилот', '')))
            table.setItem(row, 3, QTableWidgetItem(crew.get('штурман', '')))
            table.setItem(row, 4, QTableWidgetItem(crew.get('авто', '')))
            table.setItem(row, 5, QTableWidgetItem(crew.get('start_time', '')))
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.cancel_crew_start(table.currentRow(), started_crews, dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.exec()

    def cancel_crew_start(self, row, crews, dialog):
        """Отменяет старт экипажа"""
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж!")
            return
            
        crew_number = str(crews[row].get('номер', ''))
        
        for member in self.data["members"]:
            if str(member.get('номер', '')) == crew_number:
                member['started'] = False
                old_time = member['start_time']
                member['start_time'] = ""
                self.save_data()
                self.update_start_tab()
                dialog.accept()
                
                # Показываем подтверждение с информацией об экипаже
                QMessageBox.information(
                    self, 
                    "Старт отменен", 
                    f"Старт отменен для экипажа №{crew_number}:\n"
                    f"Пилот: {member['пилот']}\n"
                    f"Штурман: {member['штурман']}\n"
                    f"Время старта: {old_time}"
                )
                return

    def update_countdown(self):
        """Обновляет отсчет до старта (может показывать отрицательные значения)"""
        if "params" in self.data:
            start_time = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
            current_time = QTime.currentTime()
            
            if start_time.isValid():
                seconds = current_time.secsTo(start_time)
                abs_seconds = abs(seconds)
                
                time_str = f"{'-' if seconds < 0 else ''}{abs_seconds//3600:02d}:{(abs_seconds%3600)//60:02d}:{abs_seconds%60:02d}"
                self.countdown_label.setText(f"До старта: {time_str}")

    def show_edit_start_dialog(self):
        """Диалог редактирования времени старта (точная копия окна финиша)"""
        try:
            if not hasattr(self, 'data') or not any(m.get('started') for m in self.data.get('members', [])):
                QMessageBox.warning(self, "Ошибка", "Нет стартовавших экипажей для редактирования!")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle("Редактирование старта")
            dialog.setFixedSize(600, 400)  # Фиксированный размер как на скриншоте
            
            layout = QVBoxLayout(dialog)
            
            # Таблица с экипажами (как на скриншоте)
            table = QTableWidget()
            table.setColumnCount(5)
            table.setHorizontalHeaderLabels(["Номер", "Пилот", "Авто", "Время старта", "Выбрать"])
            table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            
            # Заполняем таблицу
            started_crews = [m for m in self.data["members"] if m.get("started")]
            table.setRowCount(len(started_crews))
            
            for row, member in enumerate(started_crews):
                table.setItem(row, 0, QTableWidgetItem(str(member.get("номер", ""))))
                table.setItem(row, 1, QTableWidgetItem(member.get("пилот", "")))
                table.setItem(row, 2, QTableWidgetItem(member.get("авто", "")))
                table.setItem(row, 3, QTableWidgetItem(member.get("start_time", "")))
                
                # Кнопка выбора
                btn = QPushButton("Выбрать")
                btn.clicked.connect(lambda _, m=member: self.show_edit_selected_start(m, dialog))
                table.setCellWidget(row, 4, btn)
            
            table.resizeColumnsToContents()
            layout.addWidget(table)

            # Информация о выбранном экипаже (как на скриншоте)
            self.selected_crew_info = QLabel("Выберите экипаж из таблицы")
            self.selected_crew_info.setWordWrap(True)
            layout.addWidget(self.selected_crew_info)

            # Поле для ввода времени (как на скриншоте)
            time_layout = QHBoxLayout()
            time_layout.addWidget(QLabel("Новое время старта:"))
            
            self.new_start_time_edit = QTimeEdit()
            self.new_start_time_edit.setDisplayFormat("HH:mm:ss")
            self.new_start_time_edit.setEnabled(False)  # Пока не выбран экипаж
            time_layout.addWidget(self.new_start_time_edit)
            
            layout.addLayout(time_layout)

            # Кнопки OK/Cancel (как на скриншоте)
            btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            btn_box.accepted.connect(lambda: self.apply_start_time(dialog))
            btn_box.rejected.connect(dialog.reject)
            layout.addWidget(btn_box)

            # Подключаем выбор экипажа
            table.itemSelectionChanged.connect(lambda: self.update_start_crew_info(table, started_crews))
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть редактирование: {str(e)}")

    def update_start_crew_info(self, table, crews):
        """Обновляет информацию о выбранном экипаже"""
        row = table.currentRow()
        if row >= 0 and row < len(crews):
            member = crews[row]
            self.selected_crew_info.setText(
                f"Редактирование старта экипажа #{member['номер']}\n"
                f"Экипаж №{member['номер']}\n"
                f"Пилот: {member.get('пилот', '')}\n"
                f"Авто: {member.get('авто', '')}\n"
                f"Текущее время старта: {member.get('start_time', '')}"
            )
            self.new_start_time_edit.setEnabled(True)
            self.new_start_time_edit.setTime(QTime.fromString(member.get("start_time", "00:00:00"), "HH:mm:ss"))

    def show_edit_selected_start(self, member, dialog):
        """Обработчик кнопки 'Выбрать'"""
        self.selected_crew_info.setText(
            f"Редактирование старта экипажа #{member['номер']}\n"
            f"Экипаж №{member['номер']}\n"
            f"Пилот: {member.get('пилот', '')}\n"
            f"Авто: {member.get('авто', '')}\n"
            f"Текущее время старта: {member.get('start_time', '')}"
        )
        self.new_start_time_edit.setEnabled(True)
        self.new_start_time_edit.setTime(QTime.fromString(member.get("start_time", "00:00:00"), "HH:mm:ss"))

    def apply_start_time(self, dialog):
        """Применяет изменения времени старта"""
        if not self.new_start_time_edit.isEnabled():
            QMessageBox.warning(dialog, "Ошибка", "Сначала выберите экипаж!")
            return
            
        new_time = self.new_start_time_edit.time().toString("HH:mm:ss")
        crew_number = self.selected_crew_info.text().split('#')[1].split('\n')[0]
        
        for member in self.data["members"]:
            if str(member.get("номер", "")) == crew_number:
                member["start_time"] = new_time
                self.save_data()
                QMessageBox.information(dialog, "Сохранено", "Время старта успешно изменено")
                dialog.accept()
                self.update_start_tab()
                return

    def update_start_countdown(self):
        """Обновляет отсчёт с отрицательными значениями после старта"""
        try:
            if "params" not in self.data:
                return
                
            start_time = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
            current_time = QTime.currentTime()
            
            if not start_time.isValid():
                return
                
            seconds = current_time.secsTo(start_time)
            abs_seconds = abs(seconds)
            
            # Форматируем время с учётом знака
            sign = "-" if seconds < 0 else ""
            time_str = f"{sign}{abs_seconds//3600:02d}:{(abs_seconds%3600)//60:02d}:{abs_seconds%60:02d}"
            
            # Обновляем label с разными стилями для разных состояний
            if seconds > 0:
                # До старта
                self.countdown_label.setText(f"До старта: {time_str}")
                self.countdown_label.setStyleSheet("color: black; font-size: 24pt; font-weight: bold;")
            elif seconds == 0:
                # Момент старта
                self.countdown_label.setText("СТАРТ!")
                self.countdown_label.setStyleSheet("color: red; font-size: 24pt; font-weight: bold;")
            else:
                # После старта
                self.countdown_label.setText(f"После старта: {time_str}")
                self.countdown_label.setStyleSheet("color: blue; font-size: 24pt; font-weight: bold;")
                
        except Exception as e:
            print(f"Ошибка обновления отсчёта: {e}")
    
    def update_current_time(self):
        """Обновляет текущее время"""
        self.current_time_label.setText(QTime.currentTime().toString("HH:mm:ss"))

    def update_start_tab(self):
        """Обновляет данные на вкладке старта с сортировкой по времени старта"""
        self.update_start_countdown()
        
        # Очищаем таблицы
        self.not_started_table.setRowCount(0)
        self.started_table.setRowCount(0)
        
        # Собираем данные для сортировки
        started_crews = []
        not_started_crews = []
        
        for member in self.data.get("members", []):
            if not member.get("registered", False):
                continue  # Пропускаем незарегистрированных
                
            crew_data = {
                "number": str(member.get("номер", "")),
                "classification": member.get("зачет", ""),
                "driver": member.get("пилот", ""),
                "navigator": member.get("штурман", ""),
                "car": member.get("авто", ""),
                "start_time": member.get("start_time", ""),
                "started": member.get("started", False),
                "stage": member.get("current_stage", "N/A") if self.data.get("logic_params", {}).get("staged", False) else "-"
            }
            
            if crew_data["started"]:
                started_crews.append(crew_data)
            else:
                not_started_crews.append(crew_data)
        
        # Сортируем стартовавших по времени старта
        started_crews.sort(key=lambda x: (
            # Сначала сортируем по времени старта
            QTime.fromString(x["start_time"], "HH:mm:ss") if x["start_time"] else QTime(23, 59, 59),
            # Затем по номеру экипажа
            int(x["number"]) if x["number"].isdigit() else float('inf')
        ))
        
        # Заполняем таблицу стартовавших (отсортированную)
        for row, crew in enumerate(started_crews):
            self.started_table.insertRow(row)
            self.started_table.setItem(row, 0, QTableWidgetItem(crew["number"]))
            self.started_table.setItem(row, 1, QTableWidgetItem(crew["classification"]))
            self.started_table.setItem(row, 2, QTableWidgetItem(crew["driver"]))
            self.started_table.setItem(row, 3, QTableWidgetItem(crew["navigator"]))
            self.started_table.setItem(row, 4, QTableWidgetItem(crew["car"]))
            self.started_table.setItem(row, 5, QTableWidgetItem(crew["start_time"]))
            self.started_table.setItem(row, 6, QTableWidgetItem(str(crew["stage"])))
        
        # Заполняем таблицу не стартовавших (сортируем по номеру)
        not_started_crews.sort(key=lambda x: (
            int(x["number"]) if x["number"].isdigit() else float('inf'),
            x["driver"]
        ))
        
        for row, crew in enumerate(not_started_crews):
            self.not_started_table.insertRow(row)
            self.not_started_table.setItem(row, 0, QTableWidgetItem(crew["number"]))
            self.not_started_table.setItem(row, 1, QTableWidgetItem(crew["classification"]))
            self.not_started_table.setItem(row, 2, QTableWidgetItem(crew["driver"]))
            self.not_started_table.setItem(row, 3, QTableWidgetItem(crew["navigator"]))
            self.not_started_table.setItem(row, 4, QTableWidgetItem(crew["car"]))

    def execute_general_start(self):
        """Автоматический старт с точным временем"""
        try:
            started_count = 0
            is_staged = self.data.get("logic_params", {}).get("staged", False)
            
            # Получаем точное время старта из параметров (формат HH:mm:ss)
            planned_start = self.data["params"]["время_старта"]
            start_time_str = planned_start if len(planned_start.split(':')) == 3 else f"{planned_start}:00"
            
            for member in self.data["members"]:
                if member.get("registered") and not member.get("started"):
                    member["started"] = True
                    member["start_time"] = start_time_str  # Используем точное время
                    
                    if is_staged:
                        member["current_stage"] = 1
                        member["stage_history"] = [{
                            "stage": 1,
                            "start_time": start_time_str
                        }]
                    
                    started_count += 1
            
            if started_count > 0:
                self.save_data()
                self.update_start_tab()
                self.show_general_start_notification(started_count, start_time_str)
                
        except Exception as e:
            print(f"[ERROR] Ошибка автоматического старта: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось выполнить автоматический старт: {str(e)}")

    def check_start_time(self):
        """Проверка времени с точностью до секунды"""
        try:
            if not self.data.get("logic_params", {}).get("common_start", False):
                return
            
            # Получаем время старта с секундами (формат HH:mm:ss)
            planned_start = self.data["params"]["время_старта"]
            start_time_str = planned_start if len(planned_start.split(':')) == 3 else f"{planned_start}:00"
            start_time = QTime.fromString(start_time_str, "HH:mm:ss")
            
            if QTime.currentTime() >= start_time:
                if not hasattr(self, '_last_auto_start') or self._last_auto_start != start_time_str:
                    self._last_auto_start = start_time_str
                    self.execute_general_start()
                    
        except Exception as e:
            print(f"[ERROR] Ошибка проверки времени: {str(e)}")

    def _tick_start_tab_updates(self):
        """Вызывается каждую секунду таймером start_tab_timer."""
        # Этот метод должен обновлять только то, что меняется каждую секунду
        # на вкладке "Старт", работая с ТЕКУЩИМИ данными self.data.
        # Он НЕ должен вызывать self.load_latest_data()!

        # Обновляем только метку обратного отсчета
        if hasattr(self, 'update_start_countdown'):
             try:
                 # Проверяем, активна ли вкладка "Старт" (опционально, для оптимизации)
                 # current_tab_text = self.tabs.tabText(self.tabs.currentIndex())
                 # if current_tab_text == "Старт":
                 self.update_start_countdown()
             except Exception as e:
                 # print(f"Ошибка при обновлении обратного отсчета старта: {e}")
                 # Можно раскомментировать для отладки, но может засорять консоль
                 pass # Игнорируем ошибки, если метка еще не создана или данные не готовы

        # Обновляем метку текущего времени (если она используется на этой вкладке)
        if hasattr(self, 'current_time_label') and hasattr(self, 'update_current_time'):
             try:
                 self.update_current_time() # Обновляем общую метку времени
             except Exception as e:
                 # print(f"Ошибка при обновлении текущего времени: {e}")
                 pass

        # ОБНОВЛЕНИЕ ТАБЛИЦ каждую секунду обычно НЕ ТРЕБУЕТСЯ и может быть неэффективным.
        # Если вам все же нужно обновлять таблицы каждую секунду (например, если
        # статус старта может измениться автоматически), вызовите здесь:
        # self.update_start_tab() # или только self.update_tables(), если update_start_tab делает лишнее
        # Но лучше обновлять таблицы только при фактическом изменении данных (старт, отмена старта).





















# -*- coding: utf-8 -*-
# Добавьте этот код внутрь вашего класса RaceApp

    # ==========================================================================
    # НОВЫЙ КОД ДЛЯ ВКЛАДКИ "ЭТАПЫ" (v2)
    # ==========================================================================

    def setup_stages_tab_v2(self):
        """Инициализация НОВОЙ вкладки Этапы."""
        self.stages_tab_widget = QWidget() # Используем новое имя виджета
        main_layout = QVBoxLayout(self.stages_tab_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)

        # --- Верхняя панель управления ---
        top_panel = QWidget()
        top_layout = QHBoxLayout(top_panel)
        top_layout.setContentsMargins(0, 0, 0, 0)

        self.stages_crew_input = QLineEdit() # Новое имя
        self.stages_crew_input.setPlaceholderText("Номер экипажа (Enter - вперед)")
        self.stages_crew_input.setFixedWidth(250)
        self.stages_crew_input.returnPressed.connect(self._handle_move_forward_input_v2)
        top_layout.addWidget(self.stages_crew_input)

        move_back_btn = QPushButton("Перевести назад")
        move_back_btn.clicked.connect(self._show_move_back_dialog_v2)
        top_layout.addWidget(move_back_btn)

        edit_time_btn = QPushButton("Изменить время СКП")
        edit_time_btn.clicked.connect(self._show_edit_skp_time_dialog_v2)
        top_layout.addWidget(edit_time_btn)

        top_layout.addStretch()
        main_layout.addWidget(top_panel)

        # --- Область с этапами и СКП ---
        self.stages_scroll_area = QScrollArea() # Новое имя
        self.stages_scroll_area.setWidgetResizable(True)
        self.stages_scroll_area.setFrameShape(QFrame.Shape.StyledPanel) # добавим рамку для наглядности

        self.stages_scroll_content = QWidget() # Новое имя
        # Установим вертикальный размер политики, чтобы избежать излишнего растяжения
        self.stages_scroll_content.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)

        # Используем QVBoxLayout для содержимого скролла
        self.stages_content_layout = QVBoxLayout(self.stages_scroll_content) # Новое имя
        self.stages_content_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.stages_content_layout.setSpacing(10)

        self.stages_scroll_area.setWidget(self.stages_scroll_content)
        main_layout.addWidget(self.stages_scroll_area)

        # --- Таймер обновления UI ---
        self.stages_update_timer = QTimer(self) # Новое имя
        self.stages_update_timer.timeout.connect(self.update_stages_tab_v2)
        self.stages_update_timer.start(1000) # Обновление каждую секунду

        # --- Добавляем вкладку (замените старый индекс, если нужно) ---
        # Найдем индекс старой вкладки "Этапы", если она есть
        old_stages_index = -1
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == "Этапы":
                old_stages_index = i
                break
        
        if old_stages_index != -1:
            self.tabs.removeTab(old_stages_index) # Удаляем старую
            self.tabs.insertTab(old_stages_index, self.stages_tab_widget, "Этапы") # Вставляем новую на то же место
            self.tabs.setCurrentIndex(old_stages_index) # Делаем ее активной
        else:
             # Если старой не было, просто добавляем в конец (или выберите нужный индекс)
             # Например, после вкладки "Старт", которая обычно имеет индекс 6 или 7
             start_tab_index = -1
             for i in range(self.tabs.count()):
                 if self.tabs.tabText(i) == "Старт":
                     start_tab_index = i
                     break
             insert_index = start_tab_index + 1 if start_tab_index != -1 else self.tabs.count()
             self.tabs.insertTab(insert_index, self.stages_tab_widget, "Этапы")


        # Первоначальное обновление
        self.update_stages_tab_v2()

    def _clear_stages_layout_v2(self):
        """Очищает layout от всех виджетов."""
        while self.stages_content_layout.count():
            child = self.stages_content_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

    def update_stages_tab_v2(self):
        """Обновляет отображение вкладки Этапы."""
        if not self.data.get("logic_params", {}).get("staged", False):
            self._clear_stages_layout_v2()
            no_stages_label = QLabel("Режим этапности отключен в настройках Логики.")
            no_stages_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.stages_content_layout.addWidget(no_stages_label)
            return

        # Сохраняем текущую прокрутку
        scroll_value = self.stages_scroll_area.verticalScrollBar().value()

        self._clear_stages_layout_v2()

        try:
            stages_info = self.data["logic_params"].get("stages", [])
            skp_settings = self.data["logic_params"].get("skp_settings", [])
            members = self.data.get("members", [])

            # Сгруппируем экипажи по их текущему местоположению
            crews_location = {} # { "stage_1": [], "skp_1": [], "stage_2": [], ... }
            for member in members:
                if not member.get("started") or member.get("finished"):
                    continue

                current_skp = member.get("current_skp")
                current_stage = member.get("current_stage", 1)

                if current_skp is not None:
                    location_key = f"skp_{current_skp}"
                else:
                    location_key = f"stage_{current_stage}"

                if location_key not in crews_location:
                    crews_location[location_key] = []
                crews_location[location_key].append(member)

            # Отрисовываем этапы и СКП
            for i, stage_info in enumerate(stages_info):
                stage_num = i + 1
                stage_name = stage_info.get('name', f'Этап {stage_num}')

                # --- Виджет Этапа ---
                stage_group = QGroupBox(f"Этап {stage_num}: {stage_name}")
                stage_layout = QVBoxLayout()
                stage_group.setLayout(stage_layout)
                stage_location_key = f"stage_{stage_num}"
                crews_on_stage = sorted(crews_location.get(stage_location_key, []), key=lambda x: int(x.get('номер', 0)))

                if not crews_on_stage:
                    stage_layout.addWidget(QLabel("Нет экипажей"))
                else:
                    for member in crews_on_stage:
                        crew_label = QLabel(f"#{member['номер']} | {member.get('пилот', '')} | {member.get('штурман', '')}")
                        stage_layout.addWidget(crew_label)

                self.stages_content_layout.addWidget(stage_group)

                # --- Виджет СКП (если не последний этап) ---
                if stage_num < len(stages_info):
                    skp_num = stage_num
                    skp_setting = next((s for s in skp_settings if s.get("number") == skp_num), None)

                    if skp_setting:
                        skp_title = f"СКП {skp_num} ({skp_setting.get('open_time','??:??')}-{skp_setting.get('close_time','??:??')})"
                    else:
                         skp_title = f"СКП {skp_num} (Нет настроек!)"

                    skp_group = QGroupBox(skp_title)
                    skp_layout = QVBoxLayout()
                    skp_group.setLayout(skp_layout)
                    skp_location_key = f"skp_{skp_num}"
                    crews_on_skp = sorted(crews_location.get(skp_location_key, []), key=lambda x: int(x.get('номер', 0)))

                    if not crews_on_skp:
                         skp_layout.addWidget(QLabel("Нет экипажей"))
                    else:
                        for member in crews_on_skp:
                            crew_widget = self._create_skp_crew_widget_v2(member, skp_setting)
                            skp_layout.addWidget(crew_widget)

                    self.stages_content_layout.addWidget(skp_group)

            self.stages_content_layout.addStretch() # Добавляем растяжение в конце

        except Exception as e:
            print(f"Ошибка при обновлении вкладки Этапы: {e}")
            traceback.print_exc()
            error_label = QLabel(f"Ошибка отображения данных: {e}")
            self.stages_content_layout.addWidget(error_label)

        # Восстанавливаем прокрутку после обновления
        QTimer.singleShot(0, lambda: self.stages_scroll_area.verticalScrollBar().setValue(scroll_value))

    def _create_skp_crew_widget_v2(self, member, skp_setting):
        """Создает виджет для отображения экипажа на СКП с таймером."""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(5, 2, 5, 2) # Компактные отступы

        # Информация об экипаже
        info_label = QLabel(f"#{member['номер']} | {member.get('пилот', '')}")
        layout.addWidget(info_label, 1) # Растягиваем инфо

        # Виджет нейтрализации
        neutral_widget = self._create_neutralization_widget_v2(member, skp_setting)
        layout.addWidget(neutral_widget)

        return widget

    def _create_neutralization_widget_v2(self, member, skp_setting):
        """Создает виджет с таймером нейтрализации."""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        progress = QProgressBar()
        progress.setFixedHeight(20) # Сделаем чуть компактнее
        progress.setFixedWidth(150)
        progress.setAlignment(Qt.AlignmentFlag.AlignCenter)
        progress.setTextVisible(True) # Показываем текст внутри

        time_label = QLabel() # Используем только для тултипа, текст будет в прогрессбаре
        time_label.setFixedWidth(0) # Скрываем label, но оставляем для toolTip

        if not skp_setting: # Если настроек СКП нет
             progress.setRange(0, 1)
             progress.setValue(0)
             progress.setFormat("Нет данных СКП")
             progress.setStyleSheet("QProgressBar { border: 1px solid grey; border-radius: 5px; text-align: center; } QProgressBar::chunk { background-color: lightgrey; }")
             layout.addWidget(progress)
             return widget

        remaining_seconds, total_limit_seconds, used_seconds = self._calculate_remaining_neutralization_v2(member, skp_setting)

        # Настройка прогресс-бара
        progress.setMaximum(total_limit_seconds if total_limit_seconds > 0 else 1) # Избегаем деления на 0
        progress.setValue(remaining_seconds if remaining_seconds >= 0 else 0)

        # Форматирование текста для прогресс-бара и тултипа
        mins, secs = divmod(remaining_seconds, 60)
        time_text = f"{int(mins)}:{int(secs):02d}"
        progress.setFormat(time_text)

        tooltip_text = f"Лимит: {total_limit_seconds // 60} мин.\n" \
                       f"Использовано: {used_seconds // 60}:{used_seconds % 60:02d}\n" \
                       f"Осталось: {time_text}"
        progress.setToolTip(tooltip_text)
        time_label.setToolTip(tooltip_text) # Дублируем на невидимый label на всякий случай

        # Настройка стилей
        if remaining_seconds <= 0:
             progress_style = "QProgressBar::chunk { background-color: red; }"
             text_color = "color: red; font-weight: bold;"
        elif remaining_seconds < 60: # Меньше минуты
             progress_style = "QProgressBar::chunk { background-color: red; }"
             text_color = "color: red;"
        elif remaining_seconds < 300: # Меньше 5 минут
             progress_style = "QProgressBar::chunk { background-color: orange; }"
             text_color = "color: black;" # На оранжевом лучше черный
        else:
             progress_style = "QProgressBar::chunk { background-color: green; }"
             text_color = "color: white;" # На зеленом лучше белый

        progress.setStyleSheet(f"""
             QProgressBar {{
                 border: 1px solid grey;
                 border-radius: 5px;
                 text-align: center;
                 {text_color} /* Цвет текста */
             }}
             {progress_style}
         """)

        layout.addWidget(progress)
        # layout.addWidget(time_label) # Label больше не нужен для отображения

        return widget

    def _calculate_remaining_neutralization_v2(self, member, skp_setting):
        """
        Рассчитывает оставшееся время нейтрализации.
        Возвращает: (remaining_seconds, total_limit_seconds, used_seconds)
        """
        neutral_type = self.data["logic_params"].get("neutralization_type", "Нет")
        current_skp_num = skp_setting.get("number")
        used_seconds = 0
        total_limit_seconds = 0

        try:
            # --- Определяем лимит времени ---
            if neutral_type == 'Нет':
                total_limit_seconds = 3600 * 24 # Условно бесконечность (сутки)
            elif neutral_type == 'Суммарная':
                total_limit_seconds = self.data["logic_params"].get("total_max_neutral_time", 0) * 60
            elif neutral_type == 'На каждом СКП':
                total_limit_seconds = skp_setting.get("max_neutral_time", 0) * 60
            else:
                 total_limit_seconds = 3600 * 24 # По умолчанию - бесконечность

            # --- Считаем использованное время ---
            if neutral_type == 'Суммарная':
                # Суммируем duration всех ЗАВЕРШЕННЫХ и НЕОТМЕНЕННЫХ СКП
                used_seconds = sum(
                    entry.get('duration', 0)
                    for entry in member.get('skp_entries', [])
                    if entry.get('exit_time') and not entry.get('cancelled', False)
                )
            # Для режима "На каждом СКП" использованное время на предыдущих не важно

            # --- Добавляем время, проведенное на ТЕКУЩЕМ активном СКП ---
            active_entry = next(
                (e for e in reversed(member.get('skp_entries', [])) # Ищем последнюю активную
                 if e.get('skp') == current_skp_num and not e.get('exit_time') and not e.get('cancelled', False)),
                None
            )

            if active_entry:
                entry_time_str = active_entry.get("entry_time")
                if entry_time_str:
                    entry_time = QTime.fromString(entry_time_str, "HH:mm:ss")
                    now_time = QTime.currentTime()
                    elapsed_on_current = entry_time.secsTo(now_time)
                    if elapsed_on_current < 0: # Переход через полночь
                        elapsed_on_current += 24 * 3600

                    if neutral_type == 'Суммарная':
                         used_seconds += elapsed_on_current
                    else: # Для "На каждом СКП" или "Нет"
                         used_seconds = elapsed_on_current # Только время на текущем

            remaining_seconds = max(0, total_limit_seconds - used_seconds)

            return remaining_seconds, total_limit_seconds, used_seconds

        except Exception as e:
            print(f"Ошибка расчета времени нейтрализации для #{member.get('номер')}: {e}")
            return 0, 0, 0 # Возвращаем нули в случае ошибки

    # --- Логика перемещения ---

    def _get_crew_by_input_v2(self, show_error=True):
        """Получает экипаж по номеру из поля ввода self.stages_crew_input."""
        number = self.stages_crew_input.text().strip()
        if not number:
            if show_error:
                QMessageBox.warning(self, "Ошибка", "Введите номер экипажа.")
            return None

        member = next((m for m in self.data.get("members", []) if str(m.get("номер")) == number), None)

        if not member:
            if show_error:
                QMessageBox.warning(self, "Ошибка", f"Экипаж с номером '{number}' не найден.")
            return None

        return member

    def _handle_move_forward_input_v2(self):
        """Обработка нажатия Enter для перемещения вперед."""
        member = self._get_crew_by_input_v2()
        if not member:
            return

        self._move_crew_forward_v2(member)
        self.stages_crew_input.clear() # Очищаем поле после ввода

    def _move_crew_forward_v2(self, member):
        """Перемещает экипаж вперед (на след. СКП или этап)."""
        if not member.get("started"):
            QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} еще не стартовал.")
            return
        if member.get("finished"):
            QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} уже финишировал.")
            return

        stages_count = len(self.data["logic_params"].get("stages", []))
        current_skp = member.get("current_skp")
        current_stage = member.get("current_stage", 1)

        target_type = None
        target_num = None

        if current_skp is not None: # Если на СКП
            if current_skp >= stages_count:
                 QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} уже на последнем СКП.")
                 return
            target_type = "stage"
            target_num = current_skp + 1
        else: # Если на этапе
             if current_stage >= stages_count:
                  # Пытаемся финишировать? Нет, это делает вкладка Финиш.
                  QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} уже на последнем этапе. Финиш на соответствующей вкладке.")
                  return
             target_type = "skp"
             target_num = current_stage # СКП имеет тот же номер, что и этап *перед* ним

        # --- Проверка доступности целевого СКП ---
        if target_type == "skp":
             if not self._can_move_to_skp_v2(target_num):
                 return # Сообщение об ошибке покажется внутри _can_move_to_skp_v2

        # --- Выполнение перемещения ---
        print(f"Перемещение вперед: #{member['номер']} с {('СКП ' + str(current_skp)) if current_skp is not None else ('Этап ' + str(current_stage))} на {target_type} {target_num}")
        self._perform_move_v2(member, target_type, target_num, is_backward_move=False)

    def _show_move_back_dialog_v2(self):
        """Показывает диалог подтверждения и выполняет перемещение назад."""
        member = self._get_crew_by_input_v2()
        if not member:
            return

        if not member.get("started"):
            QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} еще не стартовал.")
            return
        if member.get("finished"):
            QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} уже финишировал, перевод назад невозможен.")
            return

        current_skp = member.get("current_skp")
        current_stage = member.get("current_stage", 1)
        target_type = None
        target_num = None
        from_pos_str = ""
        to_pos_str = ""

        if current_skp is not None: # Если на СКП
            target_type = "stage"
            target_num = current_skp
            from_pos_str = f"СКП {current_skp}"
            to_pos_str = f"Этап {target_num}"
        else: # Если на этапе
            if current_stage <= 1:
                QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} на 1 этапе, перевод назад невозможен.")
                return
            target_type = "skp"
            target_num = current_stage - 1
            from_pos_str = f"Этап {current_stage}"
            to_pos_str = f"СКП {target_num}"

        reply = QMessageBox.question(
             self,
             "Подтверждение",
             f"Перевести экипаж #{member['номер']} ({member.get('пилот', '')})\n"
             f"с '{from_pos_str}' назад на '{to_pos_str}'?\n\n"
             f"Действие отменит запись о посещении '{from_pos_str}'.",
             QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
             QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
             print(f"Перемещение назад: #{member['номер']} с {from_pos_str} на {to_pos_str}")
             self._perform_move_v2(member, target_type, target_num, is_backward_move=True)
             self.stages_crew_input.clear() # Очищаем поле

    def _can_move_to_skp_v2(self, skp_num):
        """Проверяет, можно ли перейти на указанный СКП (открыт ли он)."""
        skp_setting = next((s for s in self.data["logic_params"].get("skp_settings", []) if s.get("number") == skp_num), None)
        if not skp_setting:
            QMessageBox.critical(self, "Ошибка", f"Настройки для СКП {skp_num} не найдены в 'logic_params'.")
            return False

        try:
            open_time = QTime.fromString(skp_setting["open_time"], "HH:mm")
            close_time = QTime.fromString(skp_setting["close_time"], "HH:mm")
            current_time = QTime.currentTime()

            if not open_time.isValid() or not close_time.isValid():
                 raise ValueError("Некорректный формат времени в настройках СКП")

            # Обработка случая, когда время закрытия на следующий день
            is_next_day_closing = close_time < open_time

            if is_next_day_closing:
                # СКП открыт, если текущее время >= времени открытия ИЛИ текущее время < времени закрытия
                if not (current_time >= open_time or current_time < close_time):
                    QMessageBox.warning(self, "СКП закрыт", f"СКП {skp_num} работает с {skp_setting['open_time']} до {skp_setting['close_time']}. Сейчас закрыто.")
                    return False
            else:
                # СКП открыт, если текущее время между временем открытия и закрытия
                if not (current_time >= open_time and current_time < close_time):
                     QMessageBox.warning(self, "СКП закрыт", f"СКП {skp_num} работает с {skp_setting['open_time']} до {skp_setting['close_time']}. Сейчас закрыто.")
                     return False

            return True
        except Exception as e:
             QMessageBox.critical(self, "Ошибка времени СКП", f"Ошибка проверки времени работы СКП {skp_num}: {e}")
             return False

    def _perform_move_v2(self, member, target_type, target_num, is_backward_move=False):
        """Основная функция перемещения экипажа (исправленная)."""
        try:
            current_time_str = QDateTime.currentDateTime().toString("HH:mm:ss")
            crew_num = member.get('номер', '?')
            current_skp = member.get("current_skp")
            current_stage = member.get("current_stage", 1)

            # Определяем текущую позицию для истории
            from_pos = {"type": "skp" if current_skp is not None else "stage",
                        "number": current_skp if current_skp is not None else current_stage}
            # Определяем целевую позицию для истории
            to_pos = {"type": target_type, "number": target_num}

            # Глубокое копирование для возможного отката (хотя полный откат сложен)
            # original_member_state = copy.deepcopy(member) # Потребует import copy

            print(f"Начало perform_move для #{crew_num}. Назад: {is_backward_move}. С {from_pos} НА {to_pos}")

            # Инициализация списков, если их нет
            member.setdefault('skp_entries', [])
            member.setdefault('stage_history', [])

            action = "" # Действие для истории

            if is_backward_move:
                action = "move_back"
                history_entry_to_cancel = None # Запись истории для отмены
                skp_entry_to_cancel = None # Запись СКП для отмены

                if from_pos["type"] == "skp": # Уходим с СКП назад на этап
                    print(f"  Движение назад: с СКП {from_pos['number']} на Этап {target_num}")
                    # Ищем последнюю АКТИВНУЮ запись этого СКП
                    skp_entry_to_cancel = next((e for e in reversed(member['skp_entries']) if e.get('skp') == from_pos['number'] and not e.get('exit_time') and not e.get('cancelled')), None)
                    if skp_entry_to_cancel:
                        skp_entry_to_cancel['cancelled'] = True
                        print(f"    Помечена отмененной запись skp_entries: {skp_entry_to_cancel}")
                        # Ищем запись в ИСТОРИИ о ВХОДЕ на этот СКП
                        history_entry_to_cancel = next((h for h in reversed(member['stage_history']) if h.get('action') == 'enter_skp' and h.get('to') == from_pos and not h.get('cancelled')), None)
                    else:
                        print(f"    ПРЕДУПРЕЖДЕНИЕ: Не найдена активная запись skp_entries для СКП {from_pos['number']} для отмены.")

                    member['current_skp'] = None
                    member['current_stage'] = target_num

                else: # Уходим с Этапа назад на СКП
                    prev_skp_num = target_num
                    print(f"  Движение назад: с Этапа {from_pos['number']} на СКП {prev_skp_num}")
                    # Ищем последнюю ЗАВЕРШЕННУЮ запись предыдущего СКП
                    skp_entry_to_cancel = next((e for e in reversed(member['skp_entries']) if e.get('skp') == prev_skp_num and e.get('exit_time') and not e.get('cancelled')), None)

                    if skp_entry_to_cancel:
                        original_entry_time = skp_entry_to_cancel.get('entry_time') # Запоминаем старое время входа
                        skp_entry_to_cancel['cancelled'] = True
                        print(f"    Помечена отмененной запись skp_entries: {skp_entry_to_cancel}")

                        # Ищем запись в ИСТОРИИ о ВЫХОДЕ с этого СКП
                        history_entry_to_cancel = next((h for h in reversed(member['stage_history']) if h.get('action') in ['manual_move', 'auto_move_timeout', 'auto_move_skp_closed'] and h.get('from') == {'type': 'skp', 'number': prev_skp_num} and not h.get('cancelled')), None)

                        # Создаем НОВУЮ АКТИВНУЮ запись для этого СКП, используя СТАРОЕ время входа
                        new_skp_entry = {
                            "skp": prev_skp_num,
                            "entry_time": original_entry_time if original_entry_time else current_time_str,
                            "exit_time": "", "duration": 0, "cancelled": False
                        }
                        member['skp_entries'].append(new_skp_entry)
                        print(f"    Создана новая активная запись skp_entries: {new_skp_entry}")
                    else:
                        print(f"    ПРЕДУПРЕЖДЕНИЕ: Не найдена завершенная запись skp_entries для СКП {prev_skp_num} для отмены. Создаем новую активную.")
                        # Если не нашли завершенную запись, просто создаем новую активную
                        new_skp_entry = {
                            "skp": prev_skp_num, "entry_time": current_time_str,
                            "exit_time": "", "duration": 0, "cancelled": False
                        }
                        member['skp_entries'].append(new_skp_entry)
                        print(f"    Создана новая активная запись skp_entries: {new_skp_entry}")


                    # Отменяем запись в истории, если нашли
                    if history_entry_to_cancel:
                        history_entry_to_cancel['cancelled'] = True
                        print(f"    Помечена отмененной запись stage_history: {history_entry_to_cancel}")

                    member['current_skp'] = target_num
                    member['current_stage'] = target_num # Этап тот же, что и СКП

            else: # Перемещение вперед
                if from_pos["type"] == "skp": # Уходим с СКП вперед на этап
                    action = "manual_move" # Или другой тип авто-перемещения
                    print(f"  Движение вперед: с СКП {from_pos['number']} на Этап {target_num}")
                    # Завершаем текущий СКП
                    if not self._complete_skp_entry_v2(member, from_pos['number'], current_time_str):
                        print(f"    ПРЕДУПРЕЖДЕНИЕ: Не удалось корректно завершить СКП {from_pos['number']} для #{crew_num}")
                    member['current_skp'] = None
                    member['current_stage'] = target_num
                else: # Уходим с Этапа вперед на СКП
                    action = "enter_skp"
                    print(f"  Движение вперед: с Этапа {from_pos['number']} на СКП {target_num}")
                    # Создаем новую запись для СКП
                    new_skp_entry = {
                        "skp": target_num, "entry_time": current_time_str,
                        "exit_time": "", "duration": 0, "cancelled": False
                    }
                    member['skp_entries'].append(new_skp_entry)
                    print(f"    Создана новая запись skp_entries: {new_skp_entry}")
                    member['current_skp'] = target_num
                    member['current_stage'] = target_num # Этап тот же, что и СКП

            # --- Добавляем запись в историю ---
            # Запись о перемещении назад сама помечается как отмененная, чтобы не влиять на будущие расчеты
            history_entry = {
                "time": current_time_str, "from": from_pos, "to": to_pos,
                "action": action, "cancelled": is_backward_move
            }
            member['stage_history'].append(history_entry)
            print(f"  Добавлена запись stage_history: {history_entry}")

            # --- Сохраняем и обновляем ---
            print(f"  Вызов _save_and_update_v2...")
            if self._save_and_update_v2("Этапы", f"{action} для #{crew_num}"):
                # Сообщение показываем только для ручных действий и ПОСЛЕ сохранения
                 if action in ["move_back", "manual_move"]:
                     QMessageBox.information(self, "Успех", f"Экипаж #{crew_num} перемещен на {target_type} {target_num}.")
                 print(f"Успешное завершение perform_move для #{crew_num}")
                 return True # Возвращаем успех
            else:
                 # Ошибка сохранения уже показана в _save_and_update_v2
                 print(f"Ошибка сохранения при perform_move для #{crew_num}")
                 # Попытка отката (очень упрощенная) - НЕ НАДЕЖНО!
                 # member.update(original_member_state) # Раскомментировать если import copy сделан
                 return False # Возвращаем неуспех

        except Exception as e:
            print(f"КРИТИЧЕСКАЯ ОШИБКА в _perform_move_v2 для #{member.get('номер')}: {e}")
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка перемещения", f"Произошла критическая ошибка: {e}")
            return False # Возвращаем неуспех

    def _complete_skp_entry_v2(self, member, skp_num, exit_time_str):
        """Завершает последнюю активную, неотмененную запись для указанного СКП."""
        try:
            active_entry = next(
                (e for e in reversed(member.get('skp_entries', []))
                 if e.get('skp') == skp_num and not e.get('exit_time') and not e.get('cancelled')),
                None
            )

            if not active_entry:
                print(f"Не найдена активная запись для СКП {skp_num} у экипажа #{member.get('номер')}")
                return False # Нечего завершать

            active_entry['exit_time'] = exit_time_str
            entry_time_str = active_entry.get('entry_time')

            if entry_time_str:
                try:
                    t_in = QTime.fromString(entry_time_str, "HH:mm:ss")
                    t_out = QTime.fromString(exit_time_str, "HH:mm:ss")
                    if t_in.isValid() and t_out.isValid():
                         duration = t_in.secsTo(t_out)
                         active_entry['duration'] = duration if duration >= 0 else duration + 86400 # Учет перехода через полночь
                    else:
                         active_entry['duration'] = 0
                         print(f"Некорректное время входа/выхода для СКП {skp_num} у #{member.get('номер')}")
                except Exception as time_calc_e:
                     active_entry['duration'] = 0
                     print(f"Ошибка расчета duration для СКП {skp_num} у #{member.get('номер')}: {time_calc_e}")
            else:
                 active_entry['duration'] = 0 # Нет времени входа - нет длительности

            print(f"Завершена запись СКП {skp_num} для #{member.get('номер')}: выход {exit_time_str}, длительность {active_entry['duration']}")
            return True
        except Exception as e:
            print(f"Ошибка в _complete_skp_entry_v2: {e}")
            return False

    # --- Редактирование времени СКП ---

    def _show_edit_skp_time_dialog_v2(self):
        """Показывает диалог редактирования времени посещения СКП (разрешено для финишировавших)."""
        member = self._get_crew_by_input_v2()
        if not member:
            return
        if not member.get("started"):
            QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} еще не стартовал.")
            return

        # ---> НАЧАЛО ИЗМЕНЕНИЯ: Блок удален или закомментирован <---
        # if member.get("finished"):
        #     QMessageBox.warning(self, "Ошибка", f"Экипаж #{member['номер']} уже финишировал.")
        #     return
        # ---> КОНЕЦ ИЗМЕНЕНИЯ <---

        # --- Создание диалога ---
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Редактирование времени СКП для экипажа #{member['номер']}")
        dialog.setMinimumSize(700, 500) # Можно немного увеличить высоту из-за возможного предупреждения
        layout = QVBoxLayout(dialog)

        # ---> ДОБАВЛЕНО: Необязательное предупреждение <---
        if member.get("finished"):
            warning_label = QLabel("<b>Внимание:</b> Экипаж уже финишировал. Изменение времени СКП повлияет на расчетное время трассы и итоговый результат.")
            warning_label.setStyleSheet("color: #E65100; padding: 5px; border: 1px solid #FFCC80; border-radius: 3px; background-color: #FFF3E0;") # Оранжевый стиль
            warning_label.setWordWrap(True)
            layout.addWidget(warning_label)
        # ---> КОНЕЦ ДОБАВЛЕНИЯ <---

        # Информация об экипаже
        info_label = QLabel(f"Пилот: {member.get('пилот', '')}\n"
                            f"Штурман: {member.get('штурман', '')}\n"
                            f"Авто: {member.get('авто', '')}")
        layout.addWidget(info_label)

        # Таблица для редактирования
        self.edit_time_table = QTableWidget() # Используем атрибут класса для доступа из других методов
        layout.addWidget(self.edit_time_table)
        self._fill_edit_time_table_v2(member) # Заполняем таблицу

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        # Сохраняем изменения, передаем member и dialog
        button_box.accepted.connect(lambda: self._save_edited_times_v2(member, dialog))
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.exec()

    def _fill_edit_time_table_v2(self, member):
        """Заполняет таблицу в диалоге редактирования времени СКП."""
        self.edit_time_table.clear()
        self.edit_time_table.setRowCount(0)
        self.edit_time_table.setColumnCount(5) # СКП, Вход, Выход, Длит., Действие/Статус
        self.edit_time_table.setHorizontalHeaderLabels(["СКП", "Время входа", "Время выхода", "Длительность", "Статус/Действие"])
        self.edit_time_table.verticalHeader().setVisible(False)

        skp_entries = member.get('skp_entries', [])
        if not skp_entries:
             self.edit_time_table.setRowCount(1)
             no_data_item = QTableWidgetItem("Нет записей о посещении СКП")
             no_data_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
             self.edit_time_table.setSpan(0, 0, 1, 5)
             self.edit_time_table.setItem(0, 0, no_data_item)
             return

        self.edit_time_table.setRowCount(len(skp_entries))
        for row, entry in enumerate(skp_entries):
            is_cancelled = entry.get('cancelled', False)
            skp_num = entry.get('skp', '?')
            entry_time = entry.get('entry_time', '')
            exit_time = entry.get('exit_time', '')
            duration = entry.get('duration', 0)

            # СКП
            item_skp = QTableWidgetItem(f"СКП {skp_num}")
            item_skp.setFlags(item_skp.flags() & ~Qt.ItemFlag.ItemIsEditable)

            # Время входа (редактируемое поле)
            edit_entry = QLineEdit(entry_time)
            edit_entry.setPlaceholderText("ЧЧ:ММ:СС")
            edit_entry.setEnabled(not is_cancelled) # Блокируем, если отменено

            # Время выхода (редактируемое поле)
            edit_exit = QLineEdit(exit_time)
            edit_exit.setPlaceholderText("ЧЧ:ММ:СС")
            edit_exit.setEnabled(not is_cancelled) # Блокируем, если отменено

            # Длительность (нередактируемое поле)
            item_duration = QTableWidgetItem(self.format_time(duration))
            item_duration.setFlags(item_duration.flags() & ~Qt.ItemFlag.ItemIsEditable)
            item_duration.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            # Статус/Действие
            status_widget = QWidget()
            status_layout = QHBoxLayout(status_widget)
            status_layout.setContentsMargins(0,0,0,0)
            status_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

            if is_cancelled:
                status_label = QLabel("Отменено")
                status_label.setStyleSheet("color: gray; font-style: italic;")
                status_layout.addWidget(status_label)
                # Серая заливка для отмененных строк
                item_skp.setBackground(QColor(230, 230, 230))
                item_duration.setBackground(QColor(230, 230, 230))
                edit_entry.setStyleSheet("background-color: #e6e6e6; color: gray;")
                edit_exit.setStyleSheet("background-color: #e6e6e6; color: gray;")
            else:
                delete_btn = QPushButton("Отменить")
                delete_btn.setStyleSheet("color: red;")
                # Передаем индекс строки и сам объект entry в lambda
                delete_btn.clicked.connect(lambda checked, r=row, e=entry: self._confirm_cancel_skp_entry_v2(r, e))
                status_layout.addWidget(delete_btn)

            self.edit_time_table.setItem(row, 0, item_skp)
            self.edit_time_table.setCellWidget(row, 1, edit_entry)
            self.edit_time_table.setCellWidget(row, 2, edit_exit)
            self.edit_time_table.setItem(row, 3, item_duration)
            self.edit_time_table.setCellWidget(row, 4, status_widget)

        self.edit_time_table.resizeColumnsToContents()
        self.edit_time_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.edit_time_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

    def _confirm_cancel_skp_entry_v2(self, row_index, entry_data):
        """Подтверждение отмены (пометки как 'cancelled') записи СКП."""
        skp_num = entry_data.get('skp', '?')
        entry_time = entry_data.get('entry_time', 'Н/Д')
        exit_time = entry_data.get('exit_time', 'Н/Д')

        reply = QMessageBox.question(
             self.edit_time_table, # Родитель - таблица
             "Подтверждение отмены",
             f"Отменить запись о посещении СКП {skp_num}?\n"
             f"Вход: {entry_time}\n"
             f"Выход: {exit_time}\n\n"
             f"Запись останется в истории, но не будет учитываться.",
             QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
             QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
             # Помечаем запись как отмененную *визуально* в таблице диалога
             # Фактическое изменение данных произойдет при сохранении
             entry_data['_marked_for_cancellation'] = True # Временный флаг
             self._mark_row_as_cancelled_in_dialog_v2(row_index)

    def _mark_row_as_cancelled_in_dialog_v2(self, row_index):
         """Визуально помечает строку как отмененную в диалоге редактирования."""
         grey_color = QColor(230, 230, 230)
         for col in [0, 3]: # СКП, Длительность
             item = self.edit_time_table.item(row_index, col)
             if item:
                 item.setBackground(grey_color)

         for col in [1, 2]: # Вход, Выход
             widget = self.edit_time_table.cellWidget(row_index, col)
             if isinstance(widget, QLineEdit):
                 widget.setEnabled(False)
                 widget.setStyleSheet("background-color: #e6e6e6; color: gray;")

         # Заменяем кнопку на статус
         status_widget = QWidget()
         status_layout = QHBoxLayout(status_widget)
         status_layout.setContentsMargins(0,0,0,0)
         status_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
         status_label = QLabel("Отменено")
         status_label.setStyleSheet("color: gray; font-style: italic;")
         status_layout.addWidget(status_label)
         self.edit_time_table.setCellWidget(row_index, 4, status_widget)

    def _save_edited_times_v2(self, member, dialog):
        """Сохраняет изменения времени СКП из диалога, включая отмену."""
        try:
            original_entries = member.get('skp_entries', [])
            if not original_entries and self.edit_time_table.rowCount() == 0:
                 dialog.accept(); return # Нечего сохранять

            if len(original_entries) != self.edit_time_table.rowCount():
                 QMessageBox.critical(dialog, "Ошибка синхронизации", "Несоответствие данных и таблицы!"); return

            new_skp_entries = [] # Собираем новый список
            history_modified = False
            validation_ok = True

            for row in range(self.edit_time_table.rowCount()):
                original_entry = original_entries[row]
                modified_entry = original_entry.copy()

                # --- Проверяем статус отмены по виджету в таблице ---
                status_widget = self.edit_time_table.cellWidget(row, 4)
                is_marked_cancelled = False
                # Ищем QLabel с текстом "Отменено" (надежнее, чем временный флаг)
                if isinstance(status_widget, QWidget) and status_widget.findChild(QLabel) and status_widget.findChild(QLabel).text() == "Отменено":
                    is_marked_cancelled = True

                if is_marked_cancelled:
                    if not modified_entry.get('cancelled'):
                        print(f"Сохранение отмены для СКП {modified_entry.get('skp')} у #{member.get('номер')}")
                        modified_entry['cancelled'] = True
                        # Очищаем время и длительность при отмене
                        modified_entry['entry_time'] = ""
                        modified_entry['exit_time'] = ""
                        modified_entry['duration'] = 0
                        history_modified = True
                        # TODO: Отменить связанные записи в stage_history (важно для консистентности!)
                        # Нужно найти запись в stage_history, соответствующую входу/выходу
                        # с этого СКП и пометить ее 'cancelled': True
                else:
                    # Если не отменено, читаем и валидируем время
                    widget_entry = self.edit_time_table.cellWidget(row, 1)
                    widget_exit = self.edit_time_table.cellWidget(row, 2)
                    if not isinstance(widget_entry, QLineEdit) or not isinstance(widget_exit, QLineEdit):
                        print(f"Предупреждение: Не найдены QLineEdit для строки {row}")
                        validation_ok = False; break

                    new_entry_time = widget_entry.text().strip()
                    new_exit_time = widget_exit.text().strip()

                    # Валидация формата
                    if new_entry_time and not self._validate_time_format_v2(new_entry_time):
                        QMessageBox.warning(dialog, "Ошибка", f"Неверный формат времени входа для СКП {modified_entry.get('skp')} (строка {row+1})."); validation_ok = False; break
                    if new_exit_time and not self._validate_time_format_v2(new_exit_time):
                        QMessageBox.warning(dialog, "Ошибка", f"Неверный формат времени выхода для СКП {modified_entry.get('skp')} (строка {row+1})."); validation_ok = False; break

                    # Валидация: выход >= вход (учитывая полночь)
                    t_in = QTime() # Создаем пустые QTime
                    t_out = QTime()
                    valid_times = False
                    if new_entry_time: t_in = QTime.fromString(new_entry_time, "HH:mm:ss")
                    if new_exit_time: t_out = QTime.fromString(new_exit_time, "HH:mm:ss")
                    if t_in.isValid() and t_out.isValid(): valid_times = True

                    # Считаем длительность, только если оба времени валидны
                    duration = 0
                    if valid_times:
                         secs_diff = t_in.secsTo(t_out)
                         duration = secs_diff if secs_diff >= 0 else secs_diff + 86400 # 24*3600 секунд в сутках
                         # Дополнительная проверка на отрицательную длительность после коррекции полуночи (не должно быть)
                         if duration < 0:
                              print(f"!!! Ошибка логики: Отрицательная длительность ({duration}) после коррекции полуночи.")
                              duration = 0 # Сбрасываем в 0 при ошибке

                    modified_entry['entry_time'] = new_entry_time
                    modified_entry['exit_time'] = new_exit_time
                    modified_entry['duration'] = duration
                    modified_entry['cancelled'] = False # Явно ставим False, если не отменяли

                modified_entry.pop('_marked_for_cancellation', None) # Удаляем временный флаг, если он был
                new_skp_entries.append(modified_entry)
                # --- Конец цикла по строкам ---

            if not validation_ok:
                print("Сохранение отменено из-за ошибки валидации.")
                return

            # Обновляем данные экипажа новым списком
            member['skp_entries'] = new_skp_entries

            # Сохраняем в файл
            if self._save_and_update_v2("Этапы (Ред. СКП)", f"Изменение времени СКП для #{member.get('номер')}"):
                QMessageBox.information(dialog, "Сохранено", "Изменения времени СКП сохранены.")
                dialog.accept()
            else:
                QMessageBox.critical(dialog, "Ошибка сохранения", "Не удалось сохранить изменения времени СКП!")

        except Exception as e:
             print(f"Ошибка сохранения измененного времени СКП: {e}")
             traceback.print_exc()
             QMessageBox.critical(dialog, "Критическая ошибка", f"Ошибка сохранения: {e}")
    # --- Таймеры проверок ---

    def start_stages_timers_v2(self):
        """Запускает таймеры, специфичные для вкладки Этапы."""
        # Таймер проверки закрытия СКП
        self.skp_closing_timer = QTimer(self)
        self.skp_closing_timer.timeout.connect(self._check_skp_closing_timer_tick_v2)
        self.skp_closing_timer.start(1000) # Проверка каждую секунду

        # Таймер проверки таймаута нейтрализации
        self.neutralization_timeout_timer = QTimer(self)
        self.neutralization_timeout_timer.timeout.connect(self._check_neutralization_timeout_timer_tick_v2)
        self.neutralization_timeout_timer.start(1000) # Проверка каждую секунду

    def stop_stages_timers_v2(self):
        """Останавливает таймеры вкладки Этапы."""
        if hasattr(self, 'skp_closing_timer'):
            self.skp_closing_timer.stop()
        if hasattr(self, 'neutralization_timeout_timer'):
            self.neutralization_timeout_timer.stop()
        if hasattr(self, 'stages_update_timer'):
            self.stages_update_timer.stop()


    def _check_skp_closing_timer_tick_v2(self):
        """Проверяет закрытие СКП и автоматически перемещает экипажи."""
        if not self.data.get("logic_params", {}).get("staged", False):
            return # Ничего не делаем, если этапность выключена

        try:
            current_time = QTime.currentTime()
            skp_settings = self.data["logic_params"].get("skp_settings", [])
            members_to_move = {} # { skp_num: { "on_skp": [], "on_prev_stage": [] } }
            needs_update = False

            for skp_setting in skp_settings:
                skp_num = skp_setting.get("number")
                close_time_str = skp_setting.get("close_time")
                if not skp_num or not close_time_str: continue

                close_time = QTime.fromString(close_time_str, "HH:mm")
                if not close_time.isValid(): continue

                # Проверяем, не закрылся ли СКП *только что* (в пределах секунды)
                # Используем abs(), так как таймер может сработать чуть раньше или позже
                if abs(current_time.secsTo(close_time)) < 2: # Даем небольшой люфт в 2 секунды
                    # Проверяем, не обрабатывали ли мы уже это закрытие
                    if not hasattr(self, '_last_processed_closing') or \
                       self._last_processed_closing.get(skp_num) != close_time_str:

                        print(f"СКП {skp_num} закрывается ({close_time_str}). Время сейчас: {current_time.toString('HH:mm:ss')}")
                        members_to_move[skp_num] = {"on_skp": [], "on_prev_stage": []}

                        # Запоминаем, что обработали это время закрытия
                        if not hasattr(self, '_last_processed_closing'):
                            self._last_processed_closing = {}
                        self._last_processed_closing[skp_num] = close_time_str

            # Собираем экипажи для перемещения
            if members_to_move:
                 for member in self.data.get("members", []):
                      if not member.get("started") or member.get("finished"): continue

                      current_skp = member.get("current_skp")
                      current_stage = member.get("current_stage", 1)

                      # Экипаж на закрывающемся СКП?
                      if current_skp is not None and current_skp in members_to_move:
                           members_to_move[current_skp]["on_skp"].append(member)
                           needs_update = True

                      # Экипаж на этапе ПЕРЕД закрывающимся СКП?
                      elif current_skp is None and current_stage in members_to_move:
                           members_to_move[current_stage]["on_prev_stage"].append(member)
                           needs_update = True

            # Выполняем перемещения
            if needs_update:
                moved_crew_details = {} # { skp_num: ["#num1", "#num2"] }
                for skp_num, crews in members_to_move.items():
                    close_time_full_str = QTime.fromString(skp_settings[skp_num-1]["close_time"], "HH:mm").toString("HH:mm:ss")
                    moved_crew_details[skp_num] = []

                    # Перемещаем тех, кто был НА СКП
                    for member in crews["on_skp"]:
                        print(f"Автоперемещение (закрытие СКП {skp_num}): #{member['номер']} с СКП {skp_num} -> Этап {skp_num + 1}")
                        # Завершаем СКП временем закрытия
                        self._complete_skp_entry_v2(member, skp_num, close_time_full_str)
                        # Перемещаем на следующий этап
                        member['current_skp'] = None
                        member['current_stage'] = skp_num + 1
                        # Добавляем историю
                        self._add_stage_history_record_v2(member,
                                                          {"type": "skp", "number": skp_num},
                                                          {"type": "stage", "number": skp_num + 1},
                                                          "auto_move_skp_closed",
                                                          time_str=close_time_full_str)
                        moved_crew_details[skp_num].append(f"#{member['номер']}")

                    # Перемещаем (пропускаем СКП) тех, кто был на ПРЕДЫДУЩЕМ этапе
                    for member in crews["on_prev_stage"]:
                         print(f"Автоперемещение (пропуск СКП {skp_num}): #{member['номер']} с Этапа {skp_num} -> Этап {skp_num + 1}")
                         # Добавляем запись о пропущенном СКП
                         member.setdefault('skp_entries', []).append({
                             "skp": skp_num,
                             "entry_time": close_time_full_str,
                             "exit_time": close_time_full_str,
                             "duration": 0,
                             "cancelled": False,
                             "skipped_due_to_closing": True # Флаг пропуска
                         })
                         # Перемещаем на следующий этап
                         member['current_skp'] = None
                         member['current_stage'] = skp_num + 1
                         # Добавляем историю
                         self._add_stage_history_record_v2(member,
                                                           {"type": "stage", "number": skp_num},
                                                           {"type": "stage", "number": skp_num + 1}, # Переход с этапа на этап
                                                           "auto_skip_skp_closed",
                                                           time_str=close_time_full_str)
                         moved_crew_details[skp_num].append(f"#{member['номер']} (пропуск)")


                # Сохраняем и обновляем UI
                if self._save_and_update_v2("Этапы", "Закрытие СКП"):
                    # Показываем уведомление
                    details = ""
                    for skp_num, crews_list in moved_crew_details.items():
                         if crews_list:
                             details += f"\nСКП {skp_num}: {', '.join(crews_list)}"
                    if details:
                        QMessageBox.information(self, "Закрытие СКП", f"Некоторые СКП закрылись. Перемещенные экипажи:{details}")
                else:
                     QMessageBox.critical(self, "Ошибка сохранения", "Не удалось сохранить изменения после закрытия СКП!")


        except Exception as e:
            print(f"Ошибка проверки закрытия СКП: {e}")
            # traceback.print_exc() # Раскомментировать для детальной отладки

    def _check_neutralization_timeout_timer_tick_v2(self):
        """Проверяет истечение времени нейтрализации."""
        if not self.data.get("logic_params", {}).get("staged", False):
            return

        try:
            crews_to_move = []
            current_time_str = QDateTime.currentDateTime().toString("HH:mm:ss")

            for member in self.data.get("members", []):
                 if not member.get("started") or member.get("finished"): continue

                 current_skp_num = member.get("current_skp")
                 if current_skp_num is None: continue # Проверяем только тех, кто на СКП

                 skp_setting = next((s for s in self.data["logic_params"].get("skp_settings", []) if s.get("number") == current_skp_num), None)
                 if not skp_setting: continue # Нет настроек для этого СКП

                 remaining_seconds, _, _ = self._calculate_remaining_neutralization_v2(member, skp_setting)

                 if remaining_seconds <= 0:
                      # Проверяем, не обрабатывали ли мы уже этот таймаут для этого экипажа
                      last_timeout_key = f"{member['номер']}_{current_skp_num}"
                      if not hasattr(self, '_last_processed_timeout') or \
                         self._last_processed_timeout.get(last_timeout_key) != current_time_str[:5]: # Проверяем с точностью до минуты

                          print(f"Таймаут нейтрализации для #{member['номер']} на СКП {current_skp_num}. Осталось: {remaining_seconds}")
                          crews_to_move.append(member)

                          # Запоминаем обработку
                          if not hasattr(self, '_last_processed_timeout'):
                              self._last_processed_timeout = {}
                          self._last_processed_timeout[last_timeout_key] = current_time_str[:5]


            if crews_to_move:
                 moved_crew_numbers = []
                 for member in crews_to_move:
                      skp_num = member["current_skp"]
                      print(f"Автоперемещение (таймаут нейтрализации): #{member['номер']} с СКП {skp_num} -> Этап {skp_num + 1}")
                      # Завершаем СКП текущим временем
                      self._complete_skp_entry_v2(member, skp_num, current_time_str)
                      # Перемещаем на следующий этап
                      member['current_skp'] = None
                      member['current_stage'] = skp_num + 1
                      # Добавляем историю
                      self._add_stage_history_record_v2(member,
                                                        {"type": "skp", "number": skp_num},
                                                        {"type": "stage", "number": skp_num + 1},
                                                        "auto_move_timeout",
                                                        time_str=current_time_str)
                      moved_crew_numbers.append(f"#{member['номер']}")

                 # Сохраняем и обновляем UI
                 if self._save_and_update_v2("Этапы", "Таймаут нейтрализации"):
                      QMessageBox.warning(self, "Таймаут нейтрализации", f"Время нейтрализации истекло. Перемещенные экипажи:\n{', '.join(moved_crew_numbers)}")
                 else:
                      QMessageBox.critical(self, "Ошибка сохранения", "Не удалось сохранить изменения после таймаута нейтрализации!")

        except Exception as e:
            print(f"Ошибка проверки таймаута нейтрализации: {e}")
            # traceback.print_exc()

    # --- Вспомогательные функции ---

    def _add_stage_history_record_v2(self, member, from_pos, to_pos, action, time_str=None, cancelled=False):
         """Добавляет запись в stage_history."""
         if time_str is None:
             time_str = QDateTime.currentDateTime().toString("HH:mm:ss")

         history_entry = {
             "time": time_str,
             "from": from_pos,
             "to": to_pos,
             "action": action,
             "cancelled": cancelled
         }
         member.setdefault('stage_history', []).append(history_entry)
         # print(f"История #{member.get('номер')}: {history_entry}") # Для отладки

    def _validate_time_format_v2(self, time_str):
         """Проверяет формат времени ЧЧ:ММ:СС."""
         if not time_str: return True # Пустое допустимо
         return re.fullmatch(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]:[0-5][0-9]$', time_str) is not None

    def _save_and_update_v2(self, source_tab_name="?", action_name="?"):
        """Общий метод для сохранения данных и обновления вкладки Этапы с проверкой."""
        print(f"Вызов сохранения из вкладки '{source_tab_name}', действие: '{action_name}'")
        save_successful = False
        try:
            # Вызываем ваш существующий метод сохранения, который должен вернуть True/False
            save_successful = self.save_data()
            if save_successful:
                print(f"Сохранение для '{action_name}' успешно.")
                self.update_stages_tab_v2() # Обновляем UI ТОЛЬКО если сохранение прошло
            else:
                print(f"Ошибка: save_data() вернул False при действии '{action_name}'!")
                QMessageBox.critical(self, "Ошибка Сохранения", f"Не удалось сохранить данные после действия:\n'{action_name}'.\nИзменения не применены.")
        except Exception as e:
            print(f"Исключение во время save_data() или update_stages_tab_v2(): {e}")
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка Сохранения/Обновления", f"Произошла ошибка:\n{e}")
            save_successful = False # Считаем неуспешным при исключении

        return save_successful


# Внутри класса RaceApp



























############## ФИНИШ ##############

    def setup_finish_tab(self):
            """Вкладка финиша участников"""
            self.finish_tab = QWidget()
            layout = QVBoxLayout(self.finish_tab)

            # --- Верхняя часть: обратный отсчет и кнопки ---
            top_layout = QHBoxLayout()

            self.closing_countdown = QLabel("До закрытия трассы: 00:00:00")
            self.closing_countdown.setStyleSheet("font-size: 24pt; font-weight: bold;")
            top_layout.addWidget(self.closing_countdown)

            # Группа кнопок управления
            btn_group = QHBoxLayout()

            self.refresh_btn = QPushButton("Обновить")
            self.refresh_btn.clicked.connect(self.update_finish_tab) # Подключаем к полному обновлению
            btn_group.addWidget(self.refresh_btn)

            self.edit_finish_btn = QPushButton("Редактировать финиш")
            self.edit_finish_btn.clicked.connect(self.show_edit_finish_dialog)
            btn_group.addWidget(self.edit_finish_btn)

            self.cancel_finish_btn = QPushButton("Отменить финиш")
            self.cancel_finish_btn.clicked.connect(self.show_cancel_finish_dialog)
            btn_group.addWidget(self.cancel_finish_btn)

            top_layout.addLayout(btn_group)
            layout.addLayout(top_layout)

            # --- Разделение на две таблицы ---
            splitter = QSplitter(Qt.Orientation.Horizontal)

            # Левая таблица: не финишировавшие (только стартовавшие)
            self.not_finished_table = QTableWidget()
            self.not_finished_table.setColumnCount(6) # Добавляем столбец Нейтрализация
            self.not_finished_table.setHorizontalHeaderLabels(
                ["Номер", "Зачет", "Пилот", "Авто", "Время старта", "Нейтрализация"]
            )
            self.not_finished_table.setSortingEnabled(True) # <--- ВКЛЮЧАЕМ СОРТИРОВКУ
            self.not_finished_table.doubleClicked.connect(self.show_finish_crew_details)
            self.not_finished_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            self.not_finished_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers) # Запрет редактирования ячеек
            self.not_finished_table.horizontalHeader().setSortIndicator(0, Qt.SortOrder.AscendingOrder) # Сортировка по номеру по умолчанию

            # Правая таблица: финишировавшие
            self.finished_table = QTableWidget()
            self.finished_table.setColumnCount(9) # Обновляем количество столбцов
            self.finished_table.setHorizontalHeaderLabels([
                "Номер", "Зачет", "Пилот", "Авто", "Время старта",
                "Время финиша", "Общее время", "Нейтрализация", "Время трассы"
            ])
            self.finished_table.setSortingEnabled(True) # <--- ВКЛЮЧАЕМ СОРТИРОВКУ
            self.finished_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            self.finished_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers) # Запрет редактирования ячеек
            # Сортировка по времени финиша по умолчанию (индекс 5)
            self.finished_table.horizontalHeader().setSortIndicator(5, Qt.SortOrder.AscendingOrder)


            # Настройка внешнего вида таблиц (общая часть)
            for table in [self.not_finished_table, self.finished_table]:
                table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
                table.horizontalHeader().setStretchLastSection(False) # Отключаем растягивание последнего
                table.setColumnWidth(2, 150) # Фиксированная ширина для пилота (пример)
                # table.setColumnWidth(3, 150) # Фиксированная ширина для авто (пример)

            splitter.addWidget(self.create_table_group("На трассе", self.not_finished_table))
            splitter.addWidget(self.create_table_group("Финишировали", self.finished_table))
            layout.addWidget(splitter)

            # --- Поле для ручного финиша ---
            self.manual_finish_group = QGroupBox("Ручной финиш")
            finish_layout = QHBoxLayout()
            self.finish_input = QLineEdit()
            self.finish_input.setPlaceholderText("Введите номер экипажа и нажмите Enter")
            self.finish_input.returnPressed.connect(self.handle_manual_finish)
            finish_layout.addWidget(self.finish_input)
            self.manual_finish_group.setLayout(finish_layout)
            layout.addWidget(self.manual_finish_group)

            # Таймер для проверки закрытия трассы (Один таймер для DNF уже есть в __init__,
            # этот можно удалить или переименовать, если логика track_closing_timer в __init__ делает то же самое)
            # self.finish_timer = QTimer()
            # self.finish_timer.timeout.connect(self.check_track_closing)
            # self.finish_timer.start(1000)

            self.tabs.addTab(self.finish_tab, "Финиш")
            self.update_finish_tab() # Первоначальное заполнение

    def update_finish_tab(self):
        """Обновляет данные на вкладке финиша"""
        self.update_closing_countdown()
        self.update_finish_tables()

    def update_closing_countdown(self):
        """Обновляет отсчет до закрытия трассы"""
        if "params" in self.data:
            closing_time = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            current_time = QTime.currentTime()
            
            if current_time < closing_time:
                remaining = current_time.secsTo(closing_time)
                self.closing_countdown.setText(
                    f"До закрытия трассы: {remaining//3600:02d}:{(remaining%3600)//60:02d}:{remaining%60:02d}"
                )
            else:
                self.closing_countdown.setText("Трасса закрыта!")

    def show_finish_crew_details(self, index):
        """Показывает детали экипажа при двойном клике в таблице финиша"""
        selected_row = self.not_finished_table.currentRow()
        if selected_row >= 0:
            number = self.not_finished_table.item(selected_row, 0).text()
            for member in self.data["members"]:
                if str(member.get("номер", "")) == number:
                    self.show_crew_finish_dialog(member)
                    break

    def show_crew_finish_dialog(self, member):
        """Диалог подтверждения финиша"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Экипаж №{member['номер']}")
        
        layout = QVBoxLayout()
        info = QLabel(
            f"Пилот: {member.get('пилот', '')}\n"
            f"Авто: {member.get('авто', '')}\n"
            f"Старт: {member.get('start_time', '')}"
        )
        layout.addWidget(info)

        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.confirm_finish(member, dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)

        dialog.setLayout(layout)
        dialog.exec()

    def handle_manual_finish(self):
        """Обработчик ручного финиша с проверкой нахождения на СКП"""
        try:
            number = self.finish_input.text().strip()
            if not number:
                return
                
            # Проверяем закрытие трассы
            # closing_time = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            # current_time = QTime.currentTime()
            
            # if current_time >= closing_time:
            #     QMessageBox.warning(self, "Ошибка", "Трасса закрыта! Финиш невозможен")
            #     self.finish_input.clear()
            #     return
                
            for member in self.data["members"]:
                if str(member.get("номер", "")) == number:
                    if not member.get("started"):
                        QMessageBox.warning(self, "Ошибка", "Экипаж еще не стартовал!")
                        self.finish_input.clear()
                        return
                        
                    if member.get("finished"):
                        QMessageBox.warning(self, "Ошибка", "Экипаж уже финишировал!")
                        self.finish_input.clear()
                        return
                        
                    if member.get("current_skp") is not None:
                        skp_num = member["current_skp"]
                        QMessageBox.warning(
                            self,
                            "Экипаж на СКП",
                            f"Экипаж #{number} не может финишировать,\n"
                            f"так как находится на СКП {skp_num}!\n\n"
                            f"Сначала завершите нейтрализацию."
                        )
                        self.finish_input.clear()
                        return
                        
                    # Если все проверки пройдены - финишируем
                    member["finished"] = True
                    member["finish_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
                    self.save_data()
                    
                    QMessageBox.information(
                        self,
                        "Финиш зарегистрирован",
                        f"Экипаж #{number} успешно финишировал!\n"
                        f"Время финиша: {member['finish_time']}"
                    )
                    
                    self.finish_input.clear()
                    self.update_finish_tables()
                    return
                    
            QMessageBox.warning(self, "Ошибка", "Экипаж с таким номером не найден!")
            self.finish_input.clear()
                
        except Exception as e:
            print(f"Ошибка при ручном финише: {e}")
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")
            self.finish_input.clear()

    def confirm_finish(self, member, dialog):
        """Подтверждение финиша с проверкой нахождения на СКП"""
        try:
            # Проверка закрытия трассы
            # closing_time = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            # if QTime.currentTime() >= closing_time:
            #     QMessageBox.warning(self, "Ошибка", "Трасса уже закрыта!")
            #     return

            # Проверка нахождения на СКП
            if member.get("current_skp") is not None:
                QMessageBox.warning(
                    self,
                    "Экипаж на СКП",
                    f"Экипаж #{member['номер']} не может финишировать,\n"
                    f"так как находится на СКП {member['current_skp']}!\n\n"
                    f"Сначала завершите нейтрализацию."
                )
                return

            # Если проверки пройдены - финишируем
            member["finished"] = True
            member["finish_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
            self.save_data()
            
            dialog.accept()
            self.update_finish_tab()
            
            QMessageBox.information(
                self,
                "Финиш зарегистрирован",
                f"Экипаж #{member['номер']} успешно финишировал!\n"
                f"Время финиша: {member['finish_time']}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def show_edit_finish_dialog(self):
        """Диалог редактирования времени финиша"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактирование финиша")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout()
        
        # Таблица с финишировавшими экипажами
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Номер", "Пилот", "Авто", "Время финиша", "Выбрать"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Заполняем таблицу
        finished_crews = [m for m in self.data["members"] if m.get("finished")]
        table.setRowCount(len(finished_crews))
        
        for row, member in enumerate(finished_crews):
            table.setItem(row, 0, QTableWidgetItem(str(member.get("номер", ""))))
            table.setItem(row, 1, QTableWidgetItem(member.get("пилот", "")))
            table.setItem(row, 2, QTableWidgetItem(member.get("авто", "")))
            table.setItem(row, 3, QTableWidgetItem(member.get("finish_time", "")))
            
            # Кнопка выбора
            btn = QPushButton("Выбрать")
            btn.clicked.connect(lambda _, m=member: self.edit_selected_finish(m, dialog))
            table.setCellWidget(row, 4, btn)
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        dialog.setLayout(layout)
        dialog.exec()

    def edit_selected_finish(self, member, parent_dialog):
        """Редактирование времени финиша для выбранного экипажа"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Редактирование финиша экипажа #{member['номер']}")
        
        layout = QVBoxLayout()
        
        # Информация об экипаже
        info = QLabel(
            f"Экипаж №{member['номер']}\n"
            f"Пилот: {member.get('пилот', '')}\n"
            f"Авто: {member.get('авто', '')}\n"
            f"Текущее время финиша: {member.get('finish_time', '')}"
        )
        layout.addWidget(info)
        
        # Поле для ввода нового времени
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel("Новое время финиша:"))
        
        self.new_finish_time = QLineEdit()
        self.new_finish_time.setPlaceholderText("ЧЧ:ММ:СС")
        self.new_finish_time.setText(member.get("finish_time", ""))
        time_layout.addWidget(self.new_finish_time)
        
        layout.addLayout(time_layout)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.apply_finish_edit(member, dialog, parent_dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.setLayout(layout)
        dialog.exec()

    def apply_finish_edit(self, member, dialog, parent_dialog):
        """Применяет изменения времени финиша"""
        new_time = self.new_finish_time.text().strip()
        
        if not self.validate_time_format(new_time):
            QMessageBox.warning(self, "Ошибка", "Неверный формат времени. Используйте ЧЧ:ММ:СС")
            return
            
        member["finish_time"] = new_time
        self.save_data()
        
        dialog.close()
        parent_dialog.close()
        self.update_finish_tab()
        
        QMessageBox.information(self, "Сохранено", "Время финиша успешно изменено")

    def show_cancel_finish_dialog(self):
        """Диалог отмены финиша"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Отмена финиша")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout()
        
        # Таблица с финишировавшими экипажами
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Номер", "Пилот", "Авто", "Время финиша", "Выбрать"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Заполняем таблицу
        finished_crews = [m for m in self.data["members"] if m.get("finished")]
        table.setRowCount(len(finished_crews))
        
        for row, member in enumerate(finished_crews):
            table.setItem(row, 0, QTableWidgetItem(str(member.get("номер", ""))))
            table.setItem(row, 1, QTableWidgetItem(member.get("пилот", "")))
            table.setItem(row, 2, QTableWidgetItem(member.get("авто", "")))
            table.setItem(row, 3, QTableWidgetItem(member.get("finish_time", "")))
            
            # Кнопка выбора
            btn = QPushButton("Выбрать")
            btn.clicked.connect(lambda _, m=member: self.cancel_selected_finish(m, dialog))
            table.setCellWidget(row, 4, btn)
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        dialog.setLayout(layout)
        dialog.exec()

    def cancel_selected_finish(self, member, dialog):
        """Отменяет финиш для выбранного экипажа"""
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите отменить финиш экипажа #{member['номер']}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            member["finished"] = False
            member["finish_time"] = ""
            self.save_data()
            
            dialog.close()
            self.update_finish_tab()
            
            QMessageBox.information(
                self,
                "Финиш отменен",
                f"Финиш экипажа #{member['номер']} успешно отменен"
            )

    def validate_time_format(self, time_str):
        """Проверяет формат времени ЧЧ:ММ:СС"""
        if not time_str:
            return False
        try:
            parts = list(map(int, time_str.split(":")))
            return (len(parts) == 3 and
                    0 <= parts[0] < 24 and
                    0 <= parts[1] < 60 and
                    0 <= parts[2] < 60)
        except ValueError:
            return False

    def check_track_closing(self):
        """Проверяет превышение лимита времени трассы (без автосохранения)"""
        if not self.data.get("logic_params", {}).get("penalty_type") == "DNF":
            return
            
        try:
            track_close_time = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            track_start_time = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
            
            # Рассчитываем базовое время трассы в секундах
            track_duration = track_start_time.secsTo(track_close_time)
            if track_duration < 0:
                track_duration += 24 * 3600
                
            current_time = QTime.currentTime()
            need_save = False
            
            for member in self.data["members"]:
                if member.get("started") and not member.get("finished"):
                    start_time = QTime.fromString(member["start_time"], "HH:mm:ss")
                    neutral_time = self.calculate_total_neutral_time(member)
                    
                    # Дедлайн с учетом нейтрализации
                    finish_deadline = start_time.addSecs(track_duration + neutral_time)
                    
                    if current_time >= finish_deadline:
                        member["finished"] = True
                        member["finish_time"] = finish_deadline.addSecs(1).toString("HH:mm:ss")
                        member["dnf_reason"] = "Превышено время трассы"
                        
                        member.setdefault("stage_history", []).append({
                            "time": QDateTime.currentDateTime().toString("HH:mm:ss"),
                            "action": "auto_finish_timeout",
                            "reason": f"Лимит: {track_duration//3600}ч + {neutral_time//60}м нейтр."
                        })
                        need_save = True
            
            # Сохраняем только если были изменения
            if need_save:
                self.save_data()
                self.update_finish_tab()
                
        except Exception as e:
            print(f"Ошибка при проверке времени трассы: {e}")

    def calculate_track_duration(self):
        """Вычисляет базовую длительность трассы в секундах"""
        start = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
        close = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
        duration = start.secsTo(close)
        return duration if duration >= 0 else duration + 24*3600

    def calculate_crew_deadline(self, member, track_duration):
        """Вычисляет дедлайн для конкретного экипажа с учетом нейтрализации"""
        start_time = QTime.fromString(member["start_time"], "HH:mm:ss")
        neutral_time = self.calculate_total_neutral_time(member)
        return start_time.addSecs(track_duration + neutral_time)

    def mark_crew_dnf(self, member, deadline):
        """Помечает экипаж как DNF с указанием времени"""
        member.update({
            "finished": True,
            "finish_time": deadline.addSecs(1).toString("HH:mm:ss"),
            "dnf_reason": "Превышено время трассы",
            "stage_history": member.get("stage_history", []) + [{
                "time": QDateTime.currentDateTime().toString("HH:mm:ss"),
                "action": "time_limit_exceeded",
                "details": {
                    "track_limit": self.data["params"]["время_закрытия_трассы"],
                    "neutralization": self.format_time(self.calculate_total_neutral_time(member))
                }
            }]
        })

    def calculate_total_time(self, member):
        """Вычисляет общее время гонки (финиш - старт) в секундах"""
        try:
            start = QTime.fromString(member.get("start_time", ""), "HH:mm:ss")
            finish = QTime.fromString(member.get("finish_time", ""), "HH:mm:ss")
            total = start.secsTo(finish)
            return total if total >= 0 else total + 24 * 3600
        except:
            return 0

    def format_time(self, seconds):
        """Форматирует секунды в ЧЧ:ММ:СС"""
        if seconds is None or seconds == 0:
            return "00:00:00"
        return f"{seconds//3600:02d}:{(seconds%3600)//60:02d}:{seconds%60:02d}"

    def check_track_timeout(self):
        """Проверка превышения времени с уведомлениями"""
        if not self.data.get("logic_params", {}).get("penalty_type") == "DNF":
            return

        try:
            track_duration = self.calculate_track_duration()
            current_time = QTime.currentTime()
            dnf_crews = []

            for member in self.data["members"]:
                if member.get("started") and not member.get("finished"):
                    deadline = self.calculate_crew_deadline(member, track_duration)
                    
                    if current_time >= deadline:
                        self.mark_crew_dnf(member, deadline)
                        dnf_crews.append((
                            member["номер"],
                            member.get("пилот", ""),
                            self.format_time(self.calculate_total_neutral_time(member))
                        ))

            if dnf_crews:
                self.show_dnf_notification(dnf_crews, track_duration)
                self.save_data()
                self.update_finish_tab()

        except Exception as e:
            print(f"Ошибка проверки времени: {str(e)}")

    def show_dnf_notification(self, crews, track_duration):
        """Показывает уведомление об автоматическом финише"""
        hours = track_duration // 3600
        minutes = (track_duration % 3600) // 60
        
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setWindowTitle("Автоматический финиш")
        msg.setText(
            f"Превышен лимит времени трассы ({hours}ч {minutes}м)!\n"
            "Следующие экипажи отмечены как DNF:"
        )
        
        # Детализированный список
        details = QTextEdit()
        details.setReadOnly(True)
        details.setText(
            "\n".join(
                f"#{num} {pilot} (нейтрализация: {neutral_time})" 
                for num, pilot, neutral_time in crews
            )
        )
        details.setMinimumSize(400, 150)
        
        # Компоновка
        layout = msg.layout()
        layout.addWidget(details, 1, 0, 1, layout.columnCount())
        
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def update_finish_tables(self):
        """Обновляет таблицы финиша с сортировкой по времени и выделением DNF"""
        # --- Выключаем сортировку на время обновления ---
        self.not_finished_table.setSortingEnabled(False)
        self.finished_table.setSortingEnabled(False)

        self.not_finished_table.setRowCount(0)
        self.finished_table.setRowCount(0)

        # Разделяем экипажи
        finished = []
        not_finished = []

        for member in self.data.get("members", []):
            if not member.get("started", False):
                continue

            if member.get("finished", False):
                finished.append(member)
            else:
                not_finished.append(member)

        # Сортировка тут больше не нужна, так как таблица будет сортировать сама

        # Заполняем таблицу финишировавших
        for member in finished:
            row = self.finished_table.rowCount()
            self.finished_table.insertRow(row)

            neutral_time = self.calculate_total_neutral_time(member)
            total_time = self.calculate_total_time(member)
            track_time = max(0, total_time - neutral_time) if total_time else 0
            is_dnf = self.is_crew_dnf(member) # Проверяем DNF статус

            # Данные с корректными типами для сортировки
            # Используем NumericTableWidgetItem для номера
            self.finished_table.setItem(row, 0, NumericTableWidgetItem(str(member.get("номер", ""))))
            self.finished_table.setItem(row, 1, QTableWidgetItem(member.get("зачет", "N/A")))
            self.finished_table.setItem(row, 2, QTableWidgetItem(member.get("пилот", "N/A")))
            self.finished_table.setItem(row, 3, QTableWidgetItem(member.get("авто", "N/A")))
            # Используем TimeTableWidgetItem для времени
            self.finished_table.setItem(row, 4, TimeTableWidgetItem(member.get("start_time", "N/A")))

            # Время финиша (для DNF указываем причину)
            finish_time_str = member.get("finish_time", "N/A")
            finish_item = TimeTableWidgetItem(finish_time_str)
            if is_dnf:
                 finish_item.setText("DNF") # Отображаем DNF
                 finish_item.setToolTip(member.get("dnf_reason", "Причина DNF не указана"))
                 finish_item.setBackground(QColor(255, 200, 200)) # Красный фон для DNF
            self.finished_table.setItem(row, 5, finish_item)

            # Временные показатели (форматируем, но используем TimeTableWidgetItem для сортировки)
            self.finished_table.setItem(row, 6, TimeTableWidgetItem(self.format_time(total_time)))
            self.finished_table.setItem(row, 7, TimeTableWidgetItem(self.format_time(neutral_time)))
            self.finished_table.setItem(row, 8, TimeTableWidgetItem(self.format_time(track_time)))

            # Выделение DNF строки
            if is_dnf:
                for col in range(self.finished_table.columnCount()):
                    existing_item = self.finished_table.item(row, col)
                    if existing_item: # Проверяем, что ячейка существует
                        existing_item.setBackground(QColor(255, 200, 200))

        # Заполняем таблицу нефинишировавших
        for member in not_finished:
            row = self.not_finished_table.rowCount()
            self.not_finished_table.insertRow(row)

            neutral_time = self.calculate_total_neutral_time(member)

            # Используем NumericTableWidgetItem для номера
            self.not_finished_table.setItem(row, 0, NumericTableWidgetItem(str(member.get("номер", ""))))
            self.not_finished_table.setItem(row, 1, QTableWidgetItem(member.get("зачет", "N/A")))
            self.not_finished_table.setItem(row, 2, QTableWidgetItem(member.get("пилот", "N/A")))
            self.not_finished_table.setItem(row, 3, QTableWidgetItem(member.get("авто", "N/A")))
            # Используем TimeTableWidgetItem для времени
            self.not_finished_table.setItem(row, 4, TimeTableWidgetItem(member.get("start_time", "N/A")))
            self.not_finished_table.setItem(row, 5, TimeTableWidgetItem(self.format_time(neutral_time)))

        # Настраиваем отображение столбцов (можно вынести в setup, если не меняется)
        self.finished_table.resizeColumnsToContents()
        self.not_finished_table.resizeColumnsToContents()

        # --- Включаем сортировку обратно ---
        self.not_finished_table.setSortingEnabled(True)
        self.finished_table.setSortingEnabled(True)


# В классе RaceApp

    def calculate_total_neutral_time(self, member):
        """Вычисляет общее время нейтрализации в секундах, ИГНОРИРУЯ ОТМЕНЕННЫЕ СКП."""
        if not member.get("skp_entries"):
            return 0

        total_neutral = 0
        print(f"Расчет нейтрализации для #{member.get('номер')}:") # Отладка
        for entry in member["skp_entries"]:
            # ---> НАЧАЛО ИЗМЕНЕНИЯ: Проверяем флаг 'cancelled' <---
            is_cancelled = entry.get("cancelled", False)
            duration = entry.get("duration", 0)
            skp_num = entry.get("skp", "?")

            if not is_cancelled:
                if isinstance(duration, (int, float)):
                    print(f"  + Учитываем СКП {skp_num}: {duration} сек") # Отладка
                    total_neutral += duration
                else:
                     print(f"  ! Пропускаем СКП {skp_num}: некорректная длительность {duration}") # Отладка
            else:
                 print(f"  - Игнорируем СКП {skp_num}: отменено") # Отладка
            # ---> КОНЕЦ ИЗМЕНЕНИЯ <---

        print(f"  Итого нейтрализация: {total_neutral} сек") # Отладка
        return total_neutral





















############## ПРОВЕРКА ##############

    def process_cp_input(self):
        """Обработка ввода номеров КП с точным совпадением"""
        text = self.cp_input.text().strip()
        if not text:
            return
            
        cp_numbers = text.split()
        cp_prefixes = ["КП", "CP", "KP"]  # Возможные префиксы
        
        for i in range(self.cp_list_widget.count()):
            item = self.cp_list_widget.item(i)
            cp_name = item.text()
            
            # Извлекаем только цифры из названия КП
            cp_num = ''.join(filter(str.isdigit, cp_name))
            
            for num in cp_numbers:
                if num == cp_num:  # Точное совпадение чисел
                    item.setCheckState(Qt.CheckState.Checked)
                    break

        self.cp_input.clear()

    def update_crew_selection(self, index):
        """Обновляет выбранный экипаж при клике"""
        self.selected_crew_row = index.row()

    def eventFilter(self, source, event):
        """Перехватываем нажатие Enter в поле ввода КП"""
        if source is self.cp_input and event.type() == QEvent.Type.KeyPress:
            if event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter:
                self.process_cp_input()
                return True  # Событие обработано, не передаем дальше
        return super().eventFilter(source, event)

    def format_time(self, seconds):
        """Форматирует секунды в ЧЧ:ММ:СС"""
        if seconds is None or seconds == 0:
            return "00:00:00"
        return f"{seconds//3600:02d}:{(seconds%3600)//60:02d}:{seconds%60:02d}"

    def calculate_time_penalty(self, member, track_time):
        """Рассчитывает штраф за превышение времени"""
        penalty_type = self.data.get("logic_params", {}).get("penalty_type")
        
        # Проверяем явный DNF (не финишировал)
        if member.get('finish_time') == "DNF":
            return "DNF"
            
        # Проверяем превышение времени трассы для типа штрафа DNF
        if penalty_type == "DNF":
            track_duration = self.calculate_track_duration()
            if track_time > track_duration:
                return "DNF"
            return 0
                
        # Логика для штрафа в баллах
        elif penalty_type == "Баллы за 1 мин":
            track_duration = self.calculate_track_duration()
            overtime = max(0, track_time - track_duration)
            
            if overtime > 0:
                penalty_value = self.data["logic_params"].get("penalty_value", 1)
                return math.ceil(overtime / 60) * penalty_value
            return 0
            
        return 0

    def calculate_track_duration(self):
        """Вычисляет базовую длительность трассы"""
        try:
            start = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
            close = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            duration = start.secsTo(close)
            return duration if duration >= 0 else duration + 24*3600
        except Exception:
            return 0

    def calculate_total_time(self, member):
        """Вычисляет общее время"""
        try:
            start_time = member.get("start_time")
            finish_time = member.get("finish_time")
            
            if not start_time or not finish_time or finish_time == "DNF":
                return 0
                
            start = QTime.fromString(start_time, "HH:mm:ss")
            finish = QTime.fromString(finish_time, "HH:mm:ss")
            total = start.secsTo(finish)
            
            return total if total >= 0 else total + 24 * 3600
        except Exception:
            return 0

    def setup_check_tab(self):
        """Вкладка проверки контрольных пунктов"""
        self.selected_crew_row = -1
        self.check_tab = QWidget()
        layout = QVBoxLayout(self.check_tab)

        # Панель управления
        control_panel = QHBoxLayout()
        
        self.refresh_check_btn = QPushButton("Обновить")
        self.refresh_check_btn.clicked.connect(self.update_check_tab)
        control_panel.addWidget(self.refresh_check_btn)
        
        self.table_mode = QComboBox()
        self.table_mode.addItems(["Краткий вид", "Полный вид", "Режим КП"])  # Добавлен новый режим
        self.table_mode.currentIndexChanged.connect(self.update_check_tab)
        control_panel.addWidget(QLabel("Режим:"))
        control_panel.addWidget(self.table_mode)
        
        control_panel.addStretch()
        layout.addLayout(control_panel)

        # Таблица экипажей
        self.check_crews_table = QTableWidget()
        self.check_crews_table.setColumnCount(10)
        self.check_crews_table.setHorizontalHeaderLabels([
            "Номер", "Зачет", "Пилот", "Штурман", "Авто", 
            "Старт", "Финиш", "Время трассы", "Баллы", "Итог"
        ])
        self.check_crews_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        header = self.check_crews_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.check_crews_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.check_crews_table.doubleClicked.connect(self.show_check_dialog)
        self.check_crews_table.clicked.connect(self.update_crew_selection)
        
        layout.addWidget(self.check_crews_table)
        self.tabs.addTab(self.check_tab, "Проверка КП")
        self.update_check_tab()

    def update_check_tab(self):
        """Обновление данных на вкладке проверки"""
        self.check_crews_table.setRowCount(0)
        if not hasattr(self, 'data') or 'members' not in self.data:
            return

        # Определяем режим отображения
        mode = self.table_mode.currentIndex()
        
        if mode == 0:  # Краткий вид
            self.setup_short_mode()
        elif mode == 1:  # Полный вид
            self.setup_full_mode()
        elif mode == 2:  # Режим КП
            self.setup_cp_mode()

    def setup_short_mode(self):
        """Настройка краткого вида таблицы"""
        self.check_crews_table.setColumnCount(10)
        self.check_crews_table.setHorizontalHeaderLabels([
            "Номер", "Зачет", "Пилот", "Штурман", "Авто", 
            "Старт", "Финиш", "Время трассы", "Баллы", "Итог"
        ])
        self.fill_basic_data()
        self.adjust_check_table_columns(False, 0)

    def show_check_dialog(self, index):
        """Диалог проверки КП для выбранного экипажа"""
        try:
            if not hasattr(self, 'data') or 'members' not in self.data:
                raise ValueError("Данные не загружены")
                
            row = index.row()
            if row < 0 or row >= self.check_crews_table.rowCount():
                raise IndexError("Неверный индекс строки")
                
            crew_number = self.check_crews_table.item(row, 0).text()
            if not crew_number:
                raise ValueError("Не найден номер экипажа")
            
            # Находим экипаж в данных
            for member in self.data['members']:
                if str(member.get('номер', '')) == crew_number:
                    self.current_check_crew = member
                    break
            else:
                QMessageBox.warning(self, "Ошибка", "Экипаж не найден!")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Проверка КП экипажа №{crew_number}")
            dialog.setMinimumSize(800, 600)
            
            layout = QVBoxLayout(dialog)

            # Информация об экипаже
            info_group = QGroupBox("Информация об экипаже")
            info_layout = QVBoxLayout()
            info_layout.addWidget(QLabel(f"Пилот: {self.current_check_crew.get('пилот', '')}"))
            info_layout.addWidget(QLabel(f"Штурман: {self.current_check_crew.get('штурман', '')}"))
            info_layout.addWidget(QLabel(f"Авто: {self.current_check_crew.get('авто', '')}"))
            info_group.setLayout(info_layout)
            layout.addWidget(info_group)

            # Поле для ввода номеров КП
            self.cp_input = QLineEdit()
            self.cp_input.setPlaceholderText("Введите номера КП через пробел и нажмите Enter")
            self.cp_input.installEventFilter(self)
            layout.addWidget(self.cp_input)

            # Создаем виджет с вкладками для этапов
            tab_widget = QTabWidget()
            
            # Список для хранения всех QListWidget
            self.cp_list_widgets = []
            
            # Вкладка "Все КП"
            all_cps_tab = QWidget()
            all_cps_layout = QVBoxLayout(all_cps_tab)
            
            # Кнопки управления
            btn_layout = QHBoxLayout()
            all_cps_btn = QPushButton("Выделить все КП")
            none_cps_btn = QPushButton("Снять все КП")
            btn_layout.addWidget(all_cps_btn)
            btn_layout.addWidget(none_cps_btn)
            all_cps_layout.addLayout(btn_layout)
            
            # Основной список КП
            self.cp_list_widget = QListWidget()
            self.cp_list_widget.itemChanged.connect(self.sync_check_states)
            self.cp_list_widgets.append(self.cp_list_widget)
            all_cps_layout.addWidget(self.cp_list_widget)
            
            tab_widget.addTab(all_cps_tab, "Все КП")

            # Вкладки для этапов
            if self.data.get('logic_params', {}).get('staged', False):
                for stage in self.data['logic_params'].get('stages', []):
                    stage_tab = QWidget()
                    stage_layout = QVBoxLayout(stage_tab)
                    
                    # Кнопки для этапа
                    stage_btn_layout = QHBoxLayout()
                    stage_select_btn = QPushButton(f"Выделить все")
                    stage_deselect_btn = QPushButton(f"Снять все")
                    stage_btn_layout.addWidget(stage_select_btn)
                    stage_btn_layout.addWidget(stage_deselect_btn)
                    stage_layout.addLayout(stage_btn_layout)
                    
                    # Список КП этапа
                    stage_cp_list = QListWidget()
                    stage_cp_list.setProperty('stage', stage['name'])
                    stage_cp_list.itemChanged.connect(self.sync_check_states)
                    self.cp_list_widgets.append(stage_cp_list)
                    stage_layout.addWidget(stage_cp_list)
                    
                    tab_widget.addTab(stage_tab, stage['name'])
                    
                    # Подключаем кнопки этапа
                    stage_select_btn.clicked.connect(
                        lambda _, s=stage['name']: self.toggle_stage_checkpoints(s, True))
                    stage_deselect_btn.clicked.connect(
                        lambda _, s=stage['name']: self.toggle_stage_checkpoints(s, False))
            
            # Подключаем кнопки "Все КП"
            all_cps_btn.clicked.connect(lambda: self.toggle_all_checkpoints(True))
            none_cps_btn.clicked.connect(lambda: self.toggle_all_checkpoints(False))
            
            layout.addWidget(tab_widget)

            # Получаем зачёт текущего экипажа
            crew_classification = self.current_check_crew.get('зачет', '')
            
            # Заполняем списки КП (только доступные для зачёта)
            if 'checkpoints' in self.data:
                taken_cps = self.current_check_crew.get('taken_cps', [])
                
                for cp in self.data['checkpoints']:
                    # Проверяем доступность КП для зачёта экипажа
                    is_available = self.is_checkpoint_available(cp, crew_classification)
                    if not is_available:
                        continue  # Пропускаем недоступные КП
                    
                    # Создаём элемент списка
                    item = QListWidgetItem(f"{cp.get('name', '')} ({cp.get('score', 0)} баллов)")
                    item.setData(Qt.ItemDataRole.UserRole, cp['name'])
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    item.setCheckState(Qt.CheckState.Checked if cp['name'] in taken_cps else Qt.CheckState.Unchecked)
                    
                    # Добавляем в основной список
                    self.cp_list_widget.addItem(item)
                    
                    # Добавляем в соответствующие этапы (если этапность включена)
                    if self.data.get('logic_params', {}).get('staged', False) and 'stages' in cp:
                        for stage_name in cp['stages']:
                            for i in range(tab_widget.count()):
                                if tab_widget.tabText(i) == stage_name:
                                    cloned_item = item.clone()
                                    tab_widget.widget(i).findChild(QListWidget).addItem(cloned_item)
                                    break

            # Корректировка времени
            time_group = QGroupBox("Корректировка времени")
            time_layout = QVBoxLayout()
            
            # Старт
            start_layout = QHBoxLayout()
            start_layout.addWidget(QLabel("Старт:"))
            self.start_time_edit = QTimeEdit()
            self.start_time_edit.setDisplayFormat("HH:mm:ss")
            if self.current_check_crew.get('start_time'):
                self.start_time_edit.setTime(QTime.fromString(self.current_check_crew['start_time'], "HH:mm:ss"))
            start_layout.addWidget(self.start_time_edit)
            time_layout.addLayout(start_layout)
            
            # Финиш
            finish_layout = QHBoxLayout()
            finish_layout.addWidget(QLabel("Финиш:"))
            self.finish_time_edit = QTimeEdit()
            self.finish_time_edit.setDisplayFormat("HH:mm:ss")
            
            # Блокируем редактирование если экипаж не финишировал
            is_finished = self.current_check_crew.get('finished', False)
            self.finish_time_edit.setEnabled(is_finished)
            
            if is_finished and self.current_check_crew.get('finish_time') and self.current_check_crew['finish_time'] != "DNF":
                self.finish_time_edit.setTime(QTime.fromString(self.current_check_crew['finish_time'], "HH:mm:ss"))
            else:
                self.finish_time_edit.setTime(QTime(0, 0, 0))
            
            finish_layout.addWidget(self.finish_time_edit)
            time_layout.addLayout(finish_layout)
            
            # Информация о нейтрализации
            if self.data.get('logic_params', {}).get('staged', False):
                time_layout.addWidget(QLabel("Редактирование времени нейтрализации осуществляется во вкладке Этапы", 
                                          alignment=Qt.AlignmentFlag.AlignRight))
            
            time_group.setLayout(time_layout)
            layout.addWidget(time_group)

            # Флаг завершения проверки
            self.check_completed_cb = QCheckBox("Проверка завершена")
            self.check_completed_cb.setChecked(self.current_check_crew.get('check_completed', False))
            layout.addWidget(self.check_completed_cb)

            # Кнопки
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
            button_box.accepted.connect(lambda: self.save_check_data(dialog))
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)

            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть проверку: {str(e)}")

    def save_check_data(self, dialog):
        """Сохранение результатов проверки с корректным форматом taken_cps"""
        # Сохраняем взятые КП (все они гарантированно доступны)
        taken_cps = []
        for i in range(self.cp_list_widget.count()):
            item = self.cp_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                taken_cps.append(item.data(Qt.ItemDataRole.UserRole))
        
        # Получаем зачет экипажа
        crew_classification = self.current_check_crew.get('зачет', '')
        
        # Собираем взятые КП
        taken_cps = []
        for i in range(self.cp_list_widget.count()):
            item = self.cp_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                cp_name = item.data(Qt.ItemDataRole.UserRole)
                
                # Проверяем доступность КП
                cp_data = next((cp for cp in self.data['checkpoints'] if cp['name'] == cp_name), None)
                if cp_data:
                    is_available = True
                    if 'classifications' in cp_data:
                        if crew_classification in cp_data['classifications']:
                            is_available = cp_data['classifications'][crew_classification]
                    
                    if is_available:
                        taken_cps.append(cp_name)
                    else:
                        # Если попытались сохранить недоступный КП
                        QMessageBox.warning(
                            dialog,
                            "Ошибка",
                            f"КП {cp_name} недоступен для зачета '{crew_classification}'\n"
                            "Он не будет сохранен в списке взятых КП"
                        )

        # Сохраняем взятые КП (только названия без баллов)
        taken_cps = []
        for i in range(self.cp_list_widget.count()):
            item = self.cp_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                # Получаем оригинальное название КП из UserRole
                cp_name = item.data(Qt.ItemDataRole.UserRole)
                taken_cps.append(cp_name)
        
        # Подсчет баллов
        total_score = 0
        if 'checkpoints' in self.data:
            for cp in self.data['checkpoints']:
                if cp.get('name') in taken_cps:
                    total_score += int(cp.get('score', 0))
        
        # Обновляем данные экипажа
        self.current_check_crew['taken_cps'] = taken_cps  # Теперь только названия КП
        self.current_check_crew['total_score'] = total_score
        
        # Обновляем время старта
        self.current_check_crew['start_time'] = self.start_time_edit.time().toString("HH:mm:ss")
        
        # Обработка времени финиша
        if self.current_check_crew.get('finished', False):
            finish_time = "DNF" if self.current_check_crew.get('finish_time') == "DNF" else ""
            
            if self.finish_time_edit.time().isValid():
                if not (self.finish_time_edit.time().hour() == 0 and 
                       self.finish_time_edit.time().minute() == 0 and 
                       self.finish_time_edit.time().second() == 0):
                    finish_time = self.finish_time_edit.time().toString("HH:mm:ss")
            
            self.current_check_crew['finish_time'] = finish_time

        # Сохраняем статус проверки
        self.current_check_crew['check_completed'] = self.check_completed_cb.isChecked()
        
        # Сохраняем данные
        self.save_data()
        self.update_check_tab()
        
        # Показываем статистику
        QMessageBox.information(
            dialog, 
            "Сохранено", 
            f"Сохранено {len(taken_cps)} КП\n"
            f"Общий балл: {total_score}\n"
            f"Время старта: {self.current_check_crew['start_time']}\n"
            f"Время финиша: {self.current_check_crew.get('finish_time', '')}"
        )
        
        dialog.accept()

    def sync_check_states(self, changed_item):
        """Синхронизация состояния чекбоксов между вкладками"""
        # Получаем имя КП и новое состояние
        cp_name = changed_item.data(Qt.ItemDataRole.UserRole)
        new_state = changed_item.checkState()
        
        # Обновляем все соответствующие items в других списках
        for list_widget in self.cp_list_widgets:
            for i in range(list_widget.count()):
                item = list_widget.item(i)
                if item.data(Qt.ItemDataRole.UserRole) == cp_name:
                    # Временно отключаем сигналы, чтобы избежать рекурсии
                    list_widget.blockSignals(True)
                    item.setCheckState(new_state)
                    list_widget.blockSignals(False)
                    break

    def toggle_all_checkpoints(self, checked):
        """Установка/снятие всех КП с синхронизацией"""
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        for list_widget in self.cp_list_widgets:
            list_widget.blockSignals(True)
            for i in range(list_widget.count()):
                list_widget.item(i).setCheckState(state)
            list_widget.blockSignals(False)
        
        # Вручную обновляем один список, чтобы сработала синхронизация
        if self.cp_list_widget.count() > 0:
            self.cp_list_widget.item(0).setCheckState(state)

    def toggle_stage_checkpoints(self, stage_name, checked):
        """Установка/снятие КП этапа с синхронизацией"""
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        for list_widget in self.cp_list_widgets:
            list_widget.blockSignals(True)
            for i in range(list_widget.count()):
                item = list_widget.item(i)
                cp_name = item.data(Qt.ItemDataRole.UserRole)
                cp_data = next((cp for cp in self.data['checkpoints'] 
                              if cp['name'] == cp_name), None)
                
                if cp_data and 'stages' in cp_data and stage_name in cp_data['stages']:
                    item.setCheckState(state)
            list_widget.blockSignals(False)
        
        # Активируем синхронизацию
        if self.cp_list_widget.count() > 0:
            self.cp_list_widget.item(0).setCheckState(self.cp_list_widget.item(0).checkState())

    def is_checkpoint_available(self, checkpoint, classification):
        """Проверяет, доступен ли КП для указанного зачёта"""
        if 'classifications' not in checkpoint:
            return True  # Если ограничений нет, КП доступен
        
        # Если зачёт явно указан в classifications, используем его значение
        if classification in checkpoint['classifications']:
            return checkpoint['classifications'][classification]
        
        # Если зачёт не указан, считаем КП доступным
        return True

    def is_cp_available(self, checkpoint, classification):
        """Проверяет доступность КП для указанного зачёта"""
        if 'classifications' not in checkpoint:
            return True
        if classification in checkpoint['classifications']:
            return checkpoint['classifications'][classification]
        return True

    def adjust_cp_columns_width(self):
        """Настройка ширины столбцов с минимальными размерами"""
        header = self.check_crews_table.horizontalHeader()
        
        # Устанавливаем режимы изменения размеров
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Базовые ширины для разных типов столбцов (в пикселях)
        base_widths = {
            'number': 50,    # Номер
            'class': 80,     # Зачет
            'name': 100,     # Пилот/Штурман
            'car': 120,      # Авто
            'points': 60,    # Баллы
            'stage': 70,     # Этап
            'cp': 40,       # КП
            'total': 80      # Итог
        }

        for col in range(self.check_crews_table.columnCount()):
            header_text = self.check_crews_table.horizontalHeaderItem(col).text()
            
            # Определяем тип столбца
            if col == 0:
                col_type = 'number'
            elif col == 1:
                col_type = 'class'
            elif col in (2, 3):  # Пилот/Штурман
                col_type = 'name'
            elif col == 4:
                col_type = 'car'
            elif col == 5:
                col_type = 'points'
            elif "Этап" in header_text:
                col_type = 'stage'
            elif "КП" in header_text or "CP" in header_text:
                col_type = 'cp'
            elif header_text == "Итог":
                col_type = 'total'
            else:
                col_type = 'cp'  # По умолчанию для неизвестных

            # Устанавливаем начальную ширину
            header.resizeSection(col, base_widths[col_type])

            # Дополнительно вычисляем минимальную ширину по содержимому
            self.check_crews_table.resizeColumnToContents(col)
            min_width = header.sectionSize(col)
            header.resizeSection(col, min(min_width, base_widths[col_type] * 2))

        # Фиксируем ширину для КП (можно прокручивать если не помещаются)
        header.setStretchLastSection(False)

    def adjust_check_table_columns(self, is_staged, skp_count):
        """Настройка ширины столбцов для вкладки 'Проверка КП'."""
        header = self.check_crews_table.horizontalHeader()
        mode = self.table_mode.currentIndex()

        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        widths = {
            "Номер": 50, "Зачет": 80, "Пилот": 150, "Штурман": 150, "Авто": 120,
            "Баллы (Общ)": 70, # <-- Изменено имя и ширина
            "Старт": 80, "Финиш": 80, "Нейтрализация (Общ)": 90,
            "Общее время": 90, "Время трассы": 90, "Баллы": 60, # Для старого столбца Баллы
            "Штраф время": 80, "Штраф КП": 70,
            "Итог (Баллы - Штрафы)": 90, # <-- Изменено имя и ширина
            "СКП Вход": 80, "СКП Выход": 80, "СКП Нейтр.": 80,
            "Этап (∑)": 70, # <--- Ширина для колонки с суммой этапа
            "КП": 45
        }

        all_cp_names = {cp.get('name') for cp in self.data.get('checkpoints', []) if cp.get('name')}
        stage_names_set = {stage.get('name') for stage in self.data.get('logic_params', {}).get('stages', [])} if is_staged else set()

        for col in range(self.check_crews_table.columnCount()):
            header_item = self.check_crews_table.horizontalHeaderItem(col)
            if not header_item: continue
            header_text = header_item.text()
            width = 80
            col_type = 'unknown'

            # Определяем тип столбца
            if header_text == "Номер": col_type, width = 'number', widths['Номер']
            elif header_text == "Зачет": col_type, width = 'class', widths['Зачет']
            elif header_text == "Пилот": col_type, width = 'name', widths['Пилот']
            elif header_text == "Штурман": col_type, width = 'name', widths['Штурман']
            elif header_text == "Авто": col_type, width = 'car', widths['Авто']
            elif header_text == "Баллы (Общ)": col_type, width = 'points', widths['Баллы (Общ)'] # <-- Обновлено
            elif header_text == "Старт": col_type, width = 'time', widths['Старт']
            elif header_text == "Финиш": col_type, width = 'time', widths['Финиш']
            elif "СКП" in header_text and "Вход" in header_text: col_type, width = 'skp_time', widths['СКП Вход']
            elif "СКП" in header_text and "Выход" in header_text: col_type, width = 'skp_time', widths['СКП Выход']
            elif "СКП" in header_text and "Нейтр" in header_text: col_type, width = 'skp_time', widths['СКП Нейтр.']
            elif header_text == "Нейтрализация (Общ)": col_type, width = 'time', widths['Нейтрализация (Общ)']
            elif header_text == "Общее время": col_type, width = 'time', widths['Общее время']
            elif header_text == "Время трассы": col_type, width = 'time', widths['Время трассы']
            elif header_text == "Баллы": col_type, width = 'points', widths['Баллы'] # Старый столбец баллов
            elif header_text == "Штраф время": col_type, width = 'penalty', widths['Штраф время']
            elif header_text == "Штраф КП": col_type, width = 'penalty', widths['Штраф КП']
            elif header_text == "Итог (Баллы - Штрафы)": col_type, width = 'total', widths['Итог (Баллы - Штрафы)'] # <-- Обновлено
            elif header_text.endswith(" (∑)") and header_text[:-4] in stage_names_set: # <-- Ловим сумму этапа
                 col_type, width = 'stage_total', widths['Этап (∑)']
            elif header_text in all_cp_names: col_type, width = 'cp', widths['КП']

            header.resizeSection(col, width) # Устанавливаем ширину

            # Применяем автоподбор для нужных колонок
            if col_type in ['name', 'car', 'class', 'unknown']:
                try:
                    self.check_crews_table.resizeColumnToContents(col)
                    header.resizeSection(col, max(width, header.sectionSize(col)))
                except Exception as e:
                    print(f"Предупреждение: не удалось resizeColumnToContents для столбца {col}: {e}")
                    header.resizeSection(col, width)

        header.setStretchLastSection(False)

        # Растягивание колонок в зависимости от режима
        # (Логика растягивания как в предыдущем ответе)
        if mode == 0: # Краткий вид
             header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch); header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
             for col in range(self.check_crews_table.columnCount()):
                  if col not in [2, 3]: header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
        elif mode == 1: # Полный вид
             header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch); header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
             for col in range(self.check_crews_table.columnCount()):
                  if col not in [2, 3]: header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
        elif mode == 2: # Режим КП
             header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch) # Авто
             for col in range(self.check_crews_table.columnCount()):
                  if col != 4: header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)

    def setup_cp_mode(self):
        """Настройка режима отображения КП с СУММАМИ ЭТАПОВ и учетом доступности КП"""
        print("Настройка 'Режим КП'...") # Отладка
        try:
            is_staged = self.data.get('logic_params', {}).get('staged', False)
            stages_info = self.data['logic_params'].get('stages', []) if is_staged else []
            stages_count = len(stages_info)
            skp_count = stages_count - 1 if stages_count > 0 else 0

            all_cps = sorted(self.data.get('checkpoints', []), key=self._extract_cp_number)

            # --- Формируем заголовки ---
            headers = ["Номер", "Зачет", "Пилот", "Штурман", "Авто", "Баллы (Общ)"]
            cp_columns = {}
            stage_total_columns = {}
            col_index = len(headers)

            stage_qcolors = [QColor(173, 216, 230), QColor(144, 238, 144),
                             QColor(255, 182, 193), QColor(255, 255, 153)]

            if is_staged:
                stages_data = {}
                for cp in all_cps:
                    cp_name = cp.get('name')
                    if not cp_name: continue
                    if 'stages' in cp:
                        for stage_name in cp['stages']:
                            if stage_name not in stages_data: stages_data[stage_name] = []
                            stages_data[stage_name].append(cp_name)

                for i, stage in enumerate(stages_info):
                    stage_name = stage['name']
                    stage_color = stage_qcolors[i % len(stage_qcolors)]
                    headers.append(f"{stage_name} (∑)")
                    stage_total_columns[stage_name] = {'col_index': col_index, 'color': stage_color}
                    col_index += 1

                    stage_cp_names = sorted(stages_data.get(stage_name, []), key=self._extract_cp_number)
                    for cp_name in stage_cp_names:
                        if cp_name not in cp_columns:
                             headers.append(cp_name)
                             cp_columns[cp_name] = {'col_index': col_index, 'stage_name': stage_name}
                             col_index += 1

                for cp in all_cps:
                     cp_name = cp.get('name')
                     if not cp_name: continue
                     if cp_name not in cp_columns and ('stages' not in cp or not cp['stages']):
                         headers.append(cp_name)
                         cp_columns[cp_name] = {'col_index': col_index, 'stage_name': None}
                         col_index += 1
            else: # Без этапности
                for cp in all_cps:
                    cp_name = cp.get('name')
                    if not cp_name: continue
                    headers.append(cp_name)
                    cp_columns[cp_name] = {'col_index': col_index, 'stage_name': None}
                    col_index += 1

            headers.append("Итог (Баллы - Штрафы)")
            total_column = col_index # Запоминаем индекс последней колонки (Итог)

            # --- Настраиваем таблицу ---
            self.check_crews_table.setColumnCount(len(headers))
            self.check_crews_table.setHorizontalHeaderLabels(headers)
            self.check_crews_table.setWordWrap(False)
            self.check_crews_table.setTextElideMode(Qt.TextElideMode.ElideRight)

            # ---> ВЫЗЫВАЕМ ЗАПОЛНЕНИЕ ДАННЫМИ <---
            self.fill_basic_data(with_extras=False, skp_count=0, # with_extras=False т.к. это не Полный вид
                                 mode_cp_columns=cp_columns,      # Передаем структуру колонок
                                 mode_stage_total_columns=stage_total_columns, # Передаем структуру колонок
                                 mode_total_column_index=total_column,         # Передаем индекс итога
                                 mode_all_cps=all_cps)                         # Передаем отсортированные КП

            # Настраиваем ширину столбцов ПОСЛЕ заполнения
            self.adjust_check_table_columns(is_staged, skp_count)

        except Exception as e:
            print(f"!!! Ошибка в setup_cp_mode: {str(e)}")
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка отображения",
                                 f"Не удалось построить таблицу в режиме КП: {e}")

    def fill_basic_data(self, with_extras=False, skp_count=0,
                        mode_cp_columns=None, mode_stage_total_columns=None,
                        mode_total_column_index=-1, mode_all_cps=None):
        """
        Заполнение данных экипажей в таблице 'Проверка КП'.
        Принимает структуру колонок из вызывающего метода setup_*.
        """
        print(f"Заполнение таблицы 'Проверка КП'. Режим: {'Полный' if with_extras else ('КП' if mode_cp_columns else 'Краткий')}") # Отладка
        self.check_crews_table.setSortingEnabled(False)
        self.check_crews_table.setRowCount(0)

        # Используем переданный список КП или получаем его, если не передан
        all_cps = mode_all_cps if mode_all_cps is not None else sorted(self.data.get('checkpoints', []), key=self._extract_cp_number)
        is_staged = self.data.get('logic_params', {}).get('staged', False) and bool(mode_stage_total_columns) # Определяем этапность по наличию колонок

        # Цвета для этапов (если нужны)
        stage_qcolors = [ QColor(173, 216, 230), QColor(144, 238, 144), QColor(255, 182, 193), QColor(255, 255, 153) ]


        for member in self.data.get('members', []):
            if not member.get('started', False): continue

            try:
                row = self.check_crews_table.rowCount()
                self.check_crews_table.insertRow(row)
                is_dnf = self.is_crew_dnf(member)
                crew_classification = member.get('зачет', '')
                taken_cps = member.get('taken_cps', [])

                # --- Заполнение базовых колонок (0-5 или 0-6) ---
                # Эти колонки ОБЩИЕ для всех режимов, кроме режима КП
                if mode_cp_columns is None: # Краткий или Полный вид
                    self.check_crews_table.setItem(row, 0, NumericTableWidgetItem(str(member.get('номер', ''))))
                    self.check_crews_table.setItem(row, 1, QTableWidgetItem(member.get('зачет', '')))
                    self.check_crews_table.setItem(row, 2, QTableWidgetItem(member.get('пилот', '')))
                    self.check_crews_table.setItem(row, 3, QTableWidgetItem(member.get('штурман', '')))
                    self.check_crews_table.setItem(row, 4, QTableWidgetItem(member.get('авто', '')))
                    self.check_crews_table.setItem(row, 5, TimeTableWidgetItem(member.get('start_time', '')))
                    # Финиш (столбец 6)
                    finish_time_str = member.get('finish_time', '')
                    if not member.get('finished', False) and finish_time_str != "DNF": finish_time_str = ""
                    finish_item = TimeTableWidgetItem(finish_time_str)
                    if is_dnf: finish_item.setText("DNF")
                    self.check_crews_table.setItem(row, 6, finish_item)
                else: # Режим КП
                    self.check_crews_table.setItem(row, 0, NumericTableWidgetItem(str(member.get('номер', ''))))
                    self.check_crews_table.setItem(row, 1, QTableWidgetItem(member.get('зачет', '')))
                    self.check_crews_table.setItem(row, 2, QTableWidgetItem(member.get('пилот', '')))
                    self.check_crews_table.setItem(row, 3, QTableWidgetItem(member.get('штурман', '')))
                    self.check_crews_table.setItem(row, 4, QTableWidgetItem(member.get('авто', '')))
                    # Общие Баллы (столбец 5)
                    total_score_member = int(member.get('total_score', 0))
                    item_total_score = NumericTableWidgetItem(str(total_score_member))
                    item_total_score.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.check_crews_table.setItem(row, 5, item_total_score)


                # --- Расчетные переменные (ЗДЕСЬ!) ---
                neutral_time_total = self.calculate_total_neutral_time(member)
                total_time = self.calculate_total_time(member)
                track_time = max(0, total_time - neutral_time_total) if total_time else 0
                total_score = int(member.get('total_score', 0))
                cp_penalty = int(member.get('cp_penalty', 0))
                time_penalty = "DNF" if is_dnf else self.calculate_time_penalty(member, track_time)
                final_result_val = "DNF" if is_dnf else total_score - (time_penalty if isinstance(time_penalty, (int, float)) else 0) - cp_penalty

                # --- Заполнение остальных столбцов ---
                if mode_cp_columns: # Режим КП
                    # Заполнение данных этапов и КП
                    if is_staged and mode_stage_total_columns:
                        for stage_name, stage_info in mode_stage_total_columns.items():
                            stage_total_col = stage_info['col_index']
                            stage_color = stage_info['color']
                            lighter_stage_color = stage_color.lighter(130)
                            current_stage_score = 0
                            for cp_name, cp_info in mode_cp_columns.items():
                                if cp_info['stage_name'] == stage_name:
                                    cp_data = next((cp for cp in all_cps if cp['name'] == cp_name), None)
                                    if cp_data and cp_name in taken_cps and self.is_cp_available(cp_data, crew_classification):
                                         current_stage_score += int(cp_data.get('score', 0))
                            stage_item = NumericTableWidgetItem(str(current_stage_score))
                            stage_item.setBackground(stage_color)
                            stage_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            self.check_crews_table.setItem(row, stage_total_col, stage_item)

                            for cp_name, cp_info in mode_cp_columns.items():
                                 if cp_info['stage_name'] == stage_name:
                                      cp_col_index = cp_info['col_index']
                                      cp_data = next((cp for cp in all_cps if cp['name'] == cp_name), None)
                                      if not cp_data: continue
                                      item = QTableWidgetItem()
                                      item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                                      is_available = self.is_cp_available(cp_data, crew_classification)
                                      if not is_available: item.setText("-"); item.setForeground(Qt.GlobalColor.gray)
                                      elif cp_name in taken_cps: item.setText(str(int(cp_data.get('score', 0))))
                                      else: item.setText("")
                                      item.setBackground(lighter_stage_color)
                                      self.check_crews_table.setItem(row, cp_col_index, item)

                        # КП без этапа
                        for cp_name, cp_info in mode_cp_columns.items():
                             if cp_info['stage_name'] is None:
                                  cp_col_index = cp_info['col_index']
                                  cp_data = next((cp for cp in all_cps if cp['name'] == cp_name), None)
                                  if not cp_data: continue
                                  item = QTableWidgetItem(); item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                                  is_available = self.is_cp_available(cp_data, crew_classification)
                                  if not is_available: item.setText("-"); item.setForeground(Qt.GlobalColor.gray)
                                  elif cp_name in taken_cps: item.setText(str(int(cp_data.get('score', 0))))
                                  else: item.setText("")
                                  self.check_crews_table.setItem(row, cp_col_index, item)

                    else: # Режим КП без этапности
                         for cp_name, cp_info in mode_cp_columns.items():
                              cp_col_index = cp_info['col_index']
                              cp_data = next((cp for cp in all_cps if cp['name'] == cp_name), None)
                              if not cp_data: continue
                              item = QTableWidgetItem(); item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                              is_available = self.is_cp_available(cp_data, crew_classification)
                              if not is_available: item.setText("-"); item.setForeground(Qt.GlobalColor.gray)
                              elif cp_name in taken_cps: item.setText(str(int(cp_data.get('score', 0))))
                              else: item.setText("")
                              self.check_crews_table.setItem(row, cp_col_index, item)

                    # Итоговый столбец (в режиме КП)
                    if mode_total_column_index != -1:
                         result_item = QTableWidgetItem(str(final_result_val))
                         result_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                         result_item.setFont(QFont("Times New Roman", 11, QFont.Weight.Bold))
                         if is_dnf: result_item.setBackground(QColor(255, 200, 200))
                         self.check_crews_table.setItem(row, mode_total_column_index, result_item)

                elif with_extras: # Полный вид
                    # ... (код заполнения СКП как раньше, используя skp_count) ...
                    skp_base_col = 7
                    # ... (заполнение skp_details) ...
                    skp_details = {}
                    if 'skp_entries' in member and skp_count > 0:
                         for entry in member['skp_entries']:
                             if not entry.get('cancelled', False):
                                 skp_num_entry = entry.get('skp')
                                 if skp_num_entry and 1 <= skp_num_entry <= skp_count:
                                     skp_details[skp_num_entry] = {'entry': entry.get('entry_time', ''), 'exit': entry.get('exit_time', ''), 'duration': entry.get('duration', 0)}
                    for skp_num in range(1, skp_count + 1):
                         entry_col = skp_base_col + (skp_num - 1) * 3; exit_col = entry_col + 1; duration_col = entry_col + 2
                         details = skp_details.get(skp_num, {})
                         self.check_crews_table.setItem(row, entry_col, TimeTableWidgetItem(details.get('entry', '')))
                         self.check_crews_table.setItem(row, exit_col, TimeTableWidgetItem(details.get('exit', '')))
                         duration_item = TimeTableWidgetItem(self.format_time(details.get('duration', 0))); duration_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                         self.check_crews_table.setItem(row, duration_col, duration_item)

                    # ... (код заполнения остальных колонок полного вида neutral_total_col до final_result_col как раньше) ...
                    neutral_total_col = skp_base_col + skp_count * 3; total_time_col = neutral_total_col + 1; track_time_col = total_time_col + 1
                    score_col = track_time_col + 1; time_penalty_col = score_col + 1; cp_penalty_col = time_penalty_col + 1; final_result_col = cp_penalty_col + 1
                    self.check_crews_table.setItem(row, neutral_total_col, TimeTableWidgetItem(self.format_time(neutral_time_total)))
                    self.check_crews_table.setItem(row, total_time_col, TimeTableWidgetItem(self.format_time(total_time)))
                    self.check_crews_table.setItem(row, track_time_col, TimeTableWidgetItem(self.format_time(track_time)))
                    self.check_crews_table.setItem(row, score_col, NumericTableWidgetItem(str(total_score)))
                    time_penalty_item = QTableWidgetItem(str(time_penalty));
                    if time_penalty == "DNF": time_penalty_item.setBackground(QColor(255, 200, 200))
                    self.check_crews_table.setItem(row, time_penalty_col, time_penalty_item)
                    self.check_crews_table.setItem(row, cp_penalty_col, NumericTableWidgetItem(str(cp_penalty)))
                    total_item = QTableWidgetItem(str(final_result_val));
                    if final_result_val == "DNF": total_item.setBackground(QColor(255, 200, 200))
                    self.check_crews_table.setItem(row, final_result_col, total_item)


                else: # Краткий вид
                    # ... (код заполнения колонок 7-9 как раньше) ...
                    track_time_col = 7; score_col = 8; final_result_col = 9
                    self.check_crews_table.setItem(row, track_time_col, TimeTableWidgetItem(self.format_time(track_time)))
                    self.check_crews_table.setItem(row, score_col, NumericTableWidgetItem(str(total_score)))
                    total_item = QTableWidgetItem(str(final_result_val));
                    if final_result_val == "DNF": total_item.setBackground(QColor(255, 200, 200))
                    self.check_crews_table.setItem(row, final_result_col, total_item)


                # --- Раскраска строки ---
                row_color = None
                if member.get('check_completed', False): row_color = QColor(200, 255, 200)
                elif is_dnf: row_color = QColor(255, 200, 200)
                if row_color:
                    for col_fill in range(self.check_crews_table.columnCount()):
                        item_fill = self.check_crews_table.item(row, col_fill)
                        if item_fill:
                             current_brush = item_fill.background()
                             if current_brush.style() == Qt.BrushStyle.NoBrush:
                                 item_fill.setBackground(row_color)

            except Exception as e_row:
                print(f"!!! Ошибка при обработке строки для экипажа {member.get('номер', 'N/A')}: {e_row}")
                traceback.print_exc()

        self.check_crews_table.setSortingEnabled(True)
        print(f"Заполнение таблицы 'Проверка КП' завершено. Строк: {self.check_crews_table.rowCount()}")

    def setup_full_mode(self):
        """Настройка полного вида таблицы с детализацией по СКП."""
        print("Настройка 'Полный вид'...") # Отладка
        try:
            # Получаем информацию об этапах и СКП
            is_staged = self.data.get('logic_params', {}).get('staged', False)
            # Используем stages_count из logic_params, если этапность включена
            stages_count = len(self.data['logic_params'].get('stages', [])) if is_staged else 0
            # Количество СКП обычно на 1 меньше числа этапов
            skp_count = stages_count - 1 if stages_count > 0 else 0

            # --- Формируем заголовки в ПРАВИЛЬНОМ порядке ---
            headers = [
                "Номер", "Зачет", "Пилот", "Штурман", "Авто", # Индексы 0-4
                "Старт", "Финиш"                               # Индексы 5-6 <--- ПРАВИЛЬНОЕ МЕСТО ДЛЯ ФИНИША
            ]

            # Добавляем столбцы для каждого СКП (начиная с индекса 7)
            for i in range(1, skp_count + 1):
                headers.extend([
                    f"СКП {i} Вход", f"СКП {i} Выход", f"СКП {i} Нейтр." # Индексы 7,8,9 для СКП 1 и т.д.
                ])

            # Добавляем остальные столбцы ПОСЛЕ СКП
            headers.extend([
                "Нейтрализация (Общ)", # Индекс: 7 + skp_count * 3
                "Общее время",
                "Время трассы",
                "Баллы",
                "Штраф время",
                "Штраф КП",
                "Итог"
            ])
            # --- КОНЕЦ ИЗМЕНЕНИЯ ПОРЯДКА ЗАГОЛОВКОВ ---

            self.check_crews_table.setColumnCount(len(headers))
            self.check_crews_table.setHorizontalHeaderLabels(headers)

            # Заполняем данными, передавая количество СКП
            # Вызов fill_basic_data остается ПРЕЖНИМ, так как он уже правильно расставляет данные
            # по индексам, которые мы только что скорректировали в заголовках.
            print("  Вызов fill_basic_data для Полного вида...") # Отладка
            self.fill_basic_data(with_extras=True, skp_count=skp_count)
            print("  fill_basic_data для Полного вида завершен.") # Отладка


            # Настройка внешнего вида столбцов (можно расширить)
            self.adjust_check_table_columns(is_staged, skp_count)
            print("'Полный вид' настроен.") # Отладка


        except Exception as e:
            print(f"!!! Ошибка в setup_full_mode: {e}") # Отладка
            traceback.print_exc()
            # Показываем базовый вид в случае ошибки
            self.setup_short_mode()
            QMessageBox.critical(self, "Ошибка отображения",
                                 "Не удалось построить полный вид таблицы с СКП.")







        
############## РЕЗУЛЬТАТЫ ##############

    def setup_results_tab(self):
        """Инициализация вкладки с итоговыми результатами"""
        self.results_tab = QWidget()
        layout = QVBoxLayout(self.results_tab)

        toolbar = QHBoxLayout()
        
        self.export_btn = QPushButton("Экспорт в Excel")
        self.export_btn.clicked.connect(self.save_results)
        toolbar.addWidget(self.export_btn)
        
        self.print_btn = QPushButton("Печать результатов")
        self.print_btn.clicked.connect(self.print_results)
        toolbar.addWidget(self.print_btn)
        
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.clicked.connect(self.update_results_tab)
        toolbar.addWidget(self.refresh_btn)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)

        # Таблица результатов
        self.results_table = QTableWidget()
        self.results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.results_table)

        # Обновление данных
        self.update_results_tab()
        self.tabs.addTab(self.results_tab, "Результаты")

    def calculate_track_time(self, member):
        """Расчет чистого времени трассы (общее время - нейтрализация)"""
        total_time = self.calculate_total_time(member)
        neutral_time = self.calculate_total_neutral_time(member)
        return max(0, total_time - neutral_time)

    def print_results(self):
        """Заглушка для функции печати результатов"""
        QMessageBox.information(
            self,
            "Больше не тыкай сюда, понятно?",
            "Эта хрень пока не работает, нажми экспорт",
            QMessageBox.StandardButton.Ok
        )

    def add_time_data(self, row, member):
        """Добавление временных показателей"""
        # Финиш
        finish_time = member.get('finish_time', '')
        if not member.get('finished', False) and finish_time != "DNF":
            finish_time = ""
        self.results_table.setItem(row, 7, QTableWidgetItem(finish_time))
        
        # Нейтрализация и время
        neutral_time = self.calculate_total_neutral_time(member)
        total_time = self.calculate_total_time(member)
        track_time = max(0, total_time - neutral_time) if total_time else 0
        
        self.results_table.setItem(row, 8, QTableWidgetItem(self.format_time(neutral_time)))
        self.results_table.setItem(row, 9, QTableWidgetItem(self.format_time(total_time)))
        self.results_table.setItem(row, 10, QTableWidgetItem(self.format_time(track_time)))

    def add_skp_data(self, row, member, skp_count):
        """Добавление данных по СКП"""
        base_col = 11  # После временных показателей
        
        for i in range(1, skp_count + 1):
            skp_data = member.get(f'skp_{i}', {})
            self.results_table.setItem(row, base_col, QTableWidgetItem(skp_data.get('entry', '')))
            self.results_table.setItem(row, base_col + 1, QTableWidgetItem(skp_data.get('exit', '')))
            self.results_table.setItem(row, base_col + 2, QTableWidgetItem(
                self.format_time(skp_data.get('neutralization', 0))))
            base_col += 3

    def add_stages_data(self, row, member):
        """Добавление данных по этапам и КП с цветовой маркировкой"""
        if not self.data.get('logic_params', {}).get('staged', False):
            return
            
        # Позиция начала этапов
        start_col = 14 + len(self.data.get('skp_points', [])) * 3  # После новых столбцов
        crew_classification = member.get('зачет', '')
        taken_cps = member.get('taken_cps', [])
        
        # Цвета для этапов
        stage_colors = [
            QColor(173, 216, 230),  # LightBlue
            QColor(144, 238, 144),   # LightGreen
            QColor(255, 182, 193),   # LightPink
            QColor(255, 255, 153)    # LightYellow
        ]
        
        for stage_idx, stage in enumerate(self.data['logic_params'].get('stages', [])):
            stage_name = stage['name']
            stage_score = 0
            cp_col = 0
            
            # Считаем баллы этапа
            for cp in self.data.get('checkpoints', []):
                if 'stages' in cp and stage_name in cp['stages']:
                    if cp['name'] in taken_cps and self.is_cp_available(cp, crew_classification):
                        stage_score += int(cp.get('score', 0))
            
            # Добавляем баллы этапа с цветом
            stage_item = QTableWidgetItem(str(stage_score))
            stage_item.setBackground(stage_colors[stage_idx % len(stage_colors)])
            stage_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.results_table.setItem(row, start_col, stage_item)
            start_col += 1
            
            # Добавляем КП этапа с цветом
            for cp in self.data.get('checkpoints', []):
                if 'stages' in cp and stage_name in cp['stages']:
                    item = QTableWidgetItem()
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    if not self.is_cp_available(cp, crew_classification):
                        item.setText("-")
                        item.setForeground(Qt.GlobalColor.gray)
                    elif cp['name'] in taken_cps:
                        item.setText(str(int(cp.get('score', 0))))
                    else:
                        item.setText("0")
                    
                    # Более светлый оттенок цвета этапа
                    item.setBackground(stage_colors[stage_idx % len(stage_colors)].lighter(130))
                    self.results_table.setItem(row, start_col + cp_col, item)
                    cp_col += 1
            
            start_col += cp_col

    def get_column_index(self, column_name):
        """Возвращает индекс столбца по его названию"""
        for col in range(self.results_table.columnCount()):
            if self.results_table.horizontalHeaderItem(col).text() == column_name:
                return col
        return -1

    def sort_crews_by_results(self):
        """Сортировка экипажей по зачетам и результатам с учетом DNF"""
        crews_by_class = {}
        
        for member in self.data['members']:
            if not member.get('started', False):
                continue
            
            # Проверяем DNF статус
            track_time = self.calculate_track_time(member)
            is_dnf = member.get('finish_time') == "DNF" or (
                self.data.get("logic_params", {}).get("penalty_type") == "DNF" and 
                track_time > self.calculate_track_duration()
            )
            
            crew_class = member.get('зачет', 'Без зачета')
            if crew_class not in crews_by_class:
                crews_by_class[crew_class] = []
            
            # Для DNF ставим минимальный приоритет сортировки
            if is_dnf:
                sort_key = (float('-inf'), float('inf'))  # Всегда в конце
            else:
                total_score = member.get('total_score', 0)
                sort_key = (-total_score, track_time if track_time > 0 else float('inf'))
            
            crews_by_class[crew_class].append({
                'member': member,
                'sort_key': sort_key,
                'is_dnf': is_dnf
            })

        # Сортируем внутри каждого зачета
        sorted_crews = []
        for crew_class in sorted(crews_by_class.keys()):
            # Сначала не-DNF экипажи, отсортированные по результатам
            non_dnf_crews = [c for c in crews_by_class[crew_class] if not c['is_dnf']]
            non_dnf_crews = sorted(non_dnf_crews, key=lambda x: x['sort_key'])
            
            # Затем DNF экипажи (они уже в конце и не сортируются)
            dnf_crews = [c for c in crews_by_class[crew_class] if c['is_dnf']]
            
            # Объединяем и нумеруем места
            class_crews = non_dnf_crews + dnf_crews
            for i, crew_data in enumerate(class_crews, 1):
                # Для DNF не указываем место (оно будет заменено на DNF позже)
                if not crew_data['is_dnf']:
                    crew_data['member']['место'] = i
                sorted_crews.append(crew_data['member'])

        return sorted_crews

    def add_basic_member_info(self, row, member, is_dnf):
        """Добавление основной информации об экипаже"""
        # Место
        place_item = QTableWidgetItem("DNF" if is_dnf else str(member.get('место', '')))
        if is_dnf:
            place_item.setBackground(Qt.GlobalColor.red)
        self.results_table.setItem(row, 0, place_item)
        
        # Остальные основные данные
        self.results_table.setItem(row, 1, QTableWidgetItem(str(member.get('номер', ''))))
        self.results_table.setItem(row, 2, QTableWidgetItem(member.get('зачет', '')))
        self.results_table.setItem(row, 3, QTableWidgetItem(member.get('пилот', '')))
        self.results_table.setItem(row, 4, QTableWidgetItem(member.get('штурман', '')))
        self.results_table.setItem(row, 5, QTableWidgetItem(member.get('авто', '')))
        self.results_table.setItem(row, 6, QTableWidgetItem(member.get('start_time', '')))

    def add_score_columns(self, row, member, is_dnf):
        """Добавление столбцов с итогом, штрафом времени и баллами"""
        # Позиция начала новых столбцов
        score_col = 11 + len(self.data.get('skp_points', [])) * 3
        
        # Рассчитываем временные показатели
        neutral_time = self.calculate_total_neutral_time(member)
        total_time = self.calculate_total_time(member)
        track_time = max(0, total_time - neutral_time) if total_time else 0
        
        # Баллы и штрафы
        total_score = member.get('total_score', 0)
        time_penalty = "DNF" if is_dnf else self.calculate_time_penalty(member, track_time)
        cp_penalty = member.get('cp_penalty', 0)
        
        # Расчет итога
        total = "DNF" if is_dnf else total_score - (time_penalty if isinstance(time_penalty, int) else 0) - cp_penalty
        
        # Итог
        total_item = QTableWidgetItem(str(total))
        if is_dnf:
            total_item.setBackground(Qt.GlobalColor.red)
        self.results_table.setItem(row, score_col, total_item)
        
        # Штраф время
        time_penalty_item = QTableWidgetItem(str(time_penalty))
        if is_dnf:
            time_penalty_item.setBackground(Qt.GlobalColor.red)
        self.results_table.setItem(row, score_col + 1, time_penalty_item)
        
        # Баллы
        self.results_table.setItem(row, score_col + 2, QTableWidgetItem(str(total_score)))

    def add_checkpoints_data(self, row, member, checkpoints, start_col):
        """Добавление данных по КП без этапности"""
        taken_cps = member.get('taken_cps', [])
        crew_classification = member.get('зачет', '')
        
        for col, cp in enumerate(checkpoints):
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            if not self.is_cp_available(cp, crew_classification):
                item.setText("-")
                item.setForeground(Qt.GlobalColor.gray)
            elif cp['name'] in taken_cps:
                item.setText(str(int(cp.get('score', 0))))
            else:
                item.setText("0")
            
            self.results_table.setItem(row, start_col + col, item)

    def is_crew_dnf(self, member):
        """Проверяет, является ли экипаж DNF"""
        track_time = self.calculate_track_time(member)
        return (member.get('finish_time') == "DNF" or 
                (self.data.get("logic_params", {}).get("penalty_type") == "DNF" and 
                track_time > self.calculate_track_duration()))

    def adjust_results_columns(self):
        """Настройка ширины столбцов"""
        header = self.results_table.horizontalHeader()
        
        # Базовые ширины столбцов
        column_widths = {
            'Место': 50,
            'Номер': 50,
            'Зачет': 80,
            'Пилот': 120,
            'Штурман': 120,
            'Авто': 100,
            'Старт': 80,
            'Финиш': 80,
            'Нейтрализация': 100,
            'Время': 80,
            'Итог': 60,
            'Штраф': 80,
            'Баллы': 60,
            'КП': 40
        }
        
        for col in range(self.results_table.columnCount()):
            header_text = self.results_table.horizontalHeaderItem(col).text()
            width = 80  # Значение по умолчанию
            
            for pattern, w in column_widths.items():
                if pattern in header_text:
                    width = w
                    break
            
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            header.resizeSection(col, width)
        
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # Пилот
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # Штурман

    def update_results_tab(self):
        """Обновление данных в таблице результатов"""
        self.results_table.clear()
        
        if not hasattr(self, 'data') or 'members' not in self.data:
            return

        # Формируем заголовки столбцов
        headers = [
            "Место", "Номер", "Зачет", "Пилот", "Штурман", "Авто", "Старт",
            "Финиш", "Нейтрализация", "Общее время", "Время трассы"
        ]

        # Добавляем СКП (если есть)
        skp_count = len(self.data.get('skp_points', []))
        for i in range(1, skp_count + 1):
            headers.extend([f"Вход СКП {i}", f"Выход СКП {i}", f"Нейтрализация СКП {i}"])

        # Добавляем основные столбцы
        headers.extend(["Итог", "Штраф время", "Баллы"])

        # Добавляем КП в зависимости от этапности
        if self.data.get('logic_params', {}).get('staged', False):
            # С этапностью - добавляем этапы и их КП
            for stage in self.data['logic_params'].get('stages', []):
                headers.append(stage['name'])
                for cp in self.data.get('checkpoints', []):
                    if 'stages' in cp and stage['name'] in cp['stages']:
                        headers.append(cp['name'])
        else:
            # Без этапности - просто все КП по порядку
            checkpoints = sorted(self.data.get('checkpoints', []), key=lambda x: x['name'])
            for cp in checkpoints:
                headers.append(cp['name'])

        self.results_table.setColumnCount(len(headers))
        self.results_table.setHorizontalHeaderLabels(headers)

        # Сортируем экипажи
        sorted_crews = self.sort_crews_by_results()

        # Заполняем данные
        self.fill_results_data(sorted_crews, skp_count)

        # Настраиваем внешний вид
        self.adjust_results_columns()

    def fill_results_data(self, crews, skp_count):
        """Заполнение таблицы данными"""
        self.results_table.setRowCount(len(crews))
        
        # Цвета для призовых мест
        place_colors = {
            1: QColor(255, 215, 0),    # Золотой
            2: QColor(192, 192, 192),  # Серебряный
            3: QColor(205, 127, 50)     # Бронзовый
        }
        
        # Определяем позицию начала КП
        base_cp_col = 14 + skp_count * 3
        
        for row, member in enumerate(crews):
            # Проверяем DNF статус
            is_dnf = self.is_crew_dnf(member)
            
            # Основные данные
            self.add_basic_member_info(row, member, is_dnf)
            
            # Временные показатели
            self.add_time_data(row, member)
            
            # Данные по СКП
            self.add_skp_data(row, member, skp_count)
            
            # Основные столбцы (итог, штраф время, баллы)
            self.add_score_columns(row, member, is_dnf)
            
            # Данные по КП (в зависимости от этапности)
            if self.data.get('logic_params', {}).get('staged', False):
                self.add_staged_checkpoints(row, member, base_cp_col)
            else:
                self.add_simple_checkpoints(row, member, base_cp_col)
            
            # Раскрашиваем призовые места (только для не-DNF)
            if not is_dnf:
                place = member.get('место', 0)
                if place in place_colors:
                    for col in range(3):  # Раскрашиваем первые 3 колонки
                        item = self.results_table.item(row, col)
                        if item:
                            item.setBackground(place_colors[place])

    def add_simple_checkpoints(self, row, member, start_col):
        """Добавление КП без этапности"""
        checkpoints = sorted(self.data.get('checkpoints', []), key=lambda x: x['name'])
        taken_cps = member.get('taken_cps', [])
        crew_classification = member.get('зачет', '')
        
        for col, cp in enumerate(checkpoints):
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            if not self.is_cp_available(cp, crew_classification):
                item.setText("-")
                item.setForeground(Qt.GlobalColor.gray)
            elif cp['name'] in taken_cps:
                item.setText(str(int(cp.get('score', 0))))
            else:
                item.setText("")  # Пустая строка вместо "0"
            
            self.results_table.setItem(row, start_col + col, item)  

    def add_staged_checkpoints(self, row, member, start_col):
        """Добавление КП с этапностью"""
        crew_classification = member.get('зачет', '')
        taken_cps = member.get('taken_cps', [])
        current_col = start_col
        
        # Цвета для этапов
        stage_colors = [
            QColor(173, 216, 230),  # LightBlue
            QColor(144, 238, 144),   # LightGreen
            QColor(255, 182, 193),   # LightPink
            QColor(255, 255, 153)    # LightYellow
        ]
        
        for stage_idx, stage in enumerate(self.data['logic_params'].get('stages', [])):
            stage_name = stage['name']
            stage_score = 0
            cp_col = 0
            
            # Считаем баллы этапа
            for cp in self.data.get('checkpoints', []):
                if 'stages' in cp and stage_name in cp['stages']:
                    if cp['name'] in taken_cps and self.is_cp_available(cp, crew_classification):
                        stage_score += int(cp.get('score', 0))
            
            # Добавляем баллы этапа с цветом
            stage_item = QTableWidgetItem(str(stage_score))
            stage_item.setBackground(stage_colors[stage_idx % len(stage_colors)])
            stage_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.results_table.setItem(row, current_col, stage_item)
            current_col += 1
            
            # Добавляем КП этапа с цветом
            for cp in self.data.get('checkpoints', []):
                if 'stages' in cp and stage_name in cp['stages']:
                    item = QTableWidgetItem()
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    if not self.is_cp_available(cp, crew_classification):
                        item.setText("-")
                        item.setForeground(Qt.GlobalColor.gray)
                    elif cp['name'] in taken_cps:
                        item.setText(str(int(cp.get('score', 0))))
                    else:
                        item.setText("")  # Пустая строка вместо "0"
                    
                    item.setBackground(stage_colors[stage_idx % len(stage_colors)].lighter(130))
                    self.results_table.setItem(row, current_col + cp_col, item)
                    cp_col += 1
            
            current_col += cp_col






############## ЭКСПОРТ В EXCEL ##############

    # В классе RaceApp

    def save_results(self):
        """Экспорт результатов в Excel файл с форматированием Times New Roman и цветами."""
        try:
            if not hasattr(self, 'data') or not self.data.get('members'):
                raise ValueError("Нет данных для экспорта")

            # Создаем имя файла на основе метаданных
            meta = self.data.get('meta', {})
            event_name = meta.get('name', 'Соревнования').replace(' ', '_')
            event_date = meta.get('date', '')

            # Добавляем текущее время к имени файла
            now = datetime.datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            file_name = f"{event_name}_{event_date}_{now}.xlsx" if event_date else f"{event_name}_{now}.xlsx"

            # Создаем папку exports если ее нет
            exports_dir = os.path.join(self.app_dir, 'exports') # Используем app_dir
            os.makedirs(exports_dir, exist_ok=True)

            file_path = os.path.join(exports_dir, file_name)

            # Создаем книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Результаты"

            # --- Стили для оформления ---
            HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Серый
            GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")   # Золотой
            SILVER_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid") # Серебряный
            BRONZE_FILL = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid") # Бронзовый
            DNF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Светло-красный для DNF

            # Цвета для этапов (в формате HEX для openpyxl)
            STAGE_COLORS_HEX = ["ADD8E6", "90EE90", "FFB6C1", "FFFF99"] # LightBlue, LightGreen, LightPink, LightYellow
            STAGE_FILLS = [PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid") for hex_color in STAGE_COLORS_HEX]

            # Функция для получения более светлого цвета этапа
            def get_lighter_stage_fill(base_fill):
                # Используем QColor для осветления, затем конвертируем обратно в HEX
                q_color = QColor(f"#{base_fill.start_color.rgb[2:]}") # Получаем HEX из ARGB
                lighter_q_color = q_color.lighter(130)
                lighter_hex = lighter_q_color.name()[1:] # Убираем #
                return PatternFill(start_color=lighter_hex, end_color=lighter_hex, fill_type="solid")

            # Границы для ячеек
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Шрифт
            default_font = Font(name='Times New Roman', size=11) # <--- ИЗМЕНЕН ШРИФТ
            bold_font = Font(name='Times New Roman', size=11, bold=True)

            # --- Запись заголовков ---
            headers = []
            # Словарь для хранения информации о колонках этапов/КП для раскраски
            column_color_info = {} # { excel_col_index: {'type': 'stage'/'cp', 'stage_index': int, 'is_stage_total': bool} }
            is_staged = self.data.get('logic_params', {}).get('staged', False)
            stages_info = self.data['logic_params'].get('stages', []) if is_staged else []
            stage_name_to_index = {stage['name']: i for i, stage in enumerate(stages_info)}

            for col in range(self.results_table.columnCount()):
                header_text = self.results_table.horizontalHeaderItem(col).text()
                headers.append(header_text)
                excel_col_index = col + 1 # Индексы openpyxl начинаются с 1

                # Записываем заголовок и базовое форматирование
                cell = ws.cell(row=1, column=excel_col_index, value=header_text)
                cell.font = bold_font # Заголовки жирным
                cell.fill = HEADER_FILL
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Сохраняем информацию для раскраски этапов/КП
                if is_staged:
                    if header_text in stage_name_to_index:
                        stage_idx = stage_name_to_index[header_text]
                        column_color_info[excel_col_index] = {'type': 'stage', 'stage_index': stage_idx, 'is_stage_total': True}
                    else:
                        # Проверяем, является ли столбец КП этапа
                        for stage_idx, stage in enumerate(stages_info):
                             stage_name = stage['name']
                             for cp in self.data.get('checkpoints', []):
                                 if header_text == cp['name'] and 'stages' in cp and stage_name in cp['stages']:
                                     column_color_info[excel_col_index] = {'type': 'cp', 'stage_index': stage_idx}
                                     break # Нашли, выходим из внутреннего цикла
                             if excel_col_index in column_color_info: break # Нашли, выходим из внешнего цикла

            # --- Запись данных ---
            for row in range(self.results_table.rowCount()):
                place_text = self.results_table.item(row, 0).text() # Получаем текст места
                is_dnf_row = (place_text == "DNF")

                for col in range(self.results_table.columnCount()):
                    item = self.results_table.item(row, col)
                    excel_col_index = col + 1
                    cell_value = item.text() if item else ""

                    cell = ws.cell(row=row + 2, column=excel_col_index, value=cell_value)
                    cell.font = default_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center') # Центрируем все по умолчанию

                    # Применяем стили в правильном порядке приоритета

                    # 1. Стиль DNF (перекрывает все остальное, кроме призовых мест)
                    if is_dnf_row and col != 0: # Не красим столбец "Место" если DNF
                         cell.fill = DNF_FILL
                    # 2. Стиль призовых мест (перекрывает DNF для столбца "Место")
                    elif col == 0:
                         if place_text == "1": cell.fill = GOLD_FILL
                         elif place_text == "2": cell.fill = SILVER_FILL
                         elif place_text == "3": cell.fill = BRONZE_FILL
                         elif place_text == "DNF": cell.fill = DNF_FILL # Явно красим DNF в столбце Место
                    # 3. Стиль этапов/КП (если не DNF и не призовое место)
                    elif excel_col_index in column_color_info:
                         color_info = column_color_info[excel_col_index]
                         stage_idx = color_info['stage_index']
                         base_fill = STAGE_FILLS[stage_idx % len(STAGE_FILLS)]

                         if color_info['type'] == 'stage' and color_info.get('is_stage_total', False):
                             cell.fill = base_fill # Цвет самого этапа (итог)
                         elif color_info['type'] == 'cp':
                             cell.fill = get_lighter_stage_fill(base_fill) # Осветленный цвет для КП этапа

            # --- Настройка ширины столбцов ---
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                header_text = headers[col-1]
                width = 15 # Базовая ширина

                # Применяем ширины на основе заголовков (пример)
                if header_text == 'Место' or header_text == 'Номер': width = 8
                elif header_text == 'Итог' or header_text == 'Баллы': width = 8
                elif header_text == 'Зачет': width = 12
                elif header_text == 'Пилот' or header_text == 'Штурман': width = 25
                elif header_text == 'Авто': width = 18
                elif 'Время' in header_text or 'Старт' in header_text or 'Финиш' in header_text or 'Нейтрализация' in header_text: width = 14
                elif 'КП' in header_text or 'CP' in header_text: width = 8 # Короткая для КП

                # Подбираем ширину по содержимому, если она больше текущей
                ws.column_dimensions[column_letter].auto_size = True
                calculated_width = ws.column_dimensions[column_letter].width
                ws.column_dimensions[column_letter].auto_size = False # Отключаем автоподбор после расчета
                ws.column_dimensions[column_letter].width = max(width, calculated_width + 2) # Берем максимум + небольшой запас

            # Замораживаем первую строку (заголовки)
            ws.freeze_panes = 'A2'

            # Сохраняем файл
            wb.save(file_path)

            # Открываем файл автоматически (может не работать на всех ОС)
            try:
                if sys.platform == "win32":
                    os.startfile(file_path)
                elif sys.platform == "darwin": # macOS
                    subprocess.call(["open", file_path])
                else: # linux variants
                    subprocess.call(["xdg-open", file_path])
            except Exception as open_err:
                print(f"Не удалось автоматически открыть файл: {open_err}")
                # Показываем путь, чтобы пользователь мог открыть вручную
                QMessageBox.information(self, "Информация", f"Файл сохранен, но не удалось открыть автоматически:\n{file_path}")


            QMessageBox.information(self, "Успех", f"Результаты экспортированы в файл:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать результаты:\n{str(e)}")
            traceback.print_exc() # Печатаем traceback для отладки


# В классе RaceApp







if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        window = RaceApp()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        print(f"Ошибка: {e}")
        input("Нажмите Enter для выхода...")






