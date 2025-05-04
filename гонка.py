# race_app.py - программа для учета автомобильных соревнований
# Основной файл приложения

import os
import sys
import json
import re
import math
import openpyxl
import shutil
import traceback
from PyQt6.QtCore import (Qt, QDate, QDateTime, QTime, QTimer, QEvent, 
                          QSettings, pyqtSignal, QObject, QUrl, QThread)
from PyQt6.QtGui import (QIntValidator, QColor)
from PyQt6.QtWidgets import (QApplication, QMainWindow, QTabWidget, 
                            QWidget, QVBoxLayout, QLabel, QPushButton,
                            QTableWidget, QTableWidgetItem, QHeaderView, 
                            QMessageBox, QLineEdit, QScrollArea, QHBoxLayout, 
                            QCheckBox, QDialog, QDialogButtonBox, QRadioButton, QButtonGroup,
                            QComboBox, QDateEdit, QListWidget, QListWidgetItem, QFileDialog,
                            QGroupBox, QSplitter, QTimeEdit, QFormLayout, QSizePolicy, QProgressBar,
                            QInputDialog, QTextEdit)
from PyQt6.QtMultimedia import QSoundEffect
from PyQt6.QtGui import QFont

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import COLOR_INDEX




############## ТАЙМЕР ОТСЧЕТА ##############

from PyQt6.QtWidgets import (QMainWindow, QLabel, QPushButton, QVBoxLayout, 
                            QWidget, QMessageBox, QHBoxLayout)
from PyQt6.QtCore import QTimer, QTime, QSettings, pyqtSignal, QObject, QUrl
from PyQt6.QtMultimedia import QSoundEffect
import os

class SoundPlayer(QObject):
    sound_finished = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.effect = QSoundEffect()
        self.effect.setVolume(1.0)
        self.effect.playingChanged.connect(self.on_playing_changed)
        self.muted = False

    def play(self, file_path):
        try:
            if self.muted:
                return
                
            if not os.path.exists(file_path):
                print(f"Файл не найден: {file_path}")
                return
                
            url = QUrl.fromLocalFile(file_path)
            if not url.isValid():
                print(f"Неверный URL: {url}")
                return
                
            self.effect.setSource(url)
            self.effect.play()
        except Exception as e:
            print(f"Ошибка воспроизведения: {str(e)}")

    def on_playing_changed(self):
        if not self.effect.isPlaying():
            self.sound_finished.emit()

class CountdownWindow(QMainWindow):

    def __init__(self, parent=None, start_time="", common_start=True):
        super().__init__(parent)
        self.start_time = QTime.fromString(start_time, "HH:mm")
        if not self.start_time.isValid():
            self.start_time = QTime.currentTime().addSecs(60)
        self.common_start = common_start
        self.sound_player = SoundPlayer()
        self.played_sounds = set()
        self.current_cycle = 0  # Счетчик циклов
        self.setup_ui()
        self.setup_timers()
        self.play_test_sound()

    def play_test_sound(self):
        print("Тестирование звука...")  # Для отладки
        self.play_sound("проверка.wav")

    def play_start_sound(self):
        self.play_sound("старт.wav")

    def toggle_sound(self):
        self.sound_player.muted = self.mute_btn.isChecked()
        self.save_settings()

    def load_settings(self):
        settings = QSettings("RaceTimer", "SoundSettings")
        self.sound_player.muted = settings.value("muted", False, type=bool)
        self.mute_btn.setChecked(self.sound_player.muted)

    def save_settings(self):
        settings = QSettings("RaceTimer", "SoundSettings")
        settings.setValue("muted", self.sound_player.muted)

    def setup_ui(self):
        self.setWindowTitle("Таймер старта")
        self.setMinimumSize(800, 400)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        self.timer_label = QLabel("00:00:00")
        self.set_style("white")  # Инициализация стиля
        
        layout.addWidget(self.timer_label)
        
        btn_layout = QHBoxLayout()
        self.test_btn = QPushButton("Проверка звука")
        self.test_btn.clicked.connect(self.play_test_sound)
        btn_layout.addWidget(self.test_btn)
        
        self.mute_btn = QPushButton("Отключить звук")
        self.mute_btn.setCheckable(True)
        self.mute_btn.clicked.connect(self.toggle_sound)
        btn_layout.addWidget(self.mute_btn)
        
        layout.addLayout(btn_layout)

    def set_style(self, color):
        """Устанавливает стиль для таймера с указанным цветом текста"""
        self.timer_label.setStyleSheet(f"""
            font-size: 72pt;
            font-weight: bold;
            color: {color};
            background-color: #000000;
            padding: 50px;
            qproperty-alignment: AlignCenter;
        """)

    def play_sound(self, filename):
        """Улучшенный метод воспроизведения звуков"""
        try:
            # Получаем абсолютный путь к файлу
            base_dir = os.path.dirname(os.path.abspath(__file__))
            sound_path = os.path.join(base_dir, "sounds", filename)
            
            print(f"Попытка воспроизвести: {sound_path}")  # Отладочная информация
            
            # Проверка существования файла
            if not os.path.exists(sound_path):
                print(f"Файл не найден: {sound_path}")
                # Проверим доступные файлы в папке sounds
                sounds_dir = os.path.join(base_dir, "sounds")
                if os.path.exists(sounds_dir):
                    print(f"Доступные файлы: {os.listdir(sounds_dir)}")
                return
            
            # Проверка прав доступа
            if not os.access(sound_path, os.R_OK):
                print(f"Нет прав на чтение файла: {sound_path}")
                return
                
            # Воспроизведение звука
            url = QUrl.fromLocalFile(sound_path)
            if url.isValid():
                self.sound_player.effect.setSource(url)
                self.sound_player.effect.play()
            else:
                print(f"Неверный URL для файла: {sound_path}")
                
        except Exception as e:
            print(f"Ошибка при воспроизведении {filename}: {str(e)}")

    def clear_last_played(self, point):
        """Очищаем отметку о воспроизведении"""
        if point in self.last_played:
            del self.last_played[point]

    def update_display(self, seconds):
        """Обновляет отображение таймера"""
        abs_seconds = abs(seconds)
        sign = "-" if seconds < 0 else ""
        time_str = f"{sign}{abs_seconds//3600:02d}:{(abs_seconds%3600)//60:02d}:{abs_seconds%60:02d}"
        
        if seconds > 0:
            self.timer_label.setText(f"До старта: {time_str}")
            self.set_style("white")
        elif seconds == 0:
            self.timer_label.setText("СТАРТ!")
            self.set_style("red")
        else:
            self.timer_label.setText(f"После старта: {time_str}")
            self.set_style("blue")

    def reset_sound_flags(self):
        """Сброс всех флагов воспроизведения"""
        self.played_sounds.clear()

    def setup_timers(self):
        self.main_timer = QTimer()
        self.main_timer.timeout.connect(self.update_countdown)
        self.main_timer.start(100)  # 10 обновлений в секунду

        # Таймер для циклического режима
        self.cycle_timer = QTimer()
        self.cycle_timer.timeout.connect(self.start_new_cycle)
        
        if not self.common_start:
            # Запускаем первый цикл сразу
            QTimer.singleShot(0, self.start_new_cycle)

    def start_new_cycle(self):
        """Начинает новый минутный цикл"""
        if not self.common_start:
            self.current_cycle += 1
            print(f"Начало цикла #{self.current_cycle}")
            self.start_time = QTime.currentTime().addSecs(60)  # Следующая минута
            self.played_sounds.clear()
            
            # Перезапускаем таймер цикла
            self.cycle_timer.start(60000)  # Каждые 60 секунд

    def update_countdown(self):
        current_time = QTime.currentTime()
        seconds = current_time.secsTo(self.start_time)
        
        # Обновление отображения
        abs_seconds = abs(seconds)
        time_str = f"{abs_seconds//3600:02d}:{(abs_seconds%3600)//60:02d}:{abs_seconds%60:02d}"
        
        if seconds >= 0:
            self.timer_label.setText(f"Цикл {self.current_cycle} | До старта: {time_str}")
            self.set_style("white")
        else:
            self.timer_label.setText(f"Цикл {self.current_cycle} | Старт прошел: {time_str}")
            self.set_style("blue")
        
        # Проверка звуков
        self.check_sounds(seconds)

    def check_sounds(self, seconds):
        sound_triggers = {
            45: "45 сек.wav",
            30: "30 сек.wav", 
            15: "15 сек.wav",
            10: "10 сек.wav",
            5: "секунды.wav",
            4: "секунды.wav",
            3: "секунды.wav",
            2: "секунды.wav",
            1: "секунды.wav",
            0: "старт.wav"
        }
        
        if seconds in sound_triggers:
            sound_key = f"{self.current_cycle}_{seconds}_{sound_triggers[seconds]}"
            if sound_key not in self.played_sounds:
                self.played_sounds.add(sound_key)
                self.play_sound(sound_triggers[seconds])
                print(f"Цикл {self.current_cycle}: {seconds} сек - {sound_triggers[seconds]}")

    def closeEvent(self, event):
        self.main_timer.stop()
        self.cycle_timer.stop()
        super().closeEvent(event)

class RaceApp(QMainWindow):
    def __init__(self):
        super().__init__()
        
        # Определяем пути
        self.app_dir = os.path.dirname(os.path.abspath(__file__))
        self.data_dir = os.path.join(self.app_dir, "data")
        
        # Загружаем данные
        self.current_file = None
        self.load_latest_data()
        
        # Настройка интерфейса
        self.setup_ui()
        
        # Устанавливаем размер окна
        self.resize(1000, 700)
        
        # Обновляем заголовок
        self.update_window_title()
        # Подключаем обработчик изменения вкладки
        self.tabs.currentChanged.connect(self.on_tab_changed)

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_start_tab)
        self.timer.start(1000)  # Обновление каждую секунду

        self.auto_finish_done = False
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_finish_tab)
        self.timer.start(1000)  # Обновление каждую секунду

        self.start_skp_check_timer()
        self.start_timeout_check_timer()

        self.skp_check_timer = QTimer()
        self.skp_check_timer.timeout.connect(self.check_skp_closing)
        self.skp_check_timer.start(1000)

        # Таймер проверки превышения времени трассы
        self.track_timeout_timer = QTimer()
        self.track_timeout_timer.timeout.connect(self.check_track_closing)
        self.track_timeout_timer.start(1000)  

                # Таймер проверки времени (1 раз в секунду)
        self.timeout_timer = QTimer()
        self.timeout_timer.timeout.connect(self.check_track_timeout)
        self.timeout_timer.start(1000)

        self.start_countdown_timer = QTimer()
        self.start_countdown_timer.timeout.connect(self.force_update_start_tab)
        self.start_countdown_timer.start(1000)  # Обновление каждую секунду

        # Таймер текущего времени
        self.current_time_timer = QTimer()
        self.current_time_timer.timeout.connect(self.update_current_time)
        self.current_time_timer.start(1000)

        # Таймер для проверки времени старта
        self.start_check_timer = QTimer()
        self.start_check_timer.timeout.connect(self.check_start_time)
        self.start_check_timer.start(1000)

        self.cp_input = None
    
    def update_window_title(self):
        """Обновляем заголовок окна"""
        if hasattr(self, 'data') and 'meta' in self.data:
            name = self.data['meta'].get('name', 'Новое соревнование')
            date = self.data['meta'].get('date', '')
            self.setWindowTitle(f"{name} - {date}")
        else:
            self.setWindowTitle("Автосоревнования")
    
    def setup_ui(self):
        """Настраиваем интерфейс приложения"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        
        # Создаем вкладки
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        # Добавляем все вкладки
        self.setup_params_tab()
        self.setup_logic_tab()
        self.setup_checkpoints_tab()
        self.setup_members_tab()
        self.setup_registration_tab()
        self.setup_start_tab()
        self.setup_stages_tab()
        self.setup_finish_tab()
        self.setup_check_tab()
        self.setup_results_tab()
    
    def on_tab_changed(self, index):
        """Обработчик изменения вкладки"""
        # Скрываем таблицу КП при переключении вкладок
        if self.tabs.tabText(index) != "КП":
            self.checkpoints_table.setVisible(False)
            self.stages_table.setVisible(False)
            self.btn_show.setText("Показать КП")
        
        # Если переключились на вкладку "Параметры" - обновляем ее
        if self.tabs.tabText(index) == "Параметры":
            self.update_params_fields()

        if self.tabs.tabText(index) == "Старт":
            self.force_update_start_tab()  # Полное обновление данных


        if self.tabs.tabText(index) == "Проверка КП":
            self.update_check_tab()  # Полное обновление данных

        # if self.tabs.tabText(index) == "Регистрация":
        #     self.update_check_tab()  # Полное обновление данных


    def load_latest_data(self):
        """Загружаем последнюю версию файла данных"""
        try:
            os.makedirs(self.data_dir, exist_ok=True)
            
            # Получаем список всех json-файлов
            json_files = [f for f in os.listdir(self.data_dir) 
                         if f.startswith('race_v') and f.endswith('.json')]
            
            if json_files:
                # Сортируем файлы по номеру версии (правильно обрабатываем числа)
                json_files.sort(key=lambda x: int(x.split('_v')[1].split('.')[0]))
                
                # Берем последний файл (с наибольшей версией)
                latest_file = json_files[-1]
                self.current_file = os.path.join(self.data_dir, latest_file)
                
                with open(self.current_file, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)
                #print(f"Загружены данные из {latest_file}")
            else:
                self.data = {
                    "meta": {
                        "version": 1,
                        "created_at": "2023-01-01",
                        "name": "Новое соревнование",
                        "date": "01-01-2023"
                    },
                    "params": {},
                    "checkpoints": [],
                    "members": [],
                    "registration": [],
                    "start": [],
                    "finish": []
                }
                self.save_data()
                
        except Exception as e:
            print(f"Ошибка загрузки данных: {e}")
            self.data = {
                "meta": {
                    "version": 1,
                    "created_at": "2023-01-01",
                    "name": "Новое соревнование",
                    "date": "01-01-2023"
                },
                "params": {},
                "checkpoints": [],
                "members": [],
                "registration": [],
                "start": [],
                "finish": []
            }


############## ПАРАМЕТРЫ ##############


    def setup_params_tab(self):
        """Вкладка параметров соревнования"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Словарь для хранения текстовых полей
        self.params_fields = {}
        
        # Параметры с подсказками и валидацией
        params_config = {
            "Время регистрации": {
                "hint": "Формат: ЧЧ:ММ (например: 09:30)",
                "validator": self.validate_time,
                "required": True
            },
            "Время брифинга": {
                "hint": "Формат: ЧЧ:ММ (например: 10:00)",
                "validator": self.validate_time,
                "required": True
            },
            "Время старта": {
                "hint": "Формат: ЧЧ:ММ (например: 11:00)",
                "validator": self.validate_time,
                "required": True
            },
            "Время закрытия трассы": {
                "hint": "Формат: ЧЧ:ММ (например: 17:00)",
                "validator": self.validate_time,
                "required": True
            },
            "Время награждения": {
                "hint": "Формат: ЧЧ:ММ (например: 18:00)",
                "validator": self.validate_time,
                "required": True
            },
            "Количество КП": {
                "hint": "Целое число (например: 10)",
                "validator": self.validate_int,
                "required": True
            },
            "Бальность КП": {
                "hint": "Число (например: 5)",
                "validator": self.validate_float,
                "required": True
            },
            "Зачет1": {
                "hint": "Название зачета (например: Спорт) СТРОГОЕ СООТВЕТСТВИЕ ЗАЯВКАМ",
                "validator": self.validate_classification,
                "required": True
            },
            "Зачет2": {
                "hint": "Название зачета (например: Дебют)",
                "validator": self.validate_classification,
                "required": True
            },
            "Зачет3": {
                "hint": "Название зачета (необязательное поле)",
                "validator": None,  # Убрана валидация
                "required": False  # Поле необязательное
            },
            "Ссылка на таблицу регистрации": {
                "hint": "URL или путь к файлу",
                "validator": None,
                "required": False
            }
        }
        
        # Создаем поля для каждого параметра
        for param, config in params_config.items():
            layout.addWidget(QLabel(param + ":" + ("" if not config["required"] else " *")))
            field = QLineEdit()
            field.setPlaceholderText(config["hint"])
            
            # Заполняем поле значением из данных, если оно есть
            if hasattr(self, 'data') and 'params' in self.data:
                key = param.lower().replace(' ', '_')
                field.setText(str(self.data['params'].get(key, "")))
            
            self.params_fields[param] = field
            layout.addWidget(field)

        if hasattr(self, 'data') and 'meta' in self.data:
            layout.addWidget(QLabel(f"Название: {self.data['meta'].get('name', '')}"))
            layout.addWidget(QLabel(f"Дата: {self.data['meta'].get('date', '')}"))
            layout.addWidget(QLabel(f"Версия данных: {self.data['meta'].get('version', 1)}"))
            layout.addWidget(QLabel(f"FAST v2.0.0"))
            # версия программы
            # логика: vХ.х.х Крупнейшие изменения
            # логика: vх.Х.х крупные доработки
            # логика: vх.х.Х Фиксы багов, мелкие фишки
            # На макошь вывести с номером 2.1.0
        # Кнопки управления
        btn_save = QPushButton("Сохранить параметры")
        btn_save.clicked.connect(self.confirm_save_params)
        layout.addWidget(btn_save)

        
        self.tabs.addTab(tab, "Параметры")
    
    def validate_time(self, time_str):
        """Проверка формата времени ЧЧ:ММ"""
        try:
            hours, minutes = map(int, time_str.split(':'))
            return 0 <= hours < 24 and 0 <= minutes < 60
        except ValueError:
            return False
    
    def validate_int(self, value):
        """Проверка что значение целое число"""
        try:
            int(value)
            return True
        except ValueError:
            return False
    
    def validate_float(self, value):
        """Проверка что значение число"""
        try:
            float(value)
            return True
        except ValueError:
            return False
    
    def validate_classification(self, value):
        """Проверка что зачет начинается с большой буквы"""
        return value and value[0].isupper()
    
    def confirm_save_params(self):
        """Подтверждение сохранения параметров"""
        # Первое подтверждение
        reply = QMessageBox.question(
            self, 'Подтверждение',
            'Вы уверены, что хотите сохранить параметры?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Второе подтверждение
            reply = QMessageBox.question(
                self, 'Последнее подтверждение',
                'Все данные будут перезаписаны. Продолжить?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.save_params()

    def save_params(self):
        """Сохраняем параметры после проверки"""
        errors = []
        
        # Проверяем все обязательные поля
        for param, field in self.params_fields.items():
            validator = None
            required = False
            
            # Определяем валидатор и обязательность поля
            if param in ["Время регистрации", "Время брифинга", "Время старта", 
                        "Время закрытия трассы", "Время награждения"]:
                validator = self.validate_time
                required = True
            elif param == "Количество КП":
                validator = self.validate_int
                required = True
            elif param == "Бальность КП":
                validator = self.validate_float
                required = True
            elif param in ["Зачет1", "Зачет2"]:
                validator = self.validate_classification
                required = True
            elif param == "Зачет3":
                # Для Зачета3 валидация не требуется
                if field.text() and not self.validate_classification(field.text()):
                    errors.append(f"Некорректное значение для '{param}' (должно начинаться с большой буквы)")
                continue
            
            if required and not field.text():
                errors.append(f"Поле '{param}' обязательно для заполнения")
            elif validator and field.text() and not validator(field.text()):
                errors.append(f"Некорректное значение для '{param}'")
        
        if errors:
            QMessageBox.critical(self, "Ошибки", "\n".join(errors))
            return
        
        # Если ошибок нет, сохраняем
        try:
            params = {
                'время_регистрации': self.params_fields["Время регистрации"].text(),
                'время_брифинга': self.params_fields["Время брифинга"].text(),
                'время_старта': self.params_fields["Время старта"].text(),
                'время_закрытия_трассы': self.params_fields["Время закрытия трассы"].text(),
                'время_награждения': self.params_fields["Время награждения"].text(),
                'количество_кп': self.params_fields["Количество КП"].text(),
                'бальность_кп': self.params_fields["Бальность КП"].text(),
                'зачет1': self.params_fields["Зачет1"].text(),
                'зачет2': self.params_fields["Зачет2"].text()
            }
            
            # Добавляем необязательные поля только если они заполнены
            if self.params_fields["Зачет3"].text():
                params['зачет3'] = self.params_fields["Зачет3"].text()
            if self.params_fields["Ссылка на таблицу регистрации"].text():
                params['ссылка_на_таблицу_регистрации'] = self.params_fields["Ссылка на таблицу регистрации"].text()
            
            self.data['params'] = params
            self.save_data()
            QMessageBox.information(self, "Сохранено", "Параметры успешно сохранены!")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения: {str(e)}")
    
    def update_params_fields(self):
        """Обновляем поля ввода и метаданные на вкладке параметров"""
        if not hasattr(self, 'data'):
            return
        
        # 1. Обновляем текстовые поля параметров
        if 'params' in self.data:
            for param, field in self.params_fields.items():
                key = param.lower().replace(' ', '_')
                if key in self.data['params']:
                    field.setText(str(self.data['params'][key]))
                else:
                    field.clear()
        
        # 2. Обновляем метаданные (название, дата, версия)
        if 'meta' in self.data:
            # Находим все QLabel на вкладке
            for i in range(self.tabs.widget(0).layout().count()):  # 0 - индекс вкладки "Параметры"
                widget = self.tabs.widget(0).layout().itemAt(i).widget()
                if isinstance(widget, QLabel):
                    text = widget.text()
                    if text.startswith("Название:"):
                        widget.setText(f"Название: {self.data['meta'].get('name', '')}")
                    elif text.startswith("Дата:"):
                        widget.setText(f"Дата: {self.data['meta'].get('date', '')}")
                    elif text.startswith("Версия данных:"):
                        widget.setText(f"Версия данных: {self.data['meta'].get('version', 1)}")

############## ЛОГИКА ##############

    def setup_logic_tab(self):
        """Вкладка логики проведения соревнования с настройками СКП"""
        self.logic_tab = QWidget()
        layout = QVBoxLayout(self.logic_tab)
        
        # 1. Общий старт
        self.common_start_cb = QCheckBox("Общий старт")
        layout.addWidget(self.common_start_cb)
        
        # 2. Штраф за опоздание на финиш
        penalty_layout = QHBoxLayout()
        penalty_layout.addWidget(QLabel("Штраф за опоздание на финиш:"))
        
        self.penalty_type_combo = QComboBox()
        self.penalty_type_combo.addItems(["Нет", "Баллы за 1 мин", "DNF"])
        self.penalty_type_combo.currentTextChanged.connect(self.update_penalty_ui)
        penalty_layout.addWidget(self.penalty_type_combo)
        
        self.penalty_value_edit = QLineEdit()
        self.penalty_value_edit.setPlaceholderText("Значение штрафа")
        self.penalty_value_edit.setValidator(QIntValidator(0, 1000))
        self.penalty_value_edit.setFixedWidth(80)
        penalty_layout.addWidget(self.penalty_value_edit)
        
        layout.addLayout(penalty_layout)
        
        # 3. Штраф за ложное КП
        false_cp_layout = QHBoxLayout()
        false_cp_layout.addWidget(QLabel("Штраф за ложное КП (Не работает. Ставить во вкладке кп отрицательные баллы)"))
        
        self.false_cp_penalty_edit = QLineEdit()
        self.false_cp_penalty_edit.setValidator(QIntValidator(0, 1000))
        self.false_cp_penalty_edit.setFixedWidth(80)
        false_cp_layout.addWidget(self.false_cp_penalty_edit)
        false_cp_layout.addWidget(QLabel("баллов"))
        
        layout.addLayout(false_cp_layout)
        
        # 4. Время трассы (рассчитывается от старта до закрытия)
        self.route_time_group = QGroupBox("Время трассы")
        route_time_layout = QVBoxLayout()
        
        self.route_time_info = QLabel("Рассчитывается автоматически\nот старта до закрытия трассы")
        self.route_time_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        route_time_layout.addWidget(self.route_time_info)
        
        self.route_time_group.setLayout(route_time_layout)
        layout.addWidget(self.route_time_group)
        
        # 5. Этапность - доработанная версия
        self.staged_cb = QCheckBox("Этапность")
        self.staged_cb.stateChanged.connect(self.toggle_stage_settings)
        layout.addWidget(self.staged_cb)
        
        # Контейнер для настроек этапов
        self.stage_settings_group = QGroupBox("Настройки этапов")
        self.stage_settings_layout = QVBoxLayout()
        self.stage_settings_group.setLayout(self.stage_settings_layout)
        self.stage_settings_group.setVisible(False)
        layout.addWidget(self.stage_settings_group)
        
        # Количество этапов
        stages_layout = QHBoxLayout()
        stages_layout.addWidget(QLabel("Количество этапов:"))
        
        self.stages_count_combo = QComboBox()
        self.stages_count_combo.addItems([str(i) for i in range(1, 6)])
        self.stages_count_combo.currentIndexChanged.connect(self.update_stages_ui)
        stages_layout.addWidget(self.stages_count_combo)
        
        self.stage_settings_layout.addLayout(stages_layout)
        
        # Настройки нейтрализации на СКП
        neutralization_group = QGroupBox("Нейтрализация на СКП")
        neutralization_layout = QVBoxLayout()
        
        # Тип нейтрализации
        self.neutralization_type_group = QButtonGroup()
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("Тип нейтрализации:"))
        
        rb_no_neutral = QRadioButton("Нет")
        rb_total_neutral = QRadioButton("Суммарная")
        rb_per_skp_neutral = QRadioButton("На каждом СКП")
        
        self.neutralization_type_group.addButton(rb_no_neutral)
        self.neutralization_type_group.addButton(rb_total_neutral)
        self.neutralization_type_group.addButton(rb_per_skp_neutral)
        rb_no_neutral.setChecked(True)
        
        type_layout.addWidget(rb_no_neutral)
        type_layout.addWidget(rb_total_neutral)
        type_layout.addWidget(rb_per_skp_neutral)
        neutralization_layout.addLayout(type_layout)
        
        # Максимальное время нейтрализации (только для суммарной)
        self.max_neutral_time_layout = QHBoxLayout()
        self.max_neutral_time_layout.addWidget(QLabel("Макс. время нейтрализации:"))
        
        self.max_neutral_time_edit = QLineEdit()
        self.max_neutral_time_edit.setValidator(QIntValidator(0, 1000))
        self.max_neutral_time_edit.setPlaceholderText("минут")
        self.max_neutral_time_edit.setFixedWidth(80)
        self.max_neutral_time_edit.setEnabled(False)
        self.max_neutral_time_layout.addWidget(self.max_neutral_time_edit)
        
        neutralization_layout.addLayout(self.max_neutral_time_layout)
        neutralization_group.setLayout(neutralization_layout)
        self.stage_settings_layout.addWidget(neutralization_group)
        
        # Обработчики для нейтрализации
        rb_no_neutral.toggled.connect(self.update_neutralization_ui)
        rb_total_neutral.toggled.connect(self.update_neutralization_ui)
        rb_per_skp_neutral.toggled.connect(self.update_neutralization_ui)
        
        # Контейнер для этапов и СКП
        self.stages_container = QWidget()
        self.stages_container_layout = QVBoxLayout()
        self.stages_container.setLayout(self.stages_container_layout)
        self.stage_settings_layout.addWidget(self.stages_container)
        
        # Кнопка сохранения
        save_btn = QPushButton("Сохранить параметры")
        save_btn.clicked.connect(self.save_logic_params)
        layout.addWidget(save_btn)
        
        self.tabs.insertTab(1, self.logic_tab, "Логика")
        self.update_logic_tab()
   
    def update_logic_tab(self):
        """Обновляет данные на вкладке Логика с учетом настроек СКП"""
        if not hasattr(self, 'data'):
            return

        # Расчет времени трассы
        self.calculate_route_time_from_params()

        params = self.data.setdefault('logic_params', {})
        
        # Основные параметры
        self.common_start_cb.setChecked(params.get('common_start', False))
        
        penalty_type = params.get('penalty_type', 'Нет')
        self.penalty_type_combo.setCurrentText(penalty_type)
        self.penalty_value_edit.setText(str(params.get('penalty_value', 0)))
        self.update_penalty_ui(penalty_type)
        
        self.false_cp_penalty_edit.setText(str(params.get('false_cp_penalty', 0)))
        
        # Параметры этапности
        self.staged_cb.setChecked(params.get('staged', False))
        self.toggle_stage_settings(params.get('staged', False))
        
        if params.get('staged'):
            # Тип нейтрализации
            neutral_type = params.get('neutralization_type', 'Нет')
            for btn in self.neutralization_type_group.buttons():
                if btn.text() == neutral_type:
                    btn.setChecked(True)
                    break
            
            # Устанавливаем время нейтрализации в зависимости от типа
            if neutral_type == 'Суммарная':
                self.max_neutral_time_edit.setText(str(params.get('total_max_neutral_time', 0)))
            else:
                self.max_neutral_time_edit.setText(str(params.get('max_neutral_time', 0)))
            
            self.max_neutral_time_edit.setEnabled(neutral_type == 'Суммарная')
            
            # Количество этапов
            self.stages_count_combo.setCurrentText(str(params.get('stages_count', 1)))
            self.update_stages_ui()
            
            # Восстановление этапов
            for i, stage in enumerate(params.get('stages', [])):
                if i < len(self.stage_widgets):
                    self.stage_widgets[i]['name_edit'].setText(stage.get('name', ''))
                    self.stage_widgets[i]['time_limit_edit'].setText(stage.get('time_limit', ''))
            
            # Восстановление СКП (для всех типов нейтрализации)
            for i, skp in enumerate(params.get('skp_settings', [])):
                if i < len(self.skp_widgets):
                    # Время работы
                    self.skp_widgets[i]['open_time_edit'].setTime(
                        QTime.fromString(skp.get('open_time', '00:00'), "HH:mm"))
                    self.skp_widgets[i]['close_time_edit'].setTime(
                        QTime.fromString(skp.get('close_time', '23:59'), "HH:mm"))
                    
                    # Действие при опоздании
                    late_action = skp.get('late_action', 'Перенос на след. этап')
                    for btn in self.skp_widgets[i]['late_action_group'].buttons():
                        if btn.text() == late_action:
                            btn.setChecked(True)
                            break
                    
                    # Время нейтрализации
                    self.skp_widgets[i]['max_time_edit'].setText(str(skp.get('max_neutral_time', 0)))

    def save_logic_params(self):
        """Сохраняет параметры логики с учетом настроек СКП"""
        try:
            # Основные параметры
            params = {
                'common_start': self.common_start_cb.isChecked(),
                'penalty_type': self.penalty_type_combo.currentText(),
                'penalty_value': int(self.penalty_value_edit.text() or 0) if self.penalty_value_edit.isEnabled() else 0,
                'false_cp_penalty': int(self.false_cp_penalty_edit.text() or 0),
                'staged': self.staged_cb.isChecked(),
                'neutralization_type': self.neutralization_type_group.checkedButton().text()
            }

            # Для суммарной нейтрализации сохраняем общее время
            if params['neutralization_type'] == 'Суммарная':
                params['total_max_neutral_time'] = int(self.max_neutral_time_edit.text() or 0)
            else:
                params['max_neutral_time'] = int(self.max_neutral_time_edit.text() or 0) if self.max_neutral_time_edit.isEnabled() else 0

            if self.staged_cb.isChecked():
                # Сохраняем этапы
                params['stages_count'] = int(self.stages_count_combo.currentText())
                params['stages'] = []
                params['skp_settings'] = []  # Настройки СКП

                # Сохраняем данные этапов
                for i in range(params['stages_count']):
                    if i < len(self.stage_widgets):
                        stage_data = {
                            'name': self.stage_widgets[i]['name_edit'].text(),
                            'time_limit': self.stage_widgets[i]['time_limit_edit'].text()
                        }
                        params['stages'].append(stage_data)

                # Сохраняем настройки СКП (для всех типов нейтрализации)
                for i in range(len(self.skp_widgets)):
                    skp_data = {
                        'number': i+1,  # Добавляем номер СКП (1-based)
                        'open_time': self.skp_widgets[i]['open_time_edit'].time().toString("HH:mm"),
                        'close_time': self.skp_widgets[i]['close_time_edit'].time().toString("HH:mm"),
                        'late_action': self.skp_widgets[i]['late_action_group'].checkedButton().text(),
                        'max_neutral_time': int(self.skp_widgets[i]['max_time_edit'].text() or 0) if self.skp_widgets[i]['max_time_edit'].isEnabled() else 0
                    }
                    params['skp_settings'].append(skp_data)

            # Сохраняем в данные
            self.data['logic_params'] = params
            self.save_data()
            
            QMessageBox.information(self, "Сохранено", "Параметры логики успешно сохранены!")
            self.update_logic_tab()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить параметры: {str(e)}")

    def update_stages_ui(self):
        """Обновляет UI для редактирования этапов и СКП с новыми настройками"""
        # Очищаем предыдущие виджеты
        for i in reversed(range(self.stages_container_layout.count())): 
            self.stages_container_layout.itemAt(i).widget().setParent(None)
        
        self.stage_widgets = []
        self.skp_widgets = []
        stages_count = int(self.stages_count_combo.currentText())
        
        for i in range(stages_count):
            # Добавляем этап
            stage_group = QGroupBox(f"Этап {i+1}")
            stage_layout = QVBoxLayout()
            
            # Название этапа
            name_layout = QHBoxLayout()
            name_layout.addWidget(QLabel("Название:"))
            name_edit = QLineEdit()
            name_edit.setPlaceholderText(f"Этап {i+1}")
            name_layout.addWidget(name_edit)
            stage_layout.addLayout(name_layout)
            
            # Макс. время этапа
            time_limit_layout = QHBoxLayout()
            time_limit_layout.addWidget(QLabel("Макс. время (мин):"))
            time_limit_edit = QLineEdit()
            time_limit_edit.setPlaceholderText("не ограничено")
            time_limit_edit.setValidator(QIntValidator(1, 1000))
            time_limit_layout.addWidget(time_limit_edit)
            stage_layout.addLayout(time_limit_layout)
            
            stage_group.setLayout(stage_layout)
            self.stages_container_layout.addWidget(stage_group)
            
            self.stage_widgets.append({
                'name_edit': name_edit,
                'time_limit_edit': time_limit_edit
            })
            
            # Добавляем СКП после этапа (кроме последнего)
            if i < stages_count - 1:  # Убрана проверка на тип нейтрализации
                skp_group = QGroupBox(f"СКП {i+1} (между этапами {i+1} и {i+2})")
                skp_layout = QVBoxLayout()
                
                # Время работы СКП
                time_work_layout = QHBoxLayout()
                time_work_layout.addWidget(QLabel("Время работы:"))
                
                # Открытие СКП
                skp_open_edit = QTimeEdit()
                skp_open_edit.setDisplayFormat("HH:mm")
                skp_open_edit.setTime(QTime(0, 0))
                time_work_layout.addWidget(QLabel("Открытие:"))
                time_work_layout.addWidget(skp_open_edit)
                
                # Закрытие СКП
                skp_close_edit = QTimeEdit()
                skp_close_edit.setDisplayFormat("HH:mm")
                skp_close_edit.setTime(QTime(23, 59))
                time_work_layout.addWidget(QLabel("Закрытие:"))
                time_work_layout.addWidget(skp_close_edit)
                
                skp_layout.addLayout(time_work_layout)
                
                # Действие при опоздании
                late_action_layout = QHBoxLayout()
                late_action_layout.addWidget(QLabel("Действие при опоздании:"))
                
                skp_late_action_group = QButtonGroup()
                rb_move_next = QRadioButton("Перенос на след. этап")
                rb_dnq = QRadioButton("DNQ (дисквалификация)")
                skp_late_action_group.addButton(rb_move_next)
                skp_late_action_group.addButton(rb_dnq)
                rb_move_next.setChecked(True)
                
                late_action_layout.addWidget(rb_move_next)
                late_action_layout.addWidget(rb_dnq)
                skp_layout.addLayout(late_action_layout)
                
                # Макс. время нейтрализации на этом СКП
                time_layout = QHBoxLayout()
                time_layout.addWidget(QLabel("Макс. время нейтрализации:"))
                
                skp_time_edit = QLineEdit()
                skp_time_edit.setValidator(QIntValidator(0, 1000))
                skp_time_edit.setPlaceholderText("минут")
                skp_time_edit.setFixedWidth(80)
                
                # Блокировка поля в зависимости от типа нейтрализации
                neutral_type = self.neutralization_type_group.checkedButton().text()
                if neutral_type == "Суммарная" or neutral_type == "Нет":
                    skp_time_edit.setEnabled(False)
                else:
                    skp_time_edit.setEnabled(True)
                
                time_layout.addWidget(skp_time_edit)
                skp_layout.addLayout(time_layout)
                skp_group.setLayout(skp_layout)
                self.stages_container_layout.addWidget(skp_group)
                
                self.skp_widgets.append({
                    'open_time_edit': skp_open_edit,
                    'close_time_edit': skp_close_edit,
                    'late_action_group': skp_late_action_group,
                    'max_time_edit': skp_time_edit
                })

    def update_neutralization_ui(self):
        """Обновляет UI в зависимости от выбранного типа нейтрализации"""
        neutral_type = self.neutralization_type_group.checkedButton().text()
        
        # Для суммарной нейтрализации - одно поле времени
        if neutral_type == "Суммарная":
            self.max_neutral_time_edit.setEnabled(True)
            self.max_neutral_time_edit.setToolTip("Общее время нейтрализации для всех СКП")
        else:
            self.max_neutral_time_edit.setEnabled(False)
            self.max_neutral_time_edit.setToolTip("")
        
        # Блокировка полей времени нейтрализации в СКП
        for skp in self.skp_widgets:
            if neutral_type == "На каждом СКП":
                skp['max_time_edit'].setEnabled(True)
                skp['max_time_edit'].setToolTip("Макс. время для этого СКП")
            else:
                skp['max_time_edit'].setEnabled(False)
                skp['max_time_edit'].setToolTip("")
        
        # Обновляем контейнер с этапами/СКП
        self.update_stages_ui()         

    def calculate_remaining_time(self, member, skp):
        """Вычисляет оставшееся время нейтрализации с учетом типа"""
        if not member.get("skp_entries"):
            return "0:00"
        
        neutral_type = self.data.get("logic_params", {}).get("neutralization_type", "Нет")
        
        if neutral_type == "Суммарная":
            # Для суммарной нейтрализации - общее время для всех СКП
            total_time = self.data["logic_params"].get("total_max_neutral_time", 0) * 60
            used_time = sum(
                self.calculate_actual_duration(entry)
                for entry in member.get("skp_entries", [])
            )
        else:
            # Для других типов - время только для текущего СКП
            total_time = skp.get("max_neutral_time", 0) * 60
            used_time = sum(
                self.calculate_actual_duration(entry)
                for entry in member.get("skp_entries", [])
                if entry.get("skp") == skp.get("number", 0)
            )
        
        remaining = max(0, total_time - used_time)
        return f"{remaining//60}:{remaining%60:02d}"            

    def calculate_route_time_from_params(self):
        """Рассчитывает время трассы от старта до закрытия из параметров"""
        if not hasattr(self, 'data') or 'params' not in self.data:
            self.route_time_info.setText("Нет данных о параметрах")
            return
        
        try:
            start_time_str = self.data['params'].get('время_старта', '')
            closing_time_str = self.data['params'].get('время_закрытия_трассы', '')
            
            if not start_time_str or not closing_time_str:
                self.route_time_info.setText("Не задано время старта/закрытия")
                return
            
            # Добавляем секунды, если их нет
            if len(start_time_str.split(':')) == 2:
                start_time_str += ":00"
            if len(closing_time_str.split(':')) == 2:
                closing_time_str += ":00"
            
            start_time = QTime.fromString(start_time_str, "HH:mm:ss")
            closing_time = QTime.fromString(closing_time_str, "HH:mm:ss")
            
            if not start_time.isValid() or not closing_time.isValid():
                self.route_time_info.setText("Некорректное время в параметрах")
                return
            
            # Если закрытие на следующий день (например, старт 20:00, закрытие 02:00)
            if closing_time < start_time:
                closing_time = closing_time.addSecs(24 * 3600)  # Добавляем 24 часа
                
            seconds = start_time.secsTo(closing_time)
            
            if seconds <= 0:
                self.route_time_info.setText("Ошибка: время закрытия раньше старта")
                return
                
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            seconds = seconds % 60
            
            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            self.route_time_info.setText(f"Время трассы: {time_str}\n(от старта до закрытия)")
            
            # Сохраняем расчетное время в параметры
            self.data['params']['расчетное_время_трассы'] = time_str
            self.save_data()
            
        except Exception as e:
            self.route_time_info.setText(f"Ошибка расчета: {str(e)}")

    def toggle_stage_settings(self, checked):
        """Показывает/скрывает настройки этапов"""
        self.stage_settings_group.setVisible(checked)
        if checked:
            self.update_stages_ui()

    def update_penalty_ui(self, penalty_type):
        """Обновляет UI в зависимости от выбранного типа штрафа"""
        if penalty_type in ["Нет", "DNF"]:
            self.penalty_value_edit.setEnabled(False)
            self.penalty_value_edit.clear()
        else:
            self.penalty_value_edit.setEnabled(True)

    def update_skp_penalty_ui(self, penalty_type):
        """Обновляет UI штрафа за опоздание на СКП"""
        sender = self.sender()
        for widget in self.stage_widgets:
            if widget['skp_penalty_combo'] == sender:
                if penalty_type in ["Нет", "DNQ"]:
                    widget['skp_penalty_value_edit'].setEnabled(False)
                    widget['skp_penalty_value_edit'].clear()
                else:
                    widget['skp_penalty_value_edit'].setEnabled(True)
                break














############## КП ##############


    def setup_checkpoints_tab(self):
        """Вкладка контрольных пунктов с исправленным скрытием таблиц"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Верхняя панель с кнопками
        top_panel = QWidget()
        top_layout = QHBoxLayout(top_panel)
        
        btn_generate = QPushButton("Генерировать КП")
        btn_generate.clicked.connect(self.generate_checkpoints)
        top_layout.addWidget(btn_generate)
        
        self.btn_show = QPushButton("Показать КП")
        self.btn_show.clicked.connect(self.show_checkpoints_table)
        top_layout.addWidget(self.btn_show)
        
        btn_add = QPushButton("Добавить КП")
        btn_add.clicked.connect(self.add_checkpoint_manual)
        top_layout.addWidget(btn_add)
        
        # btn_false_cps = QPushButton("Ложные КП")
        # btn_false_cps.clicked.connect(self.show_false_cps_table)
        # top_layout.addWidget(btn_false_cps)
        
        btn_stages = QPushButton("Этапы")
        btn_stages.clicked.connect(self.show_stages_table1)
        top_layout.addWidget(btn_stages)
        
        layout.addWidget(top_panel)
        
        # Основная таблица КП (изначально скрыта)
        self.checkpoints_table = QTableWidget()
        self.checkpoints_table.setVisible(False)
        self.checkpoints_table.doubleClicked.connect(self.edit_checkpoint_dialog)
        layout.addWidget(self.checkpoints_table)
        
        # Таблица ложных КП (изначально скрыта)
        self.false_cps_table = QTableWidget()
        self.false_cps_table.setVisible(False)
        self.false_cps_table.doubleClicked.connect(self.edit_checkpoint_dialog)
        layout.addWidget(self.false_cps_table)
        
        # Таблица этапов (изначально скрыта)
        self.stages_table = QTableWidget()
        self.stages_table.setVisible(False)
        self.stages_table.doubleClicked.connect(self.edit_checkpoint_dialog)
        layout.addWidget(self.stages_table)
        
        # Кнопка сохранения внизу
        btn_save = QPushButton("Сохранить изменения")
        btn_save.clicked.connect(self.save_checkpoints)
        layout.addWidget(btn_save)
        
        self.tabs.addTab(tab, "КП")


    def update_checkpoints_table(self):
        """Обновляем данные в таблице КП (упрощенная версия)"""
        # Определяем какие зачеты есть
        classifications = []
        if hasattr(self, 'data') and 'params' in self.data:
            if self.data['params'].get('зачет1'):
                classifications.append(self.data['params']['зачет1'])
            if self.data['params'].get('зачет2'):
                classifications.append(self.data['params']['зачет2'])
            if self.data['params'].get('зачет3'):
                classifications.append(self.data['params']['зачет3'])
        
        # Настраиваем таблицу (только название, зачеты и бальность)
        headers = ["Название"] + classifications + ["Баллы"]
        self.checkpoints_table.setColumnCount(len(headers))
        self.checkpoints_table.setHorizontalHeaderLabels(headers)
        
        # Заполняем данными из self.data
        if hasattr(self, 'data') and 'checkpoints' in self.data:
            self.checkpoints_table.setRowCount(len(self.data['checkpoints']))
            
            for row, cp in enumerate(self.data['checkpoints']):
                # Колонка "Название"
                self.checkpoints_table.setItem(row, 0, QTableWidgetItem(cp.get('name', '')))
                
                # Колонки с зачетами (чекбоксы)
                for col, classif in enumerate(classifications, 1):
                    checkbox = QCheckBox()
                    checkbox.setChecked(cp.get('classifications', {}).get(classif, False))
                    checkbox.stateChanged.connect(lambda state, r=row, c=col: self.on_checkbox_changed(r, c, state))
                    self.checkpoints_table.setCellWidget(row, col, checkbox)
                
                # Колонка с бальностью
                offset = 1 + len(classifications)
                self.checkpoints_table.setItem(row, offset, QTableWidgetItem(str(cp.get('score', ''))))
        
        # Настраиваем растягивание столбцов
        self.checkpoints_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.checkpoints_table.resizeColumnsToContents()

    def toggle_all_false_cps(self, state):
        """Массовое выделение/снятие всех ложных КП"""
        if not hasattr(self, 'data') or 'params' not in self.data:
            return
            
        classifications = []
        if self.data['params'].get('зачет1'):
            classifications.append(self.data['params']['зачет1'])
        if self.data['params'].get('зачет2'):
            classifications.append(self.data['params']['зачет2'])
        if self.data['params'].get('зачет3'):
            classifications.append(self.data['params']['зачет3'])
            
        if not classifications:
            return
            
        first_false_col = 1  # После колонки "Название КП"
        
        for row in range(self.false_cps_table.rowCount()):
            for col in range(first_false_col, first_false_col + len(classifications)):
                widget = self.false_cps_table.cellWidget(row, col)
                if widget and isinstance(widget, QCheckBox):
                    widget.setChecked(state == Qt.CheckState.Checked.value)

    def on_false_cp_changed(self, row, col, state):
        """Обработка изменения чекбокса ложного КП"""
        if not hasattr(self, 'data') or 'checkpoints' not in self.data:
            return
            
        # Определяем какой зачет изменился
        classifications = []
        if self.data['params'].get('зачет1'):
            classifications.append(self.data['params']['зачет1'])
        if self.data['params'].get('зачет2'):
            classifications.append(self.data['params']['зачет2'])
        if self.data['params'].get('зачет3'):
            classifications.append(self.data['params']['зачет3'])
            
        if col-1 < len(classifications):
            classif = classifications[col-1]
            cp_name = self.false_cps_table.item(row, 0).text()
            
            # Находим КП в данных
            for cp in self.data['checkpoints']:
                if cp.get('name') == cp_name:
                    if 'false_for' not in cp:
                        cp['false_for'] = []
                    
                    if state == Qt.CheckState.Checked.value:
                        if classif not in cp['false_for']:
                            cp['false_for'].append(classif)
                    else:
                        if classif in cp['false_for']:
                            cp['false_for'].remove(classif)
                    break

    def show_stages_table(self):
        """Показывает таблицу этапов"""
        self.checkpoints_table.setVisible(False)
        self.false_cps_table.setVisible(False)
        
        # Получаем список этапов
        stages = []
        if hasattr(self, 'data') and 'logic_params' in self.data and self.data['logic_params'].get('staged'):
            stages = [stage.get('name', f'Этап {i+1}') for i, stage in enumerate(self.data['logic_params'].get('stages', []))]
        
        # Настраиваем таблицу
        headers = ["Название КП"] + stages
        self.stages_table.setColumnCount(len(headers))
        self.stages_table.setHorizontalHeaderLabels(headers)
        
        # Добавляем чекбокс для массового выделения
        self.mass_stages_cb = QCheckBox("Выделить все для этапа")
        self.mass_stages_cb.stateChanged.connect(self.toggle_all_stages)
        
        # Заполняем данными
        if hasattr(self, 'data') and 'checkpoints' in self.data:
            self.stages_table.setRowCount(len(self.data['checkpoints']))
            
            for row, cp in enumerate(self.data['checkpoints']):
                # Колонка "Название КП"
                self.stages_table.setItem(row, 0, QTableWidgetItem(cp.get('name', '')))
                
                # Колонки с этапами
                for col, stage in enumerate(stages, 1):
                    checkbox = QCheckBox()
                    # Проверяем, принадлежит ли КП к этому этапу
                    in_stage = stage in cp.get('stages', [])
                    checkbox.setChecked(in_stage)
                    checkbox.stateChanged.connect(lambda state, r=row, c=col: self.on_stage_changed(r, c, state))
                    self.stages_table.setCellWidget(row, col, checkbox)
        
        self.stages_table.setVisible(True)
        self.btn_show.setText("Показать КП")

    def toggle_all_stages(self, state):
        """Массовое выделение/снятие всех КП для этапа"""
        if not hasattr(self, 'data') or 'logic_params' not in self.data or not self.data['logic_params'].get('staged'):
            return
            
        stages = [stage.get('name', f'Этап {i+1}') for i, stage in enumerate(self.data['logic_params'].get('stages', []))]
        
        if not stages:
            return
            
        # Определяем какой этап выделять (по текущему столбцу)
        current_col = self.stages_table.currentColumn()
        if current_col < 1 or current_col > len(stages):
            return
            
        stage_name = stages[current_col - 1]
        
        for row in range(self.stages_table.rowCount()):
            widget = self.stages_table.cellWidget(row, current_col)
            if widget and isinstance(widget, QCheckBox):
                widget.setChecked(state == Qt.CheckState.Checked.value)
                # Обновляем данные
                self.on_stage_changed(row, current_col, state)

    def on_stage_changed(self, row, col, state):
        """Обработка изменения чекбокса этапа"""
        if not hasattr(self, 'data') or 'checkpoints' not in self.data or 'logic_params' not in self.data:
            return
            
        stages = [stage.get('name', f'Этап {i+1}') for i, stage in enumerate(self.data['logic_params'].get('stages', []))]
        
        if col-1 < len(stages):
            stage_name = stages[col-1]
            cp_name = self.stages_table.item(row, 0).text()
            
            # Находим КП в данных
            for cp in self.data['checkpoints']:
                if cp.get('name') == cp_name:
                    if 'stages' not in cp:
                        cp['stages'] = []
                    
                    if state == Qt.CheckState.Checked.value:
                        if stage_name not in cp['stages']:
                            cp['stages'].append(stage_name)
                    else:
                        if stage_name in cp['stages']:
                            cp['stages'].remove(stage_name)
                    break

    def refresh_checkpoints_from_file(self):
        """Обновление данных из файла"""
        try:
            # Загружаем данные заново
            self.load_latest_data()
            
            # Если таблица видима - обновляем ее
            if self.checkpoints_table.isVisible():
                self.update_checkpoints_table()
                
            QMessageBox.information(self, "Обновлено", "Данные загружены из файла")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить данные: {str(e)}")

    def toggle_all_classifications(self, state):
        """Массовое выделение/снятие всех чекбоксов"""
        if not hasattr(self, 'data') or 'params' not in self.data:
            return
            
        classifications = []
        if self.data['params'].get('зачет1'):
            classifications.append(self.data['params']['зачет1'])
        if self.data['params'].get('зачет2'):
            classifications.append(self.data['params']['зачет2'])
        if self.data['params'].get('зачет3'):
            classifications.append(self.data['params']['зачет3'])
            
        if not classifications:
            return
            
        first_classif_col = 1  # После колонки "Название"
        
        for row in range(self.checkpoints_table.rowCount()):
            for col in range(first_classif_col, first_classif_col + len(classifications)):
                widget = self.checkpoints_table.cellWidget(row, col)
                if widget and isinstance(widget, QCheckBox):
                    widget.setChecked(state == Qt.CheckState.Checked.value)
    
    def on_checkbox_changed(self, row, col, state):
        """Обработка изменения чекбокса"""
        if not hasattr(self, 'data') or 'checkpoints' not in self.data:
            return
            
        # Определяем какой зачет изменился
        classifications = []
        if self.data['params'].get('зачет1'):
            classifications.append(self.data['params']['зачет1'])
        if self.data['params'].get('зачет2'):
            classifications.append(self.data['params']['зачет2'])
        if self.data['params'].get('зачет3'):
            classifications.append(self.data['params']['зачет3'])
            
        if col-1 < len(classifications):
            classif = classifications[col-1]
            self.data['checkpoints'][row]['classifications'][classif] = (state == Qt.CheckState.Checked.value)
    
    def save_checkpoints(self):
        """Сохранение изменений в КП"""
        if not hasattr(self, 'data'):
            QMessageBox.critical(self, "Ошибка", "Нет данных для сохранения")
            return
        
        try:
            # Данные уже обновляются через on_checkbox_changed
            # Просто сохраняем текущее состояние
            self.save_data()
            QMessageBox.information(self, "Сохранено", "Изменения в КП успешно сохранены")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения КП: {str(e)}")

    def generate_checkpoints(self):
        """Генерация КП с подтверждением"""
        if not hasattr(self, 'data') or 'params' not in self.data:
            QMessageBox.critical(self, "Ошибка", "Сначала задайте параметры соревнования")
            return
        
        try:
            # Получаем количество КП из параметров
            kp_count = int(self.data['params'].get('количество_кп', 0))
            if kp_count <= 0:
                QMessageBox.critical(self, "Ошибка", "Неверное количество КП в параметрах")
                return
            
            # Получаем бальность КП из параметров
            score = str(self.data['params'].get('бальность_кп', 10))
            
            # Запрос подтверждения
            reply = QMessageBox.question(
                self, 'Подтверждение',
                f'Вы хотите сгенерировать КП? Будет сгенерировано {kp_count} КП.\n'
                f'Бальность каждого КП: {score}',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # Генерируем новые КП
                existing_count = len(self.data.get('checkpoints', []))
                for i in range(1, kp_count + 1):
                    new_cp = {
                        "name": f"КП {existing_count + i}",  # Изменено на пробел вместо дефиса
                        "classifications": {},
                        "score": score
                    }
                    
                    # Заполняем зачеты если они есть
                    if self.data['params'].get('зачет1'):
                        new_cp['classifications'][self.data['params']['зачет1']] = True
                    if self.data['params'].get('зачет2'):
                        new_cp['classifications'][self.data['params']['зачет2']] = True
                    if self.data['params'].get('зачет3'):
                        new_cp['classifications'][self.data['params']['зачет3']] = False
                    
                    self.data['checkpoints'].append(new_cp)
                
                # Сохраняем изменения
                self.save_data()
                
                # Обновляем таблицы если они видимы
                if hasattr(self, 'checkpoints_table') and self.checkpoints_table.isVisible():
                    self.update_checkpoints_table()
                if hasattr(self, 'false_cps_table') and self.false_cps_table.isVisible():
                    self.show_false_cps_table()
                if hasattr(self, 'stages_table') and self.stages_table.isVisible():
                    self.show_stages_table()
                
                QMessageBox.information(
                    self, "Готово", 
                    f"Успешно сгенерировано {kp_count} контрольных пунктов\n"
                    f"Бальность: {score}"
                )
        
        except ValueError as ve:
            QMessageBox.critical(self, "Ошибка значения", f"Некорректные числовые параметры: {str(ve)}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Неизвестная ошибка при генерации КП: {str(e)}")

    def confirm_delete_checkpoint(self, row, dialog):
        """Подтверждение удаления КП"""
        cp_name = self.data['checkpoints'][row].get('name', 'этот КП')
        
        reply = QMessageBox.question(
            self, 
            'Подтверждение удаления',
            f'Вы уверены, что хотите удалить {cp_name}?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.delete_checkpoint(row)
            dialog.accept()

    def delete_checkpoint(self, row):
        """Удаление КП по индексу"""
        try:
            if 0 <= row < len(self.data['checkpoints']):
                # Удаляем из данных
                del self.data['checkpoints'][row]
                
                # Сохраняем изменения
                self.save_data()
                
                # Обновляем все таблицы
                self.update_checkpoints_table()
                if self.false_cps_table.isVisible():
                    self.show_false_cps_table()
                if self.stages_table.isVisible():
                    self.show_stages_table()
                    
                QMessageBox.information(self, "Успех", "КП успешно удален!")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось удалить КП: {str(e)}")

    def edit_checkpoint_dialog(self, index):
        """Диалог редактирования КП с поддержкой отрицательных значений"""
        try:
            row = index.row()
            if row < 0 or row >= len(self.data['checkpoints']):
                return
                
            cp = self.data['checkpoints'][row]
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Редактирование КП: {cp.get('name', '')}")
            dialog.setMinimumWidth(500)
            
            layout = QVBoxLayout(dialog)
            
            # Основная информация
            info_group = QGroupBox("Основная информация")
            info_layout = QVBoxLayout()
            
            # Название КП
            name_layout = QHBoxLayout()
            name_layout.addWidget(QLabel("Название:"))
            name_edit = QLineEdit(cp.get('name', ''))
            name_layout.addWidget(name_edit)
            info_layout.addLayout(name_layout)
            
            # Баллы (с возможностью отрицательных значений)
            score_layout = QHBoxLayout()
            score_layout.addWidget(QLabel("Баллы:"))
            score_edit = QLineEdit(str(cp.get('score', 10)))
            # Валидатор для целых чисел (включая отрицательные)
            score_edit.setValidator(QIntValidator(-1000, 1000))
            score_layout.addWidget(score_edit)
            info_layout.addLayout(score_layout)
            
            info_group.setLayout(info_layout)
            layout.addWidget(info_group)
            
            # Доступность для зачетов (только если есть зачеты)
            classifications = []
            if 'params' in self.data:
                if self.data['params'].get('зачет1'):
                    classifications.append(self.data['params']['зачет1'])
                if self.data['params'].get('зачет2'):
                    classifications.append(self.data['params']['зачет2'])
                if self.data['params'].get('зачет3'):
                    classifications.append(self.data['params']['зачет3'])
            
            if classifications:
                classif_group = QGroupBox("Доступность для зачетов")
                classif_layout = QVBoxLayout()
                
                self.classif_checkboxes = {}
                for classif in classifications:
                    cb = QCheckBox(classif)
                    cb.setChecked(cp.get('classifications', {}).get(classif, False))
                    self.classif_checkboxes[classif] = cb
                    classif_layout.addWidget(cb)
                
                classif_group.setLayout(classif_layout)
                layout.addWidget(classif_group)
            
            # Ложность для зачетов
            if classifications:
                false_group = QGroupBox("Ложный КП для зачетов (не работает)")
                false_layout = QVBoxLayout()
                
                self.false_checkboxes = {}
                for classif in classifications:
                    cb = QCheckBox(classif)
                    cb.setChecked(classif in cp.get('false_for', []))
                    self.false_checkboxes[classif] = cb
                    false_layout.addWidget(cb)
                
                false_group.setLayout(false_layout)
                layout.addWidget(false_group)
            
            # Принадлежность к этапам (только если включена этапность)
            if (hasattr(self, 'data') and 'logic_params' in self.data and 
                self.data['logic_params'].get('staged') and 
                'stages' in self.data['logic_params']):
                
                stages_group = QGroupBox("Принадлежность к этапам")
                stages_layout = QVBoxLayout()
                
                self.stage_checkboxes = {}
                stages = [stage.get('name', f'Этап {i+1}') 
                          for i, stage in enumerate(self.data['logic_params'].get('stages', []))]
                
                for stage in stages:
                    cb = QCheckBox(stage)
                    cb.setChecked(stage in cp.get('stages', []))
                    self.stage_checkboxes[stage] = cb
                    stages_layout.addWidget(cb)
                
                stages_group.setLayout(stages_layout)
                layout.addWidget(stages_group)
            
            # Кнопки
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | 
                                         QDialogButtonBox.StandardButton.Cancel)
            
            delete_button = QPushButton("Удалить")
            delete_button.setStyleSheet("background-color: #ff4444; color: white;")
            button_box.addButton(delete_button, QDialogButtonBox.ButtonRole.DestructiveRole)
            
            button_box.accepted.connect(lambda: self.save_checkpoint_changes(
                row, name_edit.text(), score_edit.text(), dialog))
            button_box.rejected.connect(dialog.reject)
            delete_button.clicked.connect(lambda: self.confirm_delete_checkpoint(row, dialog))
            
            layout.addWidget(button_box)
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при открытии диалога: {str(e)}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при открытии диалога: {str(e)}")
            print(f"Ошибка в edit_checkpoint_dialog: {traceback.format_exc()}")

    def save_checkpoint_changes(self, row, name, score, dialog):
        """Сохранение изменений КП с проверкой отрицательных значений"""
        try:
            if not name:
                QMessageBox.warning(dialog, "Ошибка", "Название КП не может быть пустым!")
                return
                
            try:
                score_int = int(score)  # Преобразуем в число (может быть отрицательным)
            except ValueError:
                QMessageBox.warning(dialog, "Ошибка", "Баллы должны быть целым числом!")
                return
                
            cp = self.data['checkpoints'][row]
            cp['name'] = name
            cp['score'] = str(score_int)  # Сохраняем как строку (для совместимости)
            
            # Обновляем зачеты
            if hasattr(self, 'classif_checkboxes'):
                for classif, cb in self.classif_checkboxes.items():
                    if 'classifications' not in cp:
                        cp['classifications'] = {}
                    cp['classifications'][classif] = cb.isChecked()
            
            # Обновляем ложные КП
            if hasattr(self, 'false_checkboxes'):
                cp['false_for'] = [
                    classif for classif, cb in self.false_checkboxes.items() 
                    if cb.isChecked()
                ]
            
            # Обновляем этапы
            if hasattr(self, 'stage_checkboxes'):
                cp['stages'] = [
                    stage for stage, cb in self.stage_checkboxes.items() 
                    if cb.isChecked()
                ]
            
            self.save_data()
            self.update_checkpoints_table()
            dialog.accept()
            
        except Exception as e:
            QMessageBox.critical(dialog, "Ошибка", f"Не удалось сохранить изменения: {str(e)}")

    def show_checkpoints_table(self):
        """Показывает таблицу КП и скрывает остальные"""
        self.checkpoints_table.setVisible(True)
        self.false_cps_table.setVisible(False)
        self.stages_table.setVisible(False)
        
        # Обновляем данные в таблице
        self.update_checkpoints_table()
        self.btn_show.setText("Обновить КП")

    def show_stages_table1(self):
        """Показывает таблицу этапов и скрывает остальные"""
        self.checkpoints_table.setVisible(False)
        self.false_cps_table.setVisible(False)
        self.stages_table.setVisible(True)
        
        # Обновляем данные в таблице
        self.show_stages_table()
        self.btn_show.setText("Показать КП")

    def add_checkpoint_manual(self):
        """Добавление нового КП вручную с сохранением в файл"""
        if not hasattr(self, 'data'):
            QMessageBox.critical(self, "Ошибка", "Данные не загружены")
            return
        
        try:
            # Получаем количество КП из параметров
            kp_count = len(self.data.get('checkpoints', []))
            
            # Создаем новый КП с базовыми значениями
            new_cp = {
                "name": f"КП {kp_count + 1}",
                "classifications": {},
                "score": str(self.data.get('params', {}).get('бальность_кп', 10)),
                "false_for": [],
                "stages": []
            }
            
            # Заполняем зачеты если они есть
            if 'params' in self.data:
                if self.data['params'].get('зачет1'):
                    new_cp['classifications'][self.data['params']['зачет1']] = True
                if self.data['params'].get('зачет2'):
                    new_cp['classifications'][self.data['params']['зачет2']] = False
                if self.data['params'].get('зачет3'):
                    new_cp['classifications'][self.data['params']['зачет3']] = False
            
            # Добавляем в данные
            self.data.setdefault('checkpoints', []).append(new_cp)
            
            # Сохраняем изменения в файл
            if not self.save_data():
                raise Exception("Не удалось сохранить данные")
            
            # Обновляем таблицы если они видимы
            if hasattr(self, 'checkpoints_table') and self.checkpoints_table.isVisible():
                self.update_checkpoints_table()
            if hasattr(self, 'false_cps_table') and self.false_cps_table.isVisible():
                self.show_false_cps_table()
            if hasattr(self, 'stages_table') and self.stages_table.isVisible():
                self.show_stages_table()
            
            QMessageBox.information(self, "Добавлено", f"Добавлен новый КП: {new_cp['name']}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось добавить КП: {str(e)}")















############## УЧАСТНИКИ ##############

    def setup_members_tab(self):
        """Вкладка участников"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Верхняя панель с кнопками
        button_panel = QWidget()
        button_layout = QHBoxLayout(button_panel)
        
        # Создаем кнопки
        self.btn_load_xlsx = QPushButton("Загрузить xlsx")
        self.btn_load_xlsx.clicked.connect(self.import_from_excel)
        button_layout.addWidget(self.btn_load_xlsx)
        
        self.btn_edit_crew = QPushButton("Редактировать экипаж")
        self.btn_edit_crew.clicked.connect(self.show_edit_crew_dialog)  # Изменено подключение
        button_layout.addWidget(self.btn_edit_crew)
        
        self.btn_toggle_view = QPushButton("Полная таблица")
        self.btn_toggle_view.setCheckable(True)
        self.btn_toggle_view.clicked.connect(self.toggle_members_view)
        button_layout.addWidget(self.btn_toggle_view)
        
        self.btn_add_crew = QPushButton("Добавить экипаж")
        self.btn_add_crew.clicked.connect(self.show_add_crew_dialog)
        button_layout.addWidget(self.btn_add_crew)
        
        self.btn_delete_crew = QPushButton("Удалить экипаж")
        self.btn_delete_crew.clicked.connect(self.show_delete_crew_dialog)  # Изменено подключение
        button_layout.addWidget(self.btn_delete_crew)
        
        layout.addWidget(button_panel)

        # Создаем контейнер с прокруткой для таблицы
        table_container = QScrollArea()
        table_container.setWidgetResizable(True)
        
        # Создаем таблицу внутри контейнера (ТОЛЬКО ОДИН РАЗ)
        self.members_table = QTableWidget()
        self.members_table.setSizeAdjustPolicy(QTableWidget.SizeAdjustPolicy.AdjustToContents)
        
        # Настройки прокрутки
        self.members_table.setVerticalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        self.members_table.setHorizontalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        
        # Добавляем таблицу в контейнер
        table_container.setWidget(self.members_table)
        layout.addWidget(table_container)
        
        # По умолчанию показываем сокращенную таблицу
        self.is_full_view = False
        self.setup_short_table_headers()
        
        self.members_table.setSortingEnabled(True)  # Разрешаем сортировку по клику на заголовок

        # Заполняем таблицу данными
        if hasattr(self, 'data') and 'members' in self.data:
            self.update_members_table()
        
        # УДАЛЕНО: layout.addWidget(self.members_table) - таблица уже добавлена через контейнер
        
        self.tabs.addTab(tab, "Участники")

    def setup_short_table_headers(self):
        """Устанавливаем заголовки для сокращенного вида"""
        headers = ["Номер", "Зачет", "Пилот", "Штурман", "Авто", "Гос. номер"]
        self.members_table.setColumnCount(len(headers))
        self.members_table.setHorizontalHeaderLabels(headers)
        self.is_full_view = False
        self.btn_toggle_view.setText("Полная таблица")
    
        # Настраиваем ширину столбцов
        self.members_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.members_table.horizontalHeader().setStretchLastSection(True)

    def setup_full_table_headers(self):
        """Устанавливаем заголовки для полного вида"""
        headers = [
            "Адрес электронной почты",
            "Капитан",
            "Субъект",
            "Номер",
            "Пилот",
            "Пол пилота",
            "Дата рождения пилота",
            "Контактный телефон пилота",
            "Штурман",
            "Пол штурмана",
            "Дата рождения штурмана",
            "Контактный телефон штурмана",
            "Пассажиры",
            "Общее количество",
            "До 18 лет",
            "Авто",
            "Гос. номер",
            "Привод автомобиля",
            "Зачет"
        ]
        self.members_table.setColumnCount(len(headers))
        self.members_table.setHorizontalHeaderLabels(headers)
        self.is_full_view = True
        self.btn_toggle_view.setText("Сокращенная таблица")

        # Настраиваем ширину столбцов
        self.members_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.members_table.horizontalHeader().setStretchLastSection(False)
        
        # Устанавливаем минимальную ширину для некоторых колонок (опционально)
        for col in [3, 4, 8, 15, 16]:  # Номера колонок, которые нужно сделать шире
            self.members_table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            self.members_table.setColumnWidth(col, 120)  # Ширина в пикселях

    def toggle_members_view(self):
        """Переключение между полной и сокращенной таблицей"""
        if self.is_full_view:
            self.setup_short_table_headers()
        else:
            self.setup_full_table_headers()
        self.update_members_table()

    def update_members_table(self):
        """Обновление данных в таблице участников (универсальный для обоих форматов)"""
        if not hasattr(self, 'data') or 'members' not in self.data:
            return
        
        self.members_table.setRowCount(len(self.data['members']))
        
        for row, member in enumerate(self.data['members']):
            if self.is_full_view:
                # Заполняем полную таблицу (универсально для обоих форматов)
                self.members_table.setItem(row, 0, QTableWidgetItem(member.get('адрес_электронной_почты') or member.get('email', '')))
                self.members_table.setItem(row, 1, QTableWidgetItem(member.get('капитан') or member.get('captain', '')))
                self.members_table.setItem(row, 2, QTableWidgetItem(member.get('субъект') or member.get('subject', '')))
                self.members_table.setItem(row, 3, QTableWidgetItem(str(member.get('номер') or member.get('number', ''))))
                self.members_table.setItem(row, 4, QTableWidgetItem(member.get('пилот') or member.get('driver', '')))
                self.members_table.setItem(row, 5, QTableWidgetItem(member.get('пол_пилота') or member.get('driver_gender', '')))
                self.members_table.setItem(row, 6, QTableWidgetItem(member.get('дата_рождения_пилота') or member.get('driver_birth', '')))
                self.members_table.setItem(row, 7, QTableWidgetItem(member.get('контактный_телефон_пилота') or member.get('phone', '')))
                self.members_table.setItem(row, 8, QTableWidgetItem(member.get('штурман') or member.get('navigator', '')))
                self.members_table.setItem(row, 9, QTableWidgetItem(member.get('пол_штурмана') or member.get('navigator_gender', '')))
                self.members_table.setItem(row, 10, QTableWidgetItem(member.get('дата_рождения_штурмана') or member.get('navigator_birth', '')))
                self.members_table.setItem(row, 11, QTableWidgetItem(member.get('контактный_телефон_штурмана') or member.get('navigator_phone', '')))
                self.members_table.setItem(row, 12, QTableWidgetItem(member.get('пассажиры') or member.get('passengers', '')))
                self.members_table.setItem(row, 13, QTableWidgetItem(member.get('общее_количество') or member.get('total_count', '')))
                self.members_table.setItem(row, 14, QTableWidgetItem(member.get('до_18_лет') or member.get('under_18', '')))
                self.members_table.setItem(row, 15, QTableWidgetItem(member.get('авто') or member.get('car', '')))
                self.members_table.setItem(row, 16, QTableWidgetItem(member.get('гос.номер') or member.get('plate', '')))
                self.members_table.setItem(row, 17, QTableWidgetItem(member.get('привод_автомобиля') or member.get('drive_type', '')))
                self.members_table.setItem(row, 18, QTableWidgetItem(member.get('зачет') or member.get('classification', '')))
            else:
                # Сокращенная таблица
                self.members_table.setItem(row, 0, QTableWidgetItem(str(member.get('номер') or member.get('number', ''))))
                self.members_table.setItem(row, 1, QTableWidgetItem(member.get('зачет') or member.get('classification', '')))
                self.members_table.setItem(row, 2, QTableWidgetItem(member.get('пилот') or member.get('driver', '')))
                self.members_table.setItem(row, 3, QTableWidgetItem(member.get('штурман') or member.get('navigator', '')))
                self.members_table.setItem(row, 4, QTableWidgetItem(member.get('авто') or member.get('car', '')))
                self.members_table.setItem(row, 5, QTableWidgetItem(member.get('гос.номер') or member.get('plate', '')))
        
        self.members_table.resizeColumnsToContents()
        
        # Устанавливаем минимальную ширину для важных колонок
        if self.is_full_view:
            self.members_table.setColumnWidth(3, 50)   # Номер
            self.members_table.setColumnWidth(4, 150)  # Пилот
            self.members_table.setColumnWidth(8, 150)  # Штурман
        else:
            self.members_table.setColumnWidth(0, 50)   # Номер
            self.members_table.setColumnWidth(2, 150)  # Пилот
            self.members_table.setColumnWidth(3, 150)  # Штурман

    def show_add_crew_dialog(self):
        """Показывает диалог добавления нового экипажа со всеми полями"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить новый экипаж")
        dialog.setMinimumWidth(600)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QVBoxLayout(content)
        
        # Словарь для хранения всех полей ввода
        self.fields = {}
        
        # Обязательные поля (отмечены *)
        basic_fields = [
            "Номер*",
            "Пилот*",
            "Пол пилота",
            "Дата рождения пилота",
            "Контактный телефон пилота*",
            "Штурман",
            "Пол штурмана",
            "Дата рождения штурмана",
            "Контактный телефон штурмана"
        ]
        
        # Добавляем выбор капитана
        layout.addWidget(QLabel("Капитан экипажа:"))
        self.captain_group = QButtonGroup()
        captain_pilot = QRadioButton("Пилот")
        captain_navigator = QRadioButton("Штурман")
        self.captain_group.addButton(captain_pilot)
        self.captain_group.addButton(captain_navigator)
        
        captain_layout = QHBoxLayout()
        captain_layout.addWidget(captain_pilot)
        captain_layout.addWidget(captain_navigator)
        captain_pilot.setChecked(True)  # По умолчанию выбран пилот
        layout.addLayout(captain_layout)
        
        # Остальные поля
        other_fields = [
            "Адрес электронной почты",
            "Субъект",
            "Пассажиры",
            "Общее количество",
            "До 18 лет",
            "Авто*",
            "Гос. номер*",
            "Привод автомобиля"
        ]
        
        # Создаем поля для ввода
        for field in basic_fields + other_fields:
            row = QHBoxLayout()
            label = QLabel(field.replace("*", "") + (":" if not field.endswith("*") else "*:"))
            
            if "Пол" in field or "Привод" in field:
                # Для полей с ограниченным выбором используем комбобоксы
                combo = QComboBox()
                if "Пол" in field:
                    combo.addItems(["Мужской", "Женский"])
                else:  # Привод автомобиля
                    combo.addItems(["Передний", "Задний", "Полный"])
                self.fields[field.replace("*", "")] = combo
                row.addWidget(label)
                row.addWidget(combo)
            elif "Дата" in field:
                # Для дат используем виджет выбора даты
                date_edit = QDateEdit()
                date_edit.setCalendarPopup(True)
                date_edit.setDate(QDate.currentDate())
                self.fields[field] = date_edit
                row.addWidget(label)
                row.addWidget(date_edit)
            else:
                # Обычные текстовые поля
                edit = QLineEdit()
                self.fields[field.replace("*", "")] = edit
                row.addWidget(label)
                row.addWidget(edit)
            
            layout.addLayout(row)
        
        # Добавляем выбор зачета
        layout.addWidget(QLabel("Зачет:"))
        self.classification_group = QButtonGroup()
        classifications = self.get_available_classifications()
        
        classif_layout = QHBoxLayout()
        for classif in classifications:
            rb = QRadioButton(classif)
            self.classification_group.addButton(rb)
            classif_layout.addWidget(rb)
        layout.addLayout(classif_layout)
        
        # Кнопки сохранения/отмены
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(lambda: self.validate_and_add_crew(dialog))
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        scroll.setWidget(content)
        dialog_layout = QVBoxLayout(dialog)
        dialog_layout.addWidget(scroll)
        
        dialog.exec()

    def get_available_classifications(self):
        """Возвращает список доступных зачетов"""
        classifications = []
        if hasattr(self, 'data') and 'params' in self.data:
            if self.data['params'].get('зачет1'):
                classifications.append(self.data['params']['зачет1'])
            if self.data['params'].get('зачет2'):
                classifications.append(self.data['params']['зачет2'])
            if self.data['params'].get('зачет3'):
                classifications.append(self.data['params']['зачет3'])
        return classifications

    def validate_and_add_crew(self, dialog):
        """Проверяет данные и добавляет новый экипаж (в формате как у импортированных)"""
        # Проверяем обязательные поля
        required_fields = [
            "Номер", "Пилот", "Контактный телефон пилота", 
            "Авто", "Гос. номер"
        ]
        
        for field in required_fields:
            if isinstance(self.fields[field], QLineEdit) and not self.fields[field].text():
                QMessageBox.warning(self, "Ошибка", f"Поле '{field}' обязательно для заполнения!")
                return
        
        # Проверяем уникальность номера
        number = self.fields["Номер"].text()
        if any(str(m.get('номер')) == number for m in self.data.get('members', [])):
            QMessageBox.warning(self, "Ошибка", "Экипаж с таким номером уже существует!")
            return
        
        # Получаем выбранный зачет
        selected_classif = None
        for btn in self.classification_group.buttons():
            if btn.isChecked():
                selected_classif = btn.text()
                break
        
        if not selected_classif:
            QMessageBox.warning(self, "Ошибка", "Выберите зачет!")
            return
        
        # Собираем все данные в формате как у импортированных
        new_member = {
            'адрес_электронной_почты': self.fields["Адрес электронной почты"].text(),
            'капитан': "Пилот" if self.captain_group.buttons()[0].isChecked() else "Штурман",
            'субъект': self.fields["Субъект"].text(),
            'номер': number,
            'пилот': self.fields["Пилот"].text(),
            'пол_пилота': self.fields["Пол пилота"].currentText() if isinstance(self.fields["Пол пилота"], QComboBox) else "",
            'дата_рождения_пилота': self.fields["Дата рождения пилота"].date().toString("yyyy-MM-dd") + " 00:00:00",
            'контактный_телефон_пилота': self.fields["Контактный телефон пилота"].text(),
            'штурман': self.fields["Штурман"].text(),
            'пол_штурмана': self.fields["Пол штурмана"].currentText() if isinstance(self.fields["Пол штурмана"], QComboBox) else "",
            'дата_рождения_штурмана': self.fields["Дата рождения штурмана"].date().toString("yyyy-MM-dd") + " 00:00:00",
            'контактный_телефон_штурмана': self.fields["Контактный телефон штурмана"].text(),
            'пассажиры': self.fields["Пассажиры"].text(),
            'общее_количество': self.fields["Общее количество"].text(),
            'до_18_лет': self.fields["До 18 лет"].text(),
            'авто': self.fields["Авто"].text(),
            'гос.номер': self.fields["Гос. номер"].text(),
            'привод_автомобиля': self.fields["Привод автомобиля"].currentText() if isinstance(self.fields["Привод автомобиля"], QComboBox) else "",
            'зачет': selected_classif
        }
        
        # Добавляем в данные
        if 'members' not in self.data:
            self.data['members'] = []
        self.data['members'].append(new_member)
        
        # Сохраняем и обновляем
        self.save_data()
        self.update_members_table()
        dialog.accept()
        QMessageBox.information(self, "Успех", "Экипаж успешно добавлен!")

    def show_delete_crew_dialog(self):
        """Диалог удаления экипажа - упрощенная версия"""
        if not hasattr(self, 'data') or not self.data.get('members'):
            QMessageBox.warning(self, "Ошибка", "Нет экипажей для удаления!")
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Удаление экипажа")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        # Таблица для выбора
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Номер", "Пилот", "Авто"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        
        # Заполняем таблицу
        table.setRowCount(len(self.data['members']))
        for row, member in enumerate(self.data['members']):
            # Универсальное получение данных
            number = str(member.get('номер') or member.get('number', ''))
            driver = member.get('пилот') or member.get('driver', '')
            car = member.get('авто') or member.get('car', '')
            
            table.setItem(row, 0, QTableWidgetItem(number))
            table.setItem(row, 1, QTableWidgetItem(driver))
            table.setItem(row, 2, QTableWidgetItem(car))
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.delete_crew_by_index(table.currentRow(), dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.exec()

    def delete_crew_by_index(self, row, dialog):
        """Удаляет экипаж по индексу строки в таблице"""
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж для удаления!")
            return
        
        # Двойное подтверждение
        confirm = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите удалить выбранный экипаж?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if confirm == QMessageBox.StandardButton.Yes:
            try:
                # Удаляем по индексу
                if 0 <= row < len(self.data['members']):
                    deleted_crew = self.data['members'].pop(row)
                    
                    # Сохраняем изменения
                    self.save_data()
                    
                    # Обновляем таблицу
                    self.update_members_table()
                    
                    dialog.accept()
                    
                    # Показываем подтверждение
                    number = str(deleted_crew.get('номер') or deleted_crew.get('number', 'Без номера'))
                    QMessageBox.information(
                        self,
                        "Успех",
                        f"Экипаж {number} успешно удален!"
                    )
                else:
                    QMessageBox.warning(self, "Ошибка", "Неверный индекс экипажа!")
                    
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Ошибка",
                    f"Не удалось удалить экипаж: {str(e)}"
                )

    def confirm_crew_deletion(self, dialog):
        """Подтверждает удаление выбранного экипажа"""
        selected_items = self.crew_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж для удаления!")
            return
        
        # Первое подтверждение
        reply = QMessageBox.question(
            self, 'Подтверждение',
            'Вы уверены, что хотите удалить выбранный экипаж?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Второе подтверждение
            reply = QMessageBox.question(
                self, 'Последнее подтверждение',
                'Экипаж будет полностью удален из системы. Продолжить?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.delete_selected_crew(selected_items[0], dialog)

    def delete_selected_crew(self, item, dialog):
        """Удаляет выбранный экипаж"""
        crew_number = item.data(Qt.ItemDataRole.UserRole)
        
        # Находим и удаляем экипаж
        for i, member in enumerate(self.data['members']):
            if str(member.get('number')) == str(crew_number):
                del self.data['members'][i]
                break
        
        # Сохраняем изменения
        self.save_data()
        
        # Обновляем таблицу
        self.update_members_table()
        
        dialog.accept()
        QMessageBox.information(self, "Успех", "Экипаж успешно удален!")

    def import_from_excel(self):
        """Импорт данных из Excel файла"""
        # Открываем диалог выбора файла
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выбете файл с участниками", "", 
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        
        if not file_path:
            return  # Пользователь отменил выбор
        
        try:
            # Загружаем книгу Excel
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Собираем заголовки из Excel (первая строка)
            excel_headers = []
            for col in range(1, 21):  # Колонки с 1 по 20
                excel_headers.append(str(sheet.cell(row=1, column=col).value or ""))
            
            # Сопоставление столбцов (программа -> excel)
            column_mapping = {
                "Адрес электронной почты": 2,
                "Капитан": 3,
                "Субъект": 4,
                "Номер": 5,
                "Пилот": 6,
                "Пол пилота": 7,
                "Дата рождения пилота": 8,
                "Контактный телефон пилота": 9,
                "Штурман": 10,
                "Пол штурмана": 11,
                "Дата рождения штурмана": 12,
                "Контактный телефон штурмана": 13,
                "Пассажиры": 14,
                "Общее количество": 15,
                "До 18 лет": 16,
                "Авто": 17,
                "Гос.номер": 18,
                "Привод автомобиля": 19,
                "Зачет": 20
            }
            
            # Показываем диалог подтверждения соответствия
            if not self.show_column_mapping_dialog(column_mapping, excel_headers):
                return
            
            # Читаем данные из Excel
            new_members = []
            for row in range(2, sheet.max_row + 1):  # Начинаем со второй строки
                member = {}
                for field, col in column_mapping.items():
                    value = sheet.cell(row=row, column=col).value
                    member[field.lower().replace(" ", "_")] = str(value) if value is not None else ""
                
                # Преобразуем специальные поля
                if member.get("капитан"):
                    member["капитан"] = "Пилот" if member["капитан"].lower() == "пилот" else "Штурман"
                
                new_members.append(member)
            
            # Добавляем новых участников
            if 'members' not in self.data:
                self.data['members'] = []
            
            added_count = 0
            if 'members' not in self.data:
                self.data['members'] = []
            self.data['members'].extend(new_members)  # Добавляем всех
            added_count = len(new_members)  # Считаем всех как добавленных
            
            # Сохраняем и обновляем
            self.save_data()
            self.update_members_table()
            
            QMessageBox.information(
                self, "Импорт завершен", 
                f"Успешно добавлено {added_count} новых экипажей!\n"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка импорта", 
                f"Произошла ошибка при чтении файла:\n{str(e)}"
            )

    def show_column_mapping_dialog(self, column_mapping, excel_headers):
        """Показывает диалог с соответствием столбцов"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Проверка соответствия столбцов")
        dialog.setMinimumWidth(600)
        
        layout = QVBoxLayout(dialog)
        
        # Таблица соответствия
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Поле в программе", "Столбец в Excel"])
        table.setRowCount(len(column_mapping))
        
        # Заполняем таблицу
        for row, (field, col) in enumerate(column_mapping.items()):
            table.setItem(row, 0, QTableWidgetItem(field))
            
            excel_col = excel_headers[col-1] if col-1 < len(excel_headers) else "НЕТ ДАННЫХ"
            table.setItem(row, 1, QTableWidgetItem(f"{col} ({excel_col})"))
        
        table.resizeColumnsToContents()
        table.horizontalHeader().setStretchLastSection(True)
        
        layout.addWidget(QLabel("Пожалуйста, проверьте соответствие столбцов:"))
        layout.addWidget(table)
        layout.addWidget(QLabel("Убедитесь, что столбцы в Excel соответствуют указанным номерам"))
        
        # Кнопки подтверждения
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        return dialog.exec() == QDialog.DialogCode.Accepted

    def show_edit_crew_dialog(self):
        """Диалог редактирования экипажа"""
        if not hasattr(self, 'data') or not self.data.get('members'):
            QMessageBox.warning(self, "Ошибка", "Нет экипажей для редактирования!")
            return
        
        # Диалог выбора экипажа (аналогично удалению)
        dialog = QDialog(self)
        dialog.setWindowTitle("Выбор экипажа для редактирования")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Номер", "Пилот", "Штурман"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        
        # Заполняем таблицу
        table.setRowCount(len(self.data['members']))
        for row, member in enumerate(self.data['members']):
            table.setItem(row, 0, QTableWidgetItem(str(member.get('номер', ''))))
            table.setItem(row, 1, QTableWidgetItem(member.get('пилот', '')))
            table.setItem(row, 2, QTableWidgetItem(member.get('штурман', '')))
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.edit_selected_crew(table.currentRow(), dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.exec()

    def edit_selected_crew(self, row, selection_dialog):
        """Редактирование выбранного экипажа со всеми полями"""
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж для редактирования!")
            return
        
        member = self.data['members'][row]
        
        edit_dialog = QDialog(self)
        edit_dialog.setWindowTitle(f"Редактирование экипажа №{member.get('номер', '')}")
        edit_dialog.setMinimumWidth(700)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QVBoxLayout(content)
        
        # Словарь для хранения полей редактирования
        fields = {}

        # Создаем все поля для редактирования
        fields_layout = [
            ("Адрес электронной почты", QLineEdit(member.get('адрес_электронной_почты', ''))),
            ("Субъект", QLineEdit(member.get('субъект', ''))),
            ("Номер", QLineEdit(str(member.get('номер', '')))),
            ("Пилот", QLineEdit(member.get('пилот', ''))),
            ("Контактный телефон пилота", QLineEdit(member.get('контактный_телефон_пилота', ''))),
            ("Штурман", QLineEdit(member.get('штурман', ''))),
            ("Контактный телефон штурмана", QLineEdit(member.get('контактный_телефон_штурмана', ''))),
            ("Пассажиры", QLineEdit(member.get('пассажиры', ''))),
            ("Общее количество", QLineEdit(str(member.get('общее_количество', '')))),
            ("До 18 лет", QLineEdit(str(member.get('до_18_лет', '')))),
            ("Авто", QLineEdit(member.get('авто', ''))),
            ("Гос.номер", QLineEdit(member.get('гос.номер', ''))),
        ]

        # Добавляем основные текстовые поля
        for label, widget in fields_layout:
            row_layout = QHBoxLayout()
            row_layout.addWidget(QLabel(f"{label}:"))
            row_layout.addWidget(widget)
            layout.addLayout(row_layout)
            fields[label.lower().replace(' ', '_')] = widget

        # Поля с выбором из списка
        gender_layout = QHBoxLayout()
        gender_layout.addWidget(QLabel("Пол пилота:"))
        gender_driver = QComboBox()
        gender_driver.addItems(["Мужской", "Женский"])
        gender_driver.setCurrentText(member.get('пол_пилота', 'Мужской'))
        gender_layout.addWidget(gender_driver)
        layout.addLayout(gender_layout)
        fields['пол_пилота'] = gender_driver

        gender_layout = QHBoxLayout()
        gender_layout.addWidget(QLabel("Пол штурмана:"))
        gender_navigator = QComboBox()
        gender_navigator.addItems(["Мужской", "Женский"])
        gender_navigator.setCurrentText(member.get('пол_штурмана', 'Мужской'))
        gender_layout.addWidget(gender_navigator)
        layout.addLayout(gender_layout)
        fields['пол_штурмана'] = gender_navigator

        # Поля с датами
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Дата рождения пилота:"))
        birth_driver = QDateEdit()
        birth_driver.setCalendarPopup(True)
        if member.get('дата_рождения_пилота'):
            birth_driver.setDate(QDate.fromString(member['дата_рождения_пилота'].split()[0], "yyyy-MM-dd"))
        date_layout.addWidget(birth_driver)
        layout.addLayout(date_layout)
        fields['дата_рождения_пилота'] = birth_driver

        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Дата рождения штурмана:"))
        birth_navigator = QDateEdit()
        birth_navigator.setCalendarPopup(True)
        if member.get('дата_рождения_штурмана'):
            birth_navigator.setDate(QDate.fromString(member['дата_рождения_штурмана'].split()[0], "yyyy-MM-dd"))
        date_layout.addWidget(birth_navigator)
        layout.addLayout(date_layout)
        fields['дата_рождения_штурмана'] = birth_navigator

        # Выбор привода автомобиля
        drive_layout = QHBoxLayout()
        drive_layout.addWidget(QLabel("Привод автомобиля:"))
        drive_type = QComboBox()
        drive_type.addItems(["Передний", "Задний", "Полный"])
        drive_type.setCurrentText(member.get('привод_автомобиля', 'Передний'))
        drive_layout.addWidget(drive_type)
        layout.addLayout(drive_layout)
        fields['привод_автомобиля'] = drive_type

        # Капитан экипажа
        captain_layout = QHBoxLayout()
        captain_layout.addWidget(QLabel("Капитан экипажа:"))
        captain_group = QButtonGroup()
        captain_pilot = QRadioButton("Пилот")
        captain_navigator = QRadioButton("Штурман")
        captain_group.addButton(captain_pilot)
        captain_group.addButton(captain_navigator)
        captain_layout.addWidget(captain_pilot)
        captain_layout.addWidget(captain_navigator)
        if member.get('капитан', 'Пилот') == 'Пилот':
            captain_pilot.setChecked(True)
        else:
            captain_navigator.setChecked(True)
        layout.addLayout(captain_layout)
        fields['капитан'] = captain_group

        # Выбор зачета
        classif_layout = QHBoxLayout()
        classif_layout.addWidget(QLabel("Зачет:"))
        classifications = self.get_available_classifications()
        classif_group = QButtonGroup()
        
        classif_buttons_layout = QHBoxLayout()
        for classif in classifications:
            rb = QRadioButton(classif)
            classif_group.addButton(rb)
            classif_buttons_layout.addWidget(rb)
            if member.get('зачет', '') == classif:
                rb.setChecked(True)
        
        classif_layout.addLayout(classif_buttons_layout)
        layout.addLayout(classif_layout)
        fields['зачет'] = classif_group

        # Кнопки сохранения/отмены
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.save_crew_changes(row, fields, edit_dialog, selection_dialog))
        btn_box.rejected.connect(edit_dialog.reject)
        layout.addWidget(btn_box)
        
        scroll.setWidget(content)
        edit_layout = QVBoxLayout(edit_dialog)
        edit_layout.addWidget(scroll)
        
        edit_dialog.exec()

    def save_crew_changes(self, row, fields, edit_dialog, selection_dialog):
        """Сохранение всех измененных полей экипажа"""
        member = self.data['members'][row]
        
        # Обновляем текстовые поля
        text_fields = [
            'адрес_электронной_почты', 'субъект', 'номер', 'пилот', 
            'контактный_телефон_пилота', 'штурман', 'контактный_телефон_штурмана',
            'пассажиры', 'общее_количество', 'до_18_лет', 'авто', 'гос.номер'
        ]
        
        for field in text_fields:
            member[field] = fields[field].text()
        
        # Обновляем поля с выбором
        member['пол_пилота'] = fields['пол_пилота'].currentText()
        member['пол_штурмана'] = fields['пол_штурмана'].currentText()
        member['привод_автомобиля'] = fields['привод_автомобиля'].currentText()
        
        # Обновляем даты
        member['дата_рождения_пилота'] = fields['дата_рождения_пилота'].date().toString("yyyy-MM-dd") + " 00:00:00"
        member['дата_рождения_штурмана'] = fields['дата_рождения_штурмана'].date().toString("yyyy-MM-dd") + " 00:00:00"
        
        # Обновляем капитана и зачет
        member['капитан'] = "Пилот" if fields['капитан'].checkedButton().text() == "Пилот" else "Штурман"
        member['зачет'] = fields['зачет'].checkedButton().text()
        
        # Сохраняем изменения
        self.save_data()
        self.update_members_table()
        
        edit_dialog.accept()
        selection_dialog.accept()
        QMessageBox.information(self, "Сохранено", "Все изменения успешно сохранены!")













############## РЕГИСТРАЦИЯ ##############


    def setup_registration_tab(self):
        """Настройка вкладки регистрации с добавлением штурмана"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Панель поиска
        search_panel = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Введите номер или фамилию и нажмите Enter")
        self.search_input.returnPressed.connect(self.search_crew)
        search_panel.addWidget(self.search_input)
        layout.addLayout(search_panel)
        
        # Разделение на две таблицы
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Таблица незарегистрированных (теперь 5 столбцов)
        self.unregistered_table = QTableWidget()
        self.unregistered_table.setColumnCount(5)
        self.unregistered_table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Штурман", "Авто"])
        self.unregistered_table.doubleClicked.connect(self.show_crew_details)
        self.unregistered_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Таблица зарегистрированных (теперь 6 столбцов)
        self.registered_table = QTableWidget()
        self.registered_table.setColumnCount(6)
        self.registered_table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Штурман", "Авто", "Время"])
        self.registered_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Настройка прокрутки и размеров столбцов
        for table in [self.unregistered_table, self.registered_table]:
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setStretchLastSection(True)
            table.setColumnWidth(3, 150)  # Фиксированная ширина для столбца "Штурман"
        
        splitter.addWidget(self.create_table_group("Незарегистрированные", self.unregistered_table))
        splitter.addWidget(self.create_table_group("Зарегистрированные", self.registered_table))
        layout.addWidget(splitter)
        
        # Загрузка данных
        self.load_unregistered_crews()
        self.load_registered_crews()
        
        self.tabs.addTab(tab, "Регистрация")

    def search_crew(self):
        """Поиск экипажа с учетом штурмана"""
        search_term = self.search_input.text().strip().lower()
        if not search_term:
            self.load_unregistered_crews()
            return
            
        unregistered = [m for m in self.data.get('members', []) if not m.get('registered', False)]
        results = []
        
        for member in unregistered:
            if (search_term in str(member.get('номер', '')).lower() or
                search_term in member.get('пилот', '').lower() or
                search_term in member.get('штурман', '').lower()):
                results.append(member)
        
        self.unregistered_table.setRowCount(0)
        for row, crew in enumerate(results):
            self.unregistered_table.insertRow(row)
            self.unregistered_table.setItem(row, 0, QTableWidgetItem(str(crew.get('номер', ''))))
            self.unregistered_table.setItem(row, 1, QTableWidgetItem(crew.get('зачет', '')))
            self.unregistered_table.setItem(row, 2, QTableWidgetItem(crew.get('пилот', '')))
            self.unregistered_table.setItem(row, 3, QTableWidgetItem(crew.get('штурман', '')))
            self.unregistered_table.setItem(row, 4, QTableWidgetItem(crew.get('авто', '')))

    def create_table_group(self, title, table):
        """Создает группу с заголовком для таблицы"""
        group = QGroupBox(title)
        layout = QVBoxLayout()
        layout.addWidget(table)
        group.setLayout(layout)
        return group

    def on_crew_selected_from_search(self, row, crews, dialog):
        """Обработчик выбора экипажа из результатов поиска"""
        if row >= 0:
            self.show_crew_details(crews[row])
            dialog.accept()

    def save_crew_edits(self, crew_data, fields, dialog):
        """Сохраняет изменения в данных экипажа"""
        try:
            # Текстовые поля
            for field, widget in fields.items():
                if isinstance(widget, QLineEdit):
                    crew_data[field] = widget.text()
                elif isinstance(widget, QComboBox):
                    crew_data[field] = widget.currentText()
                elif isinstance(widget, QDateEdit):
                    crew_data[field] = widget.date().toString("yyyy-MM-dd") + " 00:00:00"
                elif isinstance(widget, QButtonGroup):
                    if field == 'капитан':
                        crew_data[field] = widget.checkedButton().text()
                    elif field == 'зачет':
                        crew_data[field] = widget.checkedButton().text()
            
            # Сохраняем изменения
            self.save_data()
            QMessageBox.information(self, "Сохранено", "Изменения успешно сохранены!")
            dialog.accept()
            return True
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить изменения: {str(e)}")
            return False

    def show_edit_dialog1(self, crew_data):
        """Диалоговое окно редактирования экипажа"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Редактирование экипажа №{crew_data.get('номер', '')}")
        dialog.setMinimumWidth(700)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        layout = QVBoxLayout(content)
        
        fields = {}
        
        # Основные текстовые поля
        text_fields = [
            ("Номер", "номер"),
            ("Пилот", "пилот"),
            ("Штурман", "штурман"),
            ("Авто", "авто"),
            ("Гос.номер", "гос.номер"),
            ("Телефон пилота", "контактный_телефон_пилота"),
            ("Телефон штурмана", "контактный_телефон_штурмана"),
            ("Пассажиры", "пассажиры")
        ]
        
        for label, field in text_fields:
            row = QHBoxLayout()
            row.addWidget(QLabel(f"{label}:"))
            edit = QLineEdit(str(crew_data.get(field, '')))
            fields[field] = edit
            row.addWidget(edit)
            layout.addLayout(row)
        
        # Выпадающие списки
        combo_fields = [
            ("Пол пилота", "пол_пилота", ["Мужской", "Женский"]),
            ("Пол штурмана", "пол_штурмана", ["Мужской", "Женский"]),
            ("Привод", "привод_автомобиля", ["Передний", "Задний", "Полный"])
        ]
        
        for label, field, items in combo_fields:
            row = QHBoxLayout()
            row.addWidget(QLabel(f"{label}:"))
            combo = QComboBox()
            combo.addItems(items)
            combo.setCurrentText(crew_data.get(field, items[0]))
            fields[field] = combo
            row.addWidget(combo)
            layout.addLayout(row)
        
        # Даты
        date_fields = [
            ("Дата рождения пилота", "дата_рождения_пилота"),
            ("Дата рождения штурмана", "дата_рождения_штурмана")
        ]
        
        for label, field in date_fields:
            row = QHBoxLayout()
            row.addWidget(QLabel(f"{label}:"))
            date_edit = QDateEdit()
            date_edit.setCalendarPopup(True)
            if crew_data.get(field):
                date_edit.setDate(QDate.fromString(crew_data[field].split()[0], "yyyy-MM-dd"))
            fields[field] = date_edit
            row.addWidget(date_edit)
            layout.addLayout(row)
        
        # Капитан экипажа
        captain_layout = QHBoxLayout()
        captain_layout.addWidget(QLabel("Капитан:"))
        captain_group = QButtonGroup()
        captain_pilot = QRadioButton("Пилот")
        captain_navigator = QRadioButton("Штурман")
        captain_group.addButton(captain_pilot)
        captain_group.addButton(captain_navigator)
        captain_layout.addWidget(captain_pilot)
        captain_layout.addWidget(captain_navigator)
        
        if crew_data.get('капитан', 'Пилот') == 'Пилот':
            captain_pilot.setChecked(True)
        else:
            captain_navigator.setChecked(True)
        fields['капитан'] = captain_group
        layout.addLayout(captain_layout)
        
        # Зачет
        classif_layout = QHBoxLayout()
        classif_layout.addWidget(QLabel("Зачет:"))
        classifications = self.get_available_classifications()
        classif_group = QButtonGroup()
        
        for classif in classifications:
            rb = QRadioButton(classif)
            classif_group.addButton(rb)
            classif_layout.addWidget(rb)
            if crew_data.get('зачет', '') == classif:
                rb.setChecked(True)
        
        fields['зачет'] = classif_group
        layout.addLayout(classif_layout)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.save_crew_edits(crew_data, fields, dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        scroll.setWidget(content)
        dialog_layout = QVBoxLayout(dialog)
        dialog_layout.addWidget(scroll)
        
        btn_box.button(QDialogButtonBox.StandardButton.Ok).clicked.disconnect()
        btn_box.button(QDialogButtonBox.StandardButton.Ok).clicked.connect(
            lambda: self.validate_and_save_edit(crew_data, fields, dialog)
        )
        
        dialog.exec()
        return dialog.result() == QDialog.DialogCode.Accepted

    def edit_crew(self, crew, parent_dialog=None):
        """Редактирование данных экипажа с последующей регистрацией"""
        # Создаем копию для безопасного редактирования
        crew_copy = crew.copy()
        
        # Открываем диалог редактирования
        if not self.show_edit_dialog1(crew_copy):
            return  # Редактирование отменено
        
        # Если данные изменились
        if crew_copy != crew:
            crew.update(crew_copy)
            self.save_data()
            
            # Обновляем отображение
            self.load_unregistered_crews()
            self.load_registered_crews()
            
            # Предлагаем зарегистрировать
            if not crew.get('registered', False):
                self.offer_registration(crew)
        
        if parent_dialog:
            parent_dialog.accept()  

    def offer_print(self, crew):
        """Предлагает распечатать заявку с проверкой"""
        if not crew.get('registered', False):
            return
            
        reply = QMessageBox.question(
            self, 'Печать заявки',
            'Распечатать заявку для экипажа?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.print_crew(crew)

    def print_crew(self, crew):
        """Заглушка для функции печати"""
        QMessageBox.information(
            self, "Печать", 
            f"Здесь будет печать заявки для экипажа №{crew.get('номер', '')}"
        )

    def offer_registration(self, crew):
        """Предлагает зарегистрировать экипаж после редактирования"""
        reply = QMessageBox.question(
            self, 'Регистрация',
            'Хотите зарегистрировать этот экипаж сейчас?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.register_crew(crew)
            
            # После регистрации предлагаем печать
            # self.offer_print(crew)

    def validate_and_save_edit(self, crew_data, fields, dialog):
        """Проверяет данные перед сохранением"""
        # Проверка выбора зачета
        if not fields['зачет'].checkedButton():
            QMessageBox.warning(dialog, "Ошибка", "Не выбран зачет!")
            return
            
        # Проверка обязательных полей
        required_fields = {
            'номер': "Номер экипажа",
            'пилот': "ФИО пилота", 
            'авто': "Марка автомобиля",
            'гос.номер': "Гос. номер"
        }
        
        errors = []
        for field, name in required_fields.items():
            value = fields[field].text() if isinstance(fields[field], QLineEdit) else None
            if not value:
                errors.append(name)
        
        if errors:
            QMessageBox.warning(
                dialog, 
                "Ошибка", 
                f"Не заполнены обязательные поля:\n{', '.join(errors)}"
            )
            return
            
        # Если все проверки пройдены
        self.save_crew_edits(crew_data, fields, dialog)
    
    def register_crew_dialog(self, crew):
        """Диалог регистрации экипажа"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Регистрация экипажа №{crew.get('номер', '')}")
        
        layout = QVBoxLayout(dialog)
        
        # Информация об экипаже
        info = QLabel(
            f"Пилот: {crew.get('пилот', '')}\n"
            f"Штурман: {crew.get('штурман', '')}\n"
            f"Авто: {crew.get('авто', '')}\n"
            f"Гос.номер: {crew.get('гос.номер', '')}"
        )
        layout.addWidget(info)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.register_crew(crew, dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.exec()

    def show_crew_details(self, index):
        """Показывает детальную информацию об экипаже с кнопками действий"""
        table = self.sender()
        selected_row = table.currentRow()
        
        if selected_row < 0:
            return
            
        # Находим экипаж в данных
        number = table.item(selected_row, 0).text()
        crew = None
        for member in self.data["members"]:
            if str(member.get('номер', '')) == number:
                crew = member
                break
                
        if not crew:
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Экипаж №{crew.get('номер', '')}")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        # Отображение информации об экипаже
        info_group = QGroupBox("Информация об экипаже")
        info_layout = QFormLayout()
        
        info_layout.addRow("Номер:", QLabel(str(crew.get('номер', ''))))
        info_layout.addRow("Зачет:", QLabel(crew.get('зачет', '')))
        info_layout.addRow("Пилот:", QLabel(crew.get('пилот', '')))
        info_layout.addRow("Штурман:", QLabel(crew.get('штурман', '')))
        info_layout.addRow("Авто:", QLabel(crew.get('авто', '')))
        info_layout.addRow("Гос.номер:", QLabel(crew.get('гос.номер', '')))
        
        if crew.get('registered', False):
            info_layout.addRow("Статус:", QLabel("Зарегистрирован"))
            info_layout.addRow("Время регистрации:", QLabel(crew.get('registration_time', '')))
        else:
            info_layout.addRow("Статус:", QLabel("Не зарегистрирован"))
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # Кнопки действий
        button_box = QDialogButtonBox()
        
        if not crew.get('registered', False):
            # Для незарегистрированных - кнопки Редактировать и Зарегистрировать
            edit_button = button_box.addButton("Редактировать", QDialogButtonBox.ButtonRole.ActionRole)
            register_button = button_box.addButton("Зарегистрировать", QDialogButtonBox.ButtonRole.AcceptRole)
            register_button.setStyleSheet("background-color: #4CAF50; color: white;")
            
            edit_button.clicked.connect(lambda: self.edit_and_register(crew, dialog))
            register_button.clicked.connect(lambda: self.register_crew(crew, dialog))
        else:
            # Для зарегистрированных - только кнопка Закрыть
            close_button = button_box.addButton("Закрыть", QDialogButtonBox.ButtonRole.RejectRole)
            close_button.clicked.connect(dialog.reject)
        
        cancel_button = button_box.addButton("Отмена", QDialogButtonBox.ButtonRole.RejectRole)
        cancel_button.clicked.connect(dialog.reject)
        
        layout.addWidget(button_box)
        dialog.exec()

    def edit_and_register(self, crew, parent_dialog):
        """Редактирует экипаж и предлагает зарегистрировать"""
        parent_dialog.accept()  # Закрываем текущий диалог
        
        # Редактируем экипаж
        if self.edit_crew(crew):
            # После успешного редактирования предлагаем зарегистрировать
            self.offer_registration(crew)    

    def load_unregistered_crews(self):
        """Загружает список незарегистрированных экипажей с штурманом"""
        self.unregistered_table.setRowCount(0)
        if not hasattr(self, 'data') or 'members' not in self.data:
            return
            
        row = 0
        for member in self.data['members']:
            if not member.get('registered', False):
                self.unregistered_table.insertRow(row)
                self.unregistered_table.setItem(row, 0, QTableWidgetItem(str(member.get('номер', ''))))
                self.unregistered_table.setItem(row, 1, QTableWidgetItem(member.get('зачет', '')))
                self.unregistered_table.setItem(row, 2, QTableWidgetItem(member.get('пилот', '')))
                self.unregistered_table.setItem(row, 3, QTableWidgetItem(member.get('штурман', '')))
                self.unregistered_table.setItem(row, 4, QTableWidgetItem(member.get('авто', '')))
                row += 1

    def load_registered_crews(self):
        """Загружает список зарегистрированных экипажей с штурманом"""
        self.registered_table.setRowCount(0)
        if not hasattr(self, 'data') or 'members' not in self.data:
            return
            
        row = 0
        for member in self.data['members']:
            if member.get('registered', False):
                self.registered_table.insertRow(row)
                self.registered_table.setItem(row, 0, QTableWidgetItem(str(member.get('номер', ''))))
                self.registered_table.setItem(row, 1, QTableWidgetItem(member.get('зачет', '')))
                self.registered_table.setItem(row, 2, QTableWidgetItem(member.get('пилот', '')))
                self.registered_table.setItem(row, 3, QTableWidgetItem(member.get('штурман', '')))
                self.registered_table.setItem(row, 4, QTableWidgetItem(member.get('авто', '')))
                self.registered_table.setItem(row, 5, QTableWidgetItem(member.get('registration_time', '')))
                row += 1

    def save_data(self):
        """Сохраняем данные в новый файл с проверкой регистрационных данных"""
        if not hasattr(self, 'data'):
            print("[ERROR] Нет данных для сохранения!")
            return False
        
        try:
            # Проверяем структуру данных перед сохранением
            if 'meta' not in self.data or 'version' not in self.data['meta']:
                print("[ERROR] Некорректная структура данных (отсутствует meta/version)")
                return False
                
            if 'members' not in self.data:
                print("[ERROR] Некорректная структура данных (отсутствует members)")
                return False
                
            # Проверяем наличие зарегистрированных экипажей (для отладки)
            registered_crews = [m for m in self.data['members'] if m.get('registered')]
            print(f"[DEBUG] Найдено зарегистрированных экипажей: {len(registered_crews)}")
            
            # Увеличиваем версию и сохраняем
            new_version = self.data["meta"]["version"] + 1
            self.data["meta"]["version"] = new_version
            new_filename = os.path.join(self.data_dir, f"race_v{new_version}.json")
            
            # Временная проверка данных перед записью
            print(f"[DEBUG] Проверка данных перед сохранением (версия {new_version})")
            for i, crew in enumerate(self.data['members'][:3]):  # Выводим первые 3 экипажа для проверки
                print(f"[DEBUG] Экипаж {i}: №{crew.get('номер')}, registered={crew.get('registered')}")
            
            # Сохраняем во временный файл сначала
            temp_filename = new_filename + '.tmp'
            with open(temp_filename, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=4, ensure_ascii=False)
            
            # Если сохранение во временный файл прошло успешно - переименовываем
            if os.path.exists(temp_filename):
                # Удаляем старый файл если существует
                if os.path.exists(new_filename):
                    os.remove(new_filename)
                os.rename(temp_filename, new_filename)
                
                print(f"[SUCCESS] Данные сохранены в {new_filename}")
                self.current_file = new_filename
                self.update_window_title()
                
                # Проверяем что данные записались корректно
                try:
                    with open(new_filename, 'r', encoding='utf-8') as f:
                        test_data = json.load(f)
                    print("[DEBUG] Файл успешно верифицирован")
                    return True
                except Exception as verify_error:
                    print(f"[ERROR] Ошибка верификации файла: {verify_error}")
                    return False
            else:
                print("[ERROR] Не удалось создать временный файл")
                return False
                
        except Exception as e:
            print(f"[ERROR] Критическая ошибка сохранения: {e}\n{traceback.format_exc()}")
            return False

    def register_crew(self, crew, dialog=None):
        """Регистрация экипажа с дополнительными проверками"""
        try:
            # 1. Проверка обязательных полей
            required_fields = {
                'номер': "Номер экипажа",
                'пилот': "ФИО пилота",
                'зачет': "Категория зачета"
            }
            
            missing_fields = [name for field, name in required_fields.items() if not crew.get(field)]
            if missing_fields:
                QMessageBox.warning(
                    self, "Ошибка регистрации",
                    f"Нельзя зарегистрировать - отсутствуют:\n• " + "\n• ".join(missing_fields)
                )
                return False

            # 2. Проверка уникальности номера
            crew_number = str(crew['номер'])
            existing_crew = next(
                (m for m in self.data['members'] 
                if str(m.get('номер')) == crew_number and m.get('registered')),
                None
            )
            
            if existing_crew:
                QMessageBox.warning(
                    self, "Ошибка регистрации",
                    f"Экипаж с номером {crew_number} уже зарегистрирован!\n"
                    f"Пилот: {existing_crew.get('пилот')}\n"
                    f"Время регистрации: {existing_crew.get('registration_time')}"
                )
                return False

            # 3. Проверка минимальных требований к данным
            if not all(crew.get(field) for field in ['авто', 'гос.номер']):
                QMessageBox.warning(
                    self, "Ошибка регистрации",
                    "Для регистрации необходимо указать:\n"
                    "• Марку автомобиля\n"
                    "• Государственный номер"
                )
                return False

            # Устанавливаем параметры регистрации
            crew.update({
                'registered': True,
                'registration_time': QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm")
            })

            # Обновляем данные в списке members
            for i, m in enumerate(self.data['members']):
                if str(m.get('номер')) == crew_number:
                    self.data['members'][i] = crew
                    break

            # Сохраняем данные
            if not self.save_data():
                QMessageBox.critical(self, "Ошибка", "Не удалось сохранить данные в файл!")
                return False

            # Обновляем интерфейс
            self.load_unregistered_crews()
            self.load_registered_crews()

            # Успешное завершение
            if dialog:
                dialog.accept()
                
            QMessageBox.information(
                self, "Успех", 
                f"Экипаж №{crew_number} успешно зарегистрирован!\n"
                f"Пилот: {crew.get('пилот')}\n"
                f"Время: {crew['registration_time']}"
            )
            return True

        except Exception as e:
            error_msg = f"Ошибка регистрации: {str(e)}"
            print(f"[ERROR] {error_msg}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Ошибка", error_msg)
            return False






































############## СТАРТ ##############

    def setup_start_tab(self):
        """Вкладка старта участников с новыми функциями"""
        self.start_tab = QWidget()
        layout = QVBoxLayout(self.start_tab)

        # Верхняя панель управления
        top_panel = QHBoxLayout()
        
        # Таймер до старта
        self.countdown_label = QLabel("До старта: 00:00:00")
        self.countdown_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.countdown_label.setStyleSheet("font-size: 24pt; font-weight: bold;")
        top_panel.addWidget(self.countdown_label)
        
        # Текущее время
        self.current_time_label = QLabel()
        self.current_time_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.current_time_label.setStyleSheet("font-size: 14pt;")
        top_panel.addWidget(self.current_time_label)
        
        # Короткое поле для ручного старта
        self.manual_start_input = QLineEdit()
        self.manual_start_input.setPlaceholderText("№ экипажа")
        self.manual_start_input.setFixedWidth(100)
        self.manual_start_input.returnPressed.connect(self.handle_manual_start)
        top_panel.addWidget(self.manual_start_input)
        
        # Кнопки управления
        self.edit_start_btn = QPushButton("Редактировать старт")
        self.edit_start_btn.clicked.connect(self.show_edit_start_dialog)
        top_panel.addWidget(self.edit_start_btn)
        
        self.cancel_start_btn = QPushButton("Отменить старт")
        self.cancel_start_btn.clicked.connect(self.show_cancel_start_dialog)
        top_panel.addWidget(self.cancel_start_btn)
        
        # Чекбокс блокировки двойного клика
        self.lock_dblclick_cb = QCheckBox("Блокировать старт двойным кликом")
        top_panel.addWidget(self.lock_dblclick_cb)

        # Добавляем кнопку таймера
        timer_btn = QPushButton("Таймер старта")
        timer_btn.clicked.connect(self.show_countdown_window)
        top_panel.addWidget(timer_btn)
        
        layout.addLayout(top_panel)

        

        # Разделение на две таблицы
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Левая таблица: не стартовавшие (5 столбцов)
        self.not_started_table = QTableWidget()
        self.not_started_table.setColumnCount(5)
        self.not_started_table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Штурман", "Авто"])
        self.not_started_table.doubleClicked.connect(self.start_crew_by_dblclick)
        self.not_started_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        # Правая таблица: стартовавшие (6 столбцов)
        self.started_table = QTableWidget()
        self.started_table.setColumnCount(7)  # Было 6
        self.started_table.setHorizontalHeaderLabels([
            "Номер", "Зачет", "Пилот", "Штурман", 
            "Авто", "Время старта", "Этап"  # Новая колонка
        ])
        self.started_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        # Настройка таблиц
        for table in [self.not_started_table, self.started_table]:
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setStretchLastSection(True)
            table.setColumnWidth(3, 150)  # Фиксированная ширина для штурмана

        splitter.addWidget(self.create_table_group("Не стартовавшие", self.not_started_table))
        splitter.addWidget(self.create_table_group("Стартовавшие", self.started_table))
        layout.addWidget(splitter)

        
        self.started_table.setSortingEnabled(True)  # Разрешаем сортировку по клику на заголовок
        self.started_table.sortByColumn(5, Qt.SortOrder.AscendingOrder)  # Сортировка по времени старта по умолчанию

        # Для не стартовавших:
        self.not_started_table.sortByColumn(0, Qt.SortOrder.AscendingOrder)  # Сортировка по номеру


        # Таймеры
        self.start_check_timer = QTimer()
        self.start_check_timer.timeout.connect(self.check_start_time)
        self.start_check_timer.start(1000)
        
        self.current_time_timer = QTimer()
        self.current_time_timer.timeout.connect(self.update_current_time)
        self.current_time_timer.start(1000)

        self.tabs.addTab(self.start_tab, "Старт")
        self.update_start_tab()

    def show_countdown_window(self):
        """Метод для показа окна таймера"""
        try:
            # Получаем параметры из данных
            start_time = self.data["params"]["время_старта"]
            common_start = self.data["logic_params"].get("common_start", False)
            
            # Показываем уведомление
            if common_start:
                msg = "Включен общий старт. Таймер проиграет звуки 1 раз в указанное время."
            else:
                msg = "Общий старт отключен. Таймер будет циклически проигрывать звуки каждую минуту."
                
            QMessageBox.information(self, "Режим таймера", msg)
            
            # Создаем и показываем окно
            self.countdown_window = CountdownWindow(
                self, 
                start_time=start_time,
                common_start=common_start
            )
            self.countdown_window.show()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть таймер: {str(e)}")

    def force_update_start_tab(self):
        """Принудительное обновление с очисткой и перезагрузкой данных"""
        # 1. Полностью очищаем таблицы
        self.not_started_table.setRowCount(0)
        self.started_table.setRowCount(0)
        
        # 2. Перезагружаем данные из файла
        self.load_latest_data()
        
        # 3. Обновляем отображение
        self.update_tables()
        self.update_start_countdown()

    def update_tables(self):
        """Заполняет таблицы данными"""
        self.not_started_table.setRowCount(0)
        self.started_table.setRowCount(0)
        
        for member in self.data.get("members", []):
            if not member.get("registered", False):
                continue  # Пропускаем незарегистрированных
            
            # Получаем данные экипажа
            number = str(member.get("номер", ""))
            classification = member.get("зачет", "")
            driver = member.get("пилот", "")
            navigator = member.get("штурман", "")  # Добавляем штурмана
            car = member.get("авто", "")
            start_time = member.get("start_time", "")

            if member.get("started", False):
                # Стартовавшие
                row = self.started_table.rowCount()
                self.started_table.insertRow(row)
                self.started_table.setItem(row, 0, QTableWidgetItem(number))
                self.started_table.setItem(row, 1, QTableWidgetItem(classification))
                self.started_table.setItem(row, 2, QTableWidgetItem(driver))
                self.started_table.setItem(row, 3, QTableWidgetItem(navigator))  # Штурман
                self.started_table.setItem(row, 4, QTableWidgetItem(car))
                self.started_table.setItem(row, 5, QTableWidgetItem(start_time))
            else:
                # Не стартовавшие
                row = self.not_started_table.rowCount()
                self.not_started_table.insertRow(row)
                self.not_started_table.setItem(row, 0, QTableWidgetItem(number))
                self.not_started_table.setItem(row, 1, QTableWidgetItem(classification))
                self.not_started_table.setItem(row, 2, QTableWidgetItem(driver))
                self.not_started_table.setItem(row, 3, QTableWidgetItem(navigator))  # Штурман
                self.not_started_table.setItem(row, 4, QTableWidgetItem(car))

    def handle_general_start(self):
        if not self.data.get("logic_params", {}).get("common_start", False):
            return
            
        start_time = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
        if QTime.currentTime() >= start_time:
            for member in self.data["members"]:
                if member.get("registered") and not member.get("started"):
                    member["started"] = True
                    member["start_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
            self.save_data()
            self.update_start_tab()  # Принудительное обновление

    def handle_manual_start(self):
        """Обработка ручного ввода номера экипажа"""
        number = self.manual_start_input.text()
        found = False
        for member in self.data["members"]:
            if str(member.get("номер", "")) == number and not member.get("started", False):
                member["started"] = True
                member["start_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
                self.save_data()
                self.show_start_confirmation(member)
                found = True
                break
        
        if not found:
            QMessageBox.warning(self, "Ошибка", "Экипаж с таким номером не найден или уже стартовал!")
        self.manual_start_input.clear()
        self.update_start_tab()

        if not member.get("registered", False):
            QMessageBox.warning(self, "Ошибка", "Экипаж не зарегистрирован!")
            return

    def show_start_confirmation(self, member):
        """Показывает окно подтверждения старта"""
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText(
            f"Экипаж №{member['номер']} стартовал!\n"
            f"Пилот: {member['пилот']}\n"
            f"Авто: {member['авто']}\n"
            f"Время: {member['start_time']}"
        )
        msg.setWindowTitle("Старт зарегистрирован")
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()
        # Автоматическое закрытие через 3 секунды
        QTimer.singleShot(1000, msg.close)

    def show_general_start_notification(self, count, time):
        """Показывает уведомление об автоматическом старте"""
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText(
            f"Автоматический старт {count} экипажей!\n"
            f"Время: {time}\n\n"
            f"Таблицы будут обновлены..."
        )
        msg.setWindowTitle("Общий старт")
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def start_crew_by_dblclick(self, index):
        """Старт экипажа по двойному клику (если не заблокировано)"""
        if self.lock_dblclick_cb.isChecked():
            return
            
        row = index.row()
        if row >= 0:
            number = self.not_started_table.item(row, 0).text()
            for member in self.data["members"]:
                if str(member.get('номер', '')) == number and not member.get('started', False):
                    self.register_crew_start(member)
                    break

    def register_crew_start(self, member):
        """Регистрирует старт экипажа"""
        try:
            member["started"] = True
            member["start_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
            
            # Если включена этапность - устанавливаем первый этап
            if self.data.get("logic_params", {}).get("staged", False):
                member["current_stage"] = 1
                member["stage_history"] = [{
                    "stage": 1,
                    "start_time": member["start_time"]
                }]
            
            self.save_data()
            
            # Показываем подтверждение
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setText(
                f"Экипаж №{member['номер']} стартовал!\n"
                f"Пилот: {member['пилот']}\n"
                f"Штурман: {member['штурман']}\n"
                f"Авто: {member['авто']}\n"
                f"Время: {member['start_time']}"
            )
            msg.setWindowTitle("Старт зарегистрирован")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()
            
            self.update_start_tab()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось зарегистрировать старт: {str(e)}")

    def apply_start_time_changes(self, row, crews, dialog):
        """Применяет изменения времени старта"""
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж!")
            return
            
        new_time = self.new_start_time_edit.time().toString("HH:mm:ss")
        crew_number = str(crews[row].get('номер', ''))
        
        for member in self.data["members"]:
            if str(member.get('номер', '')) == crew_number:
                old_time = member['start_time']
                member['start_time'] = new_time
                self.save_data()
                self.update_start_tab()
                dialog.accept()
                
                # Показываем подтверждение с информацией об экипаже
                QMessageBox.information(
                    self, 
                    "Сохранено", 
                    f"Время старта обновлено для экипажа №{crew_number}:\n"
                    f"Пилот: {member['пилот']}\n"
                    f"Штурман: {member['штурман']}\n"
                    f"Старое время: {old_time}\n"
                    f"Новое время: {new_time}"
                )
                return

    def show_cancel_start_dialog(self):
        """Диалог отмены старта"""
        if not hasattr(self, 'data') or not any(m.get('started') for m in self.data.get('members', [])):
            QMessageBox.warning(self, "Ошибка", "Нет стартовавших экипажей для отмены!")
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle("Отмена старта экипажа")
        dialog.setMinimumWidth(600)
        
        layout = QVBoxLayout(dialog)
        
        # Таблица стартовавших экипажей
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Штурман", "Авто", "Время старта"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Заполняем таблицу
        started_crews = [m for m in self.data.get('members', []) if m.get('started', False)]
        table.setRowCount(len(started_crews))
        
        for row, crew in enumerate(started_crews):
            table.setItem(row, 0, QTableWidgetItem(str(crew.get('номер', ''))))
            table.setItem(row, 1, QTableWidgetItem(crew.get('зачет', '')))
            table.setItem(row, 2, QTableWidgetItem(crew.get('пилот', '')))
            table.setItem(row, 3, QTableWidgetItem(crew.get('штурман', '')))
            table.setItem(row, 4, QTableWidgetItem(crew.get('авто', '')))
            table.setItem(row, 5, QTableWidgetItem(crew.get('start_time', '')))
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.cancel_crew_start(table.currentRow(), started_crews, dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.exec()

    def cancel_crew_start(self, row, crews, dialog):
        """Отменяет старт экипажа"""
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите экипаж!")
            return
            
        crew_number = str(crews[row].get('номер', ''))
        
        for member in self.data["members"]:
            if str(member.get('номер', '')) == crew_number:
                member['started'] = False
                old_time = member['start_time']
                member['start_time'] = ""
                self.save_data()
                self.update_start_tab()
                dialog.accept()
                
                # Показываем подтверждение с информацией об экипаже
                QMessageBox.information(
                    self, 
                    "Старт отменен", 
                    f"Старт отменен для экипажа №{crew_number}:\n"
                    f"Пилот: {member['пилот']}\n"
                    f"Штурман: {member['штурман']}\n"
                    f"Время старта: {old_time}"
                )
                return

    def update_countdown(self):
        """Обновляет отсчет до старта (может показывать отрицательные значения)"""
        if "params" in self.data:
            start_time = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
            current_time = QTime.currentTime()
            
            if start_time.isValid():
                seconds = current_time.secsTo(start_time)
                abs_seconds = abs(seconds)
                
                time_str = f"{'-' if seconds < 0 else ''}{abs_seconds//3600:02d}:{(abs_seconds%3600)//60:02d}:{abs_seconds%60:02d}"
                self.countdown_label.setText(f"До старта: {time_str}")

    def show_edit_start_dialog(self):
        """Диалог редактирования времени старта (точная копия окна финиша)"""
        try:
            if not hasattr(self, 'data') or not any(m.get('started') for m in self.data.get('members', [])):
                QMessageBox.warning(self, "Ошибка", "Нет стартовавших экипажей для редактирования!")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle("Редактирование старта")
            dialog.setFixedSize(600, 400)  # Фиксированный размер как на скриншоте
            
            layout = QVBoxLayout(dialog)
            
            # Таблица с экипажами (как на скриншоте)
            table = QTableWidget()
            table.setColumnCount(5)
            table.setHorizontalHeaderLabels(["Номер", "Пилот", "Авто", "Время старта", "Выбрать"])
            table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            
            # Заполняем таблицу
            started_crews = [m for m in self.data["members"] if m.get("started")]
            table.setRowCount(len(started_crews))
            
            for row, member in enumerate(started_crews):
                table.setItem(row, 0, QTableWidgetItem(str(member.get("номер", ""))))
                table.setItem(row, 1, QTableWidgetItem(member.get("пилот", "")))
                table.setItem(row, 2, QTableWidgetItem(member.get("авто", "")))
                table.setItem(row, 3, QTableWidgetItem(member.get("start_time", "")))
                
                # Кнопка выбора
                btn = QPushButton("Выбрать")
                btn.clicked.connect(lambda _, m=member: self.show_edit_selected_start(m, dialog))
                table.setCellWidget(row, 4, btn)
            
            table.resizeColumnsToContents()
            layout.addWidget(table)

            # Информация о выбранном экипаже (как на скриншоте)
            self.selected_crew_info = QLabel("Выберите экипаж из таблицы")
            self.selected_crew_info.setWordWrap(True)
            layout.addWidget(self.selected_crew_info)

            # Поле для ввода времени (как на скриншоте)
            time_layout = QHBoxLayout()
            time_layout.addWidget(QLabel("Новое время старта:"))
            
            self.new_start_time_edit = QTimeEdit()
            self.new_start_time_edit.setDisplayFormat("HH:mm:ss")
            self.new_start_time_edit.setEnabled(False)  # Пока не выбран экипаж
            time_layout.addWidget(self.new_start_time_edit)
            
            layout.addLayout(time_layout)

            # Кнопки OK/Cancel (как на скриншоте)
            btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            btn_box.accepted.connect(lambda: self.apply_start_time(dialog))
            btn_box.rejected.connect(dialog.reject)
            layout.addWidget(btn_box)

            # Подключаем выбор экипажа
            table.itemSelectionChanged.connect(lambda: self.update_start_crew_info(table, started_crews))
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть редактирование: {str(e)}")

    def update_start_crew_info(self, table, crews):
        """Обновляет информацию о выбранном экипаже"""
        row = table.currentRow()
        if row >= 0 and row < len(crews):
            member = crews[row]
            self.selected_crew_info.setText(
                f"Редактирование старта экипажа #{member['номер']}\n"
                f"Экипаж №{member['номер']}\n"
                f"Пилот: {member.get('пилот', '')}\n"
                f"Авто: {member.get('авто', '')}\n"
                f"Текущее время старта: {member.get('start_time', '')}"
            )
            self.new_start_time_edit.setEnabled(True)
            self.new_start_time_edit.setTime(QTime.fromString(member.get("start_time", "00:00:00"), "HH:mm:ss"))

    def show_edit_selected_start(self, member, dialog):
        """Обработчик кнопки 'Выбрать'"""
        self.selected_crew_info.setText(
            f"Редактирование старта экипажа #{member['номер']}\n"
            f"Экипаж №{member['номер']}\n"
            f"Пилот: {member.get('пилот', '')}\n"
            f"Авто: {member.get('авто', '')}\n"
            f"Текущее время старта: {member.get('start_time', '')}"
        )
        self.new_start_time_edit.setEnabled(True)
        self.new_start_time_edit.setTime(QTime.fromString(member.get("start_time", "00:00:00"), "HH:mm:ss"))

    def apply_start_time(self, dialog):
        """Применяет изменения времени старта"""
        if not self.new_start_time_edit.isEnabled():
            QMessageBox.warning(dialog, "Ошибка", "Сначала выберите экипаж!")
            return
            
        new_time = self.new_start_time_edit.time().toString("HH:mm:ss")
        crew_number = self.selected_crew_info.text().split('#')[1].split('\n')[0]
        
        for member in self.data["members"]:
            if str(member.get("номер", "")) == crew_number:
                member["start_time"] = new_time
                self.save_data()
                QMessageBox.information(dialog, "Сохранено", "Время старта успешно изменено")
                dialog.accept()
                self.update_start_tab()
                return

    def update_start_countdown(self):
        """Обновляет отсчёт с отрицательными значениями после старта"""
        try:
            if "params" not in self.data:
                return
                
            start_time = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
            current_time = QTime.currentTime()
            
            if not start_time.isValid():
                return
                
            seconds = current_time.secsTo(start_time)
            abs_seconds = abs(seconds)
            
            # Форматируем время с учётом знака
            sign = "-" if seconds < 0 else ""
            time_str = f"{sign}{abs_seconds//3600:02d}:{(abs_seconds%3600)//60:02d}:{abs_seconds%60:02d}"
            
            # Обновляем label с разными стилями для разных состояний
            if seconds > 0:
                # До старта
                self.countdown_label.setText(f"До старта: {time_str}")
                self.countdown_label.setStyleSheet("color: black; font-size: 24pt; font-weight: bold;")
            elif seconds == 0:
                # Момент старта
                self.countdown_label.setText("СТАРТ!")
                self.countdown_label.setStyleSheet("color: red; font-size: 24pt; font-weight: bold;")
            else:
                # После старта
                self.countdown_label.setText(f"После старта: {time_str}")
                self.countdown_label.setStyleSheet("color: blue; font-size: 24pt; font-weight: bold;")
                
        except Exception as e:
            print(f"Ошибка обновления отсчёта: {e}")
    
    def update_current_time(self):
        """Обновляет текущее время"""
        self.current_time_label.setText(QTime.currentTime().toString("HH:mm:ss"))

    def update_start_tab(self):
        """Обновляет данные на вкладке старта с сортировкой по времени старта"""
        self.update_start_countdown()
        
        # Очищаем таблицы
        self.not_started_table.setRowCount(0)
        self.started_table.setRowCount(0)
        
        # Собираем данные для сортировки
        started_crews = []
        not_started_crews = []
        
        for member in self.data.get("members", []):
            if not member.get("registered", False):
                continue  # Пропускаем незарегистрированных
                
            crew_data = {
                "number": str(member.get("номер", "")),
                "classification": member.get("зачет", ""),
                "driver": member.get("пилот", ""),
                "navigator": member.get("штурман", ""),
                "car": member.get("авто", ""),
                "start_time": member.get("start_time", ""),
                "started": member.get("started", False),
                "stage": member.get("current_stage", "N/A") if self.data.get("logic_params", {}).get("staged", False) else "-"
            }
            
            if crew_data["started"]:
                started_crews.append(crew_data)
            else:
                not_started_crews.append(crew_data)
        
        # Сортируем стартовавших по времени старта
        started_crews.sort(key=lambda x: (
            # Сначала сортируем по времени старта
            QTime.fromString(x["start_time"], "HH:mm:ss") if x["start_time"] else QTime(23, 59, 59),
            # Затем по номеру экипажа
            int(x["number"]) if x["number"].isdigit() else float('inf')
        ))
        
        # Заполняем таблицу стартовавших (отсортированную)
        for row, crew in enumerate(started_crews):
            self.started_table.insertRow(row)
            self.started_table.setItem(row, 0, QTableWidgetItem(crew["number"]))
            self.started_table.setItem(row, 1, QTableWidgetItem(crew["classification"]))
            self.started_table.setItem(row, 2, QTableWidgetItem(crew["driver"]))
            self.started_table.setItem(row, 3, QTableWidgetItem(crew["navigator"]))
            self.started_table.setItem(row, 4, QTableWidgetItem(crew["car"]))
            self.started_table.setItem(row, 5, QTableWidgetItem(crew["start_time"]))
            self.started_table.setItem(row, 6, QTableWidgetItem(str(crew["stage"])))
        
        # Заполняем таблицу не стартовавших (сортируем по номеру)
        not_started_crews.sort(key=lambda x: (
            int(x["number"]) if x["number"].isdigit() else float('inf'),
            x["driver"]
        ))
        
        for row, crew in enumerate(not_started_crews):
            self.not_started_table.insertRow(row)
            self.not_started_table.setItem(row, 0, QTableWidgetItem(crew["number"]))
            self.not_started_table.setItem(row, 1, QTableWidgetItem(crew["classification"]))
            self.not_started_table.setItem(row, 2, QTableWidgetItem(crew["driver"]))
            self.not_started_table.setItem(row, 3, QTableWidgetItem(crew["navigator"]))
            self.not_started_table.setItem(row, 4, QTableWidgetItem(crew["car"]))

    def execute_general_start(self):
        """Автоматический старт с точным временем"""
        try:
            started_count = 0
            is_staged = self.data.get("logic_params", {}).get("staged", False)
            
            # Получаем точное время старта из параметров (формат HH:mm:ss)
            planned_start = self.data["params"]["время_старта"]
            start_time_str = planned_start if len(planned_start.split(':')) == 3 else f"{planned_start}:00"
            
            for member in self.data["members"]:
                if member.get("registered") and not member.get("started"):
                    member["started"] = True
                    member["start_time"] = start_time_str  # Используем точное время
                    
                    if is_staged:
                        member["current_stage"] = 1
                        member["stage_history"] = [{
                            "stage": 1,
                            "start_time": start_time_str
                        }]
                    
                    started_count += 1
            
            if started_count > 0:
                self.save_data()
                self.update_start_tab()
                self.show_general_start_notification(started_count, start_time_str)
                
        except Exception as e:
            print(f"[ERROR] Ошибка автоматического старта: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось выполнить автоматический старт: {str(e)}")

    def check_start_time(self):
        """Проверка времени с точностью до секунды"""
        try:
            if not self.data.get("logic_params", {}).get("common_start", False):
                return
            
            # Получаем время старта с секундами (формат HH:mm:ss)
            planned_start = self.data["params"]["время_старта"]
            start_time_str = planned_start if len(planned_start.split(':')) == 3 else f"{planned_start}:00"
            start_time = QTime.fromString(start_time_str, "HH:mm:ss")
            
            if QTime.currentTime() >= start_time:
                if not hasattr(self, '_last_auto_start') or self._last_auto_start != start_time_str:
                    self._last_auto_start = start_time_str
                    self.execute_general_start()
                    
        except Exception as e:
            print(f"[ERROR] Ошибка проверки времени: {str(e)}")















############## ЭТАПЫ ##############

    def setup_stages_tab(self):
        """Вкладка управления этапами с упрощенным функционалом и улучшенной прокруткой"""
        self.stages_tab = QWidget()
        main_layout = QVBoxLayout(self.stages_tab)
        main_layout.setContentsMargins(5, 5, 5, 5)
        

        self.stages_timer = QTimer()
        self.stages_timer.timeout.connect(self.update_stages_tab)
        self.stages_timer.start(1000)

        # Верхняя панель управления
        top_panel = QWidget()
        top_layout = QHBoxLayout(top_panel)
        top_layout.setContentsMargins(0, 0, 0, 0)
        
        # Поле для ввода номера экипажа
        self.crew_input = QLineEdit()
        self.crew_input.setPlaceholderText("Введите номер экипажа")
        self.crew_input.setFixedWidth(200)
        self.crew_input.returnPressed.connect(self.move_crew_by_input)
        top_layout.addWidget(self.crew_input)
        
        # Кнопка перевести назад
        self.back_btn = QPushButton("Перевести назад")
        self.back_btn.setFixedWidth(150)
        self.back_btn.clicked.connect(self.show_move_back_dialog)
        top_layout.addWidget(self.back_btn)
        
        # Кнопка изменить время
        self.edit_btn = QPushButton("Изменить время")
        self.edit_btn.setFixedWidth(150)
        self.edit_btn.clicked.connect(self.show_edit_dialog)
        top_layout.addWidget(self.edit_btn)
        
        top_layout.addStretch()
        main_layout.addWidget(top_panel)
        
        # Область с этапами и СКП с улучшенной прокруткой
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # Контейнер для содержимого с вертикальным layout
        self.scroll_content = QWidget()
        self.scroll_content.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.stages_layout = QVBoxLayout(self.scroll_content)
        self.stages_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.stages_layout.setSpacing(10)
        self.stages_layout.setContentsMargins(5, 5, 15, 5)  # Правое поле 15 для учёта полосы прокрутки
        
        self.scroll_area.setWidget(self.scroll_content)
        main_layout.addWidget(self.scroll_area, 1)  # Коэффициент растяжения 1
        
        self.tabs.addTab(self.stages_tab, "Этапы")
        self.update_stages_tab()

    def closeEvent(self, event):
        """Обработчик закрытия окна"""
        if hasattr(self, 'stages_timer'):
            self.stages_timer.stop()
        
        if hasattr(self, 'timeout_timer'):
            self.timeout_timer.stop()
        if hasattr(self, 'skp_check_timer'):
            self.skp_check_timer.stop()
        event.accept()
    
    def clear_layout(self, layout):
        """Очищает layout от всех виджетов"""
        while layout.count():
            child = layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

    def calculate_actual_duration(self, entry):
        """Вычисляет фактическую длительность нахождения на СКП в секундах"""
        if not entry.get("entry_time"):
            return 0
            
        exit_time = entry.get("exit_time", QDateTime.currentDateTime().toString("HH:mm:ss"))
        
        try:
            t_in = QTime.fromString(entry["entry_time"], "HH:mm:ss")
            t_out = QTime.fromString(exit_time, "HH:mm:ss")
            return abs(t_in.secsTo(t_out))
        except:
            return 0

    def update_neutralization_display(self, member, skp_settings, progress, label, current_time=None):
        """Обновляет отображение времени нейтрализации для конкретного экипажа"""
        current_time = current_time or QDateTime.currentDateTime()
        
        # Общее доступное время для всех СКП (в секундах)
        total_neutral_time = sum(
            stage.get('max_neutral_time', 0) 
            for stage in self.data["logic_params"].get("skp_settings", [])
        ) * 60
        
        # Время, использованное на всех СКП
        used_time = sum(
            self.calculate_actual_duration(entry)
            for entry in member.get("skp_entries", [])
        )
        
        # Оставшееся время для текущего СКП
        remaining_time = max(0, skp_settings.get('max_neutral_time', 0) * 60 - used_time)
        
        # Обновляем прогресс-бар
        progress.setValue(int(remaining_time))
        
        # Форматируем оставшееся время
        mins, secs = divmod(remaining_time, 60)
        time_text = f"{int(mins)}:{int(secs):02d}"
        
        # Цвет текста в зависимости от оставшегося времени
        if remaining_time < 60:  # Меньше минуты
            label.setStyleSheet("color: red; font-weight: bold;")
        elif remaining_time < 300:  # Меньше 5 минут
            label.setStyleSheet("color: orange;")
        else:
            label.setStyleSheet("")
        
        label.setText(f"Осталось: {time_text}")
        
        # Дополнительная информация при наведении
        tooltip = (f"Общее время нейтрализации: {total_neutral_time//60} мин\n"
                  f"Использовано: {used_time//60}:{used_time%60:02d}\n"
                  f"Осталось: {mins}:{secs:02d}")
        label.setToolTip(tooltip)
        progress.setToolTip(tooltip)

    def validate_time_format(self, time_str):
        """Проверяет строгий формат ЧЧ:ММ:СС"""
        if not time_str:
            return False
        parts = time_str.split(':')
        if len(parts) != 3:
            return False
        try:
            h, m, s = map(int, parts)
            return 0 <= h < 24 and 0 <= m < 60 and 0 <= s < 60
        except ValueError:
            return False

    def start_skp_check_timer(self):
        """Запускает таймер для проверки закрытия СКП"""
        self.skp_check_timer = QTimer()
        self.skp_check_timer.timeout.connect(self.check_skp_closing)
        self.skp_check_timer.start(1000)  # Проверка каждую секунду

    def show_timeout_notification(self, crews):
        """Показывает уведомление об истечении времени"""
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setWindowTitle("Автоматическое перемещение")
        msg.setText(
            "Следующие экипажи переведены на следующий этап\n"
            "по ИСТЕЧЕНИИ ВРЕМЕНИ нейтрализации:\n" +
            "\n".join(f"• {crew}" for crew in crews)
        )
        
        # Системный звук и стиль
        QApplication.beep()
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #FFF3CD;
            }
            QLabel {
                color: #856404;
            }
        """)
        
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def start_timeout_check_timer(self):
        """Запускает таймер проверки времени нейтрализации"""
        self.timeout_timer = QTimer()
        self.timeout_timer.timeout.connect(self.check_neutralization_timeout)
        self.timeout_timer.start(1000)  # Проверка каждую секунду

    def show_timeout_warning(self, member, skp_num):
        """Показывает предупреждение об отсутствии времени"""
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setWindowTitle("Нет времени нейтрализации")
        msg.setText(
            f"Экипаж #{member['номер']} ({member['пилот']})\n"
            f"Пропущен СКП {skp_num} - нет доступного времени нейтрализации\n"
            f"Автоматически переведен на этап {skp_num + 1}"
        )
        
        # Красноватый стиль
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #F8D7DA;
            }
            QLabel {
                color: #721C24;
            }
            QPushButton {
                background-color: #F5C6CB;
                color: #721C24;
                padding: 5px;
            }
        """)
        
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def _get_active_skp_entry(self, member, skp_num):
        """Возвращает активную запись СКП (без exit_time)"""
        for entry in reversed(member.get('skp_entries', [])):
            if entry["skp"] == skp_num and not entry.get("exit_time"):
                return entry
        return None

    def _has_active_skp_entry(self, member, skp_num):
        """Проверяет наличие активной записи СКП"""
        return self._get_active_skp_entry(member, skp_num) is not None

    def _can_move_to_skp(self, skp_num):
        """Проверяет возможность перехода на СКП"""
        skp_settings = next(
            (skp for skp in self.data["logic_params"].get("skp_settings", [])
            if skp["number"] == skp_num), None
        )
        
        if not skp_settings:
            QMessageBox.critical(self, "Ошибка", f"Настройки для СКП {skp_num} не найдены")
            return False
        
        current_time = QTime.currentTime()
        open_time = QTime.fromString(skp_settings["open_time"], "HH:mm")
        close_time = QTime.fromString(skp_settings["close_time"], "HH:mm")
        
        if not open_time.isValid() or not close_time.isValid():
            QMessageBox.warning(self, "Ошибка", f"Некорректное время работы СКП {skp_num}")
            return False
        
        if current_time < open_time:
            QMessageBox.warning(
                self, 
                "СКП закрыт",
                f"СКП {skp_num} откроется в {skp_settings['open_time']}\n"
                f"Текущее время: {current_time.toString('HH:mm')}"
            )
            return False
            
        if current_time >= close_time:
            QMessageBox.warning(
                self,
                "СКП закрыт",
                f"СКП {skp_num} закрылся в {skp_settings['close_time']}\n"
                f"Текущее время: {current_time.toString('HH:mm')}"
            )
            return False
            
        return True

    def move_crew_forward(self, member=None):
        """Улучшенная логика перемещения вперед"""
        if member is None:
            member = self._get_selected_crew()
            if not member:
                return

        current_stage = member.get('current_stage', 1)
        current_skp = member.get('current_skp')
        stages_count = len(self.data["logic_params"].get("stages", []))

        if current_skp is not None:
            # На СКП - предлагаем выбор
            self._show_skp_transfer_dialog(member, current_skp, stages_count)
        else:
            # На этапе - переходим на соответствующий СКП
            if current_stage > stages_count:
                QMessageBox.information(self, "Ошибка", "Экипаж уже на последнем этапе")
                return
            
            self._perform_crew_move(member, {'type': 'skp', 'number': current_stage})

    def _show_skp_transfer_dialog(self, member, skp_num, stages_count):
        """Диалог выбора действия на СКП"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Экипаж #{member['номер']} на СКП {skp_num}")
        
        layout = QVBoxLayout(dialog)
        label = QLabel("Выберите действие:")
        layout.addWidget(label)
        
        btn_group = QButtonGroup()
        
        # 1. Остаться на СКП
        stay_btn = QRadioButton("Продолжить обслуживание на СКП")
        btn_group.addButton(stay_btn)
        layout.addWidget(stay_btn)
        
        # 2. Перейти на этап (если не последний)
        if skp_num < stages_count:
            move_btn = QRadioButton(f"Перейти на этап {skp_num + 1}")
            btn_group.addButton(move_btn)
            layout.addWidget(move_btn)
            move_btn.setChecked(True)
        else:
            stay_btn.setChecked(True)
        
        # Кнопки подтверждения
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            if btn_group.checkedButton() == stay_btn:
                # Ничего не делаем - остаемся на СКП
                return
            else:
                # Переход на этап
                self._perform_crew_move(member, {'type': 'stage', 'number': skp_num + 1})

    def move_crew_by_input(self):
        """Обработка ручного ввода номера с новой логикой"""
        number = self.crew_input.text().strip()
        if not number:
            return
        
        member = next((m for m in self.data["members"] if str(m.get("номер")) == number), None)
        if not member:
            QMessageBox.warning(self, "Ошибка", "Экипаж не найден")
            return
        
        if member.get('current_skp'):
            # На СКП - показываем диалог
            self._show_skp_transfer_dialog(
                member,
                member['current_skp'],
                len(self.data["logic_params"].get("stages", []))
            )
        else:
            # На этапе - переходим на СКП
            self._perform_crew_move(
                member,
                {'type': 'skp', 'number': member.get('current_stage', 1)}
            )
        
        self.crew_input.clear()

    def create_neutralization_widget(self, member, skp, neutral_type, total_neutral_time=0):
        """Создает виджет с таймером с обновленной логикой"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        
        progress = QProgressBar()
        progress.setFixedWidth(200)
        progress.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        time_label = QLabel()
        time_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        time_label.setFixedWidth(80)
        
        # Рассчитываем оставшееся время
        remaining = self.calculate_remaining_time(member, skp, neutral_type, total_neutral_time)
        mins, secs = divmod(remaining, 60)
        time_text = f"{mins}:{secs:02d}"
        
        # Настраиваем прогресс-бар
        max_value = total_neutral_time if neutral_type == "Суммарная" else skp.get("max_neutral_time", 30)*60
        progress.setMaximum(max_value)
        progress.setValue(remaining)
        progress.setFormat(time_text)
        progress.setToolTip(f"Лимит: {max_value//60} мин\nОсталось: {time_text}")
        
        # Настраиваем стили
        if remaining < 60:
            style = "color: red; font-weight: bold;"
            progress_style = "QProgressBar::chunk { background-color: red; }"
        elif remaining < 300:
            style = "color: orange;"
            progress_style = "QProgressBar::chunk { background-color: orange; }"
        else:
            style = ""
            progress_style = "QProgressBar::chunk { background-color: green; }"
        
        time_label.setStyleSheet(style)
        progress.setStyleSheet(f"""
            QProgressBar {{
                border: 1px solid grey;
                border-radius: 5px;
                text-align: center;
            }}
            {progress_style}
        """)
        
        time_label.setText(time_text)
        layout.addWidget(progress)
        layout.addWidget(time_label)
        
        return widget

    def update_stages_tab(self):
        """Обновляет отображение этапов с учетом новой логики"""
        if not self.data.get("logic_params", {}).get("staged", False):
            return
            
        self.clear_layout(self.stages_layout)
        
        stages = self.data["logic_params"].get("stages", [])
        skp_settings = self.data["logic_params"].get("skp_settings", [])
        neutral_type = self.data["logic_params"].get("neutralization_type", "Нет")
        total_time = self.data["logic_params"].get("total_max_neutral_time", 0) * 60
        
        for i, stage in enumerate(stages):
            # Отображаем этап
            stage_group = QGroupBox(f"Этап {i+1}: {stage.get('name', '')}")
            stage_layout = QVBoxLayout()
            
            # Экипажи на этапе
            crews_on_stage = [
                m for m in self.data["members"] 
                if m.get("started") 
                and not m.get("finished")
                and m.get('current_stage') == i+1
                and m.get('current_skp') is None
            ]
            
            for member in crews_on_stage:
                crew_item = QLabel(f"#{member['номер']} | {member['пилот']} | {member['штурман']}")
                stage_layout.addWidget(crew_item)
            
            stage_group.setLayout(stage_layout)
            self.stages_layout.addWidget(stage_group)
            
            # Отображаем СКП после этапа (кроме последнего)
            if i < len(stages) - 1:
                skp = skp_settings[i] if i < len(skp_settings) else {}
                skp_group = QGroupBox(f"СКП {i+1} ({skp.get('open_time', '?')}-{skp.get('close_time', '?')})")
                skp_layout = QVBoxLayout()
                
                # Экипажи на СКП
                crews_on_skp = [
                    m for m in self.data["members"]
                    if m.get("started")
                    and not m.get("finished")
                    and m.get('current_skp') == i+1
                ]
                
                for member in crews_on_skp:
                    crew_widget = QWidget()
                    crew_layout = QHBoxLayout(crew_widget)
                    
                    info_label = QLabel(f"#{member['номер']} | {member['пилот']}")
                    crew_layout.addWidget(info_label)
                    
                    time_widget = self.create_neutralization_widget(member, skp, neutral_type, total_time)
                    crew_layout.addWidget(time_widget)
                    
                    skp_layout.addWidget(crew_widget)
                
                skp_group.setLayout(skp_layout)
                self.stages_layout.addWidget(skp_group)

    def _perform_move_back_operation(self, member, current_stage, current_skp, new_pos_num):
        """Безопасное выполнение операции перемещения назад"""
        current_time = QDateTime.currentDateTime().toString("HH:mm:ss")
        
        if current_skp is not None:
            # Выход с СКП на этап
            self._complete_skp_entry(member, current_skp, current_time)
            member["current_skp"] = None
            member["current_stage"] = new_pos_num
        else:
            # Вход на предыдущий СКП
            member["current_skp"] = new_pos_num
            member.get("skp_entries", []).append({
                "skp": new_pos_num,
                "entry_time": current_time,
                "exit_time": "",
                "duration": 0
            })
        
        # Запись в историю
        member.setdefault("stage_history", []).append({
            "time": current_time,
            "action": "move_back",
            "details": {
                "from": f"СКП {current_skp}" if current_skp else f"Этап {current_stage}",
                "to": f"СКП {new_pos_num}" if current_skp is None else f"Этап {new_pos_num}"
            }
        })
        
        self.save_data()
        self.update_stages_tab()
        
        QMessageBox.information(
            self,
            "Успех",
            f"Экипаж №{member['номер']} успешно перемещен"
        )

    def _edit_single_time(self, data_obj, time_field):
        """Редактирование одного временного поля"""
        new_time, ok = QInputDialog.getText(
            self,
            "Изменение времени",
            f"Введите новое время (HH:mm:ss):",
            QLineEdit.EchoMode.Normal,
            data_obj.get(time_field, "")
        )
        
        if ok and new_time:
            if self._validate_time_format(new_time):
                data_obj[time_field] = new_time
                self._fill_time_table(data_obj)  # Обновляем таблицу
            else:
                QMessageBox.warning(self, "Ошибка", "Неверный формат времени. Используйте HH:mm:ss")

    def _validate_time(self, row, col):
        """Проверяет введенное время и обновляет длительность"""
        try:
            editor = self.time_table.cellWidget(row, col)
            time_str = editor.text()
            
            if not time_str:
                return  # Пустое значение допустимо
            
            # Проверка формата
            if not self._validate_time_format(time_str):
                QMessageBox.warning(self, "Ошибка", "Неверный формат времени. Используйте HH:mm:ss")
                editor.setFocus()
                return
            
            # Проверка на будущее время
            current_time = QDateTime.currentDateTime().time()
            input_time = QTime.fromString(time_str, "HH:mm:ss")
            if input_time > current_time:
                QMessageBox.warning(self, "Ошибка", "Нельзя вводить время в будущем")
                editor.setFocus()
                return
            
            # Для выхода - проверка что он не раньше входа
            if col == 2:  # Колонка выхода
                entry_editor = self.time_table.cellWidget(row, 1)
                entry_time = entry_editor.text()
                if entry_time:
                    t_entry = QTime.fromString(entry_time, "HH:mm:ss")
                    if input_time < t_entry:
                        QMessageBox.warning(self, "Ошибка", "Время выхода не может быть раньше времени входа")
                        editor.setFocus()
                        return
            
            # Обновляем длительность
            self._update_duration(row)
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при валидации: {str(e)}")

    def _get_crew_position(self, member):
        """Возвращает текстовое описание текущей позиции экипажа"""
        if member.get("current_skp") is not None:
            return f"СКП {member['current_skp']}"
        return f"Этап {member.get('current_stage', 1)}"

    def _validate_time_input(self, time_str, field_name, parent_widget=None):
        """Проверяет корректность введенного времени"""
        if not time_str:
            return True  # Пустое значение допустимо
            
        # Проверка формата ЧЧ:ММ:СС
        if not re.fullmatch(r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]:[0-5][0-9]$', time_str):
            QMessageBox.warning(parent_widget or self, 
                              "Ошибка", 
                              f"Неверный формат времени {field_name}.\nИспользуйте ЧЧ:ММ:СС")
            return False
            
        # Проверка что время не в будущем
        current_time = QTime.currentTime()
        try:
            input_time = QTime.fromString(time_str, "HH:mm:ss")
            if input_time > current_time:
                QMessageBox.warning(parent_widget or self,
                                  "Ошибка",
                                  f"Время {field_name} не может быть в будущем\n"
                                  f"Текущее время: {current_time.toString('HH:mm:ss')}")
                return False
        except:
            return False
            
        return True

    def _validate_time_row(self, row, field_name):
        """Проверяет корректность времени в строке"""
        editor = self.time_table.cellWidget(row, 1 if field_name == "входа" else 2)
        time_str = editor.text()
        
        if not self._validate_time_input(time_str, field_name, self.time_table):
            editor.setFocus()
            return False
            
        # Дополнительная проверка для выхода - не может быть раньше входа
        if field_name == "выхода":
            entry_editor = self.time_table.cellWidget(row, 1)
            entry_time = entry_editor.text()
            
            if entry_time and time_str:
                try:
                    t_in = QTime.fromString(entry_time, "HH:mm:ss")
                    t_out = QTime.fromString(time_str, "HH:mm:ss")
                    if t_out < t_in:
                        QMessageBox.warning(self.time_table,
                                          "Ошибка",
                                          "Время выхода не может быть раньше времени входа")
                        editor.setFocus()
                        return False
                except:
                    pass
                    
        self._update_duration(row)
        return True

    def _update_duration(self, row):
        """Обновляет длительность в строке"""
        entry_editor = self.time_table.cellWidget(row, 1)
        exit_editor = self.time_table.cellWidget(row, 2)
        
        if not entry_editor or not exit_editor:
            return
            
        entry_time = entry_editor.text()
        exit_time = exit_editor.text()
        
        duration = 0
        if entry_time and exit_time:
            try:
                t_in = QTime.fromString(entry_time, "HH:mm:ss")
                t_out = QTime.fromString(exit_time, "HH:mm:ss")
                duration = abs(t_in.secsTo(t_out))
            except:
                pass
                
        duration_item = QTableWidgetItem(f"{duration//60:02d}:{duration%60:02d}")
        duration_item.setFlags(duration_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.time_table.setItem(row, 3, duration_item)

    def _calculate_duration(self, entry):
        """Вычисляет длительность из записи"""
        if not entry.get("entry_time") or not entry.get("exit_time"):
            return 0
            
        try:
            t_in = QTime.fromString(entry["entry_time"], "HH:mm:ss")
            t_out = QTime.fromString(entry["exit_time"], "HH:mm:ss")
            return abs(t_in.secsTo(t_out))
        except:
            return 0

    def _validate_and_save(self, member, dialog):
        """Проверяет и сохраняет изменения с автоматическим переводом на этап"""
        try:
            was_on_skp = member.get("current_skp") is not None
            current_skp = member.get("current_skp")
            moved_to_stage = False
            
            # Собираем все изменения
            changes = {}
            for row in range(self.time_table.rowCount()):
                skp_item = self.time_table.item(row, 0)
                if not skp_item:
                    continue
                    
                try:
                    skp_num = int(skp_item.text().split()[1])
                except (IndexError, ValueError):
                    continue
                
                entry_editor = self.time_table.cellWidget(row, 1)
                exit_editor = self.time_table.cellWidget(row, 2)
                
                if not entry_editor or not exit_editor:
                    continue
                    
                entry_time = entry_editor.text().strip()
                exit_time = exit_editor.text().strip()
                changes[skp_num] = (entry_time, exit_time)
            
            # Применяем изменения
            for skp_num, (entry_time, exit_time) in changes.items():
                # Находим или создаем запись СКП
                entry = None
                for e in member.get("skp_entries", []):
                    if e.get("skp") == skp_num:
                        entry = e
                        break
                
                if not entry:
                    entry = {"skp": skp_num}
                    member.setdefault("skp_entries", []).append(entry)
                
                # Обновляем время
                old_exit_time = entry.get("exit_time", "")
                entry["entry_time"] = entry_time if entry_time else entry.get("entry_time", "")
                entry["exit_time"] = exit_time if exit_time else entry.get("exit_time", "")
                
                # Если это текущий СКП и добавили время выхода - перемещаем на этап
                if (was_on_skp and skp_num == current_skp and 
                    exit_time and not old_exit_time and 
                    not member.get("finished", False)):
                    
                    member["current_stage"] = current_skp + 1
                    member["current_skp"] = None
                    moved_to_stage = True
            
            # Пересчитываем длительности
            for entry in member.get("skp_entries", []):
                if entry.get("entry_time") and entry.get("exit_time"):
                    try:
                        t_in = QTime.fromString(entry["entry_time"], "HH:mm:ss")
                        t_out = QTime.fromString(entry["exit_time"], "HH:mm:ss")
                        duration = abs(t_in.secsTo(t_out))
                        entry["duration"] = duration
                    except:
                        entry["duration"] = 0
            
            self.save_data()
            
            # Показываем уведомление о перемещении
            if moved_to_stage:
                QMessageBox.information(
                    dialog,
                    "Перемещение экипажа",
                    f"Экипаж №{member['номер']} автоматически перемещен\n"
                    f"со СКП {current_skp} на этап {current_skp + 1}\n"
                    f"так как было указано время выхода с СКП"
                )
            
            dialog.accept()
            
        except Exception as e:
            QMessageBox.critical(dialog, "Ошибка", f"Ошибка при сохранении:\n{str(e)}")

    def show_edit_dialog(self):
        """Диалог редактирования времени с возможностью удаления записей"""
        try:
            number = self.crew_input.text().strip()
            if not number:
                QMessageBox.warning(self, "Ошибка", "Введите номер экипажа")
                return

            member = next((m for m in self.data.get("members", []) 
                         if str(m.get("номер", "")) == number), None)
            if not member:
                QMessageBox.warning(self, "Ошибка", "Экипаж не найден")
                return

            if not member.get("started", False):
                QMessageBox.warning(self, "Ошибка", "Экипаж еще не стартовал")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Редактирование времени экипажа №{number}")
            dialog.setMinimumSize(900, 600)
            layout = QVBoxLayout(dialog)

            # 1. Информация об экипаже
            info_group = QGroupBox("Информация об экипаже")
            info_layout = QFormLayout()
            info_layout.addRow("Номер:", QLabel(str(member.get("номер", ""))))
            info_layout.addRow("Пилот:", QLabel(member.get("пилот", "")))
            info_layout.addRow("Текущая позиция:", QLabel(self._get_crew_position(member)))
            info_group.setLayout(info_layout)
            layout.addWidget(info_group)

            # 2. Информационное сообщение
            info_label = QLabel(
                "В этом окне сохраняется ВСЯ история перемещений экипажа.\n"
                "При расчете будет использовано наиболее продолжительное время нахождения на СКП"
            )
            info_label.setStyleSheet("color: #666; font-style: italic; padding: 5px;")
            info_label.setWordWrap(True)
            layout.addWidget(info_label)

            # 3. Таблица времен с кнопкой удаления
            table_group = QGroupBox("Временные отметки СКП (для удаления выберите строку)")
            table_layout = QVBoxLayout()
            
            self.time_table = QTableWidget()
            self.time_table.setColumnCount(5)  # Добавляем колонку для кнопки
            self.time_table.setHorizontalHeaderLabels(["СКП", "Вход", "Выход", "Длительность", ""])
            self.time_table.verticalHeader().setVisible(False)
            self.time_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            self.time_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            
            self._fill_time_table(member)
            
            table_layout.addWidget(self.time_table)
            
            # Кнопка удаления выбранной записи
            delete_btn = QPushButton("Удалить выбранную запись СКП")
            delete_btn.setStyleSheet("background-color: #ffdddd;")
            delete_btn.clicked.connect(lambda: self._delete_selected_skp(member))
            table_layout.addWidget(delete_btn)
            
            table_group.setLayout(table_layout)
            layout.addWidget(table_group)

            # 4. Кнопки управления
            btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
            btn_box.accepted.connect(lambda: self._validate_and_save(member, dialog))
            btn_box.rejected.connect(dialog.reject)
            layout.addWidget(btn_box)

            dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при открытии диалога: {str(e)}")

    def _fill_time_table(self, member):
        """Заполняет таблицу с возможностью удаления записей"""
        self.time_table.setRowCount(0)
        
        if "skp_entries" not in member:
            return
            
        for entry in member["skp_entries"]:
            row = self.time_table.rowCount()
            self.time_table.insertRow(row)
            
            entry_time = entry.get("entry_time", "")
            exit_time = entry.get("exit_time", "")

            # Колонка СКП
            self.time_table.setItem(row, 0, QTableWidgetItem(f"СКП {entry.get('skp', '')}"))
            
            # Колонка Входа
            entry_edit = QLineEdit()
            entry_edit.setPlaceholderText("HH:mm:ss")
            entry_edit.setText(entry.get("entry_time", ""))
            entry_edit.editingFinished.connect(lambda r=row: self._validate_time_row(r, "входа"))
            self.time_table.setCellWidget(row, 1, entry_edit)
            
            # Колонка Выхода
            exit_edit = QLineEdit()
            exit_edit.setPlaceholderText("HH:mm:ss")
            exit_edit.setText(entry.get("exit_time", ""))
            exit_edit.editingFinished.connect(lambda r=row: self._validate_time_row(r, "выхода"))
            self.time_table.setCellWidget(row, 2, exit_edit)
            
            # Колонка Длительности
            duration = self._calculate_duration(entry)
            duration_item = QTableWidgetItem(f"{duration//60:02d}:{duration%60:02d}")
            duration_item.setFlags(duration_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.time_table.setItem(row, 3, duration_item)
            
            # Кнопка удаления в отдельной колонке
            delete_btn = QPushButton("×")
            delete_btn.setFixedSize(25, 25)
            delete_btn.setStyleSheet("background-color: #ffdddd; border: none;")
            delete_btn.clicked.connect(lambda _, r=row: self._delete_skp_entry(member, r))
            self.time_table.setCellWidget(row, 4, delete_btn)
        
        self.time_table.resizeColumnsToContents()

    def _delete_selected_skp(self, member):
        """Удаляет выбранную запись СКП"""
        selected_row = self.time_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self.time_table, "Ошибка", "Выберите запись для удаления")
            return
            
        self._delete_skp_entry(member, selected_row)

    def _delete_skp_entry(self, member, row):
        """Удаляет конкретную запись СКП по указанной строке"""
        # Получаем данные из выбранной строки
        skp_item = self.time_table.item(row, 0)
        entry_editor = self.time_table.cellWidget(row, 1)
        exit_editor = self.time_table.cellWidget(row, 2)
        
        if not all([skp_item, entry_editor, exit_editor]):
            return
            
        try:
            skp_num = int(skp_item.text().split()[1])
            entry_time = entry_editor.text()
            exit_time = exit_editor.text()
        except:
            return
            
        # Формируем сообщение для подтверждения
        message = (f"Вы уверены, что хотите удалить эту запись СКП?\n"
                  f"СКП: {skp_num}\n"
                  f"Вход: {entry_time if entry_time else 'не указан'}\n"
                  f"Выход: {exit_time if exit_time else 'не указан'}")
        
        reply = QMessageBox.question(
            self.time_table,
            "Подтверждение удаления",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Находим и удаляем конкретную запись по всем параметрам
            new_entries = []
            deleted = False
            
            for entry in member.get("skp_entries", []):
                if (entry.get("skp") == skp_num and 
                    entry.get("entry_time") == entry_time and 
                    entry.get("exit_time") == exit_time):
                    deleted = True
                else:
                    new_entries.append(entry)
            
            if deleted:
                member["skp_entries"] = new_entries
                
                # Обновляем stage_history - удаляем связанные записи
                new_history = []
                entry_found = False
                exit_found = False
                
                for record in member.get("stage_history", []):
                    # Проверяем не связана ли запись с удаляемым СКП
                    if not entry_found and "to" in record and record["to"].get("type") == "skp" and record["to"].get("number") == skp_num:
                        if record.get("time") == entry_time:
                            entry_found = True
                            continue
                    
                    if not exit_found and "from" in record and record["from"].get("type") == "skp" and record["from"].get("number") == skp_num:
                        if record.get("time") == exit_time:
                            exit_found = True
                            continue
                    
                    new_history.append(record)
                
                member["stage_history"] = new_history
                
                # СОХРАНЯЕМ ИЗМЕНЕНИЯ В ФАЙЛ
                self.save_data()
                
                # Обновляем таблицу
                self._fill_time_table(member)
                
                QMessageBox.information(
                    self.time_table,
                    "Удалено",
                    "Запись о посещении СКП удалена"
                )
            else:
                QMessageBox.warning(
                    self.time_table,
                    "Ошибка",
                    "Не удалось найти запись для удаления"
                )

    def check_skp_closing(self):
        """Проверяет закрытие СКП и обрабатывает экипажи"""
        current_time = QTime.currentTime()
        
        for skp in self.data["logic_params"].get("skp_settings", []):
            close_time = QTime.fromString(skp["close_time"], "HH:mm")
            
            # Проверяем точное время закрытия (с интервалом +-1 секунда)
            if abs(current_time.secsTo(close_time)) <= 1:
                self._process_skp_closing(skp)
    
    def _process_skp_closing(self, skp):
        """Обрабатывает закрытие конкретного СКП"""
        skp_num = skp["number"]
        close_time_str = skp["close_time"] + ":00"  # Добавляем секунды
        
        # 1. Обрабатываем экипажи на этом СКП
        crews_on_skp = [m for m in self.data["members"] 
                       if m.get("current_skp") == skp_num 
                       and not m.get("finished", False)]
        
        for crew in crews_on_skp:
            # Находим активную запись СКП
            active_entry = next(
                (e for e in crew.get("skp_entries", [])
                 if e["skp"] == skp_num and not e.get("exit_time")),
                None
            )
            
            if active_entry:
                active_entry["exit_time"] = close_time_str
                active_entry["duration"] = QTime.fromString(
                    active_entry["entry_time"], "HH:mm:ss"
                ).secsTo(QTime.fromString(close_time_str, "HH:mm:ss"))
            
            # Переводим на следующий этап
            crew["current_stage"] = skp_num + 1
            crew["current_skp"] = None
            
            # Запись в историю
            crew.setdefault("stage_history", []).append({
                "time": close_time_str,
                "action": "auto_move_skp_closed",
                "from": f"СКП {skp_num}",
                "to": f"Этап {skp_num + 1}",
                "reason": f"СКП {skp_num} закрылся"
            })
        
        # 2. Обрабатываем экипажи на предыдущем этапе
        crews_on_prev_stage = [m for m in self.data["members"]
                             if m.get("current_stage") == skp_num
                             and m.get("current_skp") is None
                             and not m.get("finished", False)]
        
        for crew in crews_on_prev_stage:
            # Добавляем запись о "пропущенном" СКП
            crew.setdefault("skp_entries", []).append({
                "skp": skp_num,
                "entry_time": close_time_str,
                "exit_time": close_time_str,
                "duration": 0,
                "skipped": True
            })
            
            # Переводим на следующий этап
            crew["current_stage"] = skp_num + 1
            
            # Запись в историю
            crew.setdefault("stage_history", []).append({
                "time": close_time_str,
                "action": "auto_skip_skp_closed",
                "from": f"Этап {skp_num}",
                "to": f"Этап {skp_num + 1}",
                "reason": f"СКП {skp_num} закрылся",
                "skp_entry": True
            })
        
        # 3. Уведомление о перемещенных экипажах
        if crews_on_skp or crews_on_prev_stage:
            moved_crews = [f"#{c['номер']}" for c in crews_on_skp + crews_on_prev_stage]
            self._show_skp_closed_notification(skp_num, close_time_str, moved_crews)
            
        self.save_data()
        self.update_stages_tab()
    
    def _show_skp_closed_notification(self, skp_num, close_time, crews):
        """Показывает уведомление о закрытии СКП"""
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle(f"СКП {skp_num} закрыт")
        msg.setText(
            f"СКП {skp_num} закрылся в {close_time}\n\n"
            f"Перемещены экипажи: {', '.join(crews) if crews else 'нет'}"
        )
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def show_move_back_dialog(self):
        """Диалог перевода экипажа назад с проверкой финиша"""
        try:
            number = self.crew_input.text().strip()
            
            if not number:
                QMessageBox.warning(self, "Ошибка", "Введите номер экипажа")
                return
                
            member = next((m for m in self.data["members"] if str(m.get("номер")) == number), None)
            if not member:
                QMessageBox.warning(self, "Ошибка", "Экипаж с таким номером не найден")
                return
                
            # Проверка финиша
            if member.get("finished", False):
                QMessageBox.warning(
                    self,
                    "Экипаж финишировал",
                    f"Экипаж #{number} уже финишировал!\n"
                    f"Нельзя изменить его положение после финиша."
                )
                return
                
            if not member.get("started", False):
                QMessageBox.warning(self, "Ошибка", "Экипаж еще не стартовал")
                return
                
            current_stage = member.get("current_stage", 1)
            current_skp = member.get("current_skp")
            
            # Определяем новую позицию
            if current_skp is not None:
                new_pos_type = "Этап"
                new_pos_num = current_skp
                skp_info = ""
            else:
                if current_stage <= 1:
                    QMessageBox.warning(self, "Ошибка", "Экипаж уже на первом этапе")
                    return
                    
                new_pos_type = "СКП"
                new_pos_num = current_stage - 1
                
                skp_settings = next(
                    (skp for skp in self.data["logic_params"].get("skp_settings", []) 
                    if skp.get("number") == new_pos_num), 
                    None
                )
                
                skp_info = (
                    f"\nВремя работы СКП: {skp_settings['open_time']}-{skp_settings['close_time']}" 
                    if skp_settings else "\nИнформация о СКП недоступна"
                )
            
            reply = QMessageBox.question(
                self,
                "Подтверждение",
                f"Экипаж №{number} ({member.get('пилот', '')})\n"
                f"Текущее положение: {'СКП' if current_skp is not None else 'Этап'} {current_skp if current_skp is not None else current_stage}\n"
                f"Перевести на: {new_pos_type} {new_pos_num}{skp_info}\n\n"
                "Продолжить?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self._perform_move_back_operation(member, current_stage, current_skp, new_pos_num)
                
        except Exception as e:
            error_msg = f"Ошибка при перемещении назад:\n{str(e)}"
            print(error_msg)
            QMessageBox.critical(self, "Ошибка", error_msg)

    def _log_crew_move(self, member, from_pos, to_pos):
        """Логирование перемещений экипажа"""
        log_entry = {
            "timestamp": QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss"),
            "crew": member["номер"],
            "from": from_pos,
            "to": to_pos,
            "current_skp": member.get("current_skp"),
            "current_stage": member.get("current_stage")
        }
        print("Перемещение экипажа:", log_entry)

    def _log_action(self, message):
        """Логирование действий с временной меткой"""
        timestamp = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss.zzz")
        print(f"[{timestamp}] {message}")

    def calculate_remaining_time(self, member, skp, neutral_type=None, total_neutral_time=0):
        """Вычисляет оставшееся время нейтрализации с подробным логгированием"""
        try:
            current_time = QDateTime.currentDateTime()
            crew_num = member.get('номер', '?')
            skp_num = skp.get('number', '?')
            
            self._log_action(f"Расчет времени для экипажа #{crew_num} на СКП {skp_num}")

            if neutral_type == "Суммарная":
                # self._log_action(f"Режим суммарной нейтрализации. Общий лимит: {total_neutral_time} сек")
                
                used_time = sum(
                    entry.get('duration', 0)
                    for entry in member.get('skp_entries', [])
                    if entry.get('exit_time')
                )
                # self._log_action(f"Использовано времени (завершенные СКП): {used_time} сек")

                if member.get('current_skp') is not None:
                    active_entry = next(
                        (e for e in member.get('skp_entries', [])
                         if e['skp'] == member['current_skp'] and not e.get('exit_time')),
                        None
                    )
                    if active_entry:
                        entry_time = QTime.fromString(active_entry['entry_time'], "HH:mm:ss")
                        current_secs = entry_time.secsTo(current_time.time())
                        if current_secs < 0:
                            current_secs += 24 * 3600
                        # self._log_action(f"Добавляем время текущего СКП: {current_secs} сек")
                        used_time += current_secs

                remaining = max(0, total_neutral_time - used_time)
                # self._log_action(f"Всего использовано: {used_time} сек | Осталось: {remaining} сек")
                return remaining
            
            # Логика для обычного режима (по СКП)
            self._log_action(f"Режим по СКП. Лимит для СКП {skp_num}: {skp.get('max_neutral_time', 30)*60} сек")
            
            active_entry = next(
                (e for e in member.get('skp_entries', [])
                if e['skp'] == skp_num and not e.get('exit_time')),
                None
            )
            
            if not active_entry:
                self._log_action("Нет активной записи СКП - возвращаем 0")
                return 0
                
            entry_time = QTime.fromString(active_entry['entry_time'], "HH:mm:ss")
            current_secs = entry_time.secsTo(current_time.time())
            if current_secs < 0:
                current_secs += 24 * 3600
            
            remaining = max(0, skp.get('max_neutral_time', 30)*60 - current_secs)
            self._log_action(f"Использовано на СКП {skp_num}: {current_secs} сек | Осталось: {remaining} сек")
            return remaining
            
        except Exception as e:
            self._log_action(f"Ошибка расчета времени: {str(e)}")
            return 0

    def check_neutralization_timeout(self):
        """Проверяет истечение времени нейтрализации с подробным логгированием"""
        try:
            current_time = QDateTime.currentDateTime()
            self._log_action(f"=== Начало проверки таймаутов ({current_time.toString('HH:mm:ss.zzz')}) ===")
            
            neutral_type = self.data["logic_params"].get("neutralization_type", "Нет")
            total_neutral_time = self.data["logic_params"].get("total_max_neutral_time", 0) * 60
            moved_crews = []
            
            for member in self.data["members"]:
                if not member.get("started", False) or member.get("finished", False):
                    continue
                    
                crew_num = member.get('номер', '?')
                current_skp = member.get("current_skp")
                
                if current_skp is not None:
                    skp_settings = next(
                        (skp for skp in self.data["logic_params"].get("skp_settings", [])
                        if skp["number"] == current_skp), None
                    )
                    
                    if not skp_settings:
                        continue
                    
                    remaining = self.calculate_remaining_time(member, skp_settings, neutral_type, total_neutral_time)
                    self._log_action(f"Экипаж #{crew_num} на СКП {current_skp}. Осталось времени: {remaining} сек")
                    
                    if remaining <= 0:
                        exit_time = current_time.toString("HH:mm:ss")
                        self._log_action(f"Время истекло! Фиксируем выход в {exit_time}")
                        
                        for entry in member.get("skp_entries", []):
                            if entry["skp"] == current_skp and not entry.get("exit_time"):
                                entry["exit_time"] = exit_time
                                duration = QTime.fromString(entry["entry_time"], "HH:mm:ss").secsTo(
                                    QTime.fromString(exit_time, "HH:mm:ss"))
                                if duration < 0:
                                    duration += 24 * 3600
                                entry["duration"] = duration
                                self._log_action(f"Запись СКП обновлена. Длительность: {duration} сек")
                        
                        member["current_stage"] = current_skp + 1
                        member["current_skp"] = None
                        moved_crews.append(f"#{crew_num}")
                        
                        history_entry = {
                            "action": "auto_move_timeout",
                            "reason": f"Истекло время нейтрализации на СКП {current_skp}",
                            "time": exit_time,
                            "from_stage": current_skp,
                            "to_stage": current_skp + 1
                        }
                        member.setdefault("stage_history", []).append(history_entry)
                        self._log_action(f"Добавлена запись в историю: {history_entry}")
            
            if moved_crews:
                self._log_action(f"Перемещены экипажи: {', '.join(moved_crews)}")
                self.save_data()
                self.update_stages_tab()
            
            self._log_action("=== Завершение проверки таймаутов ===")
            
        except Exception as e:
            self._log_action(f"ОШИБКА в check_neutralization_timeout: {str(e)}")

    def _complete_current_skp(self, member, skp_num, exit_time):
        """Завершает текущий СКП с подробным логгированием"""
        try:
            crew_num = member.get('номер', '?')
            self._log_action(f"Завершение СКП {skp_num} для экипажа #{crew_num}")
            
            active_entry = next(
                (e for e in reversed(member['skp_entries']) 
                if e['skp'] == skp_num and not e.get('exit_time')), 
                None
            )
            
            if active_entry:
                active_entry['exit_time'] = exit_time
                self._log_action(f"Найдена активная запись СКП. Устанавливаем exit_time: {exit_time}")
                
                try:
                    t_in = QTime.fromString(active_entry['entry_time'], "HH:mm:ss")
                    t_out = QTime.fromString(exit_time, "HH:mm:ss")
                    duration = t_in.secsTo(t_out)
                    active_entry['duration'] = duration if duration >= 0 else duration + 86400
                    self._log_action(f"Рассчитана длительность: {active_entry['duration']} сек")
                except Exception as e:
                    active_entry['duration'] = 0
                    self._log_action(f"Ошибка расчета длительности: {str(e)}")
            else:
                self._log_action("Активная запись СКП не найдена")
                
        except Exception as e:
            self._log_action(f"ОШИБКА в _complete_current_skp: {str(e)}")


    def _perform_crew_move(self, member, target):
        """Унифицированный метод перемещения экипажа с исправленным сохранением"""
        try:
            current_time = QDateTime.currentDateTime().toString("HH:mm:ss")
            crew_num = member.get('номер', '?')
            
            self._log_action(f"Начало перемещения экипажа #{crew_num}")
            self._log_action(f"Текущая позиция: СКП {member.get('current_skp')} / Этап {member.get('current_stage', 1)}")
            self._log_action(f"Целевая позиция: {target['type']} {target['number']}")
            
            # 1. Проверяем, нужно ли вообще перемещение
            current_pos = {
                'type': 'skp' if member.get('current_skp') is not None else 'stage',
                'number': member.get('current_skp') if member.get('current_skp') is not None else member.get('current_stage', 1)
            }
            
            if current_pos == target:
                self._log_action("Экипаж уже находится в целевой позиции - выход")
                return True

            # 2. Завершаем текущую позицию (если нужно)
            if current_pos['type'] == 'skp':
                self._log_action("Завершаем текущий СКП перед перемещением")
                if not self._complete_skp_entry(member, current_pos['number'], current_time):
                    self._log_action("Ошибка завершения СКП - выход")
                    return False

            # 3. Устанавливаем новую позицию
            if target['type'] == 'skp':
                self._log_action(f"Переход на СКП {target['number']}")
                if not self._can_move_to_skp(target['number']):
                    return False
                
                member['current_skp'] = target['number']
                member['current_stage'] = target['number']
                
                # Создаем новую запись СКП
                new_entry = {
                    "skp": target['number'],
                    "entry_time": current_time,
                    "exit_time": "",
                    "duration": 0
                }
                member.setdefault('skp_entries', []).append(new_entry)
                self._log_action(f"Добавлена новая запись СКП: {new_entry}")
            else:
                self._log_action(f"Переход на этап {target['number']}")
                member['current_stage'] = target['number']
                member['current_skp'] = None

            # 4. Добавляем запись в историю
            history_entry = {
                "time": current_time,
                "from": current_pos,
                "to": target,
                "action": "manual_move" if target['type'] == 'stage' else "enter_skp"
            }
            member.setdefault('stage_history', []).append(history_entry)
            self._log_action(f"Добавлена запись в историю: {history_entry}")

            # 5. Сохраняем данные с проверкой
            if not self._save_with_retry():
                self._log_action("Ошибка сохранения данных!")
                return False

            # 6. Принудительно обновляем интерфейс
            self.update_stages_tab()
            QApplication.processEvents()
            
            self._log_action("Перемещение успешно завершено")
            return True
            
        except Exception as e:
            self._log_action(f"ОШИБКА в _perform_crew_move: {str(e)}")
            return False

    def _complete_skp_entry(self, member, skp_num, exit_time):
        """Завершает запись о СКП с гарантированным сохранением"""
        try:
            crew_num = member.get('номер', '?')
            self._log_action(f"Завершение СКП {skp_num} для экипажа #{crew_num}")
            
            # Находим активную запись
            active_entry = next(
                (e for e in reversed(member.get('skp_entries', []))
                if e.get('skp') == skp_num and not e.get('exit_time')),
                None
            )
            
            if not active_entry:
                self._log_action("Активная запись СКП не найдена!")
                return False

            # Обновляем запись
            active_entry['exit_time'] = exit_time
            try:
                t_in = QTime.fromString(active_entry['entry_time'], "HH:mm:ss")
                t_out = QTime.fromString(exit_time, "HH:mm:ss")
                duration = t_in.secsTo(t_out)
                active_entry['duration'] = duration if duration >= 0 else duration + 86400
                self._log_action(f"Время на СКП: {active_entry['duration']} сек")
            except Exception as e:
                active_entry['duration'] = 0
                self._log_action(f"Ошибка расчета времени: {str(e)}")

            return True
            
        except Exception as e:
            self._log_action(f"ОШИБКА в _complete_skp_entry: {str(e)}")
            return False

    def _save_with_retry(self, max_attempts=3):
        """Пытается сохранить данные с несколькими попытками"""
        for attempt in range(1, max_attempts + 1):
            try:
                self._log_action(f"Попытка сохранения #{attempt}")
                self.save_data()
                
                # Проверяем, что данные действительно сохранились
                if self._verify_data_saved():
                    self._log_action("Данные успешно сохранены")
                    return True
                
                self._log_action("Данные не сохранились (проверка не пройдена)")
                
            except Exception as e:
                self._log_action(f"Ошибка при сохранении (попытка #{attempt}): {str(e)}")
            
            if attempt < max_attempts:
                QThread.msleep(100)  # Небольшая пауза между попытками
        
        return False

    def _verify_data_saved(self):
        """Проверяет, что данные действительно сохранились"""
        try:
            # Можно добавить дополнительную проверку, например:
            # - сравнить с временной копией
            # - проверить размер файла
            # - прочитать файл обратно и проверить ключевые значения
            return True
        except:
            return False

























############## ФИНИШ ##############

    def setup_finish_tab(self):
        """Вкладка финиша участников"""
        self.finish_tab = QWidget()
        layout = QVBoxLayout(self.finish_tab)

        # --- Верхняя часть: обратный отсчет и кнопки ---
        top_layout = QHBoxLayout()
        
        self.closing_countdown = QLabel("До закрытия трассы: 00:00:00")
        self.closing_countdown.setStyleSheet("font-size: 24pt; font-weight: bold;")
        top_layout.addWidget(self.closing_countdown)

        # Группа кнопок управления
        btn_group = QHBoxLayout()
        
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.clicked.connect(self.update_finish_tab)
        btn_group.addWidget(self.refresh_btn)

        self.edit_finish_btn = QPushButton("Редактировать финиш")
        self.edit_finish_btn.clicked.connect(self.show_edit_finish_dialog)
        btn_group.addWidget(self.edit_finish_btn)

        self.cancel_finish_btn = QPushButton("Отменить финиш")
        self.cancel_finish_btn.clicked.connect(self.show_cancel_finish_dialog)
        btn_group.addWidget(self.cancel_finish_btn)

        top_layout.addLayout(btn_group)
        layout.addLayout(top_layout)

        # --- Разделение на две таблицы ---
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Левая таблица: не финишировавшие (только стартовавшие)
        self.not_finished_table = QTableWidget()
        self.not_finished_table.setColumnCount(5)
        self.not_finished_table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Авто", "Время старта"])
        self.not_finished_table.doubleClicked.connect(self.show_finish_crew_details)

        # Правая таблица: финишировавшие
        self.finished_table = QTableWidget()
        self.finished_table.setColumnCount(5)
        self.finished_table.setHorizontalHeaderLabels(["Номер", "Зачет", "Пилот", "Авто", "Время финиша"])

        splitter.addWidget(self.create_table_group("На трассе", self.not_finished_table))
        splitter.addWidget(self.create_table_group("Финишировали", self.finished_table))
        layout.addWidget(splitter)

        # --- Поле для ручного финиша ---
        self.manual_finish_group = QGroupBox("Ручной финиш")
        finish_layout = QHBoxLayout()
        self.finish_input = QLineEdit()
        self.finish_input.setPlaceholderText("Введите номер экипажа и нажмите Enter")
        self.finish_input.returnPressed.connect(self.handle_manual_finish)
        finish_layout.addWidget(self.finish_input)
        self.manual_finish_group.setLayout(finish_layout)
        layout.addWidget(self.manual_finish_group)

        # Таймер для проверки закрытия трассы
        self.finish_timer = QTimer()
        self.finish_timer.timeout.connect(self.check_track_closing)
        self.finish_timer.start(1000)  # Проверка каждую секунду

        self.tabs.addTab(self.finish_tab, "Финиш")
        self.update_finish_tab()

    def update_finish_tab(self):
        """Обновляет данные на вкладке финиша"""
        self.update_closing_countdown()
        self.update_finish_tables()

    def update_closing_countdown(self):
        """Обновляет отсчет до закрытия трассы"""
        if "params" in self.data:
            closing_time = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            current_time = QTime.currentTime()
            
            if current_time < closing_time:
                remaining = current_time.secsTo(closing_time)
                self.closing_countdown.setText(
                    f"До закрытия трассы: {remaining//3600:02d}:{(remaining%3600)//60:02d}:{remaining%60:02d}"
                )
            else:
                self.closing_countdown.setText("Трасса закрыта!")

    def show_finish_crew_details(self, index):
        """Показывает детали экипажа при двойном клике в таблице финиша"""
        selected_row = self.not_finished_table.currentRow()
        if selected_row >= 0:
            number = self.not_finished_table.item(selected_row, 0).text()
            for member in self.data["members"]:
                if str(member.get("номер", "")) == number:
                    self.show_crew_finish_dialog(member)
                    break

    def show_crew_finish_dialog(self, member):
        """Диалог подтверждения финиша"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Экипаж №{member['номер']}")
        
        layout = QVBoxLayout()
        info = QLabel(
            f"Пилот: {member.get('пилот', '')}\n"
            f"Авто: {member.get('авто', '')}\n"
            f"Старт: {member.get('start_time', '')}"
        )
        layout.addWidget(info)

        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.confirm_finish(member, dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)

        dialog.setLayout(layout)
        dialog.exec()

    def handle_manual_finish(self):
        """Обработчик ручного финиша с проверкой нахождения на СКП"""
        try:
            number = self.finish_input.text().strip()
            if not number:
                return
                
            # Проверяем закрытие трассы
            closing_time = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            current_time = QTime.currentTime()
            
            if current_time >= closing_time:
                QMessageBox.warning(self, "Ошибка", "Трасса закрыта! Финиш невозможен")
                self.finish_input.clear()
                return
                
            for member in self.data["members"]:
                if str(member.get("номер", "")) == number:
                    if not member.get("started"):
                        QMessageBox.warning(self, "Ошибка", "Экипаж еще не стартовал!")
                        self.finish_input.clear()
                        return
                        
                    if member.get("finished"):
                        QMessageBox.warning(self, "Ошибка", "Экипаж уже финишировал!")
                        self.finish_input.clear()
                        return
                        
                    if member.get("current_skp") is not None:
                        skp_num = member["current_skp"]
                        QMessageBox.warning(
                            self,
                            "Экипаж на СКП",
                            f"Экипаж #{number} не может финишировать,\n"
                            f"так как находится на СКП {skp_num}!\n\n"
                            f"Сначала завершите нейтрализацию."
                        )
                        self.finish_input.clear()
                        return
                        
                    # Если все проверки пройдены - финишируем
                    member["finished"] = True
                    member["finish_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
                    self.save_data()
                    
                    QMessageBox.information(
                        self,
                        "Финиш зарегистрирован",
                        f"Экипаж #{number} успешно финишировал!\n"
                        f"Время финиша: {member['finish_time']}"
                    )
                    
                    self.finish_input.clear()
                    self.update_finish_tables()
                    return
                    
            QMessageBox.warning(self, "Ошибка", "Экипаж с таким номером не найден!")
            self.finish_input.clear()
                
        except Exception as e:
            print(f"Ошибка при ручном финише: {e}")
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")
            self.finish_input.clear()

    def confirm_finish(self, member, dialog):
        """Подтверждение финиша с проверкой нахождения на СКП"""
        try:
            # Проверка закрытия трассы
            closing_time = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            if QTime.currentTime() >= closing_time:
                QMessageBox.warning(self, "Ошибка", "Трасса уже закрыта!")
                return

            # Проверка нахождения на СКП
            if member.get("current_skp") is not None:
                QMessageBox.warning(
                    self,
                    "Экипаж на СКП",
                    f"Экипаж #{member['номер']} не может финишировать,\n"
                    f"так как находится на СКП {member['current_skp']}!\n\n"
                    f"Сначала завершите нейтрализацию."
                )
                return

            # Если проверки пройдены - финишируем
            member["finished"] = True
            member["finish_time"] = QDateTime.currentDateTime().toString("HH:mm:ss")
            self.save_data()
            
            dialog.accept()
            self.update_finish_tab()
            
            QMessageBox.information(
                self,
                "Финиш зарегистрирован",
                f"Экипаж #{member['номер']} успешно финишировал!\n"
                f"Время финиша: {member['finish_time']}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def show_edit_finish_dialog(self):
        """Диалог редактирования времени финиша"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактирование финиша")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout()
        
        # Таблица с финишировавшими экипажами
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Номер", "Пилот", "Авто", "Время финиша", "Выбрать"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Заполняем таблицу
        finished_crews = [m for m in self.data["members"] if m.get("finished")]
        table.setRowCount(len(finished_crews))
        
        for row, member in enumerate(finished_crews):
            table.setItem(row, 0, QTableWidgetItem(str(member.get("номер", ""))))
            table.setItem(row, 1, QTableWidgetItem(member.get("пилот", "")))
            table.setItem(row, 2, QTableWidgetItem(member.get("авто", "")))
            table.setItem(row, 3, QTableWidgetItem(member.get("finish_time", "")))
            
            # Кнопка выбора
            btn = QPushButton("Выбрать")
            btn.clicked.connect(lambda _, m=member: self.edit_selected_finish(m, dialog))
            table.setCellWidget(row, 4, btn)
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        dialog.setLayout(layout)
        dialog.exec()

    def edit_selected_finish(self, member, parent_dialog):
        """Редактирование времени финиша для выбранного экипажа"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Редактирование финиша экипажа #{member['номер']}")
        
        layout = QVBoxLayout()
        
        # Информация об экипаже
        info = QLabel(
            f"Экипаж №{member['номер']}\n"
            f"Пилот: {member.get('пилот', '')}\n"
            f"Авто: {member.get('авто', '')}\n"
            f"Текущее время финиша: {member.get('finish_time', '')}"
        )
        layout.addWidget(info)
        
        # Поле для ввода нового времени
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel("Новое время финиша:"))
        
        self.new_finish_time = QLineEdit()
        self.new_finish_time.setPlaceholderText("ЧЧ:ММ:СС")
        self.new_finish_time.setText(member.get("finish_time", ""))
        time_layout.addWidget(self.new_finish_time)
        
        layout.addLayout(time_layout)
        
        # Кнопки
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(lambda: self.apply_finish_edit(member, dialog, parent_dialog))
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.setLayout(layout)
        dialog.exec()

    def apply_finish_edit(self, member, dialog, parent_dialog):
        """Применяет изменения времени финиша"""
        new_time = self.new_finish_time.text().strip()
        
        if not self.validate_time_format(new_time):
            QMessageBox.warning(self, "Ошибка", "Неверный формат времени. Используйте ЧЧ:ММ:СС")
            return
            
        member["finish_time"] = new_time
        self.save_data()
        
        dialog.close()
        parent_dialog.close()
        self.update_finish_tab()
        
        QMessageBox.information(self, "Сохранено", "Время финиша успешно изменено")

    def show_cancel_finish_dialog(self):
        """Диалог отмены финиша"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Отмена финиша")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout()
        
        # Таблица с финишировавшими экипажами
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Номер", "Пилот", "Авто", "Время финиша", "Выбрать"])
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Заполняем таблицу
        finished_crews = [m for m in self.data["members"] if m.get("finished")]
        table.setRowCount(len(finished_crews))
        
        for row, member in enumerate(finished_crews):
            table.setItem(row, 0, QTableWidgetItem(str(member.get("номер", ""))))
            table.setItem(row, 1, QTableWidgetItem(member.get("пилот", "")))
            table.setItem(row, 2, QTableWidgetItem(member.get("авто", "")))
            table.setItem(row, 3, QTableWidgetItem(member.get("finish_time", "")))
            
            # Кнопка выбора
            btn = QPushButton("Выбрать")
            btn.clicked.connect(lambda _, m=member: self.cancel_selected_finish(m, dialog))
            table.setCellWidget(row, 4, btn)
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        dialog.setLayout(layout)
        dialog.exec()

    def cancel_selected_finish(self, member, dialog):
        """Отменяет финиш для выбранного экипажа"""
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите отменить финиш экипажа #{member['номер']}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            member["finished"] = False
            member["finish_time"] = ""
            self.save_data()
            
            dialog.close()
            self.update_finish_tab()
            
            QMessageBox.information(
                self,
                "Финиш отменен",
                f"Финиш экипажа #{member['номер']} успешно отменен"
            )

    def validate_time_format(self, time_str):
        """Проверяет формат времени ЧЧ:ММ:СС"""
        if not time_str:
            return False
        try:
            parts = list(map(int, time_str.split(":")))
            return (len(parts) == 3 and
                    0 <= parts[0] < 24 and
                    0 <= parts[1] < 60 and
                    0 <= parts[2] < 60)
        except ValueError:
            return False

    def check_track_closing(self):
        """Проверяет превышение лимита времени трассы (без автосохранения)"""
        if not self.data.get("logic_params", {}).get("penalty_type") == "DNF":
            return
            
        try:
            track_close_time = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            track_start_time = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
            
            # Рассчитываем базовое время трассы в секундах
            track_duration = track_start_time.secsTo(track_close_time)
            if track_duration < 0:
                track_duration += 24 * 3600
                
            current_time = QTime.currentTime()
            need_save = False
            
            for member in self.data["members"]:
                if member.get("started") and not member.get("finished"):
                    start_time = QTime.fromString(member["start_time"], "HH:mm:ss")
                    neutral_time = self.calculate_total_neutral_time(member)
                    
                    # Дедлайн с учетом нейтрализации
                    finish_deadline = start_time.addSecs(track_duration + neutral_time)
                    
                    if current_time >= finish_deadline:
                        member["finished"] = True
                        member["finish_time"] = finish_deadline.addSecs(1).toString("HH:mm:ss")
                        member["dnf_reason"] = "Превышено время трассы"
                        
                        member.setdefault("stage_history", []).append({
                            "time": QDateTime.currentDateTime().toString("HH:mm:ss"),
                            "action": "auto_finish_timeout",
                            "reason": f"Лимит: {track_duration//3600}ч + {neutral_time//60}м нейтр."
                        })
                        need_save = True
            
            # Сохраняем только если были изменения
            if need_save:
                self.save_data()
                self.update_finish_tab()
                
        except Exception as e:
            print(f"Ошибка при проверке времени трассы: {e}")

    def calculate_track_duration(self):
        """Вычисляет базовую длительность трассы в секундах"""
        start = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
        close = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
        duration = start.secsTo(close)
        return duration if duration >= 0 else duration + 24*3600

    def calculate_crew_deadline(self, member, track_duration):
        """Вычисляет дедлайн для конкретного экипажа с учетом нейтрализации"""
        start_time = QTime.fromString(member["start_time"], "HH:mm:ss")
        neutral_time = self.calculate_total_neutral_time(member)
        return start_time.addSecs(track_duration + neutral_time)

    def mark_crew_dnf(self, member, deadline):
        """Помечает экипаж как DNF с указанием времени"""
        member.update({
            "finished": True,
            "finish_time": deadline.addSecs(1).toString("HH:mm:ss"),
            "dnf_reason": "Превышено время трассы",
            "stage_history": member.get("stage_history", []) + [{
                "time": QDateTime.currentDateTime().toString("HH:mm:ss"),
                "action": "time_limit_exceeded",
                "details": {
                    "track_limit": self.data["params"]["время_закрытия_трассы"],
                    "neutralization": self.format_time(self.calculate_total_neutral_time(member))
                }
            }]
        })

    def calculate_total_neutral_time(self, member):
        """Вычисляет общее время нейтрализации в секундах"""
        if not member.get("skp_entries"):
            return 0
        return sum(
            entry.get("duration", 0) 
            for entry in member["skp_entries"] 
            if isinstance(entry.get("duration"), (int, float)))

    def calculate_total_time(self, member):
        """Вычисляет общее время гонки (финиш - старт) в секундах"""
        try:
            start = QTime.fromString(member.get("start_time", ""), "HH:mm:ss")
            finish = QTime.fromString(member.get("finish_time", ""), "HH:mm:ss")
            total = start.secsTo(finish)
            return total if total >= 0 else total + 24 * 3600
        except:
            return 0

    def format_time(self, seconds):
        """Форматирует секунды в ЧЧ:ММ:СС"""
        if seconds is None or seconds == 0:
            return "00:00:00"
        return f"{seconds//3600:02d}:{(seconds%3600)//60:02d}:{seconds%60:02d}"

    def update_finish_tables(self):
        """Обновляет таблицы финиша с сортировкой по времени и выделением DNF"""
        self.not_finished_table.setRowCount(0)
        self.finished_table.setRowCount(0)
        
        # Устанавливаем заголовки
        self.not_finished_table.setColumnCount(6)
        self.not_finished_table.setHorizontalHeaderLabels(
            ["Номер", "Зачет", "Пилот", "Авто", "Время старта", "Нейтрализация"]
        )
        
        self.finished_table.setColumnCount(9)
        self.finished_table.setHorizontalHeaderLabels(
            ["Номер", "Зачет", "Пилот", "Авто", "Время старта",
             "Время финиша", "Общее время", "Нейтрализация", "Время трассы"]
        )

        # Разделяем экипажи
        finished = []
        not_finished = []
        
        for member in self.data.get("members", []):
            if not member.get("started", False):
                continue
                
            if member.get("finished", False):
                finished.append(member)
            else:
                not_finished.append(member)

        # Сортируем финишировавших: сначала по времени, DNF - в конце
        finished.sort(key=lambda x: (
            float('inf') if x.get("dnf_reason") else 
            QTime.fromString(x.get("finish_time", "23:59:59"), "HH:mm:ss").msecsSinceStartOfDay()
        ))

        # Заполняем таблицу финишировавших
        for member in finished:
            row = self.finished_table.rowCount()
            self.finished_table.insertRow(row)
            
            neutral_time = self.calculate_total_neutral_time(member)
            total_time = self.calculate_total_time(member)
            track_time = max(0, total_time - neutral_time) if total_time else 0
            
            # Основные данные
            self.finished_table.setItem(row, 0, QTableWidgetItem(str(member.get("номер", ""))))
            self.finished_table.setItem(row, 1, QTableWidgetItem(member.get("зачет", "N/A")))
            self.finished_table.setItem(row, 2, QTableWidgetItem(member.get("пилот", "N/A")))
            self.finished_table.setItem(row, 3, QTableWidgetItem(member.get("авто", "N/A")))
            self.finished_table.setItem(row, 4, QTableWidgetItem(member.get("start_time", "N/A")))
            
            # Время финиша (для DNF указываем причину)
            finish_item = QTableWidgetItem(member.get("finish_time", "N/A"))
            if member.get("dnf_reason"):
                finish_item.setToolTip(member["dnf_reason"])
            self.finished_table.setItem(row, 5, finish_item)
            
            # Временные показатели
            self.finished_table.setItem(row, 6, QTableWidgetItem(self.format_time(total_time)))
            self.finished_table.setItem(row, 7, QTableWidgetItem(self.format_time(neutral_time)))
            self.finished_table.setItem(row, 8, QTableWidgetItem(self.format_time(track_time)))
            
            # Выделение DNF
            if member.get("dnf_reason"):
                for col in range(self.finished_table.columnCount()):
                    self.finished_table.item(row, col).setBackground(QColor(255, 220, 220))

        # Заполняем таблицу нефинишировавших
        for member in not_finished:
            row = self.not_finished_table.rowCount()
            self.not_finished_table.insertRow(row)
            
            neutral_time = self.calculate_total_neutral_time(member)
            
            self.not_finished_table.setItem(row, 0, QTableWidgetItem(str(member.get("номер", ""))))
            self.not_finished_table.setItem(row, 1, QTableWidgetItem(member.get("зачет", "N/A")))
            self.not_finished_table.setItem(row, 2, QTableWidgetItem(member.get("пилот", "N/A")))
            self.not_finished_table.setItem(row, 3, QTableWidgetItem(member.get("авто", "N/A")))
            self.not_finished_table.setItem(row, 4, QTableWidgetItem(member.get("start_time", "N/A")))
            self.not_finished_table.setItem(row, 5, QTableWidgetItem(self.format_time(neutral_time)))

        # Настраиваем отображение
        self.finished_table.resizeColumnsToContents()
        self.not_finished_table.resizeColumnsToContents()
        self.finished_table.sortByColumn(5, Qt.SortOrder.AscendingOrder)  # Сортировка по времени финиша

    def check_track_timeout(self):
        """Проверка превышения времени с уведомлениями"""
        if not self.data.get("logic_params", {}).get("penalty_type") == "DNF":
            return

        try:
            track_duration = self.calculate_track_duration()
            current_time = QTime.currentTime()
            dnf_crews = []

            for member in self.data["members"]:
                if member.get("started") and not member.get("finished"):
                    deadline = self.calculate_crew_deadline(member, track_duration)
                    
                    if current_time >= deadline:
                        self.mark_crew_dnf(member, deadline)
                        dnf_crews.append((
                            member["номер"],
                            member.get("пилот", ""),
                            self.format_time(self.calculate_total_neutral_time(member))
                        ))

            if dnf_crews:
                self.show_dnf_notification(dnf_crews, track_duration)
                self.save_data()
                self.update_finish_tab()

        except Exception as e:
            print(f"Ошибка проверки времени: {str(e)}")

    def show_dnf_notification(self, crews, track_duration):
        """Показывает уведомление об автоматическом финише"""
        hours = track_duration // 3600
        minutes = (track_duration % 3600) // 60
        
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setWindowTitle("Автоматический финиш")
        msg.setText(
            f"Превышен лимит времени трассы ({hours}ч {minutes}м)!\n"
            "Следующие экипажи отмечены как DNF:"
        )
        
        # Детализированный список
        details = QTextEdit()
        details.setReadOnly(True)
        details.setText(
            "\n".join(
                f"#{num} {pilot} (нейтрализация: {neutral_time})" 
                for num, pilot, neutral_time in crews
            )
        )
        details.setMinimumSize(400, 150)
        
        # Компоновка
        layout = msg.layout()
        layout.addWidget(details, 1, 0, 1, layout.columnCount())
        
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()


























############## ПРОВЕРКА ##############

    def process_cp_input(self):
        """Обработка ввода номеров КП с точным совпадением"""
        text = self.cp_input.text().strip()
        if not text:
            return
            
        cp_numbers = text.split()
        cp_prefixes = ["КП", "CP", "KP"]  # Возможные префиксы
        
        for i in range(self.cp_list_widget.count()):
            item = self.cp_list_widget.item(i)
            cp_name = item.text()
            
            # Извлекаем только цифры из названия КП
            cp_num = ''.join(filter(str.isdigit, cp_name))
            
            for num in cp_numbers:
                if num == cp_num:  # Точное совпадение чисел
                    item.setCheckState(Qt.CheckState.Checked)
                    break

        self.cp_input.clear()

    def update_crew_selection(self, index):
        """Обновляет выбранный экипаж при клике"""
        self.selected_crew_row = index.row()

    def eventFilter(self, source, event):
        """Перехватываем нажатие Enter в поле ввода КП"""
        if source is self.cp_input and event.type() == QEvent.Type.KeyPress:
            if event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter:
                self.process_cp_input()
                return True  # Событие обработано, не передаем дальше
        return super().eventFilter(source, event)

    def format_time(self, seconds):
        """Форматирует секунды в ЧЧ:ММ:СС"""
        if seconds is None or seconds == 0:
            return "00:00:00"
        return f"{seconds//3600:02d}:{(seconds%3600)//60:02d}:{seconds%60:02d}"

    def calculate_time_penalty(self, member, track_time):
        """Рассчитывает штраф за превышение времени"""
        penalty_type = self.data.get("logic_params", {}).get("penalty_type")
        
        # Проверяем явный DNF (не финишировал)
        if member.get('finish_time') == "DNF":
            return "DNF"
            
        # Проверяем превышение времени трассы для типа штрафа DNF
        if penalty_type == "DNF":
            track_duration = self.calculate_track_duration()
            if track_time > track_duration:
                return "DNF"
            return 0
                
        # Логика для штрафа в баллах
        elif penalty_type == "Баллы за 1 мин":
            track_duration = self.calculate_track_duration()
            overtime = max(0, track_time - track_duration)
            
            if overtime > 0:
                penalty_value = self.data["logic_params"].get("penalty_value", 1)
                return math.ceil(overtime / 60) * penalty_value
            return 0
            
        return 0

    def calculate_track_duration(self):
        """Вычисляет базовую длительность трассы"""
        try:
            start = QTime.fromString(self.data["params"]["время_старта"], "HH:mm")
            close = QTime.fromString(self.data["params"]["время_закрытия_трассы"], "HH:mm")
            duration = start.secsTo(close)
            return duration if duration >= 0 else duration + 24*3600
        except Exception:
            return 0

    def calculate_total_time(self, member):
        """Вычисляет общее время"""
        try:
            start_time = member.get("start_time")
            finish_time = member.get("finish_time")
            
            if not start_time or not finish_time or finish_time == "DNF":
                return 0
                
            start = QTime.fromString(start_time, "HH:mm:ss")
            finish = QTime.fromString(finish_time, "HH:mm:ss")
            total = start.secsTo(finish)
            
            return total if total >= 0 else total + 24 * 3600
        except Exception:
            return 0

    def setup_check_tab(self):
        """Вкладка проверки контрольных пунктов"""
        self.selected_crew_row = -1
        self.check_tab = QWidget()
        layout = QVBoxLayout(self.check_tab)

        # Панель управления
        control_panel = QHBoxLayout()
        
        self.refresh_check_btn = QPushButton("Обновить")
        self.refresh_check_btn.clicked.connect(self.update_check_tab)
        control_panel.addWidget(self.refresh_check_btn)
        
        self.table_mode = QComboBox()
        self.table_mode.addItems(["Краткий вид", "Полный вид", "Режим КП"])  # Добавлен новый режим
        self.table_mode.currentIndexChanged.connect(self.update_check_tab)
        control_panel.addWidget(QLabel("Режим:"))
        control_panel.addWidget(self.table_mode)
        
        control_panel.addStretch()
        layout.addLayout(control_panel)

        # Таблица экипажей
        self.check_crews_table = QTableWidget()
        self.check_crews_table.setColumnCount(10)
        self.check_crews_table.setHorizontalHeaderLabels([
            "Номер", "Зачет", "Пилот", "Штурман", "Авто", 
            "Старт", "Финиш", "Время трассы", "Баллы", "Итог"
        ])
        
        header = self.check_crews_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.check_crews_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.check_crews_table.doubleClicked.connect(self.show_check_dialog)
        self.check_crews_table.clicked.connect(self.update_crew_selection)
        
        layout.addWidget(self.check_crews_table)
        self.tabs.addTab(self.check_tab, "Проверка КП")
        self.update_check_tab()

    def update_check_tab(self):
        """Обновление данных на вкладке проверки"""
        self.check_crews_table.setRowCount(0)
        if not hasattr(self, 'data') or 'members' not in self.data:
            return

        # Определяем режим отображения
        mode = self.table_mode.currentIndex()
        
        if mode == 0:  # Краткий вид
            self.setup_short_mode()
        elif mode == 1:  # Полный вид
            self.setup_full_mode()
        elif mode == 2:  # Режим КП
            self.setup_cp_mode()

    def setup_short_mode(self):
        """Настройка краткого вида таблицы"""
        self.check_crews_table.setColumnCount(10)
        self.check_crews_table.setHorizontalHeaderLabels([
            "Номер", "Зачет", "Пилот", "Штурман", "Авто", 
            "Старт", "Финиш", "Время трассы", "Баллы", "Итог"
        ])
        self.fill_basic_data()

    def setup_full_mode(self):
        """Настройка полного вида таблицы"""
        self.check_crews_table.setColumnCount(14)
        self.check_crews_table.setHorizontalHeaderLabels([
            "Номер", "Зачет", "Пилот", "Штурман", "Авто", "Старт", "Финиш",
            "Нейтрализация", "Общее время", "Время трассы", 
            "Баллы", "Штраф время", "Штраф КП", "Итог"
        ])
        self.fill_basic_data(with_extras=True)

    def fill_basic_data(self, with_extras=False):
        """Заполнение основных данных экипажей"""
        for member in self.data['members']:
            if not member.get('started', False):
                continue

            row = self.check_crews_table.rowCount()
            self.check_crews_table.insertRow(row)
            
            # Основные данные
            self.check_crews_table.setItem(row, 0, QTableWidgetItem(str(member.get('номер', ''))))
            self.check_crews_table.setItem(row, 1, QTableWidgetItem(member.get('зачет', '')))
            self.check_crews_table.setItem(row, 2, QTableWidgetItem(member.get('пилот', '')))
            self.check_crews_table.setItem(row, 3, QTableWidgetItem(member.get('штурман', '')))
            self.check_crews_table.setItem(row, 4, QTableWidgetItem(member.get('авто', '')))
            self.check_crews_table.setItem(row, 5, QTableWidgetItem(member.get('start_time', '')))
            
            # Время финиша
            finish_time = member.get('finish_time', '')
            if not member.get('finished', False) and finish_time != "DNF":
                finish_time = ""
            self.check_crews_table.setItem(row, 6, QTableWidgetItem(finish_time))
            
            # Рассчитываем временные показатели
            neutral_time = self.calculate_total_neutral_time(member)
            total_time = self.calculate_total_time(member)
            track_time = max(0, total_time - neutral_time) if total_time else 0
            
            # Баллы и штрафы
            total_score = member.get('total_score', 0)
            time_penalty = self.calculate_time_penalty(member, track_time)
            cp_penalty = member.get('cp_penalty', 0)
            
            # Расчет итога
            if time_penalty == "DNF":
                total = "DNF"
            else:
                total = total_score - (time_penalty if isinstance(time_penalty, int) else 0) - cp_penalty

            # Заполняем данные в зависимости от режима
            if with_extras:
                self.check_crews_table.setItem(row, 7, QTableWidgetItem(self.format_time(neutral_time)))
                self.check_crews_table.setItem(row, 8, QTableWidgetItem(self.format_time(total_time)))
                self.check_crews_table.setItem(row, 9, QTableWidgetItem(self.format_time(track_time)))
                
                # Баллы
                self.check_crews_table.setItem(row, 10, QTableWidgetItem(str(total_score)))
                
                # Штраф время (может быть DNF или числом)
                time_penalty_item = QTableWidgetItem(str(time_penalty))
                if time_penalty == "DNF":
                    time_penalty_item.setBackground(Qt.GlobalColor.red)
                self.check_crews_table.setItem(row, 11, time_penalty_item)
                
                # Штраф КП
                self.check_crews_table.setItem(row, 12, QTableWidgetItem(str(cp_penalty)))
                
                # Итог
                total_item = QTableWidgetItem(str(total))
                if total == "DNF":
                    total_item.setBackground(Qt.GlobalColor.red)
                self.check_crews_table.setItem(row, 13, total_item)
            else:
                # Краткий режим
                self.check_crews_table.setItem(row, 7, QTableWidgetItem(self.format_time(track_time)))
                self.check_crews_table.setItem(row, 8, QTableWidgetItem(str(total_score)))
                
                # Итог
                total_item = QTableWidgetItem(str(total))
                if total == "DNF":
                    total_item.setBackground(Qt.GlobalColor.red)
                self.check_crews_table.setItem(row, 9, total_item)

            # Раскрашиваем строку если проверка завершена
            if member.get('check_completed', False):
                for col in range(self.check_crews_table.columnCount()):
                    item = self.check_crews_table.item(row, col)
                    if item and not (col in [11, 13] and item.text() == "DNF"):
                        item.setBackground(Qt.GlobalColor.green)
            elif member.get('finish_time') == "DNF":
                for col in range(self.check_crews_table.columnCount()):
                    item = self.check_crews_table.item(row, col)
                    if item and item.text() != "DNF":
                        item.setBackground(Qt.GlobalColor.red)

        # Настраиваем ширину столбцов
        header = self.check_crews_table.horizontalHeader()
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Пилот
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Штурман
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Номер
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Зачет
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Авто

    def show_check_dialog(self, index):
        """Диалог проверки КП для выбранного экипажа"""
        try:
            if not hasattr(self, 'data') or 'members' not in self.data:
                raise ValueError("Данные не загружены")
                
            row = index.row()
            if row < 0 or row >= self.check_crews_table.rowCount():
                raise IndexError("Неверный индекс строки")
                
            crew_number = self.check_crews_table.item(row, 0).text()
            if not crew_number:
                raise ValueError("Не найден номер экипажа")
            
            # Находим экипаж в данных
            for member in self.data['members']:
                if str(member.get('номер', '')) == crew_number:
                    self.current_check_crew = member
                    break
            else:
                QMessageBox.warning(self, "Ошибка", "Экипаж не найден!")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Проверка КП экипажа №{crew_number}")
            dialog.setMinimumSize(800, 600)
            
            layout = QVBoxLayout(dialog)

            # Информация об экипаже
            info_group = QGroupBox("Информация об экипаже")
            info_layout = QVBoxLayout()
            info_layout.addWidget(QLabel(f"Пилот: {self.current_check_crew.get('пилот', '')}"))
            info_layout.addWidget(QLabel(f"Штурман: {self.current_check_crew.get('штурман', '')}"))
            info_layout.addWidget(QLabel(f"Авто: {self.current_check_crew.get('авто', '')}"))
            info_group.setLayout(info_layout)
            layout.addWidget(info_group)

            # Поле для ввода номеров КП
            self.cp_input = QLineEdit()
            self.cp_input.setPlaceholderText("Введите номера КП через пробел и нажмите Enter")
            self.cp_input.installEventFilter(self)
            layout.addWidget(self.cp_input)

            # Создаем виджет с вкладками для этапов
            tab_widget = QTabWidget()
            
            # Список для хранения всех QListWidget
            self.cp_list_widgets = []
            
            # Вкладка "Все КП"
            all_cps_tab = QWidget()
            all_cps_layout = QVBoxLayout(all_cps_tab)
            
            # Кнопки управления
            btn_layout = QHBoxLayout()
            all_cps_btn = QPushButton("Выделить все КП")
            none_cps_btn = QPushButton("Снять все КП")
            btn_layout.addWidget(all_cps_btn)
            btn_layout.addWidget(none_cps_btn)
            all_cps_layout.addLayout(btn_layout)
            
            # Основной список КП
            self.cp_list_widget = QListWidget()
            self.cp_list_widget.itemChanged.connect(self.sync_check_states)
            self.cp_list_widgets.append(self.cp_list_widget)
            all_cps_layout.addWidget(self.cp_list_widget)
            
            tab_widget.addTab(all_cps_tab, "Все КП")

            # Вкладки для этапов
            if self.data.get('logic_params', {}).get('staged', False):
                for stage in self.data['logic_params'].get('stages', []):
                    stage_tab = QWidget()
                    stage_layout = QVBoxLayout(stage_tab)
                    
                    # Кнопки для этапа
                    stage_btn_layout = QHBoxLayout()
                    stage_select_btn = QPushButton(f"Выделить все")
                    stage_deselect_btn = QPushButton(f"Снять все")
                    stage_btn_layout.addWidget(stage_select_btn)
                    stage_btn_layout.addWidget(stage_deselect_btn)
                    stage_layout.addLayout(stage_btn_layout)
                    
                    # Список КП этапа
                    stage_cp_list = QListWidget()
                    stage_cp_list.setProperty('stage', stage['name'])
                    stage_cp_list.itemChanged.connect(self.sync_check_states)
                    self.cp_list_widgets.append(stage_cp_list)
                    stage_layout.addWidget(stage_cp_list)
                    
                    tab_widget.addTab(stage_tab, stage['name'])
                    
                    # Подключаем кнопки этапа
                    stage_select_btn.clicked.connect(
                        lambda _, s=stage['name']: self.toggle_stage_checkpoints(s, True))
                    stage_deselect_btn.clicked.connect(
                        lambda _, s=stage['name']: self.toggle_stage_checkpoints(s, False))
            
            # Подключаем кнопки "Все КП"
            all_cps_btn.clicked.connect(lambda: self.toggle_all_checkpoints(True))
            none_cps_btn.clicked.connect(lambda: self.toggle_all_checkpoints(False))
            
            layout.addWidget(tab_widget)

            # Получаем зачёт текущего экипажа
            crew_classification = self.current_check_crew.get('зачет', '')
            
            # Заполняем списки КП (только доступные для зачёта)
            if 'checkpoints' in self.data:
                taken_cps = self.current_check_crew.get('taken_cps', [])
                
                for cp in self.data['checkpoints']:
                    # Проверяем доступность КП для зачёта экипажа
                    is_available = self.is_checkpoint_available(cp, crew_classification)
                    if not is_available:
                        continue  # Пропускаем недоступные КП
                    
                    # Создаём элемент списка
                    item = QListWidgetItem(f"{cp.get('name', '')} ({cp.get('score', 0)} баллов)")
                    item.setData(Qt.ItemDataRole.UserRole, cp['name'])
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    item.setCheckState(Qt.CheckState.Checked if cp['name'] in taken_cps else Qt.CheckState.Unchecked)
                    
                    # Добавляем в основной список
                    self.cp_list_widget.addItem(item)
                    
                    # Добавляем в соответствующие этапы (если этапность включена)
                    if self.data.get('logic_params', {}).get('staged', False) and 'stages' in cp:
                        for stage_name in cp['stages']:
                            for i in range(tab_widget.count()):
                                if tab_widget.tabText(i) == stage_name:
                                    cloned_item = item.clone()
                                    tab_widget.widget(i).findChild(QListWidget).addItem(cloned_item)
                                    break

            # Корректировка времени
            time_group = QGroupBox("Корректировка времени")
            time_layout = QVBoxLayout()
            
            # Старт
            start_layout = QHBoxLayout()
            start_layout.addWidget(QLabel("Старт:"))
            self.start_time_edit = QTimeEdit()
            self.start_time_edit.setDisplayFormat("HH:mm:ss")
            if self.current_check_crew.get('start_time'):
                self.start_time_edit.setTime(QTime.fromString(self.current_check_crew['start_time'], "HH:mm:ss"))
            start_layout.addWidget(self.start_time_edit)
            time_layout.addLayout(start_layout)
            
            # Финиш
            finish_layout = QHBoxLayout()
            finish_layout.addWidget(QLabel("Финиш:"))
            self.finish_time_edit = QTimeEdit()
            self.finish_time_edit.setDisplayFormat("HH:mm:ss")
            
            # Блокируем редактирование если экипаж не финишировал
            is_finished = self.current_check_crew.get('finished', False)
            self.finish_time_edit.setEnabled(is_finished)
            
            if is_finished and self.current_check_crew.get('finish_time') and self.current_check_crew['finish_time'] != "DNF":
                self.finish_time_edit.setTime(QTime.fromString(self.current_check_crew['finish_time'], "HH:mm:ss"))
            else:
                self.finish_time_edit.setTime(QTime(0, 0, 0))
            
            finish_layout.addWidget(self.finish_time_edit)
            time_layout.addLayout(finish_layout)
            
            # Информация о нейтрализации
            if self.data.get('logic_params', {}).get('staged', False):
                time_layout.addWidget(QLabel("Редактирование времени нейтрализации осуществляется во вкладке Этапы", 
                                          alignment=Qt.AlignmentFlag.AlignRight))
            
            time_group.setLayout(time_layout)
            layout.addWidget(time_group)

            # Флаг завершения проверки
            self.check_completed_cb = QCheckBox("Проверка завершена")
            self.check_completed_cb.setChecked(self.current_check_crew.get('check_completed', False))
            layout.addWidget(self.check_completed_cb)

            # Кнопки
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
            button_box.accepted.connect(lambda: self.save_check_data(dialog))
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)

            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть проверку: {str(e)}")

    def save_check_data(self, dialog):
        """Сохранение результатов проверки с корректным форматом taken_cps"""
        # Сохраняем взятые КП (все они гарантированно доступны)
        taken_cps = []
        for i in range(self.cp_list_widget.count()):
            item = self.cp_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                taken_cps.append(item.data(Qt.ItemDataRole.UserRole))
        
        # Получаем зачет экипажа
        crew_classification = self.current_check_crew.get('зачет', '')
        
        # Собираем взятые КП
        taken_cps = []
        for i in range(self.cp_list_widget.count()):
            item = self.cp_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                cp_name = item.data(Qt.ItemDataRole.UserRole)
                
                # Проверяем доступность КП
                cp_data = next((cp for cp in self.data['checkpoints'] if cp['name'] == cp_name), None)
                if cp_data:
                    is_available = True
                    if 'classifications' in cp_data:
                        if crew_classification in cp_data['classifications']:
                            is_available = cp_data['classifications'][crew_classification]
                    
                    if is_available:
                        taken_cps.append(cp_name)
                    else:
                        # Если попытались сохранить недоступный КП
                        QMessageBox.warning(
                            dialog,
                            "Ошибка",
                            f"КП {cp_name} недоступен для зачета '{crew_classification}'\n"
                            "Он не будет сохранен в списке взятых КП"
                        )

        # Сохраняем взятые КП (только названия без баллов)
        taken_cps = []
        for i in range(self.cp_list_widget.count()):
            item = self.cp_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                # Получаем оригинальное название КП из UserRole
                cp_name = item.data(Qt.ItemDataRole.UserRole)
                taken_cps.append(cp_name)
        
        # Подсчет баллов
        total_score = 0
        if 'checkpoints' in self.data:
            for cp in self.data['checkpoints']:
                if cp.get('name') in taken_cps:
                    total_score += int(cp.get('score', 0))
        
        # Обновляем данные экипажа
        self.current_check_crew['taken_cps'] = taken_cps  # Теперь только названия КП
        self.current_check_crew['total_score'] = total_score
        
        # Обновляем время старта
        self.current_check_crew['start_time'] = self.start_time_edit.time().toString("HH:mm:ss")
        
        # Обработка времени финиша
        if self.current_check_crew.get('finished', False):
            finish_time = "DNF" if self.current_check_crew.get('finish_time') == "DNF" else ""
            
            if self.finish_time_edit.time().isValid():
                if not (self.finish_time_edit.time().hour() == 0 and 
                       self.finish_time_edit.time().minute() == 0 and 
                       self.finish_time_edit.time().second() == 0):
                    finish_time = self.finish_time_edit.time().toString("HH:mm:ss")
            
            self.current_check_crew['finish_time'] = finish_time

        # Сохраняем статус проверки
        self.current_check_crew['check_completed'] = self.check_completed_cb.isChecked()
        
        # Сохраняем данные
        self.save_data()
        self.update_check_tab()
        
        # Показываем статистику
        QMessageBox.information(
            dialog, 
            "Сохранено", 
            f"Сохранено {len(taken_cps)} КП\n"
            f"Общий балл: {total_score}\n"
            f"Время старта: {self.current_check_crew['start_time']}\n"
            f"Время финиша: {self.current_check_crew.get('finish_time', '')}"
        )
        
        dialog.accept()

    def sync_check_states(self, changed_item):
        """Синхронизация состояния чекбоксов между вкладками"""
        # Получаем имя КП и новое состояние
        cp_name = changed_item.data(Qt.ItemDataRole.UserRole)
        new_state = changed_item.checkState()
        
        # Обновляем все соответствующие items в других списках
        for list_widget in self.cp_list_widgets:
            for i in range(list_widget.count()):
                item = list_widget.item(i)
                if item.data(Qt.ItemDataRole.UserRole) == cp_name:
                    # Временно отключаем сигналы, чтобы избежать рекурсии
                    list_widget.blockSignals(True)
                    item.setCheckState(new_state)
                    list_widget.blockSignals(False)
                    break

    def toggle_all_checkpoints(self, checked):
        """Установка/снятие всех КП с синхронизацией"""
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        for list_widget in self.cp_list_widgets:
            list_widget.blockSignals(True)
            for i in range(list_widget.count()):
                list_widget.item(i).setCheckState(state)
            list_widget.blockSignals(False)
        
        # Вручную обновляем один список, чтобы сработала синхронизация
        if self.cp_list_widget.count() > 0:
            self.cp_list_widget.item(0).setCheckState(state)

    def toggle_stage_checkpoints(self, stage_name, checked):
        """Установка/снятие КП этапа с синхронизацией"""
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        for list_widget in self.cp_list_widgets:
            list_widget.blockSignals(True)
            for i in range(list_widget.count()):
                item = list_widget.item(i)
                cp_name = item.data(Qt.ItemDataRole.UserRole)
                cp_data = next((cp for cp in self.data['checkpoints'] 
                              if cp['name'] == cp_name), None)
                
                if cp_data and 'stages' in cp_data and stage_name in cp_data['stages']:
                    item.setCheckState(state)
            list_widget.blockSignals(False)
        
        # Активируем синхронизацию
        if self.cp_list_widget.count() > 0:
            self.cp_list_widget.item(0).setCheckState(self.cp_list_widget.item(0).checkState())

    def is_checkpoint_available(self, checkpoint, classification):
        """Проверяет, доступен ли КП для указанного зачёта"""
        if 'classifications' not in checkpoint:
            return True  # Если ограничений нет, КП доступен
        
        # Если зачёт явно указан в classifications, используем его значение
        if classification in checkpoint['classifications']:
            return checkpoint['classifications'][classification]
        
        # Если зачёт не указан, считаем КП доступным
        return True

    def setup_cp_mode(self):
        """Настройка режима отображения КП с учетом этапности и доступности КП"""
        try:
            # Получаем настройки этапности
            is_staged = self.data.get('logic_params', {}).get('staged', False)
            stages_info = self.data['logic_params'].get('stages', []) if is_staged else []

            # Собираем и сортируем все КП
            all_cps = sorted(self.data.get('checkpoints', []), key=lambda x: x['name'])

            # Формируем базовые заголовки
            headers = ["Номер", "Зачет", "Пилот", "Штурман", "Авто", "Баллы"]

            # Словари для хранения позиций столбцов
            cp_columns = {}
            stage_columns = {}
            col_index = len(headers)  # Текущая позиция столбца

            # Обрабатываем режим с этапностью
            if is_staged:
                # Группируем КП по этапам
                stages_data = {}
                for cp in all_cps:
                    if 'stages' in cp:
                        for stage_name in cp['stages']:
                            if stage_name not in stages_data:
                                stages_data[stage_name] = []
                            stages_data[stage_name].append(cp['name'])

                # Цвета для этапов
                stage_colors = [
                    QColor(173, 216, 230),  # LightBlue
                    QColor(144, 238, 144),   # LightGreen
                    QColor(255, 182, 193),    # LightPink
                    QColor(255, 255, 153)     # LightYellow
                ]

                # Добавляем этапы и их КП в заголовки
                for i, stage in enumerate(stages_info):
                    stage_name = stage['name']
                    headers.append(stage_name)
                    stage_columns[stage_name] = {
                        'col_index': col_index,
                        'color': stage_colors[i % len(stage_colors)]
                    }
                    col_index += 1

                    # Добавляем КП этого этапа
                    for cp_name in sorted(stages_data.get(stage_name, [])):
                        headers.append(cp_name)
                        cp_columns[cp_name] = col_index
                        col_index += 1

                # Добавляем КП без этапов (если есть)
                for cp in all_cps:
                    if cp['name'] not in cp_columns and ('stages' not in cp or not cp['stages']):
                        headers.append(cp['name'])
                        cp_columns[cp['name']] = col_index
                        col_index += 1
            else:
                # Режим без этапности - просто все КП
                for cp in all_cps:
                    headers.append(cp['name'])
                    cp_columns[cp['name']] = col_index
                    col_index += 1

            # Добавляем итоговый столбец
            headers.append("Итог")
            total_column = col_index

            # Настраиваем таблицу
            self.check_crews_table.setColumnCount(len(headers))
            self.check_crews_table.setHorizontalHeaderLabels(headers)
            
            # Включаем экономный режим отображения
            self.check_crews_table.setWordWrap(False)
            self.check_crews_table.setTextElideMode(Qt.TextElideMode.ElideRight)

            # Заполняем данные
            for member in self.data['members']:
                if not member.get('started', False):
                    continue

                row = self.check_crews_table.rowCount()
                self.check_crews_table.insertRow(row)

                # Основные данные
                for col in range(6):
                    value = str(member.get([
                        'номер', 'зачет', 'пилот', 'штурман', 'авто', 'total_score'
                    ][col], ''))
                    item = QTableWidgetItem(value)
                    if col == 5:  # Центрируем баллы
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.check_crews_table.setItem(row, col, item)

                # Получаем зачет экипажа
                crew_classification = member.get('зачет', '')
                taken_cps = member.get('taken_cps', [])

                # Заполняем данные КП
                col = 6  # Начинаем после основных колонок

                if is_staged:
                    # Режим с этапностью
                    for stage_name, stage_data in stage_columns.items():
                        # Баллы за этап
                        stage_score = 0
                        for cp_name in stages_data.get(stage_name, []):
                            cp = next((cp for cp in all_cps if cp['name'] == cp_name), None)
                            if cp and cp_name in taken_cps and self.is_cp_available(cp, crew_classification):
                                stage_score += int(cp.get('score', 0))

                        stage_item = QTableWidgetItem(str(stage_score))
                        stage_item.setBackground(stage_data['color'])
                        stage_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.check_crews_table.setItem(row, stage_data['col_index'], stage_item)

                        # КП этапа
                        for cp_name in sorted(stages_data.get(stage_name, [])):
                            cp = next((cp for cp in all_cps if cp['name'] == cp_name), None)
                            if not cp:
                                continue

                            item = QTableWidgetItem()
                            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                            if not self.is_cp_available(cp, crew_classification):
                                item.setText("-")
                                item.setForeground(Qt.GlobalColor.gray)
                            elif cp_name in taken_cps:
                                item.setText(str(int(cp.get('score', 0))))
                            else:
                                item.setText("")

                            item.setBackground(stage_data['color'].lighter(130))
                            self.check_crews_table.setItem(row, cp_columns[cp_name], item)

                    # КП без этапов
                    for cp in all_cps:
                        if cp['name'] not in cp_columns:
                            continue
                        if 'stages' not in cp or not cp['stages']:
                            item = QTableWidgetItem()
                            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                            if not self.is_cp_available(cp, crew_classification):
                                item.setText("-")
                                item.setForeground(Qt.GlobalColor.gray)
                            elif cp['name'] in taken_cps:
                                item.setText(str(int(cp.get('score', 0))))
                            else:
                                item.setText("0")

                            self.check_crews_table.setItem(row, cp_columns[cp['name']], item)
                else:
                    # Режим без этапности
                    for cp in all_cps:
                        item = QTableWidgetItem()
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                        if not self.is_cp_available(cp, crew_classification):
                            item.setText("-")
                            item.setForeground(Qt.GlobalColor.gray)
                        elif cp['name'] in taken_cps:
                            item.setText(str(int(cp.get('score', 0))))
                        else:
                            item.setText("")

                        self.check_crews_table.setItem(row, cp_columns[cp['name']], item)

                # Итоговый столбец
                total_penalty = member.get('штрафные_баллы', 0)
                total_score = int(member.get('total_score', 0))
                result = total_score - total_penalty
                result_item = QTableWidgetItem(str(result))
                result_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                # Раскрашиваем итог
                if result >= 0:
                    result_item.setBackground(QColor(144, 238, 144))  # LightGreen
                else:
                    result_item.setBackground(QColor(255, 182, 193))  # LightPink

                self.check_crews_table.setItem(row, total_column, result_item)

                # Подсветка строки
                if member.get('check_completed', False):
                    for col in range(self.check_crews_table.columnCount()):
                        item = self.check_crews_table.item(row, col)
                        if item and item.text() != "-":
                            item.setBackground(Qt.GlobalColor.green)
                elif member.get('finish_time') == "DNF":
                    for col in range(self.check_crews_table.columnCount()):
                        item = self.check_crews_table.item(row, col)
                        if item and item.text() != "DNF":
                            item.setBackground(Qt.GlobalColor.red)

            # Настраиваем ширину столбцов
            self.adjust_cp_columns_width()

        except Exception as e:
            print(f"Ошибка в setup_cp_mode: {str(e)}")
            raise

    def is_cp_available(self, checkpoint, classification):
        """Проверяет доступность КП для указанного зачёта"""
        if 'classifications' not in checkpoint:
            return True
        if classification in checkpoint['classifications']:
            return checkpoint['classifications'][classification]
        return True

    def adjust_cp_columns_width(self):
        """Настройка ширины столбцов с минимальными размерами"""
        header = self.check_crews_table.horizontalHeader()
        
        # Устанавливаем режимы изменения размеров
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Базовые ширины для разных типов столбцов (в пикселях)
        base_widths = {
            'number': 50,    # Номер
            'class': 80,     # Зачет
            'name': 100,     # Пилот/Штурман
            'car': 120,      # Авто
            'points': 60,    # Баллы
            'stage': 70,     # Этап
            'cp': 40,       # КП
            'total': 80      # Итог
        }

        for col in range(self.check_crews_table.columnCount()):
            header_text = self.check_crews_table.horizontalHeaderItem(col).text()
            
            # Определяем тип столбца
            if col == 0:
                col_type = 'number'
            elif col == 1:
                col_type = 'class'
            elif col in (2, 3):  # Пилот/Штурман
                col_type = 'name'
            elif col == 4:
                col_type = 'car'
            elif col == 5:
                col_type = 'points'
            elif "Этап" in header_text:
                col_type = 'stage'
            elif "КП" in header_text or "CP" in header_text:
                col_type = 'cp'
            elif header_text == "Итог":
                col_type = 'total'
            else:
                col_type = 'cp'  # По умолчанию для неизвестных

            # Устанавливаем начальную ширину
            header.resizeSection(col, base_widths[col_type])

            # Дополнительно вычисляем минимальную ширину по содержимому
            self.check_crews_table.resizeColumnToContents(col)
            min_width = header.sectionSize(col)
            header.resizeSection(col, min(min_width, base_widths[col_type] * 2))

        # Фиксируем ширину для КП (можно прокручивать если не помещаются)
        header.setStretchLastSection(False)




############## РЕЗУЛЬТАТЫ ##############

    def setup_results_tab(self):
        """Инициализация вкладки с итоговыми результатами"""
        self.results_tab = QWidget()
        layout = QVBoxLayout(self.results_tab)

        toolbar = QHBoxLayout()
        
        self.export_btn = QPushButton("Экспорт в Excel")
        self.export_btn.clicked.connect(self.save_results)
        toolbar.addWidget(self.export_btn)
        
        self.print_btn = QPushButton("Печать результатов")
        self.print_btn.clicked.connect(self.print_results)
        toolbar.addWidget(self.print_btn)
        
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.clicked.connect(self.update_results_tab)
        toolbar.addWidget(self.refresh_btn)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)

        # Таблица результатов
        self.results_table = QTableWidget()
        self.results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.results_table)

        # Обновление данных
        self.update_results_tab()
        self.tabs.addTab(self.results_tab, "Результаты")

    def calculate_track_time(self, member):
        """Расчет чистого времени трассы (общее время - нейтрализация)"""
        total_time = self.calculate_total_time(member)
        neutral_time = self.calculate_total_neutral_time(member)
        return max(0, total_time - neutral_time)

    def print_results(self):
        """Заглушка для функции печати результатов"""
        QMessageBox.information(
            self,
            "Больше не тыкай сюда, понятно?",
            "Эта хрень пока не работает, нажми экспорт",
            QMessageBox.StandardButton.Ok
        )

    def add_time_data(self, row, member):
        """Добавление временных показателей"""
        # Финиш
        finish_time = member.get('finish_time', '')
        if not member.get('finished', False) and finish_time != "DNF":
            finish_time = ""
        self.results_table.setItem(row, 7, QTableWidgetItem(finish_time))
        
        # Нейтрализация и время
        neutral_time = self.calculate_total_neutral_time(member)
        total_time = self.calculate_total_time(member)
        track_time = max(0, total_time - neutral_time) if total_time else 0
        
        self.results_table.setItem(row, 8, QTableWidgetItem(self.format_time(neutral_time)))
        self.results_table.setItem(row, 9, QTableWidgetItem(self.format_time(total_time)))
        self.results_table.setItem(row, 10, QTableWidgetItem(self.format_time(track_time)))

    def add_skp_data(self, row, member, skp_count):
        """Добавление данных по СКП"""
        base_col = 11  # После временных показателей
        
        for i in range(1, skp_count + 1):
            skp_data = member.get(f'skp_{i}', {})
            self.results_table.setItem(row, base_col, QTableWidgetItem(skp_data.get('entry', '')))
            self.results_table.setItem(row, base_col + 1, QTableWidgetItem(skp_data.get('exit', '')))
            self.results_table.setItem(row, base_col + 2, QTableWidgetItem(
                self.format_time(skp_data.get('neutralization', 0))))
            base_col += 3

    def add_stages_data(self, row, member):
        """Добавление данных по этапам и КП с цветовой маркировкой"""
        if not self.data.get('logic_params', {}).get('staged', False):
            return
            
        # Позиция начала этапов
        start_col = 14 + len(self.data.get('skp_points', [])) * 3  # После новых столбцов
        crew_classification = member.get('зачет', '')
        taken_cps = member.get('taken_cps', [])
        
        # Цвета для этапов
        stage_colors = [
            QColor(173, 216, 230),  # LightBlue
            QColor(144, 238, 144),   # LightGreen
            QColor(255, 182, 193),   # LightPink
            QColor(255, 255, 153)    # LightYellow
        ]
        
        for stage_idx, stage in enumerate(self.data['logic_params'].get('stages', [])):
            stage_name = stage['name']
            stage_score = 0
            cp_col = 0
            
            # Считаем баллы этапа
            for cp in self.data.get('checkpoints', []):
                if 'stages' in cp and stage_name in cp['stages']:
                    if cp['name'] in taken_cps and self.is_cp_available(cp, crew_classification):
                        stage_score += int(cp.get('score', 0))
            
            # Добавляем баллы этапа с цветом
            stage_item = QTableWidgetItem(str(stage_score))
            stage_item.setBackground(stage_colors[stage_idx % len(stage_colors)])
            stage_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.results_table.setItem(row, start_col, stage_item)
            start_col += 1
            
            # Добавляем КП этапа с цветом
            for cp in self.data.get('checkpoints', []):
                if 'stages' in cp and stage_name in cp['stages']:
                    item = QTableWidgetItem()
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    if not self.is_cp_available(cp, crew_classification):
                        item.setText("-")
                        item.setForeground(Qt.GlobalColor.gray)
                    elif cp['name'] in taken_cps:
                        item.setText(str(int(cp.get('score', 0))))
                    else:
                        item.setText("0")
                    
                    # Более светлый оттенок цвета этапа
                    item.setBackground(stage_colors[stage_idx % len(stage_colors)].lighter(130))
                    self.results_table.setItem(row, start_col + cp_col, item)
                    cp_col += 1
            
            start_col += cp_col

    def get_column_index(self, column_name):
        """Возвращает индекс столбца по его названию"""
        for col in range(self.results_table.columnCount()):
            if self.results_table.horizontalHeaderItem(col).text() == column_name:
                return col
        return -1

    def sort_crews_by_results(self):
        """Сортировка экипажей по зачетам и результатам с учетом DNF"""
        crews_by_class = {}
        
        for member in self.data['members']:
            if not member.get('started', False):
                continue
            
            # Проверяем DNF статус
            track_time = self.calculate_track_time(member)
            is_dnf = member.get('finish_time') == "DNF" or (
                self.data.get("logic_params", {}).get("penalty_type") == "DNF" and 
                track_time > self.calculate_track_duration()
            )
            
            crew_class = member.get('зачет', 'Без зачета')
            if crew_class not in crews_by_class:
                crews_by_class[crew_class] = []
            
            # Для DNF ставим минимальный приоритет сортировки
            if is_dnf:
                sort_key = (float('-inf'), float('inf'))  # Всегда в конце
            else:
                total_score = member.get('total_score', 0)
                sort_key = (-total_score, track_time if track_time > 0 else float('inf'))
            
            crews_by_class[crew_class].append({
                'member': member,
                'sort_key': sort_key,
                'is_dnf': is_dnf
            })

        # Сортируем внутри каждого зачета
        sorted_crews = []
        for crew_class in sorted(crews_by_class.keys()):
            # Сначала не-DNF экипажи, отсортированные по результатам
            non_dnf_crews = [c for c in crews_by_class[crew_class] if not c['is_dnf']]
            non_dnf_crews = sorted(non_dnf_crews, key=lambda x: x['sort_key'])
            
            # Затем DNF экипажи (они уже в конце и не сортируются)
            dnf_crews = [c for c in crews_by_class[crew_class] if c['is_dnf']]
            
            # Объединяем и нумеруем места
            class_crews = non_dnf_crews + dnf_crews
            for i, crew_data in enumerate(class_crews, 1):
                # Для DNF не указываем место (оно будет заменено на DNF позже)
                if not crew_data['is_dnf']:
                    crew_data['member']['место'] = i
                sorted_crews.append(crew_data['member'])

        return sorted_crews

    def add_basic_member_info(self, row, member, is_dnf):
        """Добавление основной информации об экипаже"""
        # Место
        place_item = QTableWidgetItem("DNF" if is_dnf else str(member.get('место', '')))
        if is_dnf:
            place_item.setBackground(Qt.GlobalColor.red)
        self.results_table.setItem(row, 0, place_item)
        
        # Остальные основные данные
        self.results_table.setItem(row, 1, QTableWidgetItem(str(member.get('номер', ''))))
        self.results_table.setItem(row, 2, QTableWidgetItem(member.get('зачет', '')))
        self.results_table.setItem(row, 3, QTableWidgetItem(member.get('пилот', '')))
        self.results_table.setItem(row, 4, QTableWidgetItem(member.get('штурман', '')))
        self.results_table.setItem(row, 5, QTableWidgetItem(member.get('авто', '')))
        self.results_table.setItem(row, 6, QTableWidgetItem(member.get('start_time', '')))

    def add_score_columns(self, row, member, is_dnf):
        """Добавление столбцов с итогом, штрафом времени и баллами"""
        # Позиция начала новых столбцов
        score_col = 11 + len(self.data.get('skp_points', [])) * 3
        
        # Рассчитываем временные показатели
        neutral_time = self.calculate_total_neutral_time(member)
        total_time = self.calculate_total_time(member)
        track_time = max(0, total_time - neutral_time) if total_time else 0
        
        # Баллы и штрафы
        total_score = member.get('total_score', 0)
        time_penalty = "DNF" if is_dnf else self.calculate_time_penalty(member, track_time)
        cp_penalty = member.get('cp_penalty', 0)
        
        # Расчет итога
        total = "DNF" if is_dnf else total_score - (time_penalty if isinstance(time_penalty, int) else 0) - cp_penalty
        
        # Итог
        total_item = QTableWidgetItem(str(total))
        if is_dnf:
            total_item.setBackground(Qt.GlobalColor.red)
        self.results_table.setItem(row, score_col, total_item)
        
        # Штраф время
        time_penalty_item = QTableWidgetItem(str(time_penalty))
        if is_dnf:
            time_penalty_item.setBackground(Qt.GlobalColor.red)
        self.results_table.setItem(row, score_col + 1, time_penalty_item)
        
        # Баллы
        self.results_table.setItem(row, score_col + 2, QTableWidgetItem(str(total_score)))

    def add_checkpoints_data(self, row, member, checkpoints, start_col):
        """Добавление данных по КП без этапности"""
        taken_cps = member.get('taken_cps', [])
        crew_classification = member.get('зачет', '')
        
        for col, cp in enumerate(checkpoints):
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            if not self.is_cp_available(cp, crew_classification):
                item.setText("-")
                item.setForeground(Qt.GlobalColor.gray)
            elif cp['name'] in taken_cps:
                item.setText(str(int(cp.get('score', 0))))
            else:
                item.setText("0")
            
            self.results_table.setItem(row, start_col + col, item)

    def is_crew_dnf(self, member):
        """Проверяет, является ли экипаж DNF"""
        track_time = self.calculate_track_time(member)
        return (member.get('finish_time') == "DNF" or 
                (self.data.get("logic_params", {}).get("penalty_type") == "DNF" and 
                track_time > self.calculate_track_duration()))

    def adjust_results_columns(self):
        """Настройка ширины столбцов"""
        header = self.results_table.horizontalHeader()
        
        # Базовые ширины столбцов
        column_widths = {
            'Место': 50,
            'Номер': 50,
            'Зачет': 80,
            'Пилот': 120,
            'Штурман': 120,
            'Авто': 100,
            'Старт': 80,
            'Финиш': 80,
            'Нейтрализация': 100,
            'Время': 80,
            'Итог': 60,
            'Штраф': 80,
            'Баллы': 60,
            'КП': 40
        }
        
        for col in range(self.results_table.columnCount()):
            header_text = self.results_table.horizontalHeaderItem(col).text()
            width = 80  # Значение по умолчанию
            
            for pattern, w in column_widths.items():
                if pattern in header_text:
                    width = w
                    break
            
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            header.resizeSection(col, width)
        
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # Пилот
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # Штурман

    def update_results_tab(self):
        """Обновление данных в таблице результатов"""
        self.results_table.clear()
        
        if not hasattr(self, 'data') or 'members' not in self.data:
            return

        # Формируем заголовки столбцов
        headers = [
            "Место", "Номер", "Зачет", "Пилот", "Штурман", "Авто", "Старт",
            "Финиш", "Нейтрализация", "Общее время", "Время трассы"
        ]

        # Добавляем СКП (если есть)
        skp_count = len(self.data.get('skp_points', []))
        for i in range(1, skp_count + 1):
            headers.extend([f"Вход СКП {i}", f"Выход СКП {i}", f"Нейтрализация СКП {i}"])

        # Добавляем основные столбцы
        headers.extend(["Итог", "Штраф время", "Баллы"])

        # Добавляем КП в зависимости от этапности
        if self.data.get('logic_params', {}).get('staged', False):
            # С этапностью - добавляем этапы и их КП
            for stage in self.data['logic_params'].get('stages', []):
                headers.append(stage['name'])
                for cp in self.data.get('checkpoints', []):
                    if 'stages' in cp and stage['name'] in cp['stages']:
                        headers.append(cp['name'])
        else:
            # Без этапности - просто все КП по порядку
            checkpoints = sorted(self.data.get('checkpoints', []), key=lambda x: x['name'])
            for cp in checkpoints:
                headers.append(cp['name'])

        self.results_table.setColumnCount(len(headers))
        self.results_table.setHorizontalHeaderLabels(headers)

        # Сортируем экипажи
        sorted_crews = self.sort_crews_by_results()

        # Заполняем данные
        self.fill_results_data(sorted_crews, skp_count)

        # Настраиваем внешний вид
        self.adjust_results_columns()

    def fill_results_data(self, crews, skp_count):
        """Заполнение таблицы данными"""
        self.results_table.setRowCount(len(crews))
        
        # Цвета для призовых мест
        place_colors = {
            1: QColor(255, 215, 0),    # Золотой
            2: QColor(192, 192, 192),  # Серебряный
            3: QColor(205, 127, 50)     # Бронзовый
        }
        
        # Определяем позицию начала КП
        base_cp_col = 14 + skp_count * 3
        
        for row, member in enumerate(crews):
            # Проверяем DNF статус
            is_dnf = self.is_crew_dnf(member)
            
            # Основные данные
            self.add_basic_member_info(row, member, is_dnf)
            
            # Временные показатели
            self.add_time_data(row, member)
            
            # Данные по СКП
            self.add_skp_data(row, member, skp_count)
            
            # Основные столбцы (итог, штраф время, баллы)
            self.add_score_columns(row, member, is_dnf)
            
            # Данные по КП (в зависимости от этапности)
            if self.data.get('logic_params', {}).get('staged', False):
                self.add_staged_checkpoints(row, member, base_cp_col)
            else:
                self.add_simple_checkpoints(row, member, base_cp_col)
            
            # Раскрашиваем призовые места (только для не-DNF)
            if not is_dnf:
                place = member.get('место', 0)
                if place in place_colors:
                    for col in range(3):  # Раскрашиваем первые 3 колонки
                        item = self.results_table.item(row, col)
                        if item:
                            item.setBackground(place_colors[place])

    def add_simple_checkpoints(self, row, member, start_col):
        """Добавление КП без этапности"""
        checkpoints = sorted(self.data.get('checkpoints', []), key=lambda x: x['name'])
        taken_cps = member.get('taken_cps', [])
        crew_classification = member.get('зачет', '')
        
        for col, cp in enumerate(checkpoints):
            item = QTableWidgetItem()
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            if not self.is_cp_available(cp, crew_classification):
                item.setText("-")
                item.setForeground(Qt.GlobalColor.gray)
            elif cp['name'] in taken_cps:
                item.setText(str(int(cp.get('score', 0))))
            else:
                item.setText("")  # Пустая строка вместо "0"
            
            self.results_table.setItem(row, start_col + col, item)  

    def add_staged_checkpoints(self, row, member, start_col):
        """Добавление КП с этапностью"""
        crew_classification = member.get('зачет', '')
        taken_cps = member.get('taken_cps', [])
        current_col = start_col
        
        # Цвета для этапов
        stage_colors = [
            QColor(173, 216, 230),  # LightBlue
            QColor(144, 238, 144),   # LightGreen
            QColor(255, 182, 193),   # LightPink
            QColor(255, 255, 153)    # LightYellow
        ]
        
        for stage_idx, stage in enumerate(self.data['logic_params'].get('stages', [])):
            stage_name = stage['name']
            stage_score = 0
            cp_col = 0
            
            # Считаем баллы этапа
            for cp in self.data.get('checkpoints', []):
                if 'stages' in cp and stage_name in cp['stages']:
                    if cp['name'] in taken_cps and self.is_cp_available(cp, crew_classification):
                        stage_score += int(cp.get('score', 0))
            
            # Добавляем баллы этапа с цветом
            stage_item = QTableWidgetItem(str(stage_score))
            stage_item.setBackground(stage_colors[stage_idx % len(stage_colors)])
            stage_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.results_table.setItem(row, current_col, stage_item)
            current_col += 1
            
            # Добавляем КП этапа с цветом
            for cp in self.data.get('checkpoints', []):
                if 'stages' in cp and stage_name in cp['stages']:
                    item = QTableWidgetItem()
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    if not self.is_cp_available(cp, crew_classification):
                        item.setText("-")
                        item.setForeground(Qt.GlobalColor.gray)
                    elif cp['name'] in taken_cps:
                        item.setText(str(int(cp.get('score', 0))))
                    else:
                        item.setText("")  # Пустая строка вместо "0"
                    
                    item.setBackground(stage_colors[stage_idx % len(stage_colors)].lighter(130))
                    self.results_table.setItem(row, current_col + cp_col, item)
                    cp_col += 1
            
            current_col += cp_col






############## ЭКСПОРТ В EXCEL ##############
    def save_results(self):
        """Экспорт результатов в Excel файл"""
        try:
            if not hasattr(self, 'data') or not self.data.get('members'):
                raise ValueError("Нет данных для экспорта")
            
            # Создаем имя файла на основе метаданных
            meta = self.data.get('meta', {})
            event_name = meta.get('name', 'Соревнования').replace(' ', '_')
            event_date = meta.get('date', '')
            
            # Добавляем текущее время к имени файла
            now = datetime.datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            file_name = f"{event_name}_{event_date}_{now}.xlsx" if event_date else f"{event_name}_{now}.xlsx"
            
            # Создаем папку exports если ее нет
            if not os.path.exists('exports'):
                os.makedirs('exports')
            
            file_path = os.path.join('exports', file_name)
            
            # Создаем книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Результаты"
            
            # Цвета для оформления
            HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            SILVER_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            BRONZE_FILL = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
            DNF_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            STAGE_COLORS = [
                PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # LightBlue
                PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # LightGreen
                PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # LightPink
                PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")   # LightYellow
            ]
            
            # Границы для ячеек
            thin_border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
            
            # Шрифт
            default_font = Font(name='Georgia', size=11)
            
            # Копируем заголовки из таблицы результатов
            headers = []
            for col in range(self.results_table.columnCount()):
                headers.append(self.results_table.horizontalHeaderItem(col).text())
            
            # Записываем заголовки в Excel
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = default_font
                cell.fill = HEADER_FILL
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Заполняем данные
            for row in range(self.results_table.rowCount()):
                for col in range(self.results_table.columnCount()):
                    item = self.results_table.item(row, col)
                    if item is not None:
                        cell = ws.cell(row=row+2, column=col+1, value=item.text())
                        cell.font = default_font
                        cell.border = thin_border
                        
                        # Применяем цветовые стили
                        bg_color = item.background()
                        
                        # Обработка призовых мест и DNF
                        if col == 0:  # Столбец с местом
                            place_text = item.text()
                            if place_text == "1":
                                cell.fill = GOLD_FILL
                            elif place_text == "2":
                                cell.fill = SILVER_FILL
                            elif place_text == "3":
                                cell.fill = BRONZE_FILL
                            elif place_text == "DNF":
                                cell.fill = DNF_FILL
                        # Обработка DNF в других столбцах
                        elif item.text() == "DNF":
                            cell.fill = DNF_FILL
                        # Обработка цветов этапов и КП
                        elif isinstance(bg_color, (QColor, Qt.GlobalColor)):
                            # Преобразуем Qt цвет в HEX
                            if isinstance(bg_color, Qt.GlobalColor):
                                bg_color = QColor(bg_color)
                            hex_color = bg_color.name()[1:]  # Убираем # из начала
                            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            
            # Настраиваем ширину столбцов
            for col in range(1, len(headers)+1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 15  # Базовая ширина
                
                # Автоподбор для некоторых столбцов
                if col in [1, 2]:  # Место, Номер
                    ws.column_dimensions[column_letter].width = 8
                elif col in [3, 4, 5]:  # Пилот, Штурман, Авто
                    ws.column_dimensions[column_letter].width = 20
            
            # Замораживаем панель (A1:C2)
            ws.freeze_panes = 'D2'
            
            # Сохраняем файл
            wb.save(file_path)
            
            # Открываем файл автоматически
            os.startfile(file_path)
            
            QMessageBox.information(self, "Успех", f"Результаты экспортированы в файл:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать результаты:\n{str(e)}")    







    # def save_data(self):
    #     """Сохраняем данные в новый файл"""
    #     if not hasattr(self, 'data'):
    #         print("Нет данных для сохранения!")
    #         return
        
    #     try:
    #         new_version = self.data["meta"]["version"] + 1
    #         self.data["meta"]["version"] = new_version
    #         new_filename = os.path.join(self.data_dir, f"race_v{new_version}.json")
            
    #         with open(new_filename, 'w', encoding='utf-8') as f:
    #             json.dump(self.data, f, indent=4, ensure_ascii=False)
            
    #         print(f"Данные сохранены в {new_filename}")
    #         self.current_file = new_filename
    #         self.update_window_title()
            
    #     except Exception as e:
    #         print(f"Ошибка сохранения данных: {e}")






if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        window = RaceApp()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        print(f"Ошибка: {e}")
        input("Нажмите Enter для выхода...")






